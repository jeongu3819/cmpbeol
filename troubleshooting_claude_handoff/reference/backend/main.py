from fastapi import (
    FastAPI,
    HTTPException,
    Depends,
    Request,
    Query,
    Body,
    UploadFile,
    File as FastAPIFile,
    BackgroundTasks,
)
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy.sql import func
from sqlalchemy import false as sql_false
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, date, timedelta
import os
import json
import uuid
import logging
import threading
import httpx
import re
from zoneinfo import ZoneInfo

# =========================
# DB / ENV / AUTH
# =========================
from app.db_connections.sqlalchemy import SessionLocal, engine, Base
from app.models import (
    Project, User, Task, UserPreference, VisitLog, GroupMembership, Group,
    ProjectAiReport, ProjectAiQuery, AiSetting, SubProject as SubProjectModel, Note as NoteModel,
    NoteMention, TaskActivityMention, ProjectMember as ProjectMemberModel, UserShortcut,
    MemberGroup, MemberGroupUser,
    TaskActivity as TaskActivityModel, Attachment as AttachmentModel,
    Space, SpaceMember, SpaceJoinRequest,
    SheetTemplate, SheetExecution, SheetExecutionItem, SheetExecutionLog,
    VocItem, VocAttachment, VocVote, VocComment,
    SystemEventLog,
    CalendarEvent,
    ActivityLog,
    ProjectTaskColumn, ProjectWorkflowSnapshot, TaskStageCompletion,
    TaskComment, TaskCommentMention, TaskCommentAttachment,
    NotificationEvent, UserNotificationPreference,
)
from app.services.notifications import create_notification_event
from app.services import activity_log as actlog
from app.realtime import (
    emit_realtime_event,
    event_stream as realtime_event_stream,
    issue_ticket as realtime_issue_ticket,
    validate_ticket as realtime_validate_ticket,
    client_count as realtime_client_count,
    get_stats as realtime_get_stats,
    start_cleanup_scheduler as realtime_start_cleanup,
    stop_cleanup_scheduler as realtime_stop_cleanup,
    GLOBAL_CHANNEL as REALTIME_GLOBAL_CHANNEL,
)
from app.environment import (
    CORS_ORIGINS, SUPER_ADMIN_LOGINIDS, KST,
    MAX_ACTIVE_SPACES_PER_USER, MAX_ACTIVE_SPACES_PER_ADMIN, MAX_SPACE_CREATE_PER_DAY,
    SPACE_EMPTY_WARNING_DAYS, SPACE_EMPTY_AUTO_ARCHIVE_DAYS, SPACE_EMPTY_DELETE_AFTER_DAYS,
    SPACE_EMPTY_AUTO_DELETE_ENABLED,
    REALTIME_SYNC_ENABLED, REALTIME_MAX_CLIENTS_PER_WORKER,
)
from app.llm.dsllm_adapter import chat as dsllm_chat
from app.llm.dsllm_adapter import chat_stream as dsllm_chat_stream
from app.llm.dsllm_adapter import list_model_keys
from app.llm.dsllm_adapter import get_available_models as dsllm_get_available_models
from app.llm.dsllm_adapter import default_model_key as dsllm_default_model_key
from app.llm.dsllm_adapter import is_known_model as dsllm_is_known_model
from app.llm import router as llm_router
from app.utils.text import sanitize_llm_text, sanitize_llm_text_ai, normalize_task_blocks
from app.routers import auth, knox

# ✅ 새 테이블 자동 생성 (member_groups, member_group_users 등)
Base.metadata.create_all(bind=engine)

# ✅ 기존 테이블에 새 컬럼 추가 (ALTER TABLE)
def _run_migrations():
    from sqlalchemy import text, inspect
    insp = inspect(engine)
    # task_activities.block_type
    if "task_activities" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("task_activities")]
        if "block_type" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE task_activities ADD COLUMN block_type VARCHAR(20) NOT NULL DEFAULT "checkbox"'))
        if "checked_at" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE task_activities ADD COLUMN checked_at DATETIME'))
        # 단계 완료 체크포인트 메타 (기존 row 는 일반 체크박스 = is_stage_checkpoint 0)
        if "is_stage_checkpoint" not in cols:
            with engine.begin() as conn:
                conn.execute(text(
                    'ALTER TABLE task_activities ADD COLUMN is_stage_checkpoint TINYINT(1) NOT NULL DEFAULT 0'
                ))
        if "checkpoint_stage_id" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE task_activities ADD COLUMN checkpoint_stage_id INT'))
        if "checkpoint_required" not in cols:
            with engine.begin() as conn:
                conn.execute(text(
                    'ALTER TABLE task_activities ADD COLUMN checkpoint_required TINYINT(1) NOT NULL DEFAULT 0'
                ))

    # tasks.task_type (단발 일정 구분: normal / one_off)
    if "tasks" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("tasks")]
        if "task_type" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE tasks ADD COLUMN task_type VARCHAR(20) NOT NULL DEFAULT 'normal'"))
        # 일정 미정(TBD) 상태 — 기존 row 는 모두 FALSE(=미정 아님).
        # 기존 NULL 날짜는 "단순 누락"으로 간주하고 자동 변환하지 않는다(요구사항 §3 주의).
        if "start_date_tbd" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE tasks ADD COLUMN start_date_tbd TINYINT(1) NOT NULL DEFAULT 0"))
        if "due_date_tbd" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE tasks ADD COLUMN due_date_tbd TINYINT(1) NOT NULL DEFAULT 0"))

    # ── Space system tables + migration ──
    if "spaces" not in insp.get_table_names():
        Base.metadata.create_all(bind=engine, tables=[
            Base.metadata.tables.get("spaces"),
            Base.metadata.tables.get("space_members"),
        ])
    if "space_join_requests" not in insp.get_table_names():
        t = Base.metadata.tables.get("space_join_requests")
        if t is not None:
            Base.metadata.create_all(bind=engine, tables=[t])
    if "task_activity_mentions" not in insp.get_table_names():
        t = Base.metadata.tables.get("task_activity_mentions")
        if t is not None:
            Base.metadata.create_all(bind=engine, tables=[t])
    # calendar_events (Phase 2 — 공간 단위 단순 일정)
    if "calendar_events" not in insp.get_table_names():
        t = Base.metadata.tables.get("calendar_events")
        if t is not None:
            Base.metadata.create_all(bind=engine, tables=[t])
    # realtime_events (Realtime sync event outbox — SSE 전파용 경량 이벤트 로그)
    if "realtime_events" not in insp.get_table_names():
        t = Base.metadata.tables.get("realtime_events")
        if t is not None:
            Base.metadata.create_all(bind=engine, tables=[t])
    # Task Comment / Notification 기반 테이블 (댓글 + 알림 outbox + 사용자 알림설정)
    for _nt in (
        "task_comments", "task_comment_mentions", "task_comment_attachments",
        "notification_events", "user_notification_preferences",
    ):
        if _nt not in insp.get_table_names():
            t = Base.metadata.tables.get(_nt)
            if t is not None:
                Base.metadata.create_all(bind=engine, tables=[t])
    # user_notification_preferences: 멘션 메일 on/off 컬럼 (기존 row 는 DEFAULT 1 = 수신)
    if "user_notification_preferences" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("user_notification_preferences")]
        if "mention_email_enabled" not in cols:
            with engine.begin() as conn:
                conn.execute(text(
                    'ALTER TABLE user_notification_preferences '
                    'ADD COLUMN mention_email_enabled TINYINT(1) NOT NULL DEFAULT 1'
                ))
        if "work_note_mention_email_enabled" not in cols:
            with engine.begin() as conn:
                conn.execute(text(
                    'ALTER TABLE user_notification_preferences '
                    'ADD COLUMN work_note_mention_email_enabled TINYINT(1) NOT NULL DEFAULT 1'
                ))
    # notification_events.error_message (skipped/failed 사유 기록)
    if "notification_events" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("notification_events")]
        if "error_message" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE notification_events ADD COLUMN error_message TEXT'))
    if "projects" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("projects")]
        if "space_id" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE projects ADD COLUMN space_id INTEGER'))
        # 삭제 주체 추적 (Trash Page "내가 삭제한 항목" 필터용)
        if "deleted_by_user_id" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE projects ADD COLUMN deleted_by_user_id INTEGER'))

    # tasks.deleted_by_user_id (Trash Page "내가 삭제한 항목" 필터용)
    if "tasks" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("tasks")]
        if "deleted_by_user_id" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE tasks ADD COLUMN deleted_by_user_id INTEGER'))

    # ai_settings.provider (AI Provider 선택: dsllm / openai. NULL=env AI_PROVIDER fallback)
    if "ai_settings" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("ai_settings")]
        if "provider" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE ai_settings ADD COLUMN provider VARCHAR(20)"))

    # vision_ai_settings — Fallback Text Model + soft timeout 컬럼(기존 row 는 NULL=env fallback).
    if "vision_ai_settings" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("vision_ai_settings")]
        for _col, _ddl in (
            ("fallback_text_provider", "ALTER TABLE vision_ai_settings ADD COLUMN fallback_text_provider VARCHAR(20)"),
            ("fallback_text_model", "ALTER TABLE vision_ai_settings ADD COLUMN fallback_text_model VARCHAR(255)"),
            ("fallback_text_enabled", "ALTER TABLE vision_ai_settings ADD COLUMN fallback_text_enabled TINYINT(1)"),
            ("fallback_policy", "ALTER TABLE vision_ai_settings ADD COLUMN fallback_policy VARCHAR(40)"),
            ("soft_timeout_seconds", "ALTER TABLE vision_ai_settings ADD COLUMN soft_timeout_seconds INT"),
        ):
            if _col not in cols:
                with engine.begin() as conn:
                    conn.execute(text(_ddl))

    # ── 프로젝트별 워크플로우 컬럼 커스텀 (Slice 1) ──
    # 신규 테이블: project_task_columns / project_workflow_snapshots
    for _wf_table in ("project_task_columns", "project_workflow_snapshots", "task_stage_completions"):
        if _wf_table not in insp.get_table_names():
            t = Base.metadata.tables.get(_wf_table)
            if t is not None:
                Base.metadata.create_all(bind=engine, tables=[t])
    # project_task_columns.description (단계 세부 설명 — 기존 row는 NULL=설명 없음)
    if "project_task_columns" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("project_task_columns")]
        if "description" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE project_task_columns ADD COLUMN description TEXT'))
        # 중간 완료 체크포인트 컬럼 (기존 row는 is_checkpoint=0 / label·desc=NULL)
        if "is_checkpoint" not in cols:
            with engine.begin() as conn:
                conn.execute(text(
                    'ALTER TABLE project_task_columns '
                    'ADD COLUMN is_checkpoint TINYINT(1) NOT NULL DEFAULT 0'
                ))
        if "checkpoint_label" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE project_task_columns ADD COLUMN checkpoint_label VARCHAR(120)'))
        if "checkpoint_description" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE project_task_columns ADD COLUMN checkpoint_description TEXT'))
        # 진행률 기준 자동 이동 (CUSTOM 전용). 기존 row는 auto_advance_enabled=0 / threshold=NULL(OFF)
        if "auto_advance_enabled" not in cols:
            with engine.begin() as conn:
                conn.execute(text(
                    'ALTER TABLE project_task_columns '
                    'ADD COLUMN auto_advance_enabled TINYINT(1) NOT NULL DEFAULT 0'
                ))
        if "auto_advance_threshold" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE project_task_columns ADD COLUMN auto_advance_threshold INT'))
        # 단계 완료 조건(stage completion rule). 신규 기본=all_required_checkpoints(체크포인트 없으면 no-op),
        # 기존 progress 자동이동(auto_advance_enabled=1)은 레거시 progress_threshold 로 보존.
        if "completion_rule_type" not in cols:
            with engine.begin() as conn:
                conn.execute(text(
                    "ALTER TABLE project_task_columns "
                    "ADD COLUMN completion_rule_type VARCHAR(40) NOT NULL DEFAULT 'all_required_checkpoints'"
                ))
                conn.execute(text(
                    "UPDATE project_task_columns SET completion_rule_type='progress_threshold' "
                    "WHERE auto_advance_enabled=1"
                ))
        # 4차: 단계 완료 선언 시 자동 이동 여부. 기존 데이터는 completion_rule_type 로부터 유도
        # (manual/progress_threshold → 자동이동 OFF, 그 외 → ON).
        if "auto_move_on_complete" not in cols:
            with engine.begin() as conn:
                conn.execute(text(
                    "ALTER TABLE project_task_columns "
                    "ADD COLUMN auto_move_on_complete TINYINT(1) NOT NULL DEFAULT 1"
                ))
                conn.execute(text(
                    "UPDATE project_task_columns SET auto_move_on_complete=0 "
                    "WHERE completion_rule_type IN ('manual','progress_threshold')"
                ))
    # projects.workflow_mode / auto_progress_from_notes (기존 프로젝트는 DEFAULT/true fallback)
    if "projects" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("projects")]
        if "workflow_mode" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE projects ADD COLUMN workflow_mode VARCHAR(20) NOT NULL DEFAULT 'DEFAULT'"))
        if "auto_progress_from_notes" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE projects ADD COLUMN auto_progress_from_notes BOOLEAN NOT NULL DEFAULT 1"))
    # tasks.workflow_column_id (CUSTOM Board 표시용, NULL=status fallback) + 인덱스
    if "tasks" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("tasks")]
        if "workflow_column_id" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE tasks ADD COLUMN workflow_column_id INTEGER'))
        # tasks.parent_task_id (Task 상하 관계 — NULL=최상위) + 인덱스
        if "parent_task_id" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE tasks ADD COLUMN parent_task_id INTEGER'))
        existing_idx = {i["name"] for i in insp.get_indexes("tasks")}
        if "ix_tasks_project_workflow_col" not in existing_idx:
            with engine.begin() as conn:
                conn.execute(text("CREATE INDEX ix_tasks_project_workflow_col ON tasks (project_id, workflow_column_id)"))
        if "ix_tasks_parent_task_id" not in existing_idx:
            with engine.begin() as conn:
                conn.execute(text("CREATE INDEX ix_tasks_parent_task_id ON tasks (parent_task_id)"))

    # (General/기본공간 자동 생성 제거 — 사용자가 만든 공간만 존재)
    # 기존 General 공간이 있으면 비활성화하고 소속 프로젝트의 space_id를 NULL로 복원
    if "spaces" in insp.get_table_names():
        with engine.begin() as conn:
            row = conn.execute(text("SELECT id FROM spaces WHERE slug = 'general' AND is_active = 1 LIMIT 1")).fetchone()
            if row:
                general_id = row[0]
                conn.execute(text(f"UPDATE projects SET space_id = NULL WHERE space_id = {general_id}"))
                conn.execute(text(f"UPDATE spaces SET is_active = 0 WHERE id = {general_id}"))
                conn.execute(text(f"DELETE FROM space_members WHERE space_id = {general_id}"))

    # warned_at 컬럼 추가 (빈 공간 경고/삭제용)
    if "spaces" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("spaces")]
        if "warned_at" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE spaces ADD COLUMN warned_at DATETIME'))
        if "purpose" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE spaces ADD COLUMN purpose VARCHAR(50) DEFAULT 'project_management'"))
        # 빈 공간 안전 관리(보관 lifecycle)용 컬럼
        if "archived_at" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE spaces ADD COLUMN archived_at DATETIME"))
        if "last_activity_at" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE spaces ADD COLUMN last_activity_at DATETIME"))
        if "delete_scheduled_at" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE spaces ADD COLUMN delete_scheduled_at DATETIME"))
        if "cleanup_exempt" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE spaces ADD COLUMN cleanup_exempt BOOLEAN DEFAULT 0"))

    # user_shortcuts visibility 컬럼 추가 (개인 바로가기 공유 모델)
    if "user_shortcuts" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("user_shortcuts")]
        if "visibility" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE user_shortcuts ADD COLUMN visibility VARCHAR(20) NOT NULL DEFAULT 'PRIVATE'"))
        if "shared_user_ids" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE user_shortcuts ADD COLUMN shared_user_ids TEXT"))
        if "shared_group_ids" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE user_shortcuts ADD COLUMN shared_group_ids TEXT"))

    # ── v3.0 Sheet 운영 테이블 자동 생성 ──
    for tname in ["sheet_templates", "sheet_executions", "sheet_execution_items", "sheet_execution_logs"]:
        if tname not in insp.get_table_names():
            t = Base.metadata.tables.get(tname)
            if t is not None:
                Base.metadata.create_all(bind=engine, tables=[t])

    # sheet_executions.task_id (Task Details 에서 Check Sheet 연결)
    if "sheet_executions" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("sheet_executions")]
        if "task_id" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE sheet_executions ADD COLUMN task_id INTEGER'))

    # v3.1: sheet_templates.column_role_mapping (자동 인식 결과 저장)
    if "sheet_templates" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("sheet_templates")]
        if "column_role_mapping" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE sheet_templates ADD COLUMN column_role_mapping JSON'))
        if "structure_hash" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE sheet_templates ADD COLUMN structure_hash VARCHAR(32)'))

    # v3.2: sheet_type 및 sheet_execution_mappings 테이블
    if "sheet_templates" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("sheet_templates")]
        if "sheet_type" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE sheet_templates ADD COLUMN sheet_type VARCHAR(50) NOT NULL DEFAULT "inspection"'))

    if "sheet_executions" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("sheet_executions")]
        if "sheet_type" not in cols:
            with engine.begin() as conn:
                conn.execute(text('ALTER TABLE sheet_executions ADD COLUMN sheet_type VARCHAR(50) NOT NULL DEFAULT "inspection"'))
        # 실행본에서 숨긴 컬럼 인덱스 배열
        if "hidden_cols" not in cols:
            with engine.begin() as conn:
                if engine.dialect.name == "mysql":
                    conn.execute(text("ALTER TABLE sheet_executions ADD COLUMN hidden_cols JSON NULL"))
                else:
                    conn.execute(text("ALTER TABLE sheet_executions ADD COLUMN hidden_cols TEXT"))
        # 실행본에서 숨긴 행 인덱스 배열
        if "hidden_rows" not in cols:
            with engine.begin() as conn:
                if engine.dialect.name == "mysql":
                    conn.execute(text("ALTER TABLE sheet_executions ADD COLUMN hidden_rows JSON NULL"))
                else:
                    conn.execute(text("ALTER TABLE sheet_executions ADD COLUMN hidden_rows TEXT"))
        # 실행본 단위 구조 오버레이 (added_columns / added_rows / renamed_headers)
        if "structure_overlay" not in cols:
            with engine.begin() as conn:
                if engine.dialect.name == "mysql":
                    conn.execute(text("ALTER TABLE sheet_executions ADD COLUMN structure_overlay JSON NULL"))
                else:
                    conn.execute(text("ALTER TABLE sheet_executions ADD COLUMN structure_overlay TEXT"))

    if "sheet_execution_mappings" not in insp.get_table_names():
        from app.models import SheetExecutionMapping
        t = Base.metadata.tables.get("sheet_execution_mappings")
        if t is not None:
            Base.metadata.create_all(bind=engine, tables=[t])

    # ── 설비-배치 매핑형 보정 (legacy backfill) ──
    # 과거 업로드 시 유형 미선택으로 sheet_type 이 'inspection' 으로 남았지만, 구조에 설비/장비/
    # EQP/Equipment N 컬럼이 2개 이상 있는 시트는 사실상 설비-배치 매핑형이다. 이를 'assignment_mapping'
    # 으로 보정해, 진행률/완료율 제외·연결 약품 매칭 등 모든 화면이 sheet_type 기준으로 일관 동작하게 한다.
    if "sheet_templates" in insp.get_table_names():
        try:
            import re as _re
            _eq_col_re = _re.compile(r'^(설비|장비|eqp|equipment)\d+$', _re.IGNORECASE)

            def _has_equipment_mapping_cols(structure) -> bool:
                if not isinstance(structure, dict):
                    return False
                seen: set[str] = set()
                candidates = []
                for h in (structure.get("headers") or []):
                    if isinstance(h, dict):
                        candidates.append(h.get("value"))
                for c in (structure.get("cells") or []):
                    if isinstance(c, dict):
                        candidates.append(c.get("value"))
                for v in candidates:
                    if v is None:
                        continue
                    key = str(v).replace(" ", "").strip()
                    if not key or key in seen:
                        continue
                    if _eq_col_re.match(key):
                        seen.add(key)
                        if len(seen) >= 2:
                            return True
                return False

            with SessionLocal() as _mig_db:
                _legacy = _mig_db.query(SheetTemplate).filter(
                    SheetTemplate.sheet_type == "inspection"
                ).all()
                _fixed_ids = [
                    _t.id for _t in _legacy if _has_equipment_mapping_cols(_t.structure)
                ]
                for _t in _legacy:
                    if _t.id in _fixed_ids:
                        _t.sheet_type = "assignment_mapping"
                if _fixed_ids:
                    _mig_db.query(SheetExecution).filter(
                        SheetExecution.template_id.in_(_fixed_ids),
                        SheetExecution.sheet_type == "inspection",
                    ).update(
                        {SheetExecution.sheet_type: "assignment_mapping"},
                        synchronize_session=False,
                    )
                _mig_db.commit()
        except Exception:
            # 보정 실패는 치명적이지 않음 — 프론트가 구조 기반 폴백으로 동일하게 처리한다.
            pass

    # ── Notification System: notification_logs 테이블 자동 생성 ──
    # 외부 알림(Knox Messenger 등) 발송 이력 + 중복 방지 키 저장용.
    if "notification_logs" not in insp.get_table_names():
        from app.models import NotificationLog  # noqa: F401  (Base 등록 보장)
        t = Base.metadata.tables.get("notification_logs")
        if t is not None:
            Base.metadata.create_all(bind=engine, tables=[t])

    # ── VOC / 개선 요청: voc_items 테이블 자동 생성 ──
    if "voc_items" not in insp.get_table_names():
        t = Base.metadata.tables.get("voc_items")
        if t is not None:
            Base.metadata.create_all(bind=engine, tables=[t])
    else:
        # 기존 환경에 edit 메타 컬럼 추가 (최소 수정 이력)
        cols = [c["name"] for c in insp.get_columns("voc_items")]
        if "edited_by" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE voc_items ADD COLUMN edited_by INTEGER"))
        if "edited_at" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE voc_items ADD COLUMN edited_at DATETIME"))
        if "change_summary" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE voc_items ADD COLUMN change_summary VARCHAR(255)"))
        # 공개/비공개(visibility) 컬럼. 기존 VOC 는 안전을 위해 비공개로 유지하고,
        # 신규 작성분만 애플리케이션 기본값(public)으로 공개된다.
        if "visibility" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE voc_items ADD COLUMN visibility VARCHAR(20)"))
                conn.execute(text("UPDATE voc_items SET visibility='private' WHERE visibility IS NULL"))

    # ── VOC 첨부 이미지 테이블 ──
    if "voc_attachments" not in insp.get_table_names():
        t = Base.metadata.tables.get("voc_attachments")
        if t is not None:
            Base.metadata.create_all(bind=engine, tables=[t])
    else:
        # 댓글 이미지 구분 컬럼(comment_scoped) 추가. 기존 첨부는 모두 본문(0)으로 둔다.
        att_cols = [c["name"] for c in insp.get_columns("voc_attachments")]
        if "comment_scoped" not in att_cols:
            with engine.begin() as conn:
                conn.execute(text(
                    "ALTER TABLE voc_attachments ADD COLUMN comment_scoped BOOLEAN NOT NULL DEFAULT 0"
                ))
                conn.execute(text(
                    "UPDATE voc_attachments SET comment_scoped=0 WHERE comment_scoped IS NULL"
                ))

    # ── VOC 공감('나도 필요해요') 테이블 ──
    if "voc_votes" not in insp.get_table_names():
        t = Base.metadata.tables.get("voc_votes")
        if t is not None:
            Base.metadata.create_all(bind=engine, tables=[t])

    # ── VOC 댓글 / 관리자 답변 테이블 ──
    if "voc_comments" not in insp.get_table_names():
        t = Base.metadata.tables.get("voc_comments")
        if t is not None:
            Base.metadata.create_all(bind=engine, tables=[t])

    # v3.7: sheet_executions.task_id FK → ON DELETE SET NULL
    # Task 영구 삭제 시 SheetExecution 이력은 보존하고 task_id 만 NULL 처리되도록 한다.
    # MySQL 에서만 적용 (SQLite 는 FK 의 ON DELETE 옵션 변경이 ALTER 로 불가).
    if engine.dialect.name == "mysql" and "sheet_executions" in insp.get_table_names():
        try:
            with engine.begin() as conn:
                row = conn.execute(text("""
                    SELECT CONSTRAINT_NAME, DELETE_RULE
                    FROM information_schema.REFERENTIAL_CONSTRAINTS
                    WHERE TABLE_NAME = 'sheet_executions'
                      AND CONSTRAINT_SCHEMA = DATABASE()
                      AND REFERENCED_TABLE_NAME = 'tasks'
                """)).fetchone()
                current_name = row[0] if row else None
                current_rule = (row[1] or "").upper() if row else ""

                # task_id 컬럼은 NULL 허용으로 보장
                conn.execute(text("ALTER TABLE sheet_executions MODIFY COLUMN task_id INT NULL"))

                if current_name and current_rule != "SET NULL":
                    conn.execute(text(f"ALTER TABLE sheet_executions DROP FOREIGN KEY `{current_name}`"))
                    conn.execute(text(
                        "ALTER TABLE sheet_executions "
                        "ADD CONSTRAINT fk_sheet_executions_task "
                        "FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE SET NULL"
                    ))
                elif not current_name:
                    conn.execute(text(
                        "ALTER TABLE sheet_executions "
                        "ADD CONSTRAINT fk_sheet_executions_task "
                        "FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE SET NULL"
                    ))
        except Exception as e:
            import logging
            logging.getLogger("main").warning(
                f"sheet_executions.task_id FK ON DELETE SET NULL 마이그레이션 실패 (무시): {e}"
            )

    # v3.8: sheet_executions.project_id FK → ON DELETE SET NULL
    # Project 영구 삭제 시 SheetExecution 이력은 보존하고 project_id 만 NULL 처리되도록 한다.
    # 휴지통 만료 정리에서 Project 삭제가 FK 위반으로 실패하던 문제를 근본 해결.
    if engine.dialect.name == "mysql" and "sheet_executions" in insp.get_table_names():
        try:
            with engine.begin() as conn:
                row = conn.execute(text("""
                    SELECT CONSTRAINT_NAME, DELETE_RULE
                    FROM information_schema.REFERENTIAL_CONSTRAINTS
                    WHERE TABLE_NAME = 'sheet_executions'
                      AND CONSTRAINT_SCHEMA = DATABASE()
                      AND REFERENCED_TABLE_NAME = 'projects'
                """)).fetchone()
                current_name = row[0] if row else None
                current_rule = (row[1] or "").upper() if row else ""

                # project_id 컬럼은 NULL 허용으로 보장
                conn.execute(text("ALTER TABLE sheet_executions MODIFY COLUMN project_id INT NULL"))

                if current_name and current_rule != "SET NULL":
                    conn.execute(text(f"ALTER TABLE sheet_executions DROP FOREIGN KEY `{current_name}`"))
                    conn.execute(text(
                        "ALTER TABLE sheet_executions "
                        "ADD CONSTRAINT fk_sheet_executions_project "
                        "FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE SET NULL"
                    ))
                elif not current_name:
                    conn.execute(text(
                        "ALTER TABLE sheet_executions "
                        "ADD CONSTRAINT fk_sheet_executions_project "
                        "FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE SET NULL"
                    ))
        except Exception as e:
            import logging
            logging.getLogger("main").warning(
                f"sheet_executions.project_id FK ON DELETE SET NULL 마이그레이션 실패 (무시): {e}"
            )

    # ── 방문 통계 조회 성능용 visit_log 인덱스 (있으면 건너뜀) ──
    if "visit_log" in insp.get_table_names():
        try:
            existing_idx = {ix["name"] for ix in insp.get_indexes("visit_log")}
            wanted = {
                "idx_visit_log_timestamp": "visit_log(timestamp)",
                "idx_visit_log_dept_visit_date": "visit_log(deptname, visit_date)",
                "idx_visit_log_user_visit_date": "visit_log(user_id, visit_date)",
            }
            with engine.begin() as conn:
                for name, cols in wanted.items():
                    if name not in existing_idx:
                        conn.execute(text(f"CREATE INDEX {name} ON {cols}"))
        except Exception as e:
            import logging
            logging.getLogger("main").warning(
                f"visit_log 인덱스 생성 실패 (무시): {e}"
            )

    # ── 사용 통계(프로젝트/Task 생성 추이) 조회 성능용 created_at 인덱스 (있으면 건너뜀) ──
    for _tbl, _idx in (("projects", "idx_projects_created_at"), ("tasks", "idx_tasks_created_at")):
        if _tbl in insp.get_table_names():
            try:
                existing_idx = {ix["name"] for ix in insp.get_indexes(_tbl)}
                if _idx not in existing_idx:
                    with engine.begin() as conn:
                        conn.execute(text(f"CREATE INDEX {_idx} ON {_tbl}(created_at)"))
            except Exception as e:
                import logging
                logging.getLogger("main").warning(
                    f"{_tbl}.created_at 인덱스 생성 실패 (무시): {e}"
                )

    # ── system_event_logs 처리(resolve) 메타 컬럼 추가 ──
    if "system_event_logs" in insp.get_table_names():
        cols = [c["name"] for c in insp.get_columns("system_event_logs")]
        if "resolved_at" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE system_event_logs ADD COLUMN resolved_at DATETIME"))
        if "resolved_by" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE system_event_logs ADD COLUMN resolved_by INTEGER"))
        if "resolution_note" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE system_event_logs ADD COLUMN resolution_note TEXT"))
        if "detail_json" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE system_event_logs ADD COLUMN detail_json TEXT"))

_run_migrations()

# ═══════════════════════════════════════════════════════════
# 배포 버전 식별자.
#   회사 서버 / 로컬이 같은 코드를 실행 중인지 /api/health 와 startup 로그로
#   즉시 확인하기 위한 값. 코드 반영 후 이 값을 갱신하면 배포 여부를 검증할 수 있다.
# ═══════════════════════════════════════════════════════════
APP_VERSION = "2026-07-03-comment-mentions+delete-hide+comment-images"

app = FastAPI(title="Antigravity Schedule Platform API")

# SSO 라우터 (/api/auth/*)
app.include_router(auth.router)
app.include_router(knox.router)

# CORS
# - allow_origin_regex: dev 환경에서 LAN IP / 임의 포트의 프론트가 붙어도 CORS 헤더가
#   응답에 부착되도록 정규식으로 허용. (운영 origin 은 CORS_ORIGINS env 로 명시 관리.)
#   500 응답에도 CORSMiddleware 가 헤더를 추가하므로 브라우저의 가짜 CORS 에러 메시지를 줄일 수 있다.
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS or ["*"],
    allow_origin_regex=r"http://(localhost|127\.0\.0\.1|10(\.\d+){3}|192\.168(\.\d+){2}|172\.(1[6-9]|2\d|3[01])(\.\d+){2})(:\d+)?",
    allow_credentials=False,  # Authorization Bearer 방식이면 보통 False로 충분
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# =========================
# Visit Log Middleware
# =========================
SKIP_LOG_PREFIXES = ("/docs", "/openapi.json", "/favicon", "/static", "/health")
SKIP_LOG_PATHS = {"/", "/api/health", "/api/docs"}

@app.middleware("http")
async def visit_log_middleware(request: Request, call_next):
    response = await call_next(request)
    path = request.url.path

    # 정적/healthcheck/favicon 등 제외
    if path in SKIP_LOG_PATHS or any(path.startswith(p) for p in SKIP_LOG_PREFIXES):
        return response

    # GET /api/* 요청만 로깅 (쓰기 요청은 제외해 로그 폭증 방지)
    if request.method != "GET" or not path.startswith("/api/"):
        return response

    try:
        ip = request.headers.get("X-Forwarded-For", "").split(",")[0].strip() or request.client.host if request.client else "unknown"
        today_str = datetime.now(KST).strftime("%Y-%m-%d")

        user_id = None
        username = None
        deptname = None

        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            from app.routers.auth import load_session
            token = auth.replace("Bearer ", "").strip()
            user_info = load_session(token)
            if user_info:
                loginid = user_info.get("loginid", "")
                db = SessionLocal()
                try:
                    u = db.query(User).filter(User.loginid == loginid).first()
                    if u:
                        user_id = u.id
                        username = u.username
                        deptname = u.deptname
                    # 동일 IP는 하루 1회만 기록
                    existing = db.query(VisitLog).filter(
                        VisitLog.ip_address == ip[:50],
                        VisitLog.visit_date == today_str,
                    ).first()
                    if not existing:
                        log_entry = VisitLog(
                            ip_address=ip[:50],
                            user_id=user_id,
                            username=username,
                            deptname=deptname,
                            visit_date=today_str,
                        )
                        db.add(log_entry)
                        db.commit()
                finally:
                    db.close()
    except Exception:
        pass  # 로깅 실패가 서비스에 영향주지 않도록

    return response


# =========================
# Request ID Middleware + 전역 예외 핸들러
# =========================
# - 요청마다 request_id(REQ-YYYYMMDD-XXXXXXXX)를 생성해 ContextVar에 보관하고
#   응답 헤더 X-Request-ID 로 내려준다.
# - 처리되지 않은 예외(=대부분의 500)는 전역 핸들러에서 system_event_logs 에 요약만 적재하고,
#   사용자에게는 stack trace 대신 request_id(오류코드)만 노출한다.
@app.middleware("http")
async def request_id_middleware(request: Request, call_next):
    from app.services.system_log import new_request_id, request_id_ctx
    rid = new_request_id()
    # ContextVar: 요청 핸들러 내부(DSLLM/S3 등)에서 log_event 시 자동 참조용.
    # request.state: 전역 예외 핸들러가 finally(reset) 이후에도 안전하게 읽기 위함(scope 공유).
    token = request_id_ctx.set(rid)
    try:
        request.state.request_id = rid
    except Exception:
        pass
    try:
        response = await call_next(request)
        response.headers["X-Request-ID"] = rid
        return response
    finally:
        try:
            request_id_ctx.reset(token)
        except Exception:
            pass


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    """처리되지 않은 예외(500) 전역 핸들러.

    - 요약 + 요약 traceback(마지막 일부)만 system_event_logs 에 적재 (민감정보 미저장)
    - 사용자 응답에는 stack trace 미노출, request_id(오류코드)만 전달
    """
    from fastapi.responses import JSONResponse
    import traceback
    from app.services.system_log import log_error, current_request_id

    rid = getattr(getattr(request, "state", None), "request_id", "") or current_request_id()
    try:
        # traceback 은 길이를 제한해 요약만 저장 (마지막 2000자)
        tb = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
        tb_summary = tb[-2000:]
    except Exception:
        tb_summary = None

    # 토큰 원문 등은 절대 기록하지 않음 — endpoint/method/요약 메시지만 기록
    login_id = None
    try:
        info = try_get_user_from_token(request)
        login_id = (info or {}).get("loginid") or None
    except Exception:
        login_id = None

    log_error(
        "API",
        f"{type(exc).__name__}: {exc}",
        detail=tb_summary,
        endpoint=str(request.url.path),
        method=request.method,
        status_code=500,
        login_id=login_id,
        request_id=rid,
    )

    headers = {"X-Request-ID": rid}
    # 전역 예외는 CORSMiddleware 를 거치지 않을 수 있으므로 최소 CORS 헤더를 직접 부착
    # (allow_credentials=False 이므로 와일드카드 허용 가능)
    origin = request.headers.get("origin")
    if origin:
        headers["Access-Control-Allow-Origin"] = origin
    else:
        headers["Access-Control-Allow-Origin"] = "*"

    return JSONResponse(
        status_code=500,
        content={
            "detail": "오류가 발생했습니다. 관리자에게 문의하세요.",
            "request_id": rid,
        },
        headers=headers,
    )


def _log_dsllm_failure(message: str, *, detail=None, status_code=None, endpoint=None):
    """system_event_logs 에 DSLLM 실패 요약 기록 (프롬프트 전문은 저장하지 않음)."""
    try:
        from app.services.system_log import log_event
        log_event("ERROR", "DSLLM", message, detail=detail, status_code=status_code, endpoint=endpoint)
    except Exception:
        pass


def _std_openai_error_type(e) -> str:
    """OpenAIAdapterError → 표준 error_type(status/이름 기반). vision_diagnostics 표준값과 맞춘다."""
    st = getattr(e, "status_code", None)
    nm = getattr(e, "error_type", None) or ""
    if st == 401:
        return "authentication_error"
    if st == 403:
        return "permission_error"
    if st == 404:
        return "model_not_found"
    if st == 413:
        return "payload_too_large"
    if st == 429:
        return "model_unavailable"
    if isinstance(st, int) and 500 <= st < 600:
        return "gateway_5xx"
    if isinstance(st, int) and 400 <= st < 500:
        return "gateway_4xx"
    if nm == "APITimeoutError":
        return "read_timeout"
    if nm == "APIConnectionError":
        return "connection_error"
    return "unknown_error"


def _log_vision_diagnostic(
    level: str,
    failure_stage: str,
    error_type: str,
    *,
    provider=None,
    model=None,
    http_status=None,
    endpoint=None,
    login_id=None,
    details=None,
    retry_safe=None,
    recovery: bool = False,
    message: str = None,
):
    """Vision Report / DSLLM 실패·복구를 기존 SystemEventLog(category=VISION_AI)에 진단과 함께 기록.

    - build_diagnosis 로 요약/추정원인/권장대응/재시도안전을 결정론적으로 구성한다.
    - detail_json 에 비민감 지표만 담은 진단 JSON 을, detail 에 사람이 읽는 요약 한 줄을 저장한다.
    - §11: 한 요청에서 같은 (stage, error_type[, recovery]) 는 1건만 저장(dedupe_key).
    - log_event 는 절대 예외를 던지지 않으므로 원래 Vision 예외 처리를 방해하지 않는다.
    """
    try:
        import json as _json
        from app.services import vision_diagnostics as _vd
        from app.services.system_log import log_event

        diag = _vd.build_diagnosis(
            failure_stage, error_type,
            provider=provider, model=model, http_status=http_status,
            details=details, retry_safe=retry_safe,
        )
        diag["recovery_event"] = bool(recovery)
        msg = message or diag["message"]
        detail_line = f"{diag['summary']} / 권장: {diag['recommended_action']}"
        dedupe = f"{diag['failure_stage']}|{diag['error_type']}" + ("|recovery" if recovery else "")
        log_event(
            level, "VISION_AI", msg,
            detail=detail_line,
            detail_json=_json.dumps(diag, ensure_ascii=False),
            endpoint=endpoint,
            status_code=http_status,
            login_id=login_id,
            resolved=True if recovery else None,
            dedupe_key=dedupe,
        )
    except Exception:
        pass


@app.on_event("startup")
def startup_ensure_super_admin():
    # 배포 검증용: 실행 중인 코드의 버전/경로/CWD 를 startup 로그에 남긴다.
    #   회사 서버가 예전 코드를 돌리고 있는지 이 로그로 즉시 확인 가능.
    import logging as _startup_logging
    _slog = _startup_logging.getLogger("main")
    _slog.info("[APP_VERSION] PLAN-AI backend version=%s", APP_VERSION)
    _slog.info("[APP_PATH] main_file=%s", os.path.abspath(__file__))
    _slog.info("[APP_CWD] cwd=%s", os.getcwd())
    # 이미지 관찰 캐시 플래그 실측값 — env 반영/재시작 여부를 기동 로그로 즉시 확인(민감정보 없음).
    try:
        from app.config import settings as _vc
        _slog.info(
            "[VisionCacheConfig] observation_enabled=%s shadow_mode=%s cache_read_enabled=%s "
            "context_link_enabled=%s rollout_mode=%s canary_project_ids=%s",
            _vc.vision_image_observation_enabled, _vc.vision_image_observation_shadow_mode,
            _vc.vision_image_observation_cache_read_enabled, _vc.vision_image_context_link_enabled,
            (_vc.vision_image_cache_rollout_mode or "off"),
            (_vc.vision_image_cache_canary_project_ids or "(none)"),
        )
    except Exception:
        _startup_logging.getLogger("main").warning("[VisionCacheConfig] 로드 실패(무시)")

    # user_shortcuts 테이블 자동 생성 (없으면)
    from app.models import UserShortcut as _UserShortcut
    _UserShortcut.__table__.create(bind=engine, checkfirst=True)

    # system_event_logs 테이블 자동 생성 (없으면)
    try:
        SystemEventLog.__table__.create(bind=engine, checkfirst=True)
    except Exception as e:
        import logging
        logging.getLogger("main").warning(f"system_event_logs 테이블 생성 실패 (무시): {e}")

    db = SessionLocal()
    try:
        ensure_super_owner(db)
        # v1.2: backfill deptname -> TEAM node
        _backfill_team_nodes(db)
        # 3일 지난 삭제 항목 영구 제거
        # FK 제약 위반 등으로 정리에 실패해도 서버 부팅이 막히지 않도록 보호
        try:
            _purge_expired_trash(db)
        except Exception as e:
            import logging
            logging.getLogger("main").warning(
                f"휴지통 만료 데이터 정리 실패 (무시하고 서버 시작 계속): {e}"
            )
            try:
                db.rollback()
            except Exception:
                pass
        # 빈 공간 경고 및 자동 삭제
        _cleanup_empty_spaces(db)
    finally:
        db.close()

    # S3 백업 스케줄러 시작
    try:
        from app.services.backup_scheduler import start_backup_scheduler
        start_backup_scheduler()
    except Exception as e:
        import logging
        logging.getLogger("main").warning(f"백업 스케줄러 시작 실패 (무시): {e}")

    # realtime_events retention 정리 스케줄러 시작 (서버 시작 후 1회 + 매일 새벽)
    try:
        realtime_start_cleanup()
    except Exception as e:
        import logging
        logging.getLogger("main").warning(f"realtime cleanup 스케줄러 시작 실패 (무시): {e}")

    # 멘션 메일 알림 processor 스케줄러 시작 (notification_events outbox 소비)
    try:
        from app.services.notification_processor import start_notification_scheduler
        start_notification_scheduler()
    except Exception as e:
        import logging
        logging.getLogger("main").warning(f"알림 processor 스케줄러 시작 실패 (무시): {e}")

@app.on_event("shutdown")
def shutdown_backup_scheduler():
    try:
        from app.services.backup_scheduler import stop_backup_scheduler
        stop_backup_scheduler()
    except Exception:
        pass
    try:
        realtime_stop_cleanup()
    except Exception:
        pass
    try:
        from app.services.notification_processor import stop_notification_scheduler
        stop_notification_scheduler()
    except Exception:
        pass

def _backfill_team_nodes(db: Session):
    """deptname 값을 기반으로 groups 테이블에 TEAM 타입 노드를 자동 생성하고 users.primary_team_id를 매핑"""
    try:
        from app.models import Group as GroupModel
        all_users = db.query(User).filter(User.deptname.isnot(None), User.deptname != "").all()
        dept_names = set(u.deptname.strip() for u in all_users if u.deptname and u.deptname.strip())

        for dept in dept_names:
            existing = db.query(GroupModel).filter(GroupModel.name == dept).first()
            if not existing:
                new_group = GroupModel(name=dept, group_type="TEAM", is_active=True)
                db.add(new_group)
                db.flush()
                existing = new_group

            # Map users with this deptname to the team group
            for u in all_users:
                if u.deptname and u.deptname.strip() == dept:
                    if not u.primary_team_id:
                        u.primary_team_id = existing.id

        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[WARN] backfill team nodes failed: {e}")

# =========================
# DB Dependency
# =========================
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_active_user(request: Request, db: Session = Depends(get_db)):
    """등록된 활성 사용자만 통과. 미등록/비활성이면 403."""
    from app.routers.auth import load_session
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="로그인이 필요합니다.")
    token = auth.replace("Bearer ", "").strip()
    user_info = load_session(token)
    if not user_info:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다.")
    loginid = (user_info or {}).get("loginid", "")
    user = db.query(User).filter(User.loginid == loginid).first()
    if not user or not user.is_active:
        raise HTTPException(
            status_code=403,
            detail="접근 권한이 없습니다. 관리자에게 등록을 요청하세요."
        )
    return user

# 시간
KST = ZoneInfo("Asia/Seoul")

def _today_kst() -> date:
    return datetime.now(KST).date()

def _add_months(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    return date(y, m, 1)

def _month_window(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    end = _add_months(start, 1)
    return start, end  # [start, end)

def _parse_iso_or_ymd(s: str) -> date | None:
    if not s:
        return None
    s = str(s).strip()
    if not s:
        return None
    # "2026-02-15" or "2026-02-15T00:00:00"
    try:
        if "T" in s:
            return datetime.fromisoformat(s.replace("Z", "")).date()
        return date.fromisoformat(s[:10])
    except Exception:
        return None

def _parse_time_window_from_query(q: str, today: date) -> tuple[date, date] | None:
    q = (q or "").strip()
    if not q:
        return None

    # ── 주 단위 상대 표현 ──
    if re.search(r"(이번\s*주|이번주|금주)", q):
        # 이번주: 월요일~일요일
        monday = today - timedelta(days=today.weekday())
        return monday, monday + timedelta(days=7)
    if re.search(r"(다음\s*주|다음주|차주)", q):
        monday = today - timedelta(days=today.weekday()) + timedelta(days=7)
        return monday, monday + timedelta(days=7)
    if re.search(r"(지난\s*주|지난주|저번\s*주|저번주)", q):
        monday = today - timedelta(days=today.weekday()) - timedelta(days=7)
        return monday, monday + timedelta(days=7)

    # ── 월 단위 상대 표현 ──
    if re.search(r"(이번\s*달|이번달)", q):
        return _month_window(today.year, today.month)
    if re.search(r"(다음\s*달|다음달)", q):
        d = _add_months(date(today.year, today.month, 1), 1)
        return _month_window(d.year, d.month)
    if re.search(r"(지난\s*달|지난달|저번\s*달|저번달)", q):
        d = _add_months(date(today.year, today.month, 1), -1)
        return _month_window(d.year, d.month)

    # ── "오늘", "내일", "모레" ──
    if re.search(r"(오늘)", q):
        return today, today + timedelta(days=1)
    if re.search(r"(내일)", q):
        t = today + timedelta(days=1)
        return t, t + timedelta(days=1)
    if re.search(r"(모레)", q):
        t = today + timedelta(days=2)
        return t, t + timedelta(days=1)

    # "2026년 2월"
    m = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월", q)
    if m:
        y = int(m.group(1))
        mo = int(m.group(2))
        if 1 <= mo <= 12:
            return _month_window(y, mo)

    # "2월" (연도 없으면 올해로 가정)
    m2 = re.search(r"(\d{1,2})\s*월", q)
    if m2:
        mo = int(m2.group(1))
        if 1 <= mo <= 12:
            y = today.year
            # "내년 2월" 같은 케이스
            if "내년" in q:
                y += 1
            elif "작년" in q:
                y -= 1
            return _month_window(y, mo)

    # "2026-02-15" / "2026/02/15" / "2/15"
    m3 = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", q)
    if m3:
        y, mo, d = map(int, m3.groups())
        try:
            start = date(y, mo, d)
            return start, start + timedelta(days=1)
        except Exception:
            return None

    m4 = re.search(r"(?<!\d)(\d{1,2})/(\d{1,2})(?!\d)", q)
    if m4:
        mo, d = map(int, m4.groups())
        y = today.year
        try:
            start = date(y, mo, d)
            return start, start + timedelta(days=1)
        except Exception:
            return None

    return None

def _task_overlaps_window(task: dict, start: date, end: date) -> bool:
    sd = _parse_iso_or_ymd(task.get("start_date"))
    dd = _parse_iso_or_ymd(task.get("due_date"))

    # 둘 다 있으면 overlap 기준(기간이 걸친 Task도 포함)
    if sd and dd:
        return (sd < end) and (dd >= start)

    # 하나만 있으면 그 날짜가 window 안에 들어오면 포함
    if dd:
        return (start <= dd < end)
    if sd:
        return (start <= sd < end)

    return False

def _extract_task_ids_from_related_text(text: str) -> set[int]:
    if not text:
        return set()
    ids = set()

    # "123 / ..." 형태
    for m in re.finditer(r"^\s*(\d+)\s*/", text, flags=re.MULTILINE):
        ids.add(int(m.group(1)))

    # "ID:123" 형태
    for m in re.finditer(r"\bID\s*[:#]?\s*(\d+)\b", text, flags=re.IGNORECASE):
        ids.add(int(m.group(1)))

    return ids

# =========================
# Sidecar JSON State (추가 기능 저장소)
# - DB 스키마를 안 바꾸고도 새 기능 붙이기 위한 구조
# =========================
DATA_FILE = "data.json"
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

STATE_LOCK = threading.Lock()

DEFAULT_PERMISSIONS = {
    "post_write": "all",
    "post_edit": "all",
    "post_view": "all",
    "comment_write": "all",
    "file_view": "all",
    "file_download": "all",
}

DEFAULT_STATE = {
    # DB 보완용 메타
    "project_meta": {},   # { "project_id": {owner_id, visibility, require_approval, permissions} }
    "task_meta": {},      # { "task_id": {sub_project_id, progress} }
    "user_meta": {},      # { "user_id": {group_name} }

    # 새 기능 저장
    "sub_projects": [],   # [{id, project_id, name, description, parent_id, created_at}]
    "notes": [],          # [{id, project_id, author_id, content, created_at, updated_at}]
    "attachments": [],    # task attachments (URL 형태) [{id, task_id, url, filename, type, created_at}]
    "project_members": [],# [{project_id, user_id, role}]
    "project_files": [],  # [{id, project_id, filename, stored_name, size, uploader_id, created_at}]
    "join_requests": [],  # [{id, project_id, user_id, role, status, created_at}]
    "groups": [],         # [{id, name, created_at, description?, is_active?}]
    "shortcuts": [],      # [{...}]
    "ai_settings": {"api_url": "", "model_name": ""},
    "project_ai_queries": [],
    "ai_summaries": [],
    "search_feedback": [],
    "list_orders": {},

    # ── 사내 추천 도구 스트립 (Internal Service Recommendations) ──
    # 광고가 아닌 "업무와 연결된 추천 도구" 영역. 기본값은 반드시 OFF.
    # global flag 가 False 면 사용자 화면에 아무것도 렌더링되지 않는다.
    "internal_service_recommendations_enabled": False,
    # 추천 항목들 — 시안 2종을 seed 로 넣어두되, 위 flag 가 켜지기 전에는 노출되지 않는다.
    "internal_recommendations": [
        {
            "id": 1,
            "active": True,
            "target_space_type": "PROJECT",
            "placement": "PROJECT_DASHBOARD_TOP",
            "label": "함께 쓰면 좋은 사내 도구",
            "badge": "추천",
            "title": "문서 자동화 Hub",
            "description": "템플릿 기반 문서 생성, 승인, 배포를 빠르게 처리하세요.",
            "reason_text": "문서 협업이 많은 프로젝트에 추천",
            "cta_label": "바로가기",
            "cta_url": "",
            "secondary_label": "나중에 보기",
            "order": 0,
        },
        {
            "id": 2,
            "active": True,
            "target_space_type": "EQUIPMENT",
            "placement": "EQUIPMENT_DASHBOARD_AFTER_KPI",
            "label": "이 공간과 함께 쓰면 좋은 설비 도구",
            "badge": "추천",
            "title": "EQUIP-AI",
            "description": "실시간 설비 상태와 이력 기반 분석을 함께 확인해보세요.",
            "reason_text": "설비 상태/이력 분석에 자주 함께 사용",
            "cta_label": "바로가기",
            "cta_url": "",
            "secondary_label": "나중에 보기",
            "order": 0,
        },
    ],
    # 사용자별 숨김 상태 — { "<user_id>": { "<item_id>": {"mode": "today"|"later", "until": iso} } }
    "internal_recommendation_dismissals": {},
    # 효과 측정용 이벤트 로그 (impression/click/dismiss) — 최근 N개만 보존.
    "internal_recommendation_events": [],
}

def _deepcopy_state(obj: dict) -> dict:
    return json.loads(json.dumps(obj))

def load_state() -> Dict[str, Any]:
    with STATE_LOCK:
        if not os.path.exists(DATA_FILE):
            return _deepcopy_state(DEFAULT_STATE)

        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)

            # 마이그레이션: 누락 키 보강
            for k, v in DEFAULT_STATE.items():
                if k not in data:
                    data[k] = _deepcopy_state(v) if isinstance(v, (dict, list)) else v

            # 타입 방어
            if not isinstance(data.get("project_meta"), dict):
                data["project_meta"] = {}
            if not isinstance(data.get("task_meta"), dict):
                data["task_meta"] = {}
            if not isinstance(data.get("user_meta"), dict):
                data["user_meta"] = {}
            if not isinstance(data.get("ai_settings"), dict):
                data["ai_settings"] = {"api_url": "", "model_name": ""}

            return data
        except json.JSONDecodeError:
            return _deepcopy_state(DEFAULT_STATE)

def save_state(state: Dict[str, Any]):
    with STATE_LOCK:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, indent=2, ensure_ascii=False)

def next_id(items: list) -> int:
    if not items:
        return 1
    return max(int(item.get("id", 0)) for item in items) + 1

# 최초 파일 없으면 생성
if not os.path.exists(DATA_FILE):
    save_state(load_state())

# =========================
# Pydantic Models
# =========================
class TaskBase(BaseModel):
    project_id: int
    title: str
    description: Optional[str] = None
    status: str = "todo"
    priority: Optional[str] = "medium"
    start_date: Optional[str] = None
    due_date: Optional[str] = None
    # 일정 미정(TBD) — True 면 대응 날짜는 무시하고 NULL 로 저장한다.
    start_date_tbd: Optional[bool] = False
    due_date_tbd: Optional[bool] = False
    assignee_ids: List[int] = Field(default_factory=list)
    tags: List[str] = Field(default_factory=list)

    # 단발 일정 구분 (normal / one_off)
    task_type: Optional[str] = "normal"

    # ✅ DB Task 테이블엔 없어서 sidecar(task_meta)에 저장
    sub_project_id: Optional[int] = None
    progress: Optional[int] = 0

    # 상위 Task(Parent). None = 최상위. 하위 Task 생성 시 이 값으로 parent 지정.
    parent_task_id: Optional[int] = None
    # CUSTOM 워크플로우 컬럼(하위 Task 생성 시 부모 컬럼 상속용).
    workflow_column_id: Optional[int] = None

class TaskCreate(TaskBase):
    pass

class TaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    priority: Optional[str] = None
    start_date: Optional[str] = None
    due_date: Optional[str] = None
    start_date_tbd: Optional[bool] = None
    due_date_tbd: Optional[bool] = None
    assignee_ids: Optional[List[int]] = None
    tags: Optional[List[str]] = None
    sub_project_id: Optional[int] = None
    progress: Optional[int] = None
    # CUSTOM 워크플로우 — 전달 시 해당 컬럼의 mapped_status로 status도 동기화
    workflow_column_id: Optional[int] = None
    # 상위 Task(Parent) 변경. 명시적 null = 해제, 미전송 = 유지(exclude_unset).
    # 이 필드가 없으면 pydantic이 payload의 parent_task_id를 drop → PATCH 200이지만 미저장.
    parent_task_id: Optional[int] = None

class OneoffTaskCreate(BaseModel):
    """캘린더에서 추가하는 단발 일정. 공간별 [시스템] 단발 업무 프로젝트 하위 task 로 생성된다."""
    title: str
    description: Optional[str] = None
    start_date: Optional[str] = None
    due_date: Optional[str] = None
    status: Optional[str] = "todo"
    priority: Optional[str] = "medium"
    assignee_ids: List[int] = Field(default_factory=list)

class CalendarEventCreate(BaseModel):
    """공간 단위 단순 일정 생성 (Phase 2)."""
    title: str
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    all_day: Optional[bool] = True
    status: Optional[str] = "planned"
    event_type: Optional[str] = "general"
    assignee_id: Optional[int] = None

class CalendarEventUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    all_day: Optional[bool] = None
    status: Optional[str] = None
    event_type: Optional[str] = None
    assignee_id: Optional[int] = None

class ProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None

    # ✅ DB Project 테이블엔 없어서 sidecar(project_meta)에 저장
    owner_id: Optional[int] = 1
    visibility: Optional[str] = "private"
    require_approval: Optional[bool] = False
    permissions: Optional[Dict[str, str]] = None
    member_ids: Optional[List[int]] = None
    space_id: Optional[int] = None

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

    # ✅ sidecar
    visibility: Optional[str] = None
    require_approval: Optional[bool] = None
    permissions: Optional[Dict[str, str]] = None
    owner_id: Optional[int] = None

class UserCreate(BaseModel):
    username: str
    loginid: str
    role: Optional[str] = "member"
    avatar_color: Optional[str] = "#2955FF"
    deptname: Optional[str] = None
    mail: Optional[str] = None

class UserUpdate(BaseModel):
    username: Optional[str] = None
    role: Optional[str] = None
    avatar_color: Optional[str] = None
    is_active: Optional[bool] = None

    # ✅ DB 컬럼 없음 → sidecar user_meta에 저장
    group_name: Optional[str] = None

class LayoutUpdate(BaseModel):
    layout: Dict[str, Any]

class ListOrderUpdate(BaseModel):
    order: List[int]

class SubProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None
    parent_id: Optional[int] = None

class SubProjectUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    parent_id: Optional[int] = None

class NoteCreate(BaseModel):
    content: str

class AttachmentCreate(BaseModel):
    url: str
    filename: Optional[str] = None
    type: Optional[str] = "url"

class AttachmentUpdate(BaseModel):
    # URL 첨부 편집용. 둘 다 선택 — 들어온 값만 갱신한다.
    url: Optional[str] = None
    filename: Optional[str] = None

class MemberAdd(BaseModel):
    user_id: int
    role: Optional[str] = "member"

class MemberApproval(BaseModel):
    user_id: int
    action: str  # approve / reject

class AiSettingsUpdate(BaseModel):
    # 사내 DSLLM 접속정보는 env에서만 읽으므로 model_name만 사용.
    # api_url은 하위호환을 위해 받기만 하고 사용하지 않는다.
    model_name: str
    api_url: str | None = None
    # AI Provider 선택값 (dsllm / openai). 미지정이면 기존 provider 유지.
    provider: str | None = None


class VisionAiSettingsUpdate(BaseModel):
    """Vision AI 전용 설정 저장 — Text AI(AiSettingsUpdate)와 분리. 미지정 필드는 기존값 유지."""
    model_name: str | None = None
    provider: str | None = None  # Primary Vision Model provider (dsllm/openai)
    enabled: bool | None = None
    max_output_tokens: int | None = None
    batch_size: int | None = None
    timeout_seconds: int | None = None
    # ── Fallback Text Model (Gemma4 지연/실패 시 텍스트 전용 보고서) — Primary 와 별개 저장 ──
    fallback_text_provider: str | None = None
    fallback_text_model: str | None = None
    fallback_text_enabled: bool | None = None
    fallback_policy: str | None = None
    soft_timeout_seconds: int | None = None

class ReportRequest(BaseModel):
    project_id: int

class GroupCreate(BaseModel):
    name: str
    description: Optional[str] = None

class GroupUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None

class KnoxUserPayload(BaseModel):
    """Knox 검색으로 선택된 사용자 정보 (DB upsert용)"""
    loginid: str
    name: Optional[str] = None
    email: Optional[str] = None
    deptname: Optional[str] = None

class MemberGroupCreate(BaseModel):
    name: str
    description: Optional[str] = None
    member_user_ids: Optional[List[int]] = None
    knox_members: Optional[List[KnoxUserPayload]] = None

class MemberGroupUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    member_user_ids: Optional[List[int]] = None
    knox_members: Optional[List[KnoxUserPayload]] = None

class ShortcutCreate(BaseModel):
    name: str
    url: str
    icon_text: Optional[str] = None
    icon_color: Optional[str] = "#2955FF"
    order: Optional[int] = 0
    open_new_tab: Optional[bool] = True
    # 공유 모델 — PRIVATE / PUBLIC / SHARED_USERS / SHARED_GROUPS. 기본 PRIVATE.
    visibility: Optional[str] = "PRIVATE"
    shared_user_ids: Optional[List[int]] = None
    shared_group_ids: Optional[List[int]] = None

class ShortcutUpdate(BaseModel):
    name: Optional[str] = None
    url: Optional[str] = None
    icon_text: Optional[str] = None
    icon_color: Optional[str] = None
    order: Optional[int] = None
    open_new_tab: Optional[bool] = None
    active: Optional[bool] = None
    visibility: Optional[str] = None
    shared_user_ids: Optional[List[int]] = None
    shared_group_ids: Optional[List[int]] = None

# ── 사내 추천 도구 스트립 (Internal Service Recommendations) ──
class InternalRecommendationEvent(BaseModel):
    item_id: int
    user_id: Optional[int] = None
    type: str  # impression | click | dismiss

class InternalRecommendationFlagUpdate(BaseModel):
    enabled: bool

class InternalRecommendationUpdate(BaseModel):
    active: Optional[bool] = None
    label: Optional[str] = None
    badge: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    reason_text: Optional[str] = None
    cta_label: Optional[str] = None
    cta_url: Optional[str] = None
    secondary_label: Optional[str] = None
    order: Optional[int] = None

# 26일 추가됨
class ProjectAiQueryRequest(BaseModel):
    query: str

class SummaryFeedback(BaseModel):
    summary_id: int
    rating: int
    comment: Optional[str] = None

class SummaryCorrectionSave(BaseModel):
    summary_id: int
    corrected_text: str

# =========================
# Serializer / Helper
# =========================
def iso(dt):
    return dt.isoformat() if dt else None

def get_project_meta(state: dict, project_id: int) -> dict:
    meta = state.get("project_meta", {}).get(str(project_id), {})
    return {
        "owner_id": meta.get("owner_id", 1),
        "visibility": meta.get("visibility", "private"),
        "require_approval": bool(meta.get("require_approval", False)),
        "permissions": meta.get("permissions") or dict(DEFAULT_PERMISSIONS),
        "is_system": bool(meta.get("is_system", False)),
        "system_kind": meta.get("system_kind"),
    }

def set_project_meta(state: dict, project_id: int, values: dict):
    if "project_meta" not in state or not isinstance(state["project_meta"], dict):
        state["project_meta"] = {}
    curr = state["project_meta"].get(str(project_id), {})
    curr.update(values)
    if "permissions" in curr and curr["permissions"] is None:
        curr["permissions"] = dict(DEFAULT_PERMISSIONS)
    state["project_meta"][str(project_id)] = curr

def get_task_meta(state: dict, task_id: int) -> dict:
    meta = state.get("task_meta", {}).get(str(task_id), {})
    return {
        "sub_project_id": meta.get("sub_project_id"),
        "progress": int(meta.get("progress", 0) or 0),
    }

def set_task_meta(state: dict, task_id: int, values: dict):
    if "task_meta" not in state or not isinstance(state["task_meta"], dict):
        state["task_meta"] = {}
    curr = state["task_meta"].get(str(task_id), {})
    curr.update(values)
    # 방어
    if "progress" in curr and curr["progress"] is not None:
        try:
            curr["progress"] = max(0, min(100, int(curr["progress"])))
        except Exception:
            curr["progress"] = 0
    state["task_meta"][str(task_id)] = curr

# ── status 파생 단일 소스 ──
# 체크박스 진행률 기반 자동 status 승격 규칙을 한 곳으로 모은다.
# 규칙:
#   - hold: 수동 지정이므로 절대 자동 변경하지 않음
#   - progress >= 100 → done
#   - progress > 0 && status == todo → in_progress
#   - 그 외: 현재 status 유지 (사용자가 직접 정한 값 보존)
#   - "In Progress 50% 이상 진행"은 별도 status 값이 아니라 (status=in_progress
#     AND progress>=50) 조합으로 프론트에서 파생 표시한다. 이 함수는 해당
#     조합을 파괴하지 않는다.
def _derive_task_status(current_status: str, progress: int) -> str:
    if current_status == "hold":
        return current_status
    if progress >= 100:
        return "done"
    # v3.12: progress 가 100% 미만으로 다시 떨어지면 done → in_progress 로 강제 다운그레이드.
    #        (체크 해제 / Check Sheet 진행→미진행 / 항목 추가 등 모든 경로에서 동일하게 동작)
    if current_status == "done" and progress < 100:
        return "in_progress" if progress > 0 else "todo"
    if progress > 0 and current_status == "todo":
        return "in_progress"
    return current_status


def _compute_task_progress(activities) -> tuple[int, int]:
    """Return (progress_percent, checkbox_count) for a task's activities."""
    checkboxes = [a for a in activities if (a.block_type or "checkbox") == "checkbox"]
    total = len(checkboxes)
    if total == 0:
        return 0, 0
    checked = sum(1 for a in checkboxes if a.checked)
    return round(checked / total * 100), total


# ── 서버 시작 시 기존 task progress/status 일괄 동기화 ──
def _sync_all_task_progress():
    """기존 task들의 checkbox progress 기반 status를 일괄 보정."""
    db = SessionLocal()
    try:
        task_ids_with_cb = (
            db.query(TaskActivityModel.task_id)
            .filter(TaskActivityModel.block_type.in_(["checkbox", None]))
            .distinct()
            .all()
        )
        task_ids = {r[0] for r in task_ids_with_cb}
        if not task_ids:
            return
        state = load_state()
        fixed = 0
        for tid in task_ids:
            task = db.query(Task).filter(Task.id == tid, Task.archived_at.is_(None)).first()
            if not task or task.status == "hold":
                continue
            activities = db.query(TaskActivityModel).filter(TaskActivityModel.task_id == tid).all()
            progress, cb_total = _compute_task_progress(activities)
            if cb_total == 0:
                continue
            expected_status = _derive_task_status(task.status, progress)
            meta = get_task_meta(state, tid)
            if task.status != expected_status or meta.get("progress", 0) != progress:
                task.status = expected_status
                task.progress = progress
                set_task_meta(state, tid, {"progress": progress})
                fixed += 1
        if fixed > 0:
            db.commit()
            save_state(state)
            print(f"[startup] Synced {fixed} task(s) progress/status")
    finally:
        db.close()

_sync_all_task_progress()

def get_user_meta(state: dict, user_id: int) -> dict:
    return state.get("user_meta", {}).get(str(user_id), {})

def set_user_meta(state: dict, user_id: int, values: dict):
    if "user_meta" not in state or not isinstance(state["user_meta"], dict):
        state["user_meta"] = {}
    curr = state["user_meta"].get(str(user_id), {})
    curr.update(values)
    state["user_meta"][str(user_id)] = curr

def project_dict(p: Project, state: dict):
    meta = get_project_meta(state, p.id)
    return {
        "id": p.id,
        "name": p.name,
        "description": p.description,
        "created_at": iso(p.created_at),
        "archived_at": iso(p.archived_at),
        "owner_id": meta["owner_id"],
        "visibility": meta["visibility"],
        "require_approval": meta["require_approval"],
        "permissions": meta["permissions"],
        "space_id": p.space_id,
        "is_system": meta["is_system"],
        "system_kind": meta["system_kind"],
        "workflow_mode": getattr(p, "workflow_mode", "DEFAULT") or "DEFAULT",
        "auto_progress_from_notes": bool(getattr(p, "auto_progress_from_notes", True)),
    }

def user_dict(u: User, state: dict):
    meta = get_user_meta(state, u.id)
    return {
        "id": u.id,
        "loginid": u.loginid,
        "username": u.username,
        "role": u.role,
        "avatar_color": u.avatar_color,
        "is_active": u.is_active,
        "deptname": getattr(u, "deptname", None),
        "mail": getattr(u, "mail", None),
        "created_at": iso(u.created_at),
        "last_login_at": iso(u.last_login_at),
        "group_name": getattr(u, "group_name", None) or meta.get("group_name"),
    }

def normalize_schedule_tbd(start_date, due_date, start_tbd, due_tbd):
    """일정 미정(TBD) 규칙을 적용해 정규화된 (start_date, due_date, start_tbd, due_tbd) 를 반환.

    규칙(요구사항 §5):
      - start_tbd=True → start_date 는 NULL 로 강제
      - due_tbd=True   → due_date 는 NULL 로 강제
      - 둘 다 확정(미정 아님)이고 날짜가 모두 있으면 start_date <= due_date 검증
      - 그 외(한쪽만 있음/미정)는 모두 허용
    빈 문자열은 NULL 로 취급한다.
    """
    start_tbd = bool(start_tbd)
    due_tbd = bool(due_tbd)
    sd = (start_date or None) if not start_tbd else None
    dd = (due_date or None) if not due_tbd else None
    if isinstance(sd, str) and not sd.strip():
        sd = None
    if isinstance(dd, str) and not dd.strip():
        dd = None
    if (not start_tbd) and (not due_tbd) and sd and dd and sd > dd:
        raise HTTPException(status_code=400, detail="시작일은 마감일보다 늦을 수 없습니다.")
    return sd, dd, start_tbd, due_tbd


def task_dict(t: Task, state: dict):
    meta = get_task_meta(state, t.id)
    return {
        "id": t.id,
        "project_id": t.project_id,
        "title": t.title,
        "description": t.description,
        "status": t.status,
        "priority": t.priority,
        "task_type": getattr(t, "task_type", "normal") or "normal",
        "start_date": t.start_date,
        "due_date": t.due_date,
        "start_date_tbd": bool(getattr(t, "start_date_tbd", False)),
        "due_date_tbd": bool(getattr(t, "due_date_tbd", False)),
        "assignee_ids": t.assignee_ids or [],
        "tags": t.tags or [],
        "sub_project_id": meta.get("sub_project_id"),
        "progress": meta.get("progress", 0),
        "parent_task_id": getattr(t, "parent_task_id", None),
        "workflow_column_id": getattr(t, "workflow_column_id", None),
        "created_at": iso(t.created_at),
        "remarks": t.remarks,
        "updated_at": iso(t.updated_at),
        "archived_at": iso(t.archived_at),
    }

def calendar_event_dict(e: CalendarEvent):
    return {
        "id": e.id,
        "space_id": e.space_id,
        "title": e.title,
        "description": e.description,
        "start_date": e.start_date,
        "end_date": e.end_date,
        "all_day": bool(e.all_day),
        "status": e.status,
        "event_type": e.event_type,
        "owner_id": e.owner_id,
        "assignee_id": e.assignee_id,
        "linked_task_id": e.linked_task_id,
        "linked_project_id": e.linked_project_id,
        "created_at": iso(e.created_at),
        "updated_at": iso(e.updated_at),
    }

def get_client_ip(request: Request) -> str:
    xff = request.headers.get("x-forwarded-for")
    if xff:
        return xff.split(",")[0].strip()
    if request.client:
        return request.client.host
    return "unknown"

def try_get_user_from_token(request: Request) -> dict:
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header.replace("Bearer ", "").strip()
        from app.routers.auth import load_session  # 지연 import
        user_info = load_session(token)
        if user_info:
            return user_info
    return {}

def resolve_user_id_from_token(request: Request, db: Session) -> Optional[int]:
    user_info = try_get_user_from_token(request)
    loginid = user_info.get("loginid")
    if not loginid:
        return None
    row = db.query(User).filter(User.loginid == loginid).first()
    return row.id if row else None

def get_subprojects_from_db(db: Session, project_id: Optional[int] = None) -> List[dict]:
    """C-2: Get subprojects from DB as dicts (replaces state.get('sub_projects'))"""
    q = db.query(SubProjectModel)
    if project_id:
        q = q.filter(SubProjectModel.project_id == project_id)
    return [{
        "id": sp.id,
        "project_id": sp.project_id,
        "name": sp.name,
        "description": sp.description,
        "parent_id": sp.parent_id,
        "created_at": iso(sp.created_at),
    } for sp in q.all()]

def get_members_for_project(state: dict, project_id: int, db: Session = None) -> List[dict]:
    """C-3: Try DB first, fallback to sidecar"""
    if db:
        rows = db.query(ProjectMemberModel).filter(ProjectMemberModel.project_id == int(project_id)).all()
        if rows:
            return [{"project_id": r.project_id, "user_id": r.user_id, "role": r.role, "loginid": r.loginid, "deptname": r.deptname} for r in rows]
    return [m for m in state.get("project_members", []) if int(m.get("project_id")) == int(project_id)]

def ensure_owner_membership(state: dict, project_id: int, owner_id: Optional[int], db: Session = None):
    if not owner_id:
        return
    if db:
        existing = db.query(ProjectMemberModel).filter(
            ProjectMemberModel.project_id == int(project_id),
            ProjectMemberModel.user_id == int(owner_id),
        ).first()
        if not existing:
            owner_user = db.query(User).filter(User.id == int(owner_id)).first()
            pm = ProjectMemberModel(
                project_id=project_id, user_id=owner_id, role="owner",
                loginid=owner_user.loginid if owner_user else None,
                deptname=getattr(owner_user, "deptname", None) if owner_user else None,
            )
            db.add(pm)
            db.flush()
        return
    # Fallback: sidecar
    members = state.get("project_members", [])
    exists = any(int(m.get("project_id")) == int(project_id) and int(m.get("user_id")) == int(owner_id) for m in members)
    if not exists:
        members.append({"project_id": project_id, "user_id": owner_id, "role": "owner"})
        state["project_members"] = members

ONEOFF_PROJECT_NAME = "[시스템] 단발 업무"

def get_or_create_oneoff_project(db: Session, state: dict, space_id: int) -> Project:
    """공간별 [시스템] 단발 업무 프로젝트를 가져오거나 없으면 생성한다.

    - project_meta 에 is_system=True, system_kind="oneoff" 마커를 둔다.
    - visibility="public" 으로 두어 공간 멤버 모두가 캘린더에서 단발 일정을 볼 수 있게 한다.
      (생성 권한은 별도로 공간 멤버십으로 통제한다 — create_oneoff_task 참고)
    """
    rows = db.query(Project).filter(
        Project.space_id == space_id,
        Project.archived_at.is_(None),
    ).all()
    for p in rows:
        meta = get_project_meta(state, p.id)
        if meta.get("is_system") and meta.get("system_kind") == "oneoff":
            return p

    # 소유자: 공간 owner → 없으면 super owner → 1
    owner = db.query(SpaceMember).filter(
        SpaceMember.space_id == space_id,
        SpaceMember.role == "owner",
    ).first()
    owner_id = owner.user_id if owner else (get_super_owner_id(db) or 1)

    p = Project(
        name=ONEOFF_PROJECT_NAME,
        description="캘린더에서 추가한 단발 일정이 모이는 시스템 프로젝트입니다.",
        owner_id=owner_id,
        created_by=owner_id,
        space_id=space_id,
    )
    db.add(p)
    db.commit()
    db.refresh(p)

    set_project_meta(state, p.id, {
        "owner_id": owner_id,
        "visibility": "public",
        "require_approval": False,
        "permissions": dict(DEFAULT_PERMISSIONS),
        "is_system": True,
        "system_kind": "oneoff",
    })
    ensure_owner_membership(state, p.id, owner_id, db=db)
    db.commit()
    save_state(state)
    return p

def get_user_role(db: Session, user_id: int) -> Optional[str]:
    u = db.query(User).filter(User.id == user_id).first()
    return u.role if u else None

def is_admin_like_role(role: Optional[str]) -> bool:
    return role in {"admin", "super_admin"}  # 점진 전환용

def is_super_admin_user(db: Session, user_id: int) -> bool:
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        return False
    return bool(u.is_active) and u.role == "super_admin" and is_super_owner_loginid(u.loginid)

def check_task_edit_permission(db: Session, state: dict, project_id: int, user_id: int):
    """owner/담당자만 task 수정 가능. viewer는 불가."""
    if is_admin_like_role(get_user_role(db, user_id)):
        return True

    meta = get_project_meta(state, project_id)
    if int(meta.get("owner_id") or 0) == int(user_id):
        return True

    pm = db.query(ProjectMemberModel).filter(
        ProjectMemberModel.project_id == int(project_id),
        ProjectMemberModel.user_id == int(user_id),
    ).first()
    if pm and pm.role in ("owner", "manager", "member"):
        return True

    raise HTTPException(status_code=403, detail="Task 수정 권한이 없습니다. (viewer는 수정 불가)")

# ── 프로젝트 관리 권한 (소유자 / 담당자 / super_admin / admin) ──
# UI 라벨 매핑: "소유자" = projects.owner_id, "담당자" = project_members.role in (manager, member)
# viewer는 관리 권한 없음.
PROJECT_MANAGER_ROLES = ("owner", "manager", "member")

def can_manage_project(db: Session, state: dict, project_id: int, user_id: Optional[int]) -> bool:
    """소유자 또는 담당자(또는 super_admin/admin)이면 True. user_id가 없으면 False."""
    if not user_id:
        return False
    if is_admin_like_role(get_user_role(db, user_id)):
        return True
    meta = get_project_meta(state, project_id)
    if int(meta.get("owner_id") or 0) == int(user_id):
        return True
    pm = db.query(ProjectMemberModel).filter(
        ProjectMemberModel.project_id == int(project_id),
        ProjectMemberModel.user_id == int(user_id),
    ).first()
    if pm and pm.role in PROJECT_MANAGER_ROLES:
        return True
    return False

def require_project_management(db: Session, state: dict, project_id: int, user_id: Optional[int]):
    if not can_manage_project(db, state, project_id, user_id):
        raise HTTPException(status_code=403, detail="프로젝트 관리 권한이 없습니다. (소유자 또는 담당자만 가능)")

def require_admin(db: Session, state: dict, user_id: int):
    """super_admin 또는 admin만 허용"""
    u = db.query(User).filter(User.id == user_id).first()
    if not u or not bool(u.is_active):
        raise HTTPException(status_code=403, detail="Admin access required")

    if u.role not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

def require_admin_and_get_user(db: Session, user_id: int) -> "User":
    """require_admin 통과 후 User 객체 반환 (audit 로그의 login_id 기록용)."""
    require_admin(db, None, user_id)
    return db.query(User).filter(User.id == user_id).first()

def require_super_admin(db: Session, user_id: int):
    """super_admin만 허용 (AI settings 등)"""
    u = db.query(User).filter(User.id == user_id).first()
    if not u or not bool(u.is_active):
        raise HTTPException(status_code=403, detail="Super admin access required")
    if u.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")

def get_user_project_ids(db: Session, state: dict, user_id: int) -> set:
    pids = set()

    # C-3: DB members first
    db_memberships = db.query(ProjectMemberModel).filter(ProjectMemberModel.user_id == int(user_id)).all()
    for m in db_memberships:
        pids.add(m.project_id)

    # Fallback: sidecar members
    for m in state.get("project_members", []):
        if int(m.get("user_id")) == int(user_id):
            pids.add(int(m.get("project_id")))

    # owner
    projects = db.query(Project).filter(Project.archived_at.is_(None)).all()
    for p in projects:
        meta = get_project_meta(state, p.id)
        if int(meta.get("owner_id") or 0) == int(user_id):
            pids.add(p.id)

    return pids

def get_user_public_project_ids(db: Session, state: dict, user_id: int) -> set:
    """사용자가 속한 space의 공개 프로젝트 ID 집합 반환"""
    # 사용자가 속한 space 목록
    user_space_ids = {
        sm.space_id for sm in
        db.query(SpaceMember).filter(SpaceMember.user_id == int(user_id)).all()
    }
    if not user_space_ids:
        return set()
    # 해당 space에 속한 공개 프로젝트
    public_projects = db.query(Project).filter(
        Project.space_id.in_(user_space_ids),
        Project.archived_at.is_(None),
    ).all()
    pids = set()
    for p in public_projects:
        meta = get_project_meta(state, p.id)
        if meta.get("visibility") == "public":
            pids.add(p.id)
    return pids

def check_project_access(db: Session, state: dict, project_id: int, user_id: int):
    # admin pass
    if is_admin_like_role(get_user_role(db, user_id)):
        return True

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    meta = get_project_meta(state, p.id)

    if int(meta.get("owner_id") or 0) == int(user_id):
        return True

    # C-3: Check DB members
    db_member = db.query(ProjectMemberModel).filter(
        ProjectMemberModel.project_id == int(project_id),
        ProjectMemberModel.user_id == int(user_id),
    ).first()
    if db_member:
        return True

    if any(int(m.get("project_id")) == int(project_id) and int(m.get("user_id")) == int(user_id) for m in state.get("project_members", [])):
        return True

    # 공개 프로젝트: 같은 space 소속 사용자만 조회 허용
    if meta.get("visibility") == "public" and p.space_id:
        is_space_member = db.query(SpaceMember).filter(
            SpaceMember.space_id == p.space_id,
            SpaceMember.user_id == int(user_id),
        ).first()
        if is_space_member:
            return True

    raise HTTPException(status_code=403, detail="Access denied: you are not a member of this project")

def check_project_permission(db: Session, state: dict, project_id: int, user_id: int, permission_key: str):
    if is_admin_like_role(get_user_role(db, user_id)):
        return True

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    meta = get_project_meta(state, p.id)
    permissions = meta.get("permissions") or dict(DEFAULT_PERMISSIONS)
    perm_value = permissions.get(permission_key, "all")

    # owner pass
    if int(meta.get("owner_id") or 0) == int(user_id):
        return True

    if perm_value == "all":
        return True
    elif perm_value == "admin":
        raise HTTPException(status_code=403, detail=f"권한이 없습니다: {permission_key} 은(는) 관리자만 가능합니다")
    elif perm_value == "members_only":
        # C-3: Check DB members first
        db_member = db.query(ProjectMemberModel).filter(
            ProjectMemberModel.project_id == int(project_id),
            ProjectMemberModel.user_id == int(user_id),
        ).first()
        is_member = bool(db_member) or any(int(m.get("project_id")) == int(project_id) and int(m.get("user_id")) == int(user_id) for m in state.get("project_members", []))
        if is_member:
            return True
        raise HTTPException(status_code=403, detail=f"권한이 없습니다: {permission_key} 은(는) 프로젝트 담당자만 가능합니다")

    return True

# =========================
# Role / Super Admin Helpers
# =========================
ADMIN_ROLES = {"super_admin"}
SUPER_OWNER_LOGINID = SUPER_ADMIN_LOGINIDS  # from environment.py (env var)

def normalize_loginid(loginid: Optional[str]) -> str:
    return (loginid or "").strip().lower()

def is_admin_role(role: Optional[str]) -> bool:
    return (role or "").strip().lower() in ADMIN_ROLES

def is_super_owner_loginid(loginid: Optional[str]) -> bool:
    # NOTE: 원본 코드에 논리 오류가 있었는데(리스트 비교),
    # 사용자가 "그대로"를 원하므로 안전하게 문자열/리스트 모두 대응하도록 작성.
    if isinstance(SUPER_OWNER_LOGINID, list):
        return normalize_loginid(loginid) in [normalize_loginid(x) for x in SUPER_OWNER_LOGINID]
    return normalize_loginid(loginid) == normalize_loginid(SUPER_OWNER_LOGINID)

def get_user_role(db: Session, user_id: int) -> Optional[str]:
    u = db.query(User).filter(User.id == user_id).first()
    return u.role if u else None

def get_super_owner_user(db: Session) -> Optional[User]:
    # DB에 loginid가 소문자로 저장된다고 가정
    return db.query(User).filter(User.loginid.in_(SUPER_OWNER_LOGINID)).first()

def get_super_owner_id(db: Session) -> Optional[int]:
    u = get_super_owner_user(db)
    return u.id if u else None

def ensure_super_owner(db: Session) -> Optional[User]:
    """SUPER_ADMIN_LOGINIDS에 해당하는 계정이 있으면 super_admin + 활성 상태를 강제 보장."""
    u = db.query(User).filter(
        User.loginid.in_(SUPER_OWNER_LOGINID)
    ).first()
    if not u:
        return None

    changed = False

    if (u.role or "").strip().lower() != "super_admin":
        u.role = "super_admin"
        changed = True

    if not bool(u.is_active):
        u.is_active = True
        changed = True

    if changed:
        db.commit()
        db.refresh(u)

    return u

# =========================
# Root / Health
# =========================
@app.get("/")
def read_root():
    return {"message": "Welcome to Antigravity Schedule Platform API"}

def _health_payload():
    """헬스체크 + 배포 검증용 페이로드.

    회사 서버와 로컬이 같은 코드/같은 parser 를 실행 중인지 version/app_path/cwd 로
    바로 확인할 수 있게 한다. 민감정보는 담지 않는다.
    """
    return {
        "ok": True,
        "status": "ok",
        "version": APP_VERSION,
        "env": os.environ.get("APP_ENV") or os.environ.get("ENV") or "unknown",
        "app_path": os.path.abspath(__file__),
        "cwd": os.getcwd(),
    }

@app.get("/health")
def health_check():
    return _health_payload()

# nginx 운영 경로(`location /api/` 만 백엔드로 프록시)에서도 헬스체크가 닿도록
# `/api/health` 별칭을 둔다. 직접(backend) 점검은 /health, nginx 경유 점검은 /api/health.
@app.get("/api/health")
def api_health_check():
    return _health_payload()

@app.get("/api/files/sheet-images/{filename}")
def get_sheet_image(filename: str):
    """로컬에 저장된 시트 이미지 반환"""
    import os
    from fastapi.responses import FileResponse
    file_path = os.path.join(SHEET_IMAGES_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Image not found")
    return FileResponse(file_path)

# =========================
# Combined Data Endpoint
# =========================
@app.get("/api/data")
def get_all_data(user_id: Optional[int] = None, db: Session = Depends(get_db)):
    state = load_state()

    projects_db = db.query(Project).filter(Project.archived_at.is_(None)).all()
    users_db = db.query(User).all()
    tasks_db = db.query(Task).filter(Task.archived_at.is_(None)).all()
    prefs_rows = db.query(UserPreference).all()

    projects = [project_dict(p, state) for p in projects_db]
    users = [user_dict(u, state) for u in users_db]
    tasks = [task_dict(t, state) for t in tasks_db]

    # user_id 필터 (새 프론트 호환)
    if user_id:
        role = get_user_role(db, user_id)
        is_admin = is_admin_like_role(role)
        if not is_admin:
            authorized_pids = get_user_project_ids(db, state, user_id)
            # 같은 space 소속 공개 프로젝트도 허용
            authorized_pids |= get_user_public_project_ids(db, state, user_id)

            projects = [p for p in projects if p["id"] in authorized_pids]
            tasks = [
                t for t in tasks
                if t.get("project_id") in authorized_pids and (
                    not t.get("assignee_ids") or user_id in (t.get("assignee_ids") or [])
                )
            ]

    prefs = {}
    for pr in prefs_rows:
        prefs[str(pr.user_id)] = {"layout": pr.layout}

    return {
        # DB core
        "projects": projects,
        "users": users,
        "tasks": tasks,
        "activity_logs": [],  # 미구현 유지
        "user_preferences": prefs,

        # C-2: sub_projects from DB
        "sub_projects": get_subprojects_from_db(db),
        "notes": state.get("notes", []),
        "attachments": state.get("attachments", []),
        "project_members": state.get("project_members", []),
        "project_files": state.get("project_files", []),
        "join_requests": state.get("join_requests", []),
        "groups": state.get("groups", []),
        "shortcuts": state.get("shortcuts", []),

        # 레거시 호환용 (예전 main.py에서 쓰던 키)
        "roadmap_items": [],
    }

# =========================
# User Endpoints (DB + user_meta)
# =========================
@app.get("/api/users")
def get_users(db: Session = Depends(get_db)):
    state = load_state()
    users = db.query(User).all()
    return {"users": [user_dict(u, state) for u in users]}

@app.post("/api/users")
def create_user(user: UserCreate, db: Session = Depends(get_db)):
    state = load_state()

    exists = db.query(User).filter(User.loginid == user.loginid).first()
    if exists:
        raise HTTPException(status_code=400, detail="Login ID already exists")

    # mail이 없으면 loginid@samsung.com으로 자동 생성
    mail = user.mail or f"{user.loginid}@samsung.com"

    u = User(
        loginid=user.loginid,
        username=user.username,
        role=user.role or "member",
        avatar_color=user.avatar_color or "#2955FF",
        deptname=user.deptname,
        mail=mail,
        is_active=True,
    )
    db.add(u)
    db.commit()
    db.refresh(u)

    return user_dict(u, state)

@app.patch("/api/users/{user_id}")
def update_user(user_id: int, updates: UserUpdate, db: Session = Depends(get_db)):
    state = load_state()
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    data = updates.model_dump(exclude_unset=True)

    # DB fields
    for k in ["username", "role", "avatar_color", "is_active"]:
        if k in data:
            setattr(u, k, data[k])

    # sidecar field
    if "group_name" in data:
        set_user_meta(state, user_id, {"group_name": data["group_name"]})
        save_state(state)

    db.commit()
    db.refresh(u)
    return user_dict(u, state)

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db)):
    state = load_state()

    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    if (u.loginid or "").lower() in [x.lower() for x in SUPER_OWNER_LOGINID]:
        raise HTTPException(status_code=400, detail="Super owner 계정은 삭제할 수 없습니다.")

    # tasks.assignee_ids(JSON)에서 제거
    tasks = db.query(Task).filter(Task.archived_at.is_(None)).all()
    for t in tasks:
        ids = list(t.assignee_ids or [])
        if user_id in ids:
            ids.remove(user_id)
            t.assignee_ids = ids

    # preferences 삭제
    pref = db.query(UserPreference).filter(UserPreference.user_id == user_id).first()
    if pref:
        db.delete(pref)

    # sidecar 정리
    state["project_members"] = [m for m in state.get("project_members", []) if int(m.get("user_id")) != user_id]
    state["join_requests"] = [jr for jr in state.get("join_requests", []) if int(jr.get("user_id")) != user_id]
    state["notes"] = [n for n in state.get("notes", []) if int(n.get("author_id", -1)) != user_id]
    state.get("user_meta", {}).pop(str(user_id), None)

    db.delete(u)
    db.commit()
    save_state(state)
    return {"message": "User deleted"}

@app.post("/api/visit")
def write_visit_log(request: Request, db: Session = Depends(get_db)):
    user = try_get_user_from_token(request)
    ip = get_client_ip(request)
    uid = resolve_user_id_from_token(request, db)
    today_str = datetime.now(KST).strftime("%Y-%m-%d")

    # 동일 IP는 하루 1회만 기록
    existing = db.query(VisitLog).filter(
        VisitLog.ip_address == ip,
        VisitLog.visit_date == today_str,
    ).first()
    if existing:
        return {"message": "already logged today", "ip": ip}

    row = VisitLog(
        ip_address=ip,
        deptname=user.get("deptname"),
        username=user.get("username"),
        user_id=uid,
        visit_date=today_str,
    )
    db.add(row)
    db.commit()
    return {"message": "logged", "ip": ip}

# =========================
# User Preferences (DB)
# =========================
@app.get("/api/users/{user_id}/preferences")
def get_user_preferences(user_id: int, db: Session = Depends(get_db)):
    pref = db.query(UserPreference).filter(UserPreference.user_id == user_id).first()
    if not pref:
        return {"layout": None}
    return {"layout": pref.layout}

@app.put("/api/users/{user_id}/preferences/layout")
def save_user_layout(user_id: int, body: LayoutUpdate, db: Session = Depends(get_db)):
    pref = db.query(UserPreference).filter(UserPreference.user_id == user_id).first()
    if not pref:
        pref = UserPreference(user_id=user_id, layout=body.layout)
        db.add(pref)
    else:
        pref.layout = body.layout

    db.commit()
    return {"message": "Layout saved", "layout": body.layout}

@app.post("/api/users/{user_id}/hidden-projects/{project_id}")
def toggle_hidden_project(user_id: int, project_id: int, db: Session = Depends(get_db)):
    """사용자별 프로젝트 숨기기/보이기 토글"""
    pref = db.query(UserPreference).filter(UserPreference.user_id == user_id).first()
    if not pref:
        pref = UserPreference(user_id=user_id, layout={})
        db.add(pref)
        db.flush()

    layout = dict(pref.layout or {})
    hidden = list(layout.get("hidden_projects", []))

    if project_id in hidden:
        hidden.remove(project_id)
        action = "shown"
    else:
        hidden.append(project_id)
        action = "hidden"

    layout["hidden_projects"] = hidden
    pref.layout = layout
    db.commit()
    return {"action": action, "hidden_projects": hidden}

@app.get("/api/users/{user_id}/hidden-projects")
def get_hidden_projects(user_id: int, db: Session = Depends(get_db)):
    """사용자별 숨긴 프로젝트 목록"""
    pref = db.query(UserPreference).filter(UserPreference.user_id == user_id).first()
    hidden = (pref.layout or {}).get("hidden_projects", []) if pref else []
    return {"hidden_projects": hidden}

# =========================
# Project Endpoints (DB + project_meta + members sidecar)
# =========================
@app.get("/api/projects")
def get_projects(user_id: Optional[int] = None, space_id: Optional[int] = None, db: Session = Depends(get_db), _active: User = Depends(get_active_user)):
    state = load_state()
    q = db.query(Project).filter(Project.archived_at.is_(None))
    # Filter by space if provided
    if space_id:
        q = q.filter(Project.space_id == space_id)
    rows = q.all()
    projects = [project_dict(p, state) for p in rows]

    # 시스템 프로젝트([시스템] 단발 업무 등)는 일반 프로젝트 목록에서 숨긴다.
    # 단발 일정은 캘린더에서만 다루며 stats/overview 의 task 집계에는 그대로 포함된다.
    projects = [p for p in projects if not p.get("is_system")]

    if user_id:
        accessible_project_ids = get_user_project_ids(db, state, user_id)
        accessible_project_ids |= get_user_public_project_ids(db, state, user_id)
        projects = [
            p for p in projects
            if p["id"] in accessible_project_ids
        ]

    return {"projects": projects}

@app.post("/api/projects")
def create_project(
    project: ProjectCreate,
    suppress_realtime: bool = Query(False),
    db: Session = Depends(get_db),
):
    state = load_state()

    # 프로젝트는 반드시 유효한 공간에 소속되어야 함
    if not project.space_id:
        raise HTTPException(400, "프로젝트를 생성하려면 먼저 공간이 필요합니다.")
    space = db.query(Space).filter(Space.id == project.space_id, Space.is_active == True).first()
    if not space:
        raise HTTPException(400, "유효하지 않은 공간입니다. 프로젝트를 생성하려면 먼저 공간이 필요합니다.")

    # C-4: Calculate owner before creating project
    default_owner_id = project.owner_id
    if not default_owner_id:
        default_owner_id = get_super_owner_id(db) or 1

    p = Project(
        name=project.name,
        description=project.description,
        owner_id=default_owner_id,
        created_by=default_owner_id,
        space_id=project.space_id,
    )
    db.add(p)
    db.commit()
    db.refresh(p)

    set_project_meta(state, p.id, {
        "owner_id": default_owner_id,
        "visibility": project.visibility or "private",
        "require_approval": bool(project.require_approval or False),
        "permissions": project.permissions or dict(DEFAULT_PERMISSIONS),
    })

    # C-3: Owner membership via DB
    ensure_owner_membership(state, p.id, default_owner_id, db=db)

    for mid in (project.member_ids or []):
        if mid == default_owner_id:
            continue
        existing = db.query(ProjectMemberModel).filter(
            ProjectMemberModel.project_id == p.id,
            ProjectMemberModel.user_id == int(mid),
        ).first()
        if not existing:
            mu = db.query(User).filter(User.id == int(mid)).first()
            pm = ProjectMemberModel(
                project_id=p.id, user_id=mid, role="member",
                loginid=mu.loginid if mu else None,
                deptname=getattr(mu, "deptname", None) if mu else None,
            )
            db.add(pm)

    # 대량 import 중에는 개별 이벤트를 만들지 않고, import 완료 후 bulk event 1개만 발행한다.
    if not suppress_realtime:
        emit_realtime_event(
            db, space_id=p.space_id, event_type="project_created",
            entity_type="project", entity_id=p.id, project_id=p.id,
            actor_user_id=default_owner_id,
        )
    db.commit()

    save_state(state)
    return project_dict(p, state)

@app.patch("/api/projects/{project_id}")
def update_project(
    project_id: int,
    updates: ProjectUpdate,
    caller_user_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    state = load_state()
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    data = updates.model_dump(exclude_unset=True)

    # 프로젝트 관리 권한(소유자 OR 담당자 OR super_admin/admin)
    require_project_management(db, state, project_id, caller_user_id)

    _proj_old = {"name": p.name, "description": p.description}
    if "name" in data:
        p.name = data["name"]
    if "description" in data:
        p.description = data["description"]

    # sidecar meta — owner_id 변경은 소유자/관리자만 (담당자는 소유자 양도 불가)
    meta_updates = {}
    for k in ["visibility", "require_approval", "permissions", "owner_id"]:
        if k in data:
            meta_updates[k] = data[k]
    if "owner_id" in meta_updates:
        meta = get_project_meta(state, project_id)
        caller = db.query(User).filter(User.id == caller_user_id).first() if caller_user_id else None
        is_super_admin = caller and caller.role == "super_admin"
        if not is_super_admin and int(meta.get("owner_id") or 0) != int(caller_user_id or 0):
            raise HTTPException(status_code=403, detail="소유자 이전은 현재 소유자만 가능합니다.")
    if meta_updates:
        set_project_meta(state, project_id, meta_updates)
        if "owner_id" in meta_updates:
            ensure_owner_membership(state, project_id, meta_updates["owner_id"], db=db)

    # 변경 이력 기록 (이름/설명 + 소유자 변경)
    _pchanges, _paction, _pmsg = actlog.diff_changes(
        _proj_old, data, ["name", "description"],
    )
    if "owner_id" in meta_updates:
        _pchanges = (_pchanges or []) + [{
            "field": "owner_id", "label": "소유자",
            "before": None, "after": meta_updates.get("owner_id"),
        }]
        _pmsg = (_pmsg + ", 소유자 변경") if _pmsg else "소유자 변경"
    if _pchanges:
        actlog.log_activity(
            db, user_id=caller_user_id, action=(_paction or actlog.ACTION_UPDATED),
            entity_type=actlog.ENTITY_PROJECT, entity_id=p.id,
            message=_pmsg, meta={"changes": _pchanges},
        )

    emit_realtime_event(
        db, space_id=p.space_id, event_type="project_updated",
        entity_type="project", entity_id=p.id, project_id=p.id,
        actor_user_id=caller_user_id,
    )
    db.commit()
    db.refresh(p)
    save_state(state)
    return project_dict(p, state)

@app.delete("/api/projects/{project_id}")
def delete_project(project_id: int, request: Request, db: Session = Depends(get_db)):
    state = load_state()
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    # 프로젝트 삭제: 소유자 또는 담당자(또는 admin/super_admin)만 가능
    caller_info = try_get_user_from_token(request)
    if caller_info:
        caller_loginid = caller_info.get("loginid", "")
        caller = db.query(User).filter(User.loginid == caller_loginid).first()
        meta = get_project_meta(state, project_id)
        owner_id = p.owner_id or meta.get("owner_id")
        if owner_id and caller:
            if not can_manage_project(db, state, project_id, caller.id):
                raise HTTPException(status_code=403, detail="프로젝트 삭제 권한이 없습니다. (소유자 또는 담당자만 가능)")

    now = datetime.now()
    _del_actor = caller.id if (caller_info and caller) else None
    p.archived_at = now
    p.deleted_by_user_id = _del_actor

    # 관련 task soft delete
    tasks = db.query(Task).filter(Task.project_id == project_id).all()
    for t in tasks:
        if not t.archived_at:
            t.archived_at = now
            t.deleted_by_user_id = _del_actor

    actlog.log_activity(
        db, user_id=_del_actor, action=actlog.ACTION_DELETED,
        entity_type=actlog.ENTITY_PROJECT, entity_id=p.id,
        message="프로젝트 삭제", meta={"name": p.name},
    )
    emit_realtime_event(
        db, space_id=p.space_id, event_type="project_deleted",
        entity_type="project", entity_id=p.id, project_id=p.id,
        actor_user_id=_del_actor,
    )
    db.commit()
    return {"message": "Project deleted"}

@app.post("/api/projects/{project_id}/restore")
def restore_project(project_id: int, request: Request = None, db: Session = Depends(get_db)):
    p = db.query(Project).filter(Project.id == project_id).first()
    if not p or not p.archived_at:
        raise HTTPException(status_code=404, detail="Archived project not found")

    # 본인이 삭제한 항목만 복원 가능 (admin/super_admin은 전체 허용)
    caller_id = resolve_user_id_from_token(request, db) if request else None
    caller_role = get_user_role(db, caller_id) if caller_id else None
    is_admin = caller_role in ("admin", "super_admin")
    if not is_admin and (p.deleted_by_user_id is None or p.deleted_by_user_id != caller_id):
        raise HTTPException(status_code=403, detail="본인이 삭제한 항목만 복원할 수 있습니다.")

    p.archived_at = None
    p.deleted_by_user_id = None
    # 관련 task도 함께 복원
    for t in db.query(Task).filter(Task.project_id == project_id).all():
        if t.archived_at:
            t.archived_at = None
            t.deleted_by_user_id = None
    db.commit()
    return {"message": "Project restored"}

@app.get("/api/trash")
def get_trash(request: Request = None, scope: str = Query("mine"), db: Session = Depends(get_db)):
    """Trash Page 조회.

    기본(scope=mine): 현재 로그인 사용자가 삭제한 project/task만 반환한다.
    다른 사용자가 삭제한 항목, 그리고 deleted_by_user_id 가 NULL 인 legacy 삭제 항목은
    일반 사용자에게 노출하지 않는다(backend 레벨 차단).
    scope=all: admin/super_admin 전용. 전체 삭제 항목을 반환한다.
    """
    state = load_state()
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if not caller_id:
        # 인증 정보가 없으면 어떤 삭제 항목도 노출하지 않는다.
        return {"projects": [], "tasks": [], "scope": "mine"}

    caller_role = get_user_role(db, caller_id)
    is_admin = caller_role in ("admin", "super_admin")
    effective_scope = "all" if (scope == "all" and is_admin) else "mine"

    pq = db.query(Project).filter(Project.archived_at.isnot(None))
    tq = db.query(Task).filter(Task.archived_at.isnot(None))
    if effective_scope == "mine":
        pq = pq.filter(Project.deleted_by_user_id == caller_id)
        tq = tq.filter(Task.deleted_by_user_id == caller_id)
    archived_projects = pq.all()
    archived_tasks = tq.all()

    projects_out = []
    for p in archived_projects:
        task_count = db.query(Task).filter(Task.project_id == p.id).count()
        meta = get_project_meta(state, p.id)
        projects_out.append({
            "id": p.id,
            "name": p.name,
            "description": p.description,
            "archived_at": iso(p.archived_at),
            "task_count": task_count,
            "owner_id": meta.get("owner_id"),
        })

    # 개별 삭제된 task (프로젝트 자체는 살아있는 경우)
    active_project_ids = {p.id for p in db.query(Project).filter(Project.archived_at.is_(None)).all()}
    tasks_out = []
    for t in archived_tasks:
        if t.project_id in active_project_ids:
            tasks_out.append({
                "id": t.id,
                "title": t.title,
                "project_id": t.project_id,
                "archived_at": iso(t.archived_at),
            })

    return {"projects": projects_out, "tasks": tasks_out, "scope": effective_scope}

def _purge_expired_trash(db: Session):
    """3일 지난 archived 항목 영구 삭제."""
    cutoff = datetime.now() - timedelta(days=3)
    state = load_state()
    changed = False

    # 프로젝트 영구 삭제
    expired_projects = db.query(Project).filter(
        Project.archived_at.isnot(None), Project.archived_at < cutoff
    ).all()
    for p in expired_projects:
        pid = p.id
        # 관련 task 영구 삭제
        db.query(TaskActivityModel).filter(
            TaskActivityModel.task_id.in_(
                db.query(Task.id).filter(Task.project_id == pid)
            )
        ).delete(synchronize_session=False)
        db.query(AttachmentModel).filter(
            AttachmentModel.task_id.in_(
                db.query(Task.id).filter(Task.project_id == pid)
            )
        ).delete(synchronize_session=False)
        # Task 삭제 전 SheetExecution.task_id 연결 해제 (시트 실행 이력은 보존)
        # FK가 ON DELETE SET NULL 로 설정되지 않은 환경에서도 Task bulk delete가 막히지 않도록 수동 NULL 처리
        db.query(SheetExecution).filter(
            SheetExecution.task_id.in_(
                db.query(Task.id).filter(Task.project_id == pid)
            )
        ).update({SheetExecution.task_id: None}, synchronize_session=False)
        # Project 삭제 전 SheetExecution.project_id 도 NULL 처리 (이력 보존, FK 위반 방지)
        # fk_sheet_executions_project 가 ON DELETE SET NULL 로 설정되지 않은 환경 보호.
        db.query(SheetExecution).filter(
            SheetExecution.project_id == pid
        ).update({SheetExecution.project_id: None}, synchronize_session=False)
        db.query(Task).filter(Task.project_id == pid).delete(synchronize_session=False)
        db.query(SubProjectModel).filter(SubProjectModel.project_id == pid).delete()
        project_notes = db.query(NoteModel).filter(NoteModel.project_id == pid).all()
        for pn in project_notes:
            db.query(NoteMention).filter(NoteMention.note_id == pn.id).delete()
        db.query(NoteModel).filter(NoteModel.project_id == pid).delete()
        db.query(ProjectMemberModel).filter(ProjectMemberModel.project_id == pid).delete()

        # sidecar cleanup
        for key in ["sub_projects", "notes", "project_members", "join_requests"]:
            state[key] = [x for x in state.get(key, []) if int(x.get("project_id", 0)) != pid]
        for pf in [f for f in state.get("project_files", []) if int(f.get("project_id", 0)) == pid]:
            fp = os.path.join(UPLOAD_DIR, str(pid), pf.get("stored_name", ""))
            if os.path.exists(fp):
                try: os.remove(fp)
                except Exception: pass
        state["project_files"] = [f for f in state.get("project_files", []) if int(f.get("project_id", 0)) != pid]
        state.get("project_meta", {}).pop(str(pid), None)

        db.delete(p)
        changed = True

    # 개별 task 영구 삭제 (프로젝트는 살아있지만 task만 삭제된 경우)
    expired_tasks = db.query(Task).filter(
        Task.archived_at.isnot(None), Task.archived_at < cutoff
    ).all()
    for t in expired_tasks:
        db.query(TaskActivityModel).filter(TaskActivityModel.task_id == t.id).delete()
        db.query(AttachmentModel).filter(AttachmentModel.task_id == t.id).delete()
        # SheetExecution 이력은 보존하고 task_id 만 NULL 로 연결 해제
        db.query(SheetExecution).filter(SheetExecution.task_id == t.id).update(
            {SheetExecution.task_id: None}, synchronize_session=False
        )
        db.delete(t)
        changed = True

    db.commit()
    if changed:
        save_state(state)


def _space_activity(db: Session, space_id: int) -> dict:
    """단일 공간의 활동 카운트/빈 공간 판정.

    - project_count: 시스템 프로젝트([시스템] 단발 업무) 제외한 '실제' 프로젝트 수
    - task_count: 공간 내 모든 프로젝트(시스템 포함, 단발 일정 task 포함)의 비보관 task 수
    - checksheet_count: SheetExecution(space_id)
    - calendar_count: CalendarEvent(space_id, 비보관)
    - is_empty(완전 빈 공간): 위 4개 카운트가 모두 0
    - has_project: 실제 프로젝트가 1개 이상
    VOC 는 공간에 묶이지 않는 전역 데이터라 카운트하지 않는다.
    """
    project_count = db.query(func.count(Project.id)).filter(
        Project.space_id == space_id,
        Project.archived_at.is_(None),
        Project.name != ONEOFF_PROJECT_NAME,
    ).scalar() or 0

    task_count = db.query(func.count(Task.id)).join(
        Project, Task.project_id == Project.id
    ).filter(
        Project.space_id == space_id,
        Task.archived_at.is_(None),
    ).scalar() or 0

    checksheet_count = db.query(func.count(SheetExecution.id)).filter(
        SheetExecution.space_id == space_id,
    ).scalar() or 0

    calendar_count = db.query(func.count(CalendarEvent.id)).filter(
        CalendarEvent.space_id == space_id,
        CalendarEvent.archived_at.is_(None),
    ).scalar() or 0

    member_count = db.query(func.count(SpaceMember.id)).filter(
        SpaceMember.space_id == space_id,
    ).scalar() or 0

    # 마지막 활동일: 공간 생성일 + 각 엔티티 최신 시각의 최대값 (Phase 1 집계 방식)
    candidates = []
    space_obj = db.query(Space).filter(Space.id == space_id).first()
    if space_obj and space_obj.created_at:
        candidates.append(space_obj.created_at)
    for val in (
        db.query(func.max(Project.created_at)).filter(Project.space_id == space_id).scalar(),
        db.query(func.max(Task.updated_at)).join(Project, Task.project_id == Project.id).filter(Project.space_id == space_id).scalar(),
        db.query(func.max(SheetExecution.started_at)).filter(SheetExecution.space_id == space_id).scalar(),
        db.query(func.max(CalendarEvent.updated_at)).filter(CalendarEvent.space_id == space_id).scalar(),
    ):
        if val is not None:
            candidates.append(val)
    last_activity_at = max(candidates) if candidates else None

    is_empty = (project_count == 0 and task_count == 0 and checksheet_count == 0 and calendar_count == 0)
    return {
        "project_count": int(project_count),
        "task_count": int(task_count),
        "checksheet_count": int(checksheet_count),
        "calendar_count": int(calendar_count),
        "member_count": int(member_count),
        "last_activity_at": last_activity_at,
        "is_empty": is_empty,
        "has_project": project_count > 0,
    }


def _is_system_space(space: Space) -> bool:
    """자동 정리에서 제외할 시스템/기본 공간 여부."""
    return (space.slug or "") in ("general",)


def _cleanup_empty_spaces(db: Session):
    """완전 빈 공간 경고 → 자동 보관(soft) lifecycle.

    안전 정책 (사용자 확정):
    - '빈 공간' = 완전 빈 공간(project/task/checksheet/calendar 모두 0). 프로젝트만 없는 건 대상 아님.
    - 생성 후 SPACE_EMPTY_WARNING_DAYS 경과 → warned_at 기록 + 삭제 예정일 설정 + 로그.
    - warned_at 후 (AUTO_ARCHIVE_DAYS - WARNING_DAYS) 경과 → 자동 보관(is_active=False, archived_at).
      ⚠️ 멤버(SpaceMember)는 절대 자동 삭제하지 않는다. (복구 가능하도록 보존)
    - 자동 hard-delete 는 SPACE_EMPTY_AUTO_DELETE_ENABLED=True 일 때만. 기본 False → 자동 삭제 없음.
    - cleanup_exempt / 시스템 공간은 건너뛴다.
    """
    from app.services.system_log import log_event

    now = datetime.utcnow()
    warn_cutoff = now - timedelta(days=SPACE_EMPTY_WARNING_DAYS)
    archive_after = max(0, SPACE_EMPTY_AUTO_ARCHIVE_DAYS - SPACE_EMPTY_WARNING_DAYS)
    archive_cutoff = now - timedelta(days=archive_after)

    active_spaces = db.query(Space).filter(Space.is_active == True).all()
    for space in active_spaces:
        if _is_system_space(space) or bool(getattr(space, "cleanup_exempt", False)):
            continue

        act = _space_activity(db, space.id)
        # 마지막 활동일 컬럼 갱신(집계값 캐시)
        space.last_activity_at = act["last_activity_at"]

        if not act["is_empty"]:
            # 활동이 생기면 경고/삭제예정 해제
            if space.warned_at is not None:
                space.warned_at = None
            if space.delete_scheduled_at is not None:
                space.delete_scheduled_at = None
            continue

        # 완전 빈 공간
        if not (space.created_at and space.created_at < warn_cutoff):
            continue  # 아직 경고 기간 미도래

        owner = db.query(SpaceMember).filter(
            SpaceMember.space_id == space.id, SpaceMember.role == "owner"
        ).first()
        owner_id = owner.user_id if owner else None
        days_empty = (now - space.created_at).days if space.created_at else None

        if space.warned_at is None:
            # 1) 경고 기록
            space.warned_at = now
            space.delete_scheduled_at = now + timedelta(days=SPACE_EMPTY_DELETE_AFTER_DAYS)
            log_event(
                "INFO", "SPACE", "Empty space warning sent",
                detail=f"space_id={space.id} owner_id={owner_id} days_empty={days_empty}",
            )
        elif space.warned_at < archive_cutoff:
            # 2) 자동 보관 (멤버 보존)
            space.is_active = False
            space.archived_at = now
            log_event(
                "WARNING", "SPACE", "Empty space archived",
                detail=f"space_id={space.id} owner_id={owner_id} days_empty={days_empty}",
            )
            # 3) 자동 hard-delete: 기본 비활성. 켜져 있어도 보관 직후가 아닌, 별도 검증 후에만.
            #    (Phase 1 정책상 자동 삭제는 수행하지 않는다.)

    db.commit()

# =========================
# Project Members / Join Requests (C-3: DB + sidecar fallback)
# =========================
@app.get("/api/projects/{project_id}/members")
def get_project_members(project_id: int, assignable_only: bool = False, db: Session = Depends(get_db)):
    state = load_state()
    members = get_members_for_project(state, project_id, db=db)
    users_map = {u.id: u for u in db.query(User).all()}

    enriched = []
    for m in members:
        uid = int(m.get("user_id"))
        u = users_map.get(uid)
        # assignable_only: viewer cannot be assigned to tasks
        if assignable_only and m.get("role") == "viewer":
            continue
        enriched.append({
            **m,
            "username": u.username if u else "Unknown",
            "avatar_color": (u.avatar_color if u else "#ccc"),
            "deptname": u.deptname if u else None,
            "mail": u.mail if u else None,
            "loginid": u.loginid if u else None,
        })
    return {"members": enriched}

@app.post("/api/projects/{project_id}/members")
def add_project_member(
    project_id: int,
    member: MemberAdd,
    caller_user_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    state = load_state()

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    u = db.query(User).filter(User.id == member.user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    meta = get_project_meta(state, project_id)
    require_approval = bool(meta.get("require_approval", False))

    # 자기 자신을 참여요청 모드로 등록하는 경우는 권한 체크 제외 (기존 join-request 경로 호환)
    is_self_request = require_approval and int(caller_user_id or 0) == int(member.user_id)
    if not is_self_request:
        require_project_management(db, state, project_id, caller_user_id)

    # 공간 멤버십 검증: 대상 사용자는 프로젝트가 속한 공간의 멤버여야 함
    if p.space_id:
        in_space = db.query(SpaceMember).filter(
            SpaceMember.space_id == p.space_id,
            SpaceMember.user_id == member.user_id,
        ).first()
        if not in_space:
            raise HTTPException(
                status_code=400,
                detail="이 사용자는 해당 공간의 멤버가 아니므로 프로젝트에 추가할 수 없습니다.",
            )

    # C-3: Check DB for existing membership
    existing = db.query(ProjectMemberModel).filter(
        ProjectMemberModel.project_id == project_id,
        ProjectMemberModel.user_id == member.user_id,
    ).first()
    if existing:
        return {"message": "Already a member", "status": "exists"}

    if require_approval:
        join_requests = state.get("join_requests", [])
        if any(
            int(jr.get("project_id")) == project_id and
            int(jr.get("user_id")) == member.user_id and
            jr.get("status") == "pending"
            for jr in join_requests
        ):
            raise HTTPException(status_code=400, detail="이미 참여 요청이 있습니다")

        new_request = {
            "id": next_id(join_requests),
            "project_id": project_id,
            "user_id": member.user_id,
            "role": member.role or "member",
            "status": "pending",
            "created_at": datetime.now().isoformat(),
        }
        join_requests.append(new_request)
        state["join_requests"] = join_requests
        save_state(state)
        return {"message": "참여 요청이 등록되었습니다. 관리자 승인 후 참여 가능합니다.", "status": "pending"}

    # C-3: Add to DB (denormalize loginid/deptname from user)
    pm = ProjectMemberModel(
        project_id=project_id,
        user_id=member.user_id,
        role=member.role or "member",
        loginid=u.loginid,
        deptname=getattr(u, "deptname", None),
    )
    db.add(pm)
    db.commit()
    # Also keep sidecar in sync for backward compat
    sidecar_members = state.get("project_members", [])
    already_in_sidecar = any(
        int(m.get("project_id")) == project_id and int(m.get("user_id")) == member.user_id
        for m in sidecar_members
    )
    if not already_in_sidecar:
        sidecar_members.append({"project_id": project_id, "user_id": member.user_id, "role": member.role or "member"})
        state["project_members"] = sidecar_members
        save_state(state)
    return {"message": "Member added"}

@app.patch("/api/projects/{project_id}/members/{target_user_id}/role")
def update_project_member_role(
    project_id: int, target_user_id: int,
    body: dict, user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """프로젝트 내 멤버 역할(manager/member/viewer) 변경. 소유자/담당자/super_admin만 가능."""
    state = load_state()
    caller = db.query(User).filter(User.id == user_id).first()
    if not caller:
        raise HTTPException(status_code=403, detail="권한 없음")
    if not can_manage_project(db, state, project_id, user_id):
        raise HTTPException(status_code=403, detail="프로젝트 역할 변경 권한이 없습니다.")

    new_role = body.get("role", "member")
    if new_role not in ("member", "manager", "viewer"):
        raise HTTPException(status_code=400, detail="유효하지 않은 역할입니다.")

    pm = db.query(ProjectMemberModel).filter(
        ProjectMemberModel.project_id == project_id,
        ProjectMemberModel.user_id == target_user_id,
    ).first()
    if not pm:
        raise HTTPException(status_code=404, detail="프로젝트 멤버가 아닙니다.")
    if pm.role == "owner":
        raise HTTPException(status_code=400, detail="Owner 역할은 변경할 수 없습니다.")
    pm.role = new_role
    db.commit()
    return {"message": f"역할이 {new_role}로 변경되었습니다."}

@app.delete("/api/projects/{project_id}/members/{user_id}")
def remove_project_member(
    project_id: int,
    user_id: int,
    caller_user_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    state = load_state()
    # 자기 자신 탈퇴는 허용, 그 외에는 소유자/담당자만 가능
    if not (caller_user_id and int(caller_user_id) == int(user_id)):
        require_project_management(db, state, project_id, caller_user_id)

    # 소유자(project.owner_id)는 멤버 제거로 내릴 수 없음
    meta = get_project_meta(state, project_id)
    if int(meta.get("owner_id") or 0) == int(user_id):
        raise HTTPException(status_code=400, detail="프로젝트 소유자는 제거할 수 없습니다.")

    # C-3: Remove from DB
    db.query(ProjectMemberModel).filter(
        ProjectMemberModel.project_id == project_id,
        ProjectMemberModel.user_id == user_id,
    ).delete()
    db.commit()
    # Also remove from sidecar
    state["project_members"] = [
        m for m in state.get("project_members", [])
        if not (int(m.get("project_id")) == project_id and int(m.get("user_id")) == user_id)
    ]
    save_state(state)
    return {"message": "Member removed"}

@app.post("/api/projects/{project_id}/join-request")
def request_join(project_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    state = load_state()
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    if any(int(m.get("project_id")) == project_id and int(m.get("user_id")) == user_id for m in state.get("project_members", [])):
        raise HTTPException(status_code=400, detail="이미 프로젝트 멤버입니다")

    join_requests = state.get("join_requests", [])
    if any(int(jr.get("project_id")) == project_id and int(jr.get("user_id")) == user_id and jr.get("status") == "pending" for jr in join_requests):
        raise HTTPException(status_code=400, detail="이미 참여 요청이 있습니다")

    new_request = {
        "id": next_id(join_requests),
        "project_id": project_id,
        "user_id": user_id,
        "role": "member",
        "status": "pending",
        "created_at": datetime.now().isoformat(),
    }
    join_requests.append(new_request)
    state["join_requests"] = join_requests
    save_state(state)
    return {"message": "참여 요청이 등록되었습니다", "request": new_request}

@app.get("/api/projects/{project_id}/join-requests")
def get_join_requests(project_id: int, db: Session = Depends(get_db)):
    state = load_state()
    reqs = [jr.copy() for jr in state.get("join_requests", []) if int(jr.get("project_id")) == project_id]
    users_map = {u.id: u for u in db.query(User).all()}
    for jr in reqs:
        u = users_map.get(int(jr.get("user_id")))
        jr["username"] = u.username if u else "Unknown"
        jr["avatar_color"] = u.avatar_color if u else "#ccc"
    return {"join_requests": reqs}

@app.post("/api/projects/{project_id}/join-requests/approve")
def approve_join_request(project_id: int, body: MemberApproval):
    state = load_state()
    join_requests = state.get("join_requests", [])

    target = None
    for jr in join_requests:
        if int(jr.get("project_id")) == project_id and int(jr.get("user_id")) == body.user_id and jr.get("status") == "pending":
            target = jr
            break

    if not target:
        raise HTTPException(status_code=404, detail="참여 요청을 찾을 수 없습니다")

    if body.action == "approve":
        target["status"] = "approved"
        members = state.get("project_members", [])
        exists = any(int(m.get("project_id")) == project_id and int(m.get("user_id")) == body.user_id for m in members)
        if not exists:
            members.append({
                "project_id": project_id,
                "user_id": body.user_id,
                "role": target.get("role", "member"),
            })
        state["project_members"] = members
        state["join_requests"] = join_requests
        save_state(state)
        return {"message": "참여가 승인되었습니다"}

    elif body.action == "reject":
        target["status"] = "rejected"
        state["join_requests"] = join_requests
        save_state(state)
        return {"message": "참여가 거부되었습니다"}

    raise HTTPException(status_code=400, detail="action must be 'approve' or 'reject'")

# =========================
# Task Endpoints (DB + task_meta)
# =========================
@app.get("/api/tasks")
def get_tasks(
    project_id: Optional[int] = None,
    assignee_id: Optional[int] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _active: User = Depends(get_active_user),
):
    state = load_state()

    q = db.query(Task).filter(Task.archived_at.is_(None))
    if project_id:
        q = q.filter(Task.project_id == project_id)

    rows = q.all()
    tasks = [task_dict(t, state) for t in rows]

    if assignee_id:
        tasks = [t for t in tasks if assignee_id in (t.get("assignee_ids") or [])]

    # user 권한 기반 필터: 멤버인 프로젝트의 모든 task + public 프로젝트 task
    if user_id:
        user_project_ids = get_user_project_ids(db, state, user_id)
        user_project_ids |= get_user_public_project_ids(db, state, user_id)

        tasks = [
            t for t in tasks
            if t.get("project_id") in user_project_ids
        ]

    # Add attachment_count from DB
    task_ids = [t["id"] for t in tasks]
    if task_ids:
        att_counts = dict(
            db.query(AttachmentModel.task_id, func.count(AttachmentModel.id))
            .filter(AttachmentModel.task_id.in_(task_ids))
            .group_by(AttachmentModel.task_id)
            .all()
        )
        for t in tasks:
            t["attachment_count"] = att_counts.get(t["id"], 0)

        # 하위(Child) Task 수 — Board 부모 카드의 "하위 N개" 뱃지용 (배치 집계로 N+1 방지)
        child_counts = dict(
            db.query(Task.parent_task_id, func.count(Task.id))
            .filter(Task.parent_task_id.in_(task_ids), Task.archived_at.is_(None))
            .group_by(Task.parent_task_id)
            .all()
        )
        for t in tasks:
            t["child_task_count"] = child_counts.get(t["id"], 0)

        # 현재 단계 완료 선언 여부 — Board/TaskDrawer "단계 완료" 표시용
        comp_rows = db.query(
            TaskStageCompletion.task_id, TaskStageCompletion.stage_id,
        ).filter(
            TaskStageCompletion.task_id.in_(task_ids),
            TaskStageCompletion.completed == True,   # noqa: E712
        ).all()
        comp_set = {(tid, sid) for tid, sid in comp_rows}
        for t in tasks:
            t["current_stage_completed"] = (t["id"], t.get("workflow_column_id")) in comp_set
    else:
        for t in tasks:
            t["attachment_count"] = 0
            t["child_task_count"] = 0
            t["current_stage_completed"] = False

    # 담당자/서브프로젝트 표시명 enrich (Board 검색이 별도 로딩 없이 task payload 만으로 동작).
    # 화면(TaskCard 등)에 이미 노출되는 수준의 이름/login_id 만 포함, 배치 조회로 N+1 방지.
    assignee_id_set = set()
    sub_pid_set = set()
    for t in tasks:
        for aid in (t.get("assignee_ids") or []):
            assignee_id_set.add(int(aid))
        if t.get("sub_project_id"):
            sub_pid_set.add(int(t["sub_project_id"]))
    user_name_map: dict[int, tuple] = {}
    if assignee_id_set:
        for u in db.query(User).filter(User.id.in_(assignee_id_set)).all():
            user_name_map[u.id] = (u.username, u.loginid)
    sub_name_map: dict[int, str] = {}
    if sub_pid_set:
        for sp in db.query(SubProjectModel).filter(SubProjectModel.id.in_(sub_pid_set)).all():
            sub_name_map[sp.id] = sp.name
    for t in tasks:
        names, logins = [], []
        for aid in (t.get("assignee_ids") or []):
            nm = user_name_map.get(int(aid))
            if nm:
                if nm[0]:
                    names.append(nm[0])
                if nm[1]:
                    logins.append(nm[1])
        t["assignee_names"] = names
        t["assignee_login_ids"] = logins
        spid = t.get("sub_project_id")
        t["sub_project_name"] = sub_name_map.get(int(spid)) if spid else None

    # CUSTOM 워크플로우 라벨 enrich (N+1 방지 — 관련 프로젝트의 active 컬럼을 단일 쿼리로 조회)
    proj_ids = {t.get("project_id") for t in tasks if t.get("project_id")}
    if proj_ids:
        col_rows = (
            db.query(ProjectTaskColumn)
            .filter(ProjectTaskColumn.project_id.in_(proj_ids), ProjectTaskColumn.active == True)
            .all()
        )
        col_by_id = {c.id: c for c in col_rows}
        for t in tasks:
            col = col_by_id.get(t.get("workflow_column_id"))
            if col is not None:
                t["workflow_label"] = col.label
                t["workflow_mapped_status"] = col.mapped_status
                t["workflow_color"] = col.color
            else:
                t["workflow_label"] = None
                t["workflow_mapped_status"] = None
                t["workflow_color"] = None

    return {"tasks": tasks}

def _task_board_key(task) -> tuple:
    """Board 단계 비교용 키.

    커스텀 워크플로우면 workflow_column_id 로, 아니면 canonical status 로 비교한다.
    (mapped_status 만 비교하면 커스텀 모드에서 서로 다른 컬럼이 같게 판정될 수 있으므로
     실제 컬럼 id 를 우선한다.)
    """
    col = getattr(task, "workflow_column_id", None)
    if col is not None:
        return ("col", int(col))
    return ("status", getattr(task, "status", None))


def _get_task_descendant_ids(db: Session, task_id: int) -> set:
    """task_id 의 모든 하위(자식/손자/...) Task id 집합 (task_id 자신은 제외).

    다단계 트리를 BFS 로 내려가며 수집한다. 잘못된 기존 데이터로 사이클이 있어도
    visited 로 무한루프를 방어한다. 순환 차단 검증과 단계/Sub Project 전파에 재사용.
    """
    descendants: set = set()
    visited: set = {int(task_id)}
    frontier = [int(task_id)]
    while frontier:
        rows = db.query(Task.id).filter(
            Task.parent_task_id.in_(frontier), Task.archived_at.is_(None)
        ).all()
        next_frontier = []
        for (cid,) in rows:
            cid = int(cid)
            if cid in visited:
                continue
            visited.add(cid)
            descendants.add(cid)
            next_frontier.append(cid)
        frontier = next_frontier
    return descendants


def _validate_parent_task(
    db: Session,
    state: dict,
    project_id: int,
    task_id: Optional[int],
    parent_id: int,
    child_sub_project_id: Optional[int],
    child_board_key: Optional[tuple] = None,
) -> Task:
    """상위 Task(Parent) 지정 검증 — 다단계 허용/순환/교차 프로젝트/같은 단계/Sub Project 방지. 실패 시 HTTPException(400).

    다단계(재귀형) 계층을 허용한다: 하위 Task 도 다시 다른 Task 의 상위가 될 수 있다.
    - parent_id == task_id → 자기 자신 불가
    - parent_id 가 task_id 의 하위(자식/손자/...)면 순환이므로 불가
    - parent 존재 + 미보관(archived_at is None) + 같은 project
    - 같은 Board 단계(child_board_key 지정 시)
    - Sub Project 조건: 상/하위 둘 다 Sub Project 가 있으면 반드시 같은 Sub Project
      (한쪽이라도 없으면 허용). Sub Project 값은 사이드카 meta 기준.
    """
    if task_id is not None and int(parent_id) == int(task_id):
        raise HTTPException(status_code=400, detail="자신 또는 자신의 하위 작업을 상위 작업으로 지정할 수 없습니다.")
    parent = db.query(Task).filter(Task.id == parent_id, Task.archived_at.is_(None)).first()
    if not parent:
        raise HTTPException(status_code=400, detail="상위 Task를 찾을 수 없습니다.")
    if parent.project_id != project_id:
        raise HTTPException(status_code=400, detail="상위 Task는 같은 프로젝트의 Task여야 합니다.")
    # 순환 차단: parent 가 이 Task 의 하위(자식/손자/...) 면 A→B→C→A 순환이 되므로 금지.
    if task_id is not None and int(parent_id) in _get_task_descendant_ids(db, int(task_id)):
        raise HTTPException(
            status_code=400,
            detail="자신 또는 자신의 하위 작업을 상위 작업으로 지정할 수 없습니다.",
        )
    # 같은 Board 단계(컬럼/상태) 조건 — 다른 단계의 Task 는 상/하위로 묶을 수 없음.
    if child_board_key is not None and _task_board_key(parent) != child_board_key:
        raise HTTPException(
            status_code=400,
            detail="상위 작업은 같은 Board 단계에 있는 작업만 선택할 수 있습니다.",
        )
    # Sub Project 조건 (둘 다 있으면 같아야 함)
    parent_sub_project_id = get_task_meta(state, parent.id).get("sub_project_id")
    if child_sub_project_id is not None and parent_sub_project_id is not None:
        if int(child_sub_project_id) != int(parent_sub_project_id):
            raise HTTPException(
                status_code=400,
                detail="상위 Task와 하위 Task가 모두 Sub Project에 속한 경우 같은 Sub Project 안에서만 연결할 수 있습니다.",
            )
    return parent


@app.post("/api/tasks")
def create_task(
    task: TaskCreate,
    suppress_realtime: bool = Query(False),
    request: Request = None,
    db: Session = Depends(get_db),
):
    state = load_state()

    p = db.query(Project).filter(Project.id == task.project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=400, detail="Project not found")

    # viewer cannot create tasks
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if caller_id:
        check_task_edit_permission(db, state, task.project_id, caller_id)

    # 상위 Task(Parent) 검증 + 기본값 상속 (미지정 시 부모의 sub_project/status 사용)
    parent = None
    parent_id = getattr(task, "parent_task_id", None)
    eff_status = task.status
    eff_sub_project_id = getattr(task, "sub_project_id", None)
    if parent_id is not None:
        # 검증은 요청된 sub_project 기준(상속 전) — 명시적으로 다른 Sub Project를 주면 차단
        parent = _validate_parent_task(
            db, state, task.project_id, None, parent_id, eff_sub_project_id
        )
        # 하위 Task 는 반드시 상위 Task 와 같은 Board 단계(status)에 놓인다.
        eff_status = parent.status
        if eff_sub_project_id is None:
            # 미지정 시 부모의 Sub Project 상속 (부모 meta 기준)
            eff_sub_project_id = get_task_meta(state, parent.id).get("sub_project_id")

    _sd, _dd, _sd_tbd, _dd_tbd = normalize_schedule_tbd(
        task.start_date, task.due_date,
        getattr(task, "start_date_tbd", False), getattr(task, "due_date_tbd", False),
    )
    t = Task(
        project_id=task.project_id,
        title=task.title,
        description=task.description,
        status=eff_status,
        priority=task.priority or "medium",
        task_type=getattr(task, "task_type", "normal") or "normal",
        start_date=_sd,
        due_date=_dd,
        start_date_tbd=_sd_tbd,
        due_date_tbd=_dd_tbd,
        assignee_ids=task.assignee_ids or [],
        tags=task.tags or [],
        parent_task_id=(parent.id if parent is not None else None),
        archived_at=None,
    )
    db.add(t)
    db.commit()
    db.refresh(t)

    # CUSTOM 워크플로우 프로젝트면 status 기준으로 표시 컬럼을 배정 (없으면 보드에서 안 보임).
    # 단, 상위 Task 가 있으면 상위와 같은 컬럼에 정확히 붙도록 상위의 workflow_column_id 를 그대로 상속한다.
    if (getattr(p, "workflow_mode", "DEFAULT") or "DEFAULT") == "CUSTOM":
        if parent is not None and getattr(parent, "workflow_column_id", None) is not None:
            t.workflow_column_id = parent.workflow_column_id
            db.commit()
            db.refresh(t)
        else:
            cols = _get_workflow_columns(db, p.id, include_inactive=False)
            col = _pick_column_for_status(cols, t.status)
            if col is not None:
                t.workflow_column_id = col.id
                db.commit()
                db.refresh(t)

    set_task_meta(state, t.id, {
        "sub_project_id": eff_sub_project_id,
        "progress": task.progress if task.progress is not None else 0,
    })
    save_state(state)

    actlog.log_activity(
        db, user_id=caller_id, action=actlog.ACTION_CREATED,
        entity_type=actlog.ENTITY_TASK, entity_id=t.id,
        message="Task 생성", meta={"title": t.title},
    )
    # 대량 import 중에는 개별 이벤트를 만들지 않고, import 완료 후 bulk event 1개만 발행한다.
    if p.space_id and not suppress_realtime:
        emit_realtime_event(
            db, space_id=p.space_id, event_type="task_created",
            entity_type="task", entity_id=t.id, project_id=t.project_id, task_id=t.id,
            actor_user_id=caller_id,
        )
    db.commit()

    return task_dict(t, state)

@app.post("/api/spaces/{space_id}/oneoff-tasks")
def create_oneoff_task(
    space_id: int,
    body: OneoffTaskCreate,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """캘린더 단발 일정 생성. 공간별 [시스템] 단발 업무 프로젝트 하위 task 로 만든다.

    권한은 프로젝트 멤버십이 아니라 공간 멤버십으로 통제한다
    (단발 업무 프로젝트는 공간 멤버 누구나 추가 가능).
    """
    state = load_state()

    space = db.query(Space).filter(Space.id == space_id, Space.is_active == True).first()
    if not space:
        raise HTTPException(404, "Space not found")

    if not is_admin_like_role(get_user_role(db, user_id)):
        member = db.query(SpaceMember).filter(
            SpaceMember.space_id == space_id,
            SpaceMember.user_id == user_id,
        ).first()
        if not member:
            raise HTTPException(403, "이 공간에 단발 일정을 추가할 권한이 없습니다.")

    title = (body.title or "").strip()
    if not title:
        raise HTTPException(400, "제목을 입력해주세요.")

    proj = get_or_create_oneoff_project(db, state, space_id)

    t = Task(
        project_id=proj.id,
        title=title,
        description=body.description,
        status=body.status or "todo",
        priority=body.priority or "medium",
        task_type="one_off",
        start_date=body.start_date or body.due_date,
        due_date=body.due_date,
        assignee_ids=body.assignee_ids or [],
        tags=[],
        archived_at=None,
    )
    db.add(t)
    db.commit()
    db.refresh(t)

    set_task_meta(state, t.id, {"progress": 0})
    save_state(state)

    emit_realtime_event(
        db, space_id=space_id, event_type="task_created",
        entity_type="task", entity_id=t.id, project_id=proj.id, task_id=t.id,
        actor_user_id=user_id,
    )
    db.commit()

    return task_dict(t, state)

# =========================
# Calendar Events (Phase 2)
# =========================
@app.get("/api/spaces/{space_id}/calendar-events")
def list_calendar_events(
    space_id: int,
    user_id: int = Query(...),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """공간의 단순 일정 목록. date_from/date_to(YYYY-MM-DD)로 기간 필터(선택)."""
    space = db.query(Space).filter(Space.id == space_id, Space.is_active == True).first()
    if not space:
        raise HTTPException(404, "Space not found")
    _require_space_member(db, space_id, user_id)

    q = db.query(CalendarEvent).filter(
        CalendarEvent.space_id == space_id,
        CalendarEvent.archived_at.is_(None),
    )
    # 기간 겹침 필터: 이벤트의 [start,end] 가 [date_from,date_to] 와 겹치면 포함.
    # 날짜는 YYYY-MM-DD 문자열이라 사전식 비교가 곧 날짜 비교다.
    if date_to:
        q = q.filter(func.coalesce(CalendarEvent.start_date, CalendarEvent.end_date) <= date_to)
    if date_from:
        q = q.filter(func.coalesce(CalendarEvent.end_date, CalendarEvent.start_date) >= date_from)
    rows = q.all()
    return {"events": [calendar_event_dict(e) for e in rows]}

@app.post("/api/spaces/{space_id}/calendar-events")
def create_calendar_event(
    space_id: int,
    body: CalendarEventCreate,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    space = db.query(Space).filter(Space.id == space_id, Space.is_active == True).first()
    if not space:
        raise HTTPException(404, "Space not found")
    _require_space_member(db, space_id, user_id)

    title = (body.title or "").strip()
    if not title:
        raise HTTPException(400, "제목을 입력해주세요.")

    # 담당자는 이 공간의 멤버만 지정 가능 (미지정 허용)
    _ensure_assignee_in_space(db, space_id, body.assignee_id)

    e = CalendarEvent(
        space_id=space_id,
        title=title,
        description=body.description,
        start_date=body.start_date or body.end_date,
        end_date=body.end_date or body.start_date,
        all_day=True if body.all_day is None else bool(body.all_day),
        status=body.status or "planned",
        event_type=body.event_type or "general",
        owner_id=user_id,
        assignee_id=body.assignee_id,
    )
    db.add(e)
    db.commit()
    db.refresh(e)
    actlog.log_activity(
        db, user_id=user_id, action=actlog.ACTION_CREATED,
        entity_type=actlog.ENTITY_CALENDAR_EVENT, entity_id=e.id,
        message="일정 생성", meta={"title": e.title},
    )
    emit_realtime_event(
        db, space_id=space_id, event_type="calendar_event_created",
        entity_type="calendar_event", entity_id=e.id,
        actor_user_id=user_id,
    )
    db.commit()
    return calendar_event_dict(e)

@app.patch("/api/calendar-events/{event_id}")
def update_calendar_event(
    event_id: int,
    updates: CalendarEventUpdate,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    e = db.query(CalendarEvent).filter(
        CalendarEvent.id == event_id,
        CalendarEvent.archived_at.is_(None),
    ).first()
    if not e:
        raise HTTPException(404, "Event not found")
    _require_space_member(db, e.space_id, user_id)

    data = updates.model_dump(exclude_unset=True)
    # 담당자를 '새로운' 사용자로 바꾸는 경우에만 공간 멤버 검증.
    # (기존 데이터가 이미 비멤버여도 그 값을 그대로 두는 저장은 막지 않아 호환성 유지)
    if "assignee_id" in data and data["assignee_id"] != e.assignee_id:
        _ensure_assignee_in_space(db, e.space_id, data["assignee_id"])

    _ev_fields = ["title", "description", "start_date", "end_date", "all_day", "status", "event_type", "assignee_id"]
    _ev_old = {k: getattr(e, k, None) for k in _ev_fields}
    for k in _ev_fields:
        if k in data:
            setattr(e, k, data[k])
    e.updated_at = datetime.now()

    _evchanges, _evaction, _evmsg = actlog.diff_changes(_ev_old, data, _ev_fields)
    if _evchanges:
        actlog.log_activity(
            db, user_id=user_id, action=_evaction,
            entity_type=actlog.ENTITY_CALENDAR_EVENT, entity_id=e.id,
            message=_evmsg, meta={"changes": _evchanges},
        )

    emit_realtime_event(
        db, space_id=e.space_id, event_type="calendar_event_updated",
        entity_type="calendar_event", entity_id=e.id,
        actor_user_id=user_id,
    )
    db.commit()
    db.refresh(e)
    return calendar_event_dict(e)

@app.delete("/api/calendar-events/{event_id}")
def delete_calendar_event(
    event_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    e = db.query(CalendarEvent).filter(
        CalendarEvent.id == event_id,
        CalendarEvent.archived_at.is_(None),
    ).first()
    if not e:
        raise HTTPException(404, "Event not found")
    _require_space_member(db, e.space_id, user_id)
    e.archived_at = datetime.now()
    actlog.log_activity(
        db, user_id=user_id, action=actlog.ACTION_DELETED,
        entity_type=actlog.ENTITY_CALENDAR_EVENT, entity_id=e.id,
        message="일정 삭제", meta={"title": e.title},
    )
    emit_realtime_event(
        db, space_id=e.space_id, event_type="calendar_event_deleted",
        entity_type="calendar_event", entity_id=e.id,
        actor_user_id=user_id,
    )
    db.commit()
    return {"message": "Event deleted"}

# =========================
# Realtime sync (SSE)
# =========================
@app.post("/api/events/ticket")
def create_events_ticket(
    space_id: int = Query(...),
    request: Request = None,
    db: Session = Depends(get_db),
):
    """EventSource 접속용 단발성 ticket 발급.

    - 로그인 사용자만(미로그인 401)
    - 해당 space 멤버만(비멤버 403) — admin/super_admin 은 통과
    - 세션 토큰을 URL 에 노출하지 않기 위한 단발성 ticket 방식
    """
    if not REALTIME_SYNC_ENABLED:
        raise HTTPException(503, "실시간 동기화가 비활성화되어 있습니다.")
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if not caller_id:
        raise HTTPException(401, "로그인이 필요합니다.")
    # 전역 채널(0)은 ticket 대상이 아니다. 실제 space 멤버십만 검증한다.
    if space_id and space_id > 0:
        _require_space_member(db, space_id, caller_id)
    else:
        raise HTTPException(400, "유효하지 않은 공간입니다.")
    ticket, ttl = realtime_issue_ticket(caller_id, space_id)
    return {"ticket": ticket, "expires_in": ttl}


@app.get("/api/events/stream")
async def events_stream(
    space_id: int = Query(...),
    ticket: Optional[str] = Query(None),
    last_event_id: Optional[int] = Query(None),
    request: Request = None,
):
    """같은 space 변경사항을 SSE 로 전파. (broadcaster 가 채널당 1 poller 로 묶어 부하 제어)

    인증: ticket 우선, 없으면 Bearer fallback. 권한: 해당 space 멤버 재확인(우회 차단).
    DB session 은 짧게 열고 닫는다(SSE 동안 session 을 잡지 않음 — 누수 방지).
    """
    if not REALTIME_SYNC_ENABLED:
        raise HTTPException(503, "실시간 동기화가 비활성화되어 있습니다.")

    user_id = None
    if ticket:
        user_id = realtime_validate_ticket(ticket, space_id)

    # 멤버십 재확인은 짧게 연 session 으로 처리하고 즉시 닫는다.
    db = SessionLocal()
    try:
        if not user_id:
            caller_id = resolve_user_id_from_token(request, db) if request else None
            if caller_id:
                user_id = caller_id
        if not user_id:
            raise HTTPException(401, "유효하지 않은 티켓입니다. 다시 로그인해주세요.")
        if space_id and space_id > 0:
            _require_space_member(db, space_id, user_id)
        else:
            raise HTTPException(400, "유효하지 않은 공간입니다.")
    finally:
        db.close()

    if realtime_client_count() >= REALTIME_MAX_CLIENTS_PER_WORKER:
        raise HTTPException(503, "동시 접속이 많습니다. 잠시 후 다시 시도해주세요.")

    # Last-Event-ID: 헤더 우선, query param fallback (프록시/브라우저 차이 대비)
    raw_leid = (request.headers.get("Last-Event-ID") if request else None)
    if raw_leid is None and last_event_id is not None:
        raw_leid = last_event_id
    try:
        leid = int(raw_leid) if raw_leid not in (None, "") else 0
    except (TypeError, ValueError):
        leid = 0

    headers = {
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
        "Connection": "keep-alive",
    }
    return StreamingResponse(
        realtime_event_stream(request, user_id, space_id, leid),
        media_type="text/event-stream",
        headers=headers,
    )


@app.get("/api/admin/realtime-status")
def admin_realtime_status(user_id: int = Query(...), db: Session = Depends(get_db)):
    """운영 모니터링: 현재 SSE client 수 / space별 client 수 / 설정값."""
    if not is_admin_like_role(get_user_role(db, user_id)):
        raise HTTPException(403, "관리자만 접근할 수 있습니다.")
    return realtime_get_stats()


# bulk event payload 에 허용하는 숫자 카운트 키 (민감정보 없음).
_REALTIME_BULK_SUMMARY_KEYS = {
    "projects_created",
    "tasks_created",
    "notes_created",
    "text_notes_created",
    "checklist_notes_created",
}
_REALTIME_BULK_EVENT_TYPES = {"task_bulk_changed", "project_bulk_changed", "space_bulk_changed"}


class RealtimeBulkEventCreate(BaseModel):
    event_type: Optional[str] = "space_bulk_changed"
    project_id: Optional[int] = None
    summary: Optional[Dict[str, Any]] = None


@app.post("/api/spaces/{space_id}/realtime/bulk-event")
def create_realtime_bulk_event(
    space_id: int,
    body: RealtimeBulkEventCreate,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """대량 작업(CSV/XLSX import 등) 완료 후 bulk event 1개를 발행한다.

    개별 create 요청은 suppress_realtime=true 로 이벤트를 만들지 않고,
    호출자가 import 성공 후 이 endpoint 로 단일 bulk event 만 남긴다.
    payload 에는 민감정보(title/description/note 내용)를 넣지 않고 숫자 카운트만 허용한다.
    """
    if not REALTIME_SYNC_ENABLED:
        return {"emitted": False, "reason": "disabled"}

    space = db.query(Space).filter(Space.id == space_id, Space.is_active == True).first()
    if not space:
        raise HTTPException(404, "Space not found")
    _require_space_member(db, space_id, user_id)

    event_type = body.event_type if body.event_type in _REALTIME_BULK_EVENT_TYPES else "space_bulk_changed"

    # summary 는 화이트리스트 숫자 카운트만 통과 (텍스트/제목/내용 주입 차단)
    summary = {}
    for k, v in (body.summary or {}).items():
        if k not in _REALTIME_BULK_SUMMARY_KEYS:
            continue
        try:
            summary[k] = max(0, int(v))
        except (TypeError, ValueError):
            continue

    payload = {"summary": summary} if summary else None
    emit_realtime_event(
        db, space_id=space_id, event_type=event_type,
        entity_type="space", entity_id=space_id, project_id=body.project_id,
        actor_user_id=user_id, payload=payload,
    )
    db.commit()
    return {"emitted": True, "event_type": event_type, "summary": summary}


@app.get("/api/activity-logs")
def get_activity_logs(
    entity_type: str = Query(...),
    entity_id: int = Query(...),
    limit: int = Query(50),
    db: Session = Depends(get_db),
):
    """특정 아이템(project / task / calendar_event)의 변경 이력을 최신순으로 반환.

    누가(user_name) / 언제(created_at) / 무엇을(message, meta) 바꿨는지 표시용.
    """
    if entity_type not in (actlog.ENTITY_PROJECT, actlog.ENTITY_TASK, actlog.ENTITY_CALENDAR_EVENT):
        raise HTTPException(400, "invalid entity_type")
    safe_limit = max(1, min(int(limit or 50), 200))

    rows = (
        db.query(ActivityLog)
        .filter(ActivityLog.entity_type == entity_type, ActivityLog.entity_id == entity_id)
        .order_by(ActivityLog.created_at.desc(), ActivityLog.id.desc())
        .limit(safe_limit)
        .all()
    )

    user_ids = {r.user_id for r in rows if r.user_id}
    users = {}
    if user_ids:
        for u in db.query(User).filter(User.id.in_(user_ids)).all():
            users[u.id] = u

    out = []
    for r in rows:
        u = users.get(r.user_id) if r.user_id else None
        out.append({
            "id": r.id,
            "action": r.action,
            "entity_type": r.entity_type,
            "entity_id": r.entity_id,
            "user_id": r.user_id,
            "user_name": (u.username if u else None),
            "message": r.message,
            "meta": r.meta,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return {"logs": out}

@app.post("/api/calendar-events/{event_id}/convert-to-task")
def convert_event_to_task(
    event_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """단순 일정을 [시스템] 단발 업무 프로젝트의 one_off task 로 전환하고 연결한다.
    이미 전환된 경우 기존 연결 task 를 그대로 반환한다."""
    state = load_state()
    e = db.query(CalendarEvent).filter(
        CalendarEvent.id == event_id,
        CalendarEvent.archived_at.is_(None),
    ).first()
    if not e:
        raise HTTPException(404, "Event not found")
    _require_space_member(db, e.space_id, user_id)

    # 이미 연결된 task 가 살아있으면 재사용 (중복 전환 방지)
    if e.linked_task_id:
        existing = db.query(Task).filter(
            Task.id == e.linked_task_id,
            Task.archived_at.is_(None),
        ).first()
        if existing:
            return {"event": calendar_event_dict(e), "task": task_dict(existing, state)}

    proj = get_or_create_oneoff_project(db, state, e.space_id)
    t = Task(
        project_id=proj.id,
        title=e.title,
        description=e.description,
        status="todo",
        priority="medium",
        task_type="one_off",
        start_date=e.start_date or e.end_date,
        due_date=e.end_date or e.start_date,
        assignee_ids=[e.assignee_id] if e.assignee_id else [],
        tags=[],
        archived_at=None,
    )
    db.add(t)
    db.commit()
    db.refresh(t)

    set_task_meta(state, t.id, {"progress": 0})
    save_state(state)

    e.linked_task_id = t.id
    e.linked_project_id = proj.id
    e.updated_at = datetime.now()
    emit_realtime_event(
        db, space_id=e.space_id, event_type="task_created",
        entity_type="task", entity_id=t.id, project_id=proj.id, task_id=t.id,
        actor_user_id=user_id,
    )
    db.commit()
    db.refresh(e)

    return {"event": calendar_event_dict(e), "task": task_dict(t, state)}

@app.patch("/api/tasks/{task_id}")
def update_task(task_id: int, updates: TaskUpdate, request: Request = None, db: Session = Depends(get_db)):
    state = load_state()
    t = db.query(Task).filter(Task.id == task_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")

    # viewer cannot update tasks
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if caller_id:
        check_task_edit_permission(db, state, t.project_id, caller_id)

    data = updates.model_dump(exclude_unset=True)

    # DB fields
    _hist_fields = ["title", "description", "status", "priority", "start_date", "due_date",
                    "start_date_tbd", "due_date_tbd", "assignee_ids", "tags", "remarks"]
    _old_vals = {k: getattr(t, k, None) for k in _hist_fields}
    for k in _hist_fields:
        if k in data:
            setattr(t, k, data[k])

    # 일정 미정(TBD) 정규화 — 스케줄 관련 필드가 하나라도 오면, 병합된 최종 값 기준으로
    # 미정 여부에 따라 날짜를 NULL 처리하고 start<=due 순서를 검증한다.
    if any(k in data for k in ("start_date", "due_date", "start_date_tbd", "due_date_tbd")):
        _sd, _dd, _sd_tbd, _dd_tbd = normalize_schedule_tbd(
            t.start_date, t.due_date, t.start_date_tbd, t.due_date_tbd,
        )
        t.start_date, t.due_date, t.start_date_tbd, t.due_date_tbd = _sd, _dd, _sd_tbd, _dd_tbd
        # 이력/이벤트 diff 가 정규화된 실제 저장값을 반영하도록 data 도 동기화
        data["start_date"] = _sd
        data["due_date"] = _dd

    # progress도 DB 컬럼과 사이드카 meta에 함께 쓴다 (두 곳의 진실이
    # 어긋나서 체크박스/드래그 간 진행률이 튀는 현상 방지).
    if "progress" in data and data["progress"] is not None:
        try:
            t.progress = max(0, min(100, int(data["progress"])))
        except Exception:
            pass

    # CUSTOM 워크플로우 컬럼 이동 — workflow_column_id가 오면 status를
    # 해당 컬럼의 mapped_status로 함께 동기화한다 (canonical status 유지 정책).
    if "workflow_column_id" in data:
        new_col_id = data["workflow_column_id"]
        t.workflow_column_id = new_col_id
        if new_col_id is not None:
            col = db.query(ProjectTaskColumn).filter(
                ProjectTaskColumn.id == new_col_id,
                ProjectTaskColumn.project_id == t.project_id,
            ).first()
            if col is not None and col.mapped_status:
                # status 변경 이력이 남도록 data에도 반영 (_old_vals엔 원래 status가 이미 캡처됨)
                if col.mapped_status != t.status:
                    data["status"] = col.mapped_status
                t.status = col.mapped_status

    # ── Sub Project 변경 값(사이드카 meta) 미리 계산 — 상/하위 Sub Project 조건 검증에 사용 ──
    _curr_sub = get_task_meta(state, task_id).get("sub_project_id")
    _new_sub = data["sub_project_id"] if "sub_project_id" in data else _curr_sub

    # 이 Task 가 이미 하위 Task 인데(상위 연결 유지) 단계(status/컬럼)만 바꾸려 하면 차단.
    # 하위 Task 는 상위와 같은 Board 단계에 있어야 하므로, 단계 변경은 상위 연결 해제 후에만 허용.
    _stage_changed = ("status" in data) or ("workflow_column_id" in data)
    if _stage_changed and "parent_task_id" not in data and t.parent_task_id is not None:
        _p = db.query(Task).filter(Task.id == t.parent_task_id, Task.archived_at.is_(None)).first()
        if _p is not None and _task_board_key(_p) != _task_board_key(t):
            raise HTTPException(
                status_code=400,
                detail="이 Task는 상위 Task에 연결되어 있습니다. 다른 단계로 이동하려면 먼저 상위 Task 연결을 해제해주세요.",
            )

    # 하위 Task 가 상위와 다른 Sub Project 로 바뀌려 하면 차단(상위 연결 해제 후에만 허용).
    if ("sub_project_id" in data and "parent_task_id" not in data
            and t.parent_task_id is not None and _new_sub is not None):
        _p = db.query(Task).filter(Task.id == t.parent_task_id, Task.archived_at.is_(None)).first()
        _p_sub = get_task_meta(state, _p.id).get("sub_project_id") if _p else None
        if _p_sub is not None and int(_new_sub) != int(_p_sub):
            raise HTTPException(
                status_code=400,
                detail="하위 Task는 상위 Task와 다른 Sub Project로 변경할 수 없습니다. 먼저 상위 Task 연결을 해제해주세요.",
            )

    # 상위 Task(Parent) 변경 — null=해제, int=검증 후 지정 (단일 레벨/순환/같은단계/Sub Project 방지)
    if "parent_task_id" in data:
        new_parent_id = data["parent_task_id"]
        if new_parent_id is None:
            t.parent_task_id = None
        else:
            parent = _validate_parent_task(
                db, state, t.project_id, t.id, new_parent_id, _new_sub,
                child_board_key=_task_board_key(t),
            )
            t.parent_task_id = parent.id

    # 상위 Task 가 다른 Board 단계로 이동하면 하위 Task 전체(자식/손자/...)도 같은 단계로 함께 이동
    # (상위·하위는 같은 Board 단계라는 불변식을 다단계에서도 유지).
    if _stage_changed:
        _desc_ids = _get_task_descendant_ids(db, task_id)
        if _desc_ids:
            for _ch in db.query(Task).filter(Task.id.in_(_desc_ids), Task.archived_at.is_(None)).all():
                _ch.status = t.status
                _ch.workflow_column_id = t.workflow_column_id
                _ch.updated_at = datetime.now()

    # 상위 Task 의 Sub Project 가 바뀌면 하위 Task 전체도 같은 Sub Project 로 함께 이동(구조 유지).
    if "sub_project_id" in data and _new_sub is not None:
        for _ch_id in _get_task_descendant_ids(db, task_id):
            set_task_meta(state, _ch_id, {"sub_project_id": _new_sub})

    t.updated_at = datetime.now()

    # sidecar fields
    meta_updates = {}
    if "sub_project_id" in data:
        meta_updates["sub_project_id"] = data["sub_project_id"]
    if "progress" in data:
        meta_updates["progress"] = data["progress"]
    if meta_updates:
        set_task_meta(state, task_id, meta_updates)
        save_state(state)

    # 변경 이력 기록 (실제 바뀐 필드만)
    _changes, _action, _msg = actlog.diff_changes(_old_vals, data, _hist_fields)
    if _changes:
        actlog.log_activity(
            db, user_id=caller_id, action=_action,
            entity_type=actlog.ENTITY_TASK, entity_id=t.id,
            message=_msg, meta={"changes": _changes},
        )

    _t_sid = db.query(Project.space_id).filter(Project.id == t.project_id).scalar()
    if _t_sid:
        emit_realtime_event(
            db, space_id=_t_sid, event_type="task_updated",
            entity_type="task", entity_id=t.id, project_id=t.project_id, task_id=t.id,
            actor_user_id=caller_id,
        )
    db.commit()
    db.refresh(t)
    # 사용자가 직접 status 또는 progress를 변경한 경우 자동 동기화를
    # 건너뛰어 수동 변경(드래그/명시적 진행률 입력)을 우선한다. 이렇게 해야
    # 드래그 직후 _sync_task_progress가 체크박스 기준으로 다시 덮어써서
    # 상태가 되돌아가는 현상이 발생하지 않는다.
    if not ({"status", "progress", "workflow_column_id"} & data.keys()):
        _sync_task_progress(db, task_id)
        db.refresh(t)
    state = load_state()
    return task_dict(t, state)

class TaskParentPatch(BaseModel):
    """상위 작업(Parent) 연결/해제 전용. Board/Graph Drag&Drop, 컨텍스트 메뉴에서 사용.
    parent_task_id=null → 연결 해제, int → 검증 후 연결."""
    parent_task_id: Optional[int] = None


@app.patch("/api/tasks/{task_id}/parent")
def update_task_parent(
    task_id: int,
    body: TaskParentPatch,
    request: Request = None,
    db: Session = Depends(get_db),
):
    """상위 작업 연결/해제 — Task 수정 API 와 동일한 검증(권한/같은 프로젝트/같은 Board 단계/순환/Sub Project)을
    한 곳에서 수행하는 경량 endpoint. 전체 Task PATCH 를 보내지 않고 계층만 빠르게 바꾼다.

    성공 시 갱신된 task_dict 를 반환하며, 변경 이력(task_parent_changed / task_parent_removed)을 남긴다.
    """
    t = db.query(Task).filter(Task.id == task_id, Task.archived_at.is_(None)).first()
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")

    state = load_state()
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if caller_id:
        check_task_edit_permission(db, state, t.project_id, caller_id)

    old_parent_id = getattr(t, "parent_task_id", None)
    new_parent_id = body.parent_task_id

    if new_parent_id is None:
        t.parent_task_id = None
        _subtype = "task_parent_removed"
    else:
        _curr_sub = get_task_meta(state, t.id).get("sub_project_id")
        parent = _validate_parent_task(
            db, state, t.project_id, t.id, new_parent_id, _curr_sub,
            child_board_key=_task_board_key(t),
        )
        # 하위 작업은 상위와 같은 Board 단계에 놓인다(불변식 유지) — 상위의 status/컬럼으로 정렬.
        t.parent_task_id = parent.id
        t.status = parent.status
        t.workflow_column_id = getattr(parent, "workflow_column_id", None)
        _subtype = "task_parent_changed"

    t.updated_at = datetime.now()

    actlog.log_activity(
        db, user_id=caller_id, action=actlog.ACTION_UPDATED,
        entity_type=actlog.ENTITY_TASK, entity_id=t.id,
        message=("상위 작업 연결 해제" if new_parent_id is None else "상위 작업 연결"),
        meta={
            "subtype": _subtype,
            "old_parent_task_id": old_parent_id,
            "new_parent_task_id": new_parent_id,
            "changed_by": caller_id,
        },
    )
    _t_sid = db.query(Project.space_id).filter(Project.id == t.project_id).scalar()
    if _t_sid:
        emit_realtime_event(
            db, space_id=_t_sid, event_type="task_updated",
            entity_type="task", entity_id=t.id, project_id=t.project_id, task_id=t.id,
            actor_user_id=caller_id,
        )
    db.commit()
    db.refresh(t)
    return task_dict(t, load_state())


@app.delete("/api/tasks/{task_id}")
def delete_task(task_id: int, request: Request = None, db: Session = Depends(get_db)):
    t = db.query(Task).filter(Task.id == task_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")

    # viewer cannot delete tasks
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if caller_id:
        check_task_edit_permission(db, load_state(), t.project_id, caller_id)

    t.archived_at = datetime.now()
    t.deleted_by_user_id = caller_id
    # 상위 Task 보관 시 하위 Task는 삭제하지 않고 최상위로 승격(고아화/보드에서 사라짐 방지).
    db.query(Task).filter(Task.parent_task_id == task_id).update(
        {Task.parent_task_id: None}, synchronize_session=False
    )
    actlog.log_activity(
        db, user_id=caller_id, action=actlog.ACTION_DELETED,
        entity_type=actlog.ENTITY_TASK, entity_id=t.id,
        message="Task 삭제", meta={"title": t.title},
    )
    _t_sid = db.query(Project.space_id).filter(Project.id == t.project_id).scalar()
    if _t_sid:
        emit_realtime_event(
            db, space_id=_t_sid, event_type="task_deleted",
            entity_type="task", entity_id=t.id, project_id=t.project_id, task_id=t.id,
            actor_user_id=caller_id,
        )
    db.commit()
    return {"message": "Task deleted"}

@app.post("/api/tasks/{task_id}/restore")
def restore_task(task_id: int, request: Request = None, db: Session = Depends(get_db)):
    t = db.query(Task).filter(Task.id == task_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")

    # 본인이 삭제한 항목만 복원 가능 (admin/super_admin은 전체 허용)
    caller_id = resolve_user_id_from_token(request, db) if request else None
    caller_role = get_user_role(db, caller_id) if caller_id else None
    is_admin = caller_role in ("admin", "super_admin")
    if not is_admin and (t.deleted_by_user_id is None or t.deleted_by_user_id != caller_id):
        raise HTTPException(status_code=403, detail="본인이 삭제한 항목만 복원할 수 있습니다.")

    t.archived_at = None
    t.deleted_by_user_id = None
    db.commit()
    return {"message": "Task restored"}


class TaskMoveIn(BaseModel):
    target_project_id: int
    # 대상 Project 의 Sub Project(선택). None = 프로젝트 직속.
    target_sub_project_id: Optional[int] = None


@app.post("/api/tasks/{task_id}/move")
def move_task(task_id: int, body: TaskMoveIn, request: Request = None, db: Session = Depends(get_db)):
    """Task 를 같은 Space 안의 다른 Project 로 이동한다.

    정책(1차):
    - 같은 Space 안의 Project 간 이동만 허용(다른 Space 이동은 2차 보류).
    - 현재 Project 수정 권한 + 대상 Project 수정 권한 둘 다 필요(viewer 불가).
    - sub_project_id 는 대상 Project 기준으로 재선택(미선택 = null).
    - status/workflow 는 대상 Project 의 첫 단계로 배치.
    - parent_task_id 는 null 로 초기화(cross-project parent 방지).
    - 이 Task 의 하위 Task 들은 원래 Project 에 남기고 최상위로 승격(cross-project 방지).
    - title/description/work_notes/comments/attachments/history 는 task_id 기준이라 그대로 유지된다.
    """
    state = load_state()
    t = db.query(Task).filter(Task.id == task_id, Task.archived_at.is_(None)).first()
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")

    caller_id = resolve_user_id_from_token(request, db) if request else None
    if not caller_id:
        raise HTTPException(status_code=401, detail="인증이 필요합니다.")

    from_project_id = t.project_id
    target_project_id = int(body.target_project_id)
    if target_project_id == int(from_project_id):
        raise HTTPException(status_code=400, detail="이미 해당 Project 에 있는 Task 입니다.")

    # 권한 ①: 현재 Project 에서 Task 수정 권한 (viewer 불가 → 403)
    check_task_edit_permission(db, state, from_project_id, caller_id)

    src = db.query(Project).filter(Project.id == from_project_id, Project.archived_at.is_(None)).first()
    dst = db.query(Project).filter(Project.id == target_project_id, Project.archived_at.is_(None)).first()
    if not dst:
        raise HTTPException(status_code=404, detail="대상 Project 를 찾을 수 없습니다.")

    # 같은 Space 제약 (1차 범위)
    if not src or not src.space_id or not dst.space_id or int(src.space_id) != int(dst.space_id):
        raise HTTPException(status_code=400, detail="같은 Space 안의 Project 로만 이동할 수 있습니다.")

    # 권한 ②: 대상 Project 에서도 Task 생성/관리 권한 (viewer 불가 → 403)
    check_task_edit_permission(db, state, target_project_id, caller_id)

    # 대상 Sub Project 검증 (선택 시 반드시 대상 Project 소속이어야 함)
    target_sub_project_id = body.target_sub_project_id
    if target_sub_project_id is not None:
        sp = db.query(SubProjectModel).filter(
            SubProjectModel.id == int(target_sub_project_id),
            SubProjectModel.project_id == target_project_id,
        ).first()
        if not sp:
            raise HTTPException(status_code=400, detail="선택한 Sub Project 가 대상 Project 에 속해있지 않습니다.")

    from_sub_project_id = get_task_meta(state, task_id).get("sub_project_id")

    # ── 이동 처리 ──
    t.project_id = target_project_id
    t.parent_task_id = None  # cross-project parent 방지

    # 이 Task 의 하위 Task 들은 원래 Project 에 남기고 최상위로 승격(고아화/보드 이탈 방지)
    db.query(Task).filter(Task.parent_task_id == task_id, Task.archived_at.is_(None)).update(
        {Task.parent_task_id: None}, synchronize_session=False
    )

    # status/workflow 는 대상 Project 의 첫 단계로 배치
    dst_mode = (getattr(dst, "workflow_mode", "DEFAULT") or "DEFAULT")
    if dst_mode == "CUSTOM":
        cols = _get_workflow_columns(db, target_project_id, include_inactive=False)
        first_col = cols[0] if cols else None
        if first_col is not None:
            t.workflow_column_id = first_col.id
            t.status = first_col.mapped_status or "todo"
        else:
            t.workflow_column_id = None
            t.status = "todo"
    else:
        t.workflow_column_id = None
        t.status = "todo"

    t.updated_at = datetime.now()

    # Sub Project(사이드카 meta) 재설정
    set_task_meta(state, task_id, {"sub_project_id": target_sub_project_id})
    save_state(state)

    # 이동 이력 (ActivityLog 재사용)
    actlog.log_activity(
        db, user_id=caller_id, action=actlog.ACTION_MOVED,
        entity_type=actlog.ENTITY_TASK, entity_id=t.id,
        message="Task 이동",
        meta={
            "event": "task_moved_project",
            "from_project_id": from_project_id,
            "to_project_id": target_project_id,
            "from_sub_project_id": from_sub_project_id,
            "to_sub_project_id": target_sub_project_id,
        },
    )

    # Realtime sync — 같은 Space 이므로 이벤트 1개로 양쪽 Board(from/to)를 invalidate.
    #   payload 에 from/to project_id 를 담아 프론트가 두 프로젝트 뷰를 함께 갱신하도록 한다.
    if src.space_id:
        emit_realtime_event(
            db, space_id=src.space_id, event_type="task_moved",
            entity_type="task", entity_id=t.id, project_id=target_project_id, task_id=t.id,
            actor_user_id=caller_id,
            payload={"from_project_id": from_project_id, "to_project_id": target_project_id},
        )

    db.commit()
    db.refresh(t)
    state = load_state()
    return task_dict(t, state)

# =========================
# Workflow Columns (프로젝트별 커스텀 Board 단계) — Slice 1
# =========================

class WorkflowColumnIn(BaseModel):
    id: Optional[int] = None
    label: str
    description: Optional[str] = None
    color: Optional[str] = None
    sort_order: int
    # 'hold' 힌트만 의미있게 사용 — 나머지는 순서로 자동 계산
    mapped_status: Optional[str] = None
    # 중간 완료 체크포인트 (보조 지표 — 최종 완료/Completion 계산과 무관)
    is_checkpoint: Optional[bool] = None
    checkpoint_label: Optional[str] = None
    checkpoint_description: Optional[str] = None
    # 진행률 기준 자동 이동 (CUSTOM 전용, 단계별). 기본 OFF. (레거시/고급 progress_threshold 규칙용)
    auto_advance_enabled: Optional[bool] = None
    auto_advance_threshold: Optional[int] = None
    # 단계 완료 조건: all_required_checkpoints / manual / progress_threshold (레거시)
    completion_rule_type: Optional[str] = None
    # 4차: 단계 완료 선언 시 다음 단계로 자동 이동할지 여부
    auto_move_on_complete: Optional[bool] = None

class WorkflowColumnsSave(BaseModel):
    columns: List[WorkflowColumnIn]

class WorkflowModeChange(BaseModel):
    target_mode: str  # DEFAULT / CUSTOM


def _sanitize_column_description(s: Optional[str]) -> Optional[str]:
    """단계 세부 설명을 plain text 로 정리. HTML 태그 제거 + 길이 제한(200자).

    빈 문자열/공백만 있으면 None(=설명 없음)으로 저장한다.
    """
    if not s:
        return None
    plain = re.sub(r"<[^>]+>", "", s)            # HTML 태그 제거
    plain = plain.replace("\r\n", "\n").strip()
    if not plain:
        return None
    return plain[:200]


def _wf_column_dict(c: ProjectTaskColumn) -> dict:
    return {
        "id": c.id,
        "project_id": c.project_id,
        "key": c.key,
        "label": c.label,
        "description": getattr(c, "description", None),
        "mapped_status": c.mapped_status,
        "sort_order": c.sort_order,
        "color": c.color,
        "is_final_done": bool(c.is_final_done),
        "active": bool(c.active),
        "is_checkpoint": bool(getattr(c, "is_checkpoint", False)),
        "checkpoint_label": getattr(c, "checkpoint_label", None),
        "checkpoint_description": getattr(c, "checkpoint_description", None),
        "auto_advance_enabled": bool(getattr(c, "auto_advance_enabled", False)),
        "auto_advance_threshold": getattr(c, "auto_advance_threshold", None),
        "completion_rule_type": getattr(c, "completion_rule_type", None) or "manual",
        "auto_move_on_complete": bool(getattr(c, "auto_move_on_complete", True)),
    }


def _apply_positional_mapping(ordered: List[ProjectTaskColumn]) -> None:
    """sort_order 정렬된 컬럼 리스트에 대해 mapped_status/is_final_done 자동 계산.

    - 첫 컬럼=todo / 마지막=done / 중간=in_progress (컬럼 '순서' 기준)
    - mapped_status=='hold' 인 컬럼은 흐름에서 제외하고 hold 유지
    - 흐름 컬럼이 1개뿐이면 그 컬럼이 곧 완료(done)
    """
    flow = [c for c in ordered if (c.mapped_status or "") != "hold"]
    n = len(flow)
    for i, c in enumerate(flow):
        if n == 1:
            c.mapped_status, c.is_final_done = "done", True
        elif i == 0:
            c.mapped_status, c.is_final_done = "todo", False
        elif i == n - 1:
            c.mapped_status, c.is_final_done = "done", True
        else:
            c.mapped_status, c.is_final_done = "in_progress", False
    for c in ordered:
        if (c.mapped_status or "") == "hold":
            c.is_final_done = False


def _get_workflow_columns(db: Session, project_id: int, include_inactive: bool = False) -> List[ProjectTaskColumn]:
    q = db.query(ProjectTaskColumn).filter(ProjectTaskColumn.project_id == project_id)
    if not include_inactive:
        q = q.filter(ProjectTaskColumn.active == True)
    return q.order_by(ProjectTaskColumn.sort_order.asc(), ProjectTaskColumn.id.asc()).all()


def _workflow_label_map(db: Session, project_id: int) -> dict:
    """{workflow_column_id: ProjectTaskColumn} — Roadmap/Graph/AI Report 라벨 오버레이용.

    CUSTOM 여부와 무관하게 active 컬럼만 매핑한다. 표시 여부는 프론트가 mode로 판단한다.
    """
    return {c.id: c for c in _get_workflow_columns(db, project_id, include_inactive=False)}


def _pick_column_for_status(cols: List[ProjectTaskColumn], status: str) -> Optional[ProjectTaskColumn]:
    """status를 표시할 컬럼 선택 — 첫 todo / 첫 in_progress / done(=마지막) / hold."""
    if not cols:
        return None
    if status == "done":
        finals = [c for c in cols if c.is_final_done]
        if finals:
            return finals[-1]
        dones = [c for c in cols if c.mapped_status == "done"]
        return dones[-1] if dones else cols[-1]
    if status == "hold":
        holds = [c for c in cols if c.mapped_status == "hold"]
        if holds:
            return holds[0]
        return cols[0]
    match = [c for c in cols if c.mapped_status == status]
    if match:
        return match[0]
    # in_progress fallback: 중간 컬럼
    if status == "in_progress":
        flow = [c for c in cols if c.mapped_status != "hold"]
        if len(flow) >= 3:
            return flow[len(flow) // 2]
        if len(flow) >= 2:
            return flow[1]
    return cols[0]


def _is_task_completed_for_checkpoint(
    task_order: int, task_progress: int, task_status: str,
    checkpoint_order: int, checkpoint_is_final: bool,
) -> bool:
    """해당 Task 가 체크포인트 단계를 '완료' 했는지 판단.

    - 체크포인트보다 뒤 단계에 있는 Task = 완료 (단계를 통과함).
    - 같은 단계에 있는 Task = 내부 progress 100%(또는 done) 일 때만 완료.
    - 앞 단계에 있는 Task = 미완료.
    - 단, 체크포인트가 최종 완료 컬럼(is_final_done)이면 도달=완료로 보아 기존 done 정책 유지.

    주의: '도달(>=)'이 아니라 '완료' 기준 — 같은 단계의 진행 중 Task 를 완료로 세면 안 된다.
    """
    if task_order > checkpoint_order:
        return True
    if task_order < checkpoint_order:
        return False
    # 같은 단계 — 최종 완료 컬럼은 도달=완료, 그 외에는 내부 progress 100% 일 때만 완료
    if checkpoint_is_final:
        return True
    if (task_status or "") == "done":
        return True
    return (task_progress or 0) >= 100


def _compute_completion_checkpoints(
    db: Session, project_id: int, cols: Optional[List[ProjectTaskColumn]] = None
) -> List[dict]:
    """중간 완료 체크포인트별 진행 지표를 계산한다 (보조 지표).

    - 최종 완료율/Dashboard Completion %/프로젝트 완료 여부와 무관 (is_final_done 로직 불변).
    - is_checkpoint=True 컬럼만 대상. 없으면 빈 리스트.
    - active(보관/삭제 제외) Task 기준. 분모(totalTaskCount)는 프로젝트 active Task 수로 공통.
    - Task의 현재 단계는 workflow_column_id로 판단하고, 없으면 status로 fallback.
    - 통과 기준: 체크포인트 단계를 '완료'한 Task (뒤 단계 = 완료, 같은 단계 = progress 100%).
      → _is_task_completed_for_checkpoint 참고. 프론트 completionCheckpoints.ts 와 동일 공식.
    """
    if cols is None:
        cols = _get_workflow_columns(db, project_id, include_inactive=False)
    checkpoint_cols = [c for c in cols if getattr(c, "is_checkpoint", False)]
    if not checkpoint_cols:
        return []

    sort_by_col_id = {c.id: c.sort_order for c in cols}
    tasks = db.query(Task).filter(
        Task.project_id == project_id, Task.archived_at.is_(None)
    ).all()
    total = len(tasks)

    # (단계 sort_order, 내부 progress, canonical status) 튜플로 Task 위치/진척을 캡처
    task_pos: List[tuple[int, int, str]] = []
    for t in tasks:
        so = sort_by_col_id.get(getattr(t, "workflow_column_id", None))
        if so is None:  # 컬럼 미배치 task는 status로 위치 추정
            fallback = _pick_column_for_status(cols, t.status or "todo")
            so = fallback.sort_order if fallback is not None else 0
        prog = t.progress if getattr(t, "progress", None) is not None else 0
        task_pos.append((so, prog, t.status or "todo"))

    result: List[dict] = []
    for c in sorted(checkpoint_cols, key=lambda x: x.sort_order):
        is_final = bool(getattr(c, "is_final_done", False))
        passed = sum(
            1 for (so, prog, st) in task_pos
            if _is_task_completed_for_checkpoint(so, prog, st, c.sort_order, is_final)
        )
        percent = round(passed / total * 100) if total else 0
        if total == 0 or passed == 0:
            status = "waiting"
        elif passed >= total:
            status = "completed"
        else:
            status = "in_progress"
        label = (getattr(c, "checkpoint_label", None) or "").strip() or f"{c.label} 완료"
        result.append({
            "checkpointColumnId": c.id,
            "checkpointLabel": label,
            "checkpointDescription": getattr(c, "checkpoint_description", None) or None,
            "targetColumnLabel": c.label,
            "targetSortOrder": c.sort_order,
            "passedTaskCount": passed,
            "totalTaskCount": total,
            "progressPercent": percent,
            "status": status,
        })
    return result


def _stage_checkpoint_counts(db: Session, task: "Task", stage_id: Optional[int] = None) -> tuple[int, int]:
    """현재(또는 지정) 단계의 '필수 단계 완료 체크포인트' 완료 수/전체 수.

    반환 (done, total). total==0 이면 이 단계에 필수 체크포인트가 없음(자동 이동 대상 아님).
    Task 자신의 작업노트(TaskActivity) 중 is_stage_checkpoint & checkpoint_required 이고
    checkpoint_stage_id 가 해당 단계인 항목만 대상으로 한다.
    """
    if task is None:
        return (0, 0)
    sid = stage_id if stage_id is not None else getattr(task, "workflow_column_id", None)
    if sid is None:
        return (0, 0)
    reqs = db.query(TaskActivityModel).filter(
        TaskActivityModel.task_id == task.id,
        TaskActivityModel.is_stage_checkpoint == True,   # noqa: E712
        TaskActivityModel.checkpoint_required == True,   # noqa: E712
        TaskActivityModel.checkpoint_stage_id == sid,
    ).all()
    total = len(reqs)
    done = sum(1 for a in reqs if a.checked)
    return (done, total)


def _advance_task_one_stage(
    db: Session, task: "Task", cur, cols: List["ProjectTaskColumn"],
    actor_user_id: Optional[int], log_message: str, log_meta: dict,
) -> Optional[dict]:
    """Task(+하위 Task)를 다음 flow 단계로 1단계 이동시키는 공용 로직.

    - hold 단계 제외 flow 순서에서 현재 다음 단계로 이동. 마지막이면 None(이동 없음).
    - 상위 이동 시 하위 Task 도 같은 단계로 함께 이동(같은 Board 단계 정책 유지).
    - 이력(ActivityLog ACTION_MOVED) + realtime(task_auto_advanced) 기록 후 commit.
    """
    flow = [c for c in cols if (c.mapped_status or "") != "hold"]
    idx = next((i for i, c in enumerate(flow) if c.id == cur.id), None)
    if idx is None or idx >= len(flow) - 1:
        return None
    nxt = flow[idx + 1]

    children = db.query(Task).filter(
        Task.parent_task_id == task.id, Task.archived_at.is_(None)
    ).all()

    from_col_id, from_status = cur.id, task.status
    now = datetime.now()
    task.workflow_column_id = nxt.id
    task.status = nxt.mapped_status or task.status
    task.updated_at = now
    for ch in children:
        ch.workflow_column_id = nxt.id
        ch.status = nxt.mapped_status or ch.status
        ch.updated_at = now

    meta = {
        "project_id": task.project_id,
        "from_column_id": from_col_id,
        "to_column_id": nxt.id,
        "from_status": from_status,
        "to_status": task.status,
        "moved_child_task_ids": [ch.id for ch in children],
        "moved_child_task_count": len(children),
    }
    meta.update(log_meta or {})
    try:
        actlog.log_activity(
            db, user_id=actor_user_id, action=actlog.ACTION_MOVED,
            entity_type=actlog.ENTITY_TASK, entity_id=task.id,
            message=log_message, meta=meta,
        )
    except Exception:  # noqa: BLE001
        pass

    try:
        sid = db.query(Project.space_id).filter(Project.id == task.project_id).scalar()
        if sid:
            emit_realtime_event(
                db, space_id=sid, event_type="task_auto_advanced",
                entity_type="task", entity_id=task.id,
                project_id=task.project_id, task_id=task.id,
                actor_user_id=actor_user_id,
            )
    except Exception:  # noqa: BLE001
        pass

    db.commit()
    return {
        "task_id": task.id,
        "from_column_id": from_col_id,
        "to_column_id": nxt.id,
        "to_column_label": nxt.label,
        "to_status": task.status,
        "moved_child_task_count": len(children),
    }


def _resolve_advance_context(db: Session, task: "Task"):
    """자동 이동 공통 전제 검증 → (proj, cols, cur) 또는 None.

    CUSTOM 모드 + 컬럼 존재 + Task 가 현재 단계에 배치돼 있어야 한다. 하위 Task 는 단독 이동
    금지이므로 이 함수는 상위/일반 Task 만 대상(호출부에서 child 여부를 먼저 처리)."""
    if task is None or getattr(task, "archived_at", None) is not None:
        return None
    proj = db.query(Project).filter(Project.id == task.project_id).first()
    if proj is None or (getattr(proj, "workflow_mode", "DEFAULT") or "DEFAULT") != "CUSTOM":
        return None
    cols = _get_workflow_columns(db, task.project_id, include_inactive=False)
    if not cols:
        return None
    cur = next((c for c in cols if c.id == getattr(task, "workflow_column_id", None)), None)
    if cur is None:
        return None
    return proj, cols, cur


def maybe_auto_advance_task_by_progress(
    db: Session,
    task: "Task",
    actor_user_id: Optional[int] = None,
    reason: str = "progress_threshold_reached",
) -> Optional[dict]:
    """(레거시/고급) 진행률 기준 자동 이동 — completion_rule_type=='progress_threshold' 인 단계만.

    - 하위 Task 로 호출되면 상위 Task 를 대신 평가(단독 이동 금지).
    - 상위 Task 본인 + 모든 활성 하위 Task progress 가 threshold 이상일 때만 다음 단계로 1단계 이동.
    """
    if task is None or getattr(task, "archived_at", None) is not None:
        return None
    if getattr(task, "parent_task_id", None) is not None:
        parent = db.query(Task).filter(
            Task.id == task.parent_task_id, Task.archived_at.is_(None)
        ).first()
        if parent is None:
            return None
        return maybe_auto_advance_task_by_progress(db, parent, actor_user_id, reason)

    ctx = _resolve_advance_context(db, task)
    if ctx is None:
        return None
    _, cols, cur = ctx
    if (getattr(cur, "completion_rule_type", None) or "") != "progress_threshold":
        return None
    if not bool(getattr(cur, "auto_advance_enabled", False)):
        return None
    threshold = getattr(cur, "auto_advance_threshold", None)
    if threshold is None or not (1 <= int(threshold) <= 100):
        return None
    threshold = int(threshold)
    if (task.progress or 0) < threshold:
        return None
    children = db.query(Task).filter(
        Task.parent_task_id == task.id, Task.archived_at.is_(None)
    ).all()
    if any((ch.progress or 0) < threshold for ch in children):
        return None

    flow = [c for c in cols if (c.mapped_status or "") != "hold"]
    idx = next((i for i, c in enumerate(flow) if c.id == cur.id), None)
    nxt = flow[idx + 1] if (idx is not None and idx < len(flow) - 1) else None
    result = _advance_task_one_stage(
        db, task, cur, cols, actor_user_id,
        log_message=(
            f"진행률 {task.progress}%가 기준({threshold}%)을 충족하여 "
            f"'{cur.label}' → '{nxt.label if nxt else '-'}' 단계로 자동 이동"
        ),
        log_meta={
            "event": "task_auto_advanced_by_progress",
            "reason": reason,
            "progress": task.progress,
            "threshold": threshold,
        },
    )
    if result is not None:
        result["progress"] = task.progress
        result["threshold"] = threshold
    return result


def maybe_auto_advance_task_by_checkpoints(
    db: Session,
    task: "Task",
    actor_user_id: Optional[int] = None,
) -> Optional[dict]:
    """단계 완료 체크포인트 기준 자동 이동 — completion_rule_type=='all_required_checkpoints' 인 단계만.

    - 하위 Task 는 단독 이동 금지 → 상위 Task 와 함께 이동하므로 child 로 호출되면 no-op.
    - 현재 단계에 연결된 '필수 단계 완료 체크포인트'가 1개 이상 있고 모두 완료되면 다음 단계로 1단계 이동.
    - 체크포인트 완료 이벤트(체크박스 체크)에서만 호출한다. 저장 버튼으로는 호출하지 않는다.
    """
    if task is None or getattr(task, "archived_at", None) is not None:
        return None
    # 하위 Task 는 상위와 함께 이동 → 단독 평가하지 않음
    if getattr(task, "parent_task_id", None) is not None:
        return None
    ctx = _resolve_advance_context(db, task)
    if ctx is None:
        return None
    _, cols, cur = ctx
    if (getattr(cur, "completion_rule_type", None) or "") != "all_required_checkpoints":
        return None

    done, total = _stage_checkpoint_counts(db, task, cur.id)
    if total == 0 or done < total:
        # 필수 체크포인트가 없거나(=자동 이동 안 함) 아직 다 완료되지 않음
        return None

    req_ids = [
        a.id for a in db.query(TaskActivityModel.id).filter(
            TaskActivityModel.task_id == task.id,
            TaskActivityModel.is_stage_checkpoint == True,   # noqa: E712
            TaskActivityModel.checkpoint_required == True,   # noqa: E712
            TaskActivityModel.checkpoint_stage_id == cur.id,
        ).all()
    ]
    flow = [c for c in cols if (c.mapped_status or "") != "hold"]
    idx = next((i for i, c in enumerate(flow) if c.id == cur.id), None)
    nxt = flow[idx + 1] if (idx is not None and idx < len(flow) - 1) else None
    result = _advance_task_one_stage(
        db, task, cur, cols, actor_user_id,
        log_message=(
            f"필수 완료 체크포인트({total}개)를 모두 충족하여 "
            f"'{cur.label}' → '{nxt.label if nxt else '-'}' 단계로 자동 이동"
        ),
        log_meta={
            "event": "task_auto_advanced_by_checkpoints",
            "reason": "stage_checkpoints_completed",
            "from_stage_id": cur.id,
            "to_stage_id": (nxt.id if nxt else None),
            "completed_checkpoint_ids": req_ids,
            "moved_by": "system",
        },
    )
    if result is not None:
        result["completed_checkpoint_count"] = total
    return result


def _default_columns_spec(db: Session, project_id: int) -> List[dict]:
    """DEFAULT→CUSTOM 최초 전환 시 생성할 기본 컬럼. hold 사용 중이면 Hold 추가."""
    has_hold = db.query(Task).filter(
        Task.project_id == project_id, Task.archived_at.is_(None), Task.status == "hold"
    ).first() is not None
    spec = [
        {"label": "To Do", "mapped_status": "todo"},
        {"label": "In Progress", "mapped_status": "in_progress"},
        {"label": "Done", "mapped_status": "done"},
    ]
    if has_hold:
        spec.append({"label": "Hold", "mapped_status": "hold"})
    return spec


@app.get("/api/projects/{project_id}/workflow-columns")
def get_workflow_columns(project_id: int, db: Session = Depends(get_db)):
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(404, "Project not found")
    cols = _get_workflow_columns(db, project_id, include_inactive=False)
    return {
        "mode": getattr(p, "workflow_mode", "DEFAULT") or "DEFAULT",
        "auto_progress_from_notes": bool(getattr(p, "auto_progress_from_notes", True)),
        "columns": [_wf_column_dict(c) for c in cols],
        "completion_checkpoints": _compute_completion_checkpoints(db, project_id, cols),
    }


@app.get("/api/projects/{project_id}/completion-checkpoints")
def get_completion_checkpoints(project_id: int, db: Session = Depends(get_db)):
    """중간 완료 체크포인트 진행 지표 (live). Task 이동에 즉시 반영되는 보조 지표.

    최종 완료율/Completion %와는 독립적이며, 체크포인트 컬럼이 없으면 빈 리스트.
    """
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(404, "Project not found")
    return {"completion_checkpoints": _compute_completion_checkpoints(db, project_id)}


@app.put("/api/projects/{project_id}/workflow-columns")
def save_workflow_columns(
    project_id: int, body: WorkflowColumnsSave,
    request: Request = None, db: Session = Depends(get_db),
):
    """Drawer 저장용 bulk upsert — 추가/이름변경/순서변경/삭제(soft)를 한 번에 처리.

    저장 시 sort_order 순으로 mapped_status/is_final_done 자동 계산.
    삭제(=기존엔 있는데 body에 없는 컬럼) 시 해당 컬럼에 Task가 있으면 409.
    """
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(404, "Project not found")
    state = load_state()
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if caller_id:
        check_task_edit_permission(db, state, project_id, caller_id)

    existing = {c.id: c for c in _get_workflow_columns(db, project_id, include_inactive=True)}
    incoming_ids = {c.id for c in body.columns if c.id is not None}
    # reorder로 컬럼의 mapped_status가 바뀌면 그 컬럼에 속한 Task의 canonical status도
    # 동기화해야 하므로 변경 전 값을 스냅샷한다 (예: 중간 컬럼이 마지막=done 으로 이동).
    old_mapped = {cid: c.mapped_status for cid, c in existing.items()}

    # 삭제 대상 = 기존 active 컬럼 중 body에 없는 것 → Task 있으면 차단, 없으면 soft delete
    for col in existing.values():
        if col.active and col.id not in incoming_ids:
            task_cnt = db.query(Task).filter(
                Task.workflow_column_id == col.id, Task.archived_at.is_(None)
            ).count()
            if task_cnt > 0:
                raise HTTPException(
                    409,
                    f"'{col.label}' 단계에 Task가 있어 삭제할 수 없습니다. 먼저 다른 단계로 이동해 주세요.",
                )
            col.active = False

    # upsert
    saved: List[ProjectTaskColumn] = []
    for item in sorted(body.columns, key=lambda x: x.sort_order):
        is_cp = bool(item.is_checkpoint)
        cp_label = (item.checkpoint_label or "").strip()[:120] or None if is_cp else None
        cp_desc = _sanitize_column_description(item.checkpoint_description) if is_cp else None
        # 자동 이동: enabled 이고 threshold 가 1~100 일 때만 유효. 아니면 OFF + threshold NULL.
        aa_enabled = bool(item.auto_advance_enabled)
        aa_threshold = item.auto_advance_threshold
        if aa_enabled and aa_threshold is not None and 1 <= int(aa_threshold) <= 100:
            aa_threshold = int(aa_threshold)
        else:
            aa_enabled = False
            aa_threshold = None
        # 단계 완료 조건. 미지정 시 레거시 progress 설정 유무로 추정.
        rule = (item.completion_rule_type or "").strip()
        if rule not in ("all_required_checkpoints", "manual", "progress_threshold"):
            rule = "progress_threshold" if aa_enabled else "all_required_checkpoints"
        # progress_threshold 규칙이 아니면 progress 자동이동 필드는 클리어(데이터 정합)
        if rule != "progress_threshold":
            aa_enabled = False
            aa_threshold = None
        # 4차 주 컨트롤: auto_move_on_complete. 지정되면 completion_rule_type 도 동기화(back-compat).
        amoc = item.auto_move_on_complete
        if amoc is None:
            amoc = (rule == "all_required_checkpoints")
        else:
            amoc = bool(amoc)
            rule = "all_required_checkpoints" if amoc else "manual"
            aa_enabled = False
            aa_threshold = None
        if item.id is not None and item.id in existing:
            col = existing[item.id]
            col.label = item.label
            col.description = _sanitize_column_description(item.description)
            col.color = item.color
            col.sort_order = item.sort_order
            col.active = True
            col.is_checkpoint = is_cp
            col.checkpoint_label = cp_label
            col.checkpoint_description = cp_desc
            col.auto_advance_enabled = aa_enabled
            col.auto_advance_threshold = aa_threshold
            col.completion_rule_type = rule
            col.auto_move_on_complete = amoc
            if item.mapped_status == "hold":
                col.mapped_status = "hold"
        else:
            col = ProjectTaskColumn(
                project_id=project_id,
                label=item.label,
                description=_sanitize_column_description(item.description),
                color=item.color,
                sort_order=item.sort_order,
                mapped_status=("hold" if item.mapped_status == "hold" else "in_progress"),
                active=True,
                is_checkpoint=is_cp,
                checkpoint_label=cp_label,
                checkpoint_description=cp_desc,
                auto_advance_enabled=aa_enabled,
                auto_advance_threshold=aa_threshold,
                completion_rule_type=rule,
                auto_move_on_complete=amoc,
            )
            db.add(col)
        saved.append(col)

    db.flush()
    saved.sort(key=lambda c: c.sort_order)
    _apply_positional_mapping(saved)

    # mapped_status가 바뀐 컬럼에 속한 Task들의 canonical status를 함께 동기화
    # (마지막 컬럼이 바뀌는 reorder 등 → Dashboard/통계 done 카운트 정합성 유지)
    for col in saved:
        prev = old_mapped.get(col.id)
        if prev is not None and prev != col.mapped_status:
            db.query(Task).filter(
                Task.workflow_column_id == col.id, Task.archived_at.is_(None)
            ).update({Task.status: col.mapped_status}, synchronize_session=False)

    db.commit()
    cols = _get_workflow_columns(db, project_id, include_inactive=False)
    return {"columns": [_wf_column_dict(c) for c in cols]}


class StageCompletionIn(BaseModel):
    completed: bool = True


def _get_task_stage_completion(db: Session, task_id: int, stage_id: int) -> Optional["TaskStageCompletion"]:
    return db.query(TaskStageCompletion).filter(
        TaskStageCompletion.task_id == task_id,
        TaskStageCompletion.stage_id == stage_id,
    ).first()


@app.post("/api/tasks/{task_id}/stage-completion")
def declare_stage_completion(
    task_id: int, body: StageCompletionIn,
    request: Request = None, db: Session = Depends(get_db),
):
    """사용자 '단계 완료 선언' — 현재 단계 완료를 체크/해제한다.

    - completed=True: 완료 선언(stage_completion_declared) 기록.
      * 현재 단계 auto_move_on_complete=True 이고 다음 단계가 있으면 다음 단계로 1단계 자동 이동
        (auto_stage_transition, moved_by=system). child Task 는 상위와 함께 이동.
      * 아니면 현재 단계 유지(stage_completion_no_transition).
    - completed=False: 완료 선언 해제(현재 단계 유지).
    - 저장 버튼이 아니라 이 완료 체크 이벤트에서만 자동 이동을 검사한다. 한 번에 1단계.
    """
    state = load_state()
    t = db.query(Task).filter(Task.id == task_id, Task.archived_at.is_(None)).first()
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if caller_id:
        check_task_edit_permission(db, state, t.project_id, caller_id)

    proj = db.query(Project).filter(Project.id == t.project_id).first()
    if proj is None or (getattr(proj, "workflow_mode", "DEFAULT") or "DEFAULT") != "CUSTOM":
        raise HTTPException(status_code=400, detail="커스텀 워크플로우 프로젝트에서만 단계 완료를 사용할 수 있습니다.")

    stage_id = getattr(t, "workflow_column_id", None)
    cols = _get_workflow_columns(db, t.project_id, include_inactive=False)
    cur = next((c for c in cols if c.id == stage_id), None)
    if cur is None:
        raise HTTPException(status_code=400, detail="Task 가 워크플로우 단계에 배치되어 있지 않습니다.")

    # 완료 선언 upsert
    rec = _get_task_stage_completion(db, t.id, cur.id)
    now = datetime.now()
    if rec is None:
        rec = TaskStageCompletion(task_id=t.id, stage_id=cur.id)
        db.add(rec)
    rec.completed = bool(body.completed)
    rec.completed_at = now if body.completed else None
    rec.actor_user_id = caller_id
    db.flush()

    # 다음 flow 단계 존재 여부 (hold 제외)
    flow = [c for c in cols if (c.mapped_status or "") != "hold"]
    idx = next((i for i, c in enumerate(flow) if c.id == cur.id), None)
    has_next = idx is not None and idx < len(flow) - 1
    is_child = getattr(t, "parent_task_id", None) is not None

    auto_advanced = None
    if body.completed:
        # 완료 선언 이력
        try:
            actlog.log_activity(
                db, user_id=caller_id, action=actlog.ACTION_UPDATED,
                entity_type=actlog.ENTITY_TASK, entity_id=t.id,
                message=f"'{cur.label}' 단계 완료를 선언했습니다.",
                meta={
                    "event": "stage_completion_declared",
                    "stage_id": cur.id,
                    "checkpoint_id": rec.id,
                    "actor_user_id": caller_id,
                },
            )
        except Exception:  # noqa: BLE001
            pass

        # 자동 이동: 현재 단계 auto_move_on_complete + 다음 단계 존재 + child 아님
        if bool(getattr(cur, "auto_move_on_complete", True)) and has_next and not is_child:
            auto_advanced = _advance_task_one_stage(
                db, t, cur, cols, caller_id,
                log_message=f"단계 완료 선언으로 '{cur.label}' → '{flow[idx + 1].label}' 단계로 자동 이동",
                log_meta={
                    "event": "auto_stage_transition",
                    "reason": "user_declared_stage_complete",
                    "from_stage_id": cur.id,
                    "to_stage_id": flow[idx + 1].id,
                    "trigger_checkpoint_id": rec.id,
                    "moved_by": "system",
                },
            )
        else:
            # 이동 없이 완료만 선언(설정 OFF / 마지막 단계 / child)
            try:
                actlog.log_activity(
                    db, user_id=caller_id, action=actlog.ACTION_UPDATED,
                    entity_type=actlog.ENTITY_TASK, entity_id=t.id,
                    message=(
                        f"'{cur.label}' 단계 최종 완료 처리" if not has_next
                        else f"'{cur.label}' 단계 완료(현재 단계 유지)"
                    ),
                    meta={
                        "event": "stage_completion_no_transition",
                        "stage_id": cur.id,
                        "checkpoint_id": rec.id,
                        "is_last_stage": (not has_next),
                    },
                )
            except Exception:  # noqa: BLE001
                pass

    db.commit()
    db.refresh(t)
    result = task_dict(t, state)
    # 이동 후 현재 단계 기준으로 완료 상태 재계산
    cur_after = t.workflow_column_id
    rec_after = _get_task_stage_completion(db, t.id, cur_after) if cur_after else None
    result["current_stage_completed"] = bool(rec_after.completed) if rec_after else False
    if auto_advanced is not None:
        result["auto_advanced"] = auto_advanced
    return result


def _build_mode_preview(db: Session, project_id: int, target_mode: str) -> dict:
    """모드 전환 미리보기 — 쓰기 없음."""
    target_mode = (target_mode or "").upper()
    if target_mode == "CUSTOM":
        existing = _get_workflow_columns(db, project_id, include_inactive=True)
        if existing:
            restored = [c for c in existing]
            restored.sort(key=lambda c: c.sort_order)
            cols_preview = [
                {"label": c.label, "mapped_status": c.mapped_status, "restored": True}
                for c in restored
            ]
        else:
            cols_preview = [
                {"label": s["label"], "mapped_status": s["mapped_status"], "restored": False}
                for s in _default_columns_spec(db, project_id)
            ]
        # status별 task 배치 카운트
        rows = db.query(Task.status, func.count(Task.id)).filter(
            Task.project_id == project_id, Task.archived_at.is_(None)
        ).group_by(Task.status).all()
        status_counts = {s: c for s, c in rows}
        return {
            "target_mode": "CUSTOM",
            "columns": cols_preview,
            "task_buckets": status_counts,
            "auto_progress_note": "커스텀 모드에서는 작업노트 체크율에 따른 자동 상태 이동이 비활성화됩니다. 체크율은 진행률 표시용으로만 사용됩니다.",
        }
    return {
        "target_mode": "DEFAULT",
        "columns": [],
        "task_buckets": {},
        "auto_progress_note": "기본 모드에서는 작업노트 체크율에 따른 자동 상태 이동이 다시 활성화됩니다.",
    }


@app.post("/api/projects/{project_id}/workflow-mode/preview")
def preview_workflow_mode(
    project_id: int, body: WorkflowModeChange, db: Session = Depends(get_db),
):
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(404, "Project not found")
    return _build_mode_preview(db, project_id, body.target_mode)


@app.put("/api/projects/{project_id}/workflow-mode")
def change_workflow_mode(
    project_id: int, body: WorkflowModeChange,
    request: Request = None, db: Session = Depends(get_db),
):
    """모드 실제 적용 — 스냅샷 저장 후 컬럼 생성/복원 + Task 매핑. status는 보존."""
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(404, "Project not found")
    state = load_state()
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if caller_id:
        check_task_edit_permission(db, state, project_id, caller_id)

    target_mode = (body.target_mode or "").upper()
    if target_mode not in ("DEFAULT", "CUSTOM"):
        raise HTTPException(400, "target_mode must be DEFAULT or CUSTOM")

    # 변경 직전 스냅샷 저장 (되돌리기용)
    prev_cols = _get_workflow_columns(db, project_id, include_inactive=True)
    snap = ProjectWorkflowSnapshot(
        project_id=project_id,
        previous_mode=getattr(p, "workflow_mode", "DEFAULT") or "DEFAULT",
        next_mode=target_mode,
        snapshot_json={
            "workflow_mode": getattr(p, "workflow_mode", "DEFAULT") or "DEFAULT",
            "auto_progress_from_notes": bool(getattr(p, "auto_progress_from_notes", True)),
            "columns": [_wf_column_dict(c) for c in prev_cols],
        },
        created_by=caller_id,
    )
    db.add(snap)

    if target_mode == "CUSTOM":
        active_cols = [c for c in prev_cols if c.active]
        if not active_cols:
            inactive = [c for c in prev_cols if not c.active]
            if inactive:
                # 이전 CUSTOM 컬럼 복원
                for c in inactive:
                    c.active = True
                active_cols = inactive
            else:
                # 최초 전환 — 기본 컬럼 생성
                for i, s in enumerate(_default_columns_spec(db, project_id)):
                    col = ProjectTaskColumn(
                        project_id=project_id, label=s["label"],
                        mapped_status=s["mapped_status"], sort_order=i, active=True,
                    )
                    db.add(col)
                db.flush()
                active_cols = _get_workflow_columns(db, project_id, include_inactive=False)
        active_cols.sort(key=lambda c: c.sort_order)
        _apply_positional_mapping(active_cols)
        db.flush()

        # Task 매핑 — workflow_column_id 없는 task만 status 기준 배치 (status 보존)
        tasks = db.query(Task).filter(
            Task.project_id == project_id, Task.archived_at.is_(None)
        ).all()
        for t in tasks:
            if t.workflow_column_id is None:
                col = _pick_column_for_status(active_cols, t.status)
                if col is not None:
                    t.workflow_column_id = col.id

        p.workflow_mode = "CUSTOM"
        p.auto_progress_from_notes = False
    else:
        # CUSTOM→DEFAULT — 컬럼 hard delete 금지(보존), status/매핑 그대로
        p.workflow_mode = "DEFAULT"
        p.auto_progress_from_notes = True

    db.commit()
    db.refresh(snap)
    cols = _get_workflow_columns(db, project_id, include_inactive=False)
    return {
        "mode": p.workflow_mode,
        "auto_progress_from_notes": bool(p.auto_progress_from_notes),
        "columns": [_wf_column_dict(c) for c in cols],
        "snapshot_id": snap.id,
    }


@app.post("/api/projects/{project_id}/workflow-mode/undo")
def undo_workflow_mode(
    project_id: int, request: Request = None, db: Session = Depends(get_db),
):
    """가장 최근 스냅샷 기준으로 모드+컬럼 설정 복원 (1차: 설정 복구까지)."""
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(404, "Project not found")
    state = load_state()
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if caller_id:
        check_task_edit_permission(db, state, project_id, caller_id)

    snap = db.query(ProjectWorkflowSnapshot).filter(
        ProjectWorkflowSnapshot.project_id == project_id
    ).order_by(ProjectWorkflowSnapshot.created_at.desc(), ProjectWorkflowSnapshot.id.desc()).first()
    if not snap or not snap.snapshot_json:
        raise HTTPException(404, "복원할 워크플로우 스냅샷이 없습니다.")

    data = snap.snapshot_json
    p.workflow_mode = data.get("workflow_mode", "DEFAULT") or "DEFAULT"
    p.auto_progress_from_notes = bool(data.get("auto_progress_from_notes", True))

    snap_cols = {c["id"]: c for c in (data.get("columns") or []) if c.get("id") is not None}
    current = {c.id: c for c in _get_workflow_columns(db, project_id, include_inactive=True)}
    # 스냅샷에 있던 컬럼은 그 상태로 복원
    for cid, cdata in snap_cols.items():
        col = current.get(cid)
        if col is not None:
            col.label = cdata.get("label", col.label)
            col.description = cdata.get("description", col.description)
            col.mapped_status = cdata.get("mapped_status", col.mapped_status)
            col.sort_order = cdata.get("sort_order", col.sort_order)
            col.color = cdata.get("color", col.color)
            col.is_final_done = bool(cdata.get("is_final_done", col.is_final_done))
            col.active = bool(cdata.get("active", col.active))
    # 스냅샷 이후 생성된 컬럼은 비활성화
    for cid, col in current.items():
        if cid not in snap_cols:
            col.active = False

    db.delete(snap)
    db.commit()
    cols = _get_workflow_columns(db, project_id, include_inactive=False)
    return {
        "mode": p.workflow_mode,
        "auto_progress_from_notes": bool(p.auto_progress_from_notes),
        "columns": [_wf_column_dict(c) for c in cols],
    }

# =========================
# Task Attachments (URL attachment / sidecar)
# =========================
@app.get("/api/tasks/{task_id}/attachments")
def get_task_attachments(task_id: int):
    state = load_state()
    attachments = [a for a in state.get("attachments", []) if int(a.get("task_id")) == task_id]
    return {"attachments": attachments}

@app.post("/api/tasks/{task_id}/attachments")
def create_attachment(task_id: int, attachment: AttachmentCreate, db: Session = Depends(get_db)):
    state = load_state()

    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    attachments = state.get("attachments", [])
    new_att = attachment.model_dump()
    new_att["id"] = next_id(attachments)
    new_att["task_id"] = task_id
    new_att["created_at"] = datetime.now().isoformat()

    attachments.append(new_att)
    state["attachments"] = attachments
    save_state(state)
    return new_att

@app.patch("/api/attachments/{attachment_id}")
def update_attachment(
    attachment_id: int,
    payload: AttachmentUpdate,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """URL 첨부 편집 (이름/주소). 파일 첨부는 편집 대상이 아니다.

    권한: Task 수정 권한과 동일(check_task_edit_permission). 프론트 버튼 숨김과
    별개로 백엔드에서도 반드시 검증한다 → 권한 없으면 403.
    """
    state = load_state()
    att = next((a for a in state.get("attachments", []) if int(a.get("id")) == attachment_id), None)
    if not att:
        raise HTTPException(status_code=404, detail="Attachment not found")

    # 파일 첨부는 편집 불가 (이름/주소 개념이 URL 첨부와 다름)
    if att.get("type") == "file":
        raise HTTPException(status_code=400, detail="파일 첨부는 편집할 수 없습니다.")

    # 권한 검증: 연결된 Task 의 프로젝트 기준
    task_id = att.get("task_id")
    task = db.query(Task).filter(Task.id == int(task_id)).first() if task_id else None
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    check_task_edit_permission(db, state, task.project_id, user_id)

    # URL 주소 갱신 (빈 값이면 기존 유지). http/https 형태 권장.
    if payload.url is not None:
        new_url = payload.url.strip()
        if new_url:
            if not (new_url.startswith("http://") or new_url.startswith("https://")):
                raise HTTPException(status_code=400, detail="URL은 http:// 또는 https:// 로 시작해야 합니다.")
            att["url"] = new_url

    # 표시 이름 갱신 (빈 값이면 기존 이름 유지)
    if payload.filename is not None:
        new_name = payload.filename.strip()
        if new_name:
            att["filename"] = new_name

    save_state(state)

    # DB 에도 레코드가 있으면 동기화 (best-effort)
    try:
        db_att = db.query(AttachmentModel).filter(AttachmentModel.id == attachment_id).first()
        if db_att:
            db_att.url = att.get("url")
            db_att.filename = att.get("filename")
            db.commit()
    except Exception:
        db.rollback()

    return att


@app.delete("/api/attachments/{attachment_id}")
def delete_attachment(attachment_id: int, db: Session = Depends(get_db)):
    """첨부파일 삭제. S3 + 로컬 + DB + sidecar 모두에서 삭제."""
    state = load_state()
    att = next((a for a in state.get("attachments", []) if int(a.get("id")) == attachment_id), None)

    # file 타입일 때만 S3/로컬 파일 정리 (S3 import를 여기서만 수행)
    if att and att.get("type") == "file" and att.get("stored_name"):
        task_id = att.get("task_id")
        try:
            from app.utils.s3.s3_utils import delete_from_s3, is_s3_configured, get_attachment_s3_key
            if is_s3_configured():
                s3_key = att.get("s3_key", "")
                if not s3_key and task_id:
                    s3_key = get_attachment_s3_key(
                        att.get("filename", ""), att["stored_name"], "task", int(task_id)
                    )
                if s3_key:
                    delete_from_s3(s3_key)
        except Exception as e:
            logging.getLogger("main").warning(f"S3 파일 삭제 실패 (무시): {e}")

        # 로컬 파일도 있으면 삭제
        if task_id:
            file_path = os.path.join(UPLOAD_DIR, f"tasks/{task_id}", att["stored_name"])
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass

    # sidecar에서 삭제
    state["attachments"] = [a for a in state.get("attachments", []) if int(a.get("id")) != attachment_id]
    save_state(state)

    # DB에서도 삭제 (URL 첨부 등 DB에 레코드가 있을 수 있음)
    try:
        db.query(AttachmentModel).filter(AttachmentModel.id == attachment_id).delete()
        db.commit()
    except Exception:
        db.rollback()

    return {"message": "Attachment deleted"}


def _log_upload_rejected(context_type, entity_id, user_id, filename, reason: str):
    """업로드 제한 위반을 운영 로그에 간단히 남긴다 (파일 내용은 저장하지 않음).

    context_type: 'task' / 'project' / 'voc' / 'space-image' 등.
    """
    try:
        from app.services.system_log import log_warning
        endpoint_map = {
            "task": f"/api/tasks/{entity_id}/files",
            "project": f"/api/projects/{entity_id}/files",
            "voc": f"/api/voc/{entity_id}/attachments",
            "space-image": f"/api/spaces/{entity_id}/images",
        }
        log_warning(
            "API",
            f"첨부 업로드 거부: {reason}",
            detail=f"context={context_type}, entity_id={entity_id}, user_id={user_id}, filename={filename!r}",
            endpoint=endpoint_map.get(context_type, f"/api/{context_type}/{entity_id}"),
            method="POST",
            status_code=400,
            user_id=user_id if user_id and user_id > 0 else None,
        )
    except Exception:
        pass


@app.get("/api/upload-policy")
def get_upload_policy_endpoint():
    """프론트가 즉시 검증에 사용할 업로드 제한 정책 (일반 첨부 + 인라인 이미지)."""
    from app.utils.upload_policy import get_upload_policies
    return get_upload_policies()


@app.post("/api/tasks/{task_id}/files")
async def upload_task_file(
    task_id: int,
    file: UploadFile = FastAPIFile(...),
    user_id: int = Query(default=1),
    request: Request = None,
    db: Session = Depends(get_db),
):
    """Upload a file attachment to a task. S3 우선, S3 실패 시 로컬 저장.

    예외는 모두 try/except 로 감싸 traceback 을 서버 로그에 남기고
    HTTPException(500, detail=...) 으로 변환한다. 프론트가 원인을 알 수 있도록
    detail 에 예외 타입/메시지를 포함하며, CORSMiddleware 가 응답 헤더를 부착한다.
    """
    from app.utils.s3.s3_utils import upload_attachment_bytes_to_s3, is_s3_configured
    from app.utils.upload_policy import GENERAL_ATTACHMENT_POLICY, UploadValidationError
    import logging as _logging
    import traceback as _traceback

    log = _logging.getLogger("main")
    saved_to_local = False  # noqa: F841 — 진단용
    try:
        state = load_state()
        t = db.query(Task).filter(Task.id == task_id).first()
        if not t:
            raise HTTPException(status_code=404, detail="Task not found")

        # Permission check: viewer cannot upload
        if user_id and user_id > 0:
            check_task_edit_permission(db, state, t.project_id, user_id)

        # 업로드 정책 1차 검증: 확장자/MIME (파일을 읽기 전에 차단)
        try:
            GENERAL_ATTACHMENT_POLICY.validate_type(file.filename, getattr(file, "content_type", "") or "")
        except UploadValidationError as ve:
            _log_upload_rejected("task", task_id, user_id, file.filename, str(ve))
            raise HTTPException(status_code=400, detail=str(ve))

        ext = os.path.splitext(file.filename or "")[1]
        stored_name = f"{uuid.uuid4().hex}{ext}"
        original_filename = file.filename or stored_name

        contents = await file.read()
        if not contents:
            raise HTTPException(status_code=400, detail="빈 파일입니다.")

        # 업로드 정책 2차 검증: 파일 크기 / task 첨부 개수 / task 총 용량 (S3 업로드 전)
        existing_files = [
            a for a in state.get("attachments", [])
            if int(a.get("task_id") or 0) == task_id and a.get("type") == "file"
        ]
        existing_count = len(existing_files)
        existing_total = sum(int(a.get("size") or 0) for a in existing_files)
        try:
            GENERAL_ATTACHMENT_POLICY.validate(
                filename=original_filename,
                size_bytes=len(contents),
                content_type=getattr(file, "content_type", "") or "",
                existing_count=existing_count,
                existing_total_bytes=existing_total,
            )
        except UploadValidationError as ve:
            _log_upload_rejected("task", task_id, user_id, original_filename, str(ve))
            raise HTTPException(status_code=400, detail=str(ve))

        s3_key = ""

        if is_s3_configured():
            # S3 업로드 시도 (s3fs 가 외부에서 raise 하더라도 이 try 가 잡는다)
            try:
                s3_result = upload_attachment_bytes_to_s3(
                    data=contents,
                    original_filename=original_filename,
                    stored_name=stored_name,
                    context_type="task",
                    context_id=task_id,
                )
            except Exception as s3e:
                log.warning(f"S3 업로드 호출 자체 예외: {s3e}")
                s3_result = {"success": False, "s3_key": "", "error": str(s3e)}

            if s3_result.get("success"):
                s3_key = s3_result["s3_key"]
            else:
                # S3 실패 → 로컬에 저장 (데이터 유실 방지)
                log.warning(f"S3 업로드 실패, 로컬 저장으로 전환: {s3_result.get('error')}")
                task_dir = os.path.join(UPLOAD_DIR, "tasks", str(task_id))
                os.makedirs(task_dir, exist_ok=True)
                with open(os.path.join(task_dir, stored_name), "wb") as f:
                    f.write(contents)
                saved_to_local = True
        else:
            # S3 미설정 → 로컬 저장
            task_dir = os.path.join(UPLOAD_DIR, "tasks", str(task_id))
            os.makedirs(task_dir, exist_ok=True)
            with open(os.path.join(task_dir, stored_name), "wb") as f:
                f.write(contents)
            saved_to_local = True

        # 이미지면 width/height/sha256/format 캡처 (AI readiness/중복식별 기반).
        from app.utils.image_meta import compute_image_meta
        _img_meta = compute_image_meta(contents, getattr(file, "content_type", "") or "", original_filename)

        attachments = state.get("attachments", [])
        new_att = {
            "id": next_id(attachments),
            "task_id": task_id,
            "url": f"/api/tasks/{task_id}/files/{stored_name}/download",
            "filename": original_filename,
            "stored_name": stored_name,
            "type": "file",
            "size": len(contents),
            "size_bytes": len(contents),
            "content_type": getattr(file, "content_type", "") or "",
            "uploaded_by": user_id,
            "created_at": datetime.now().isoformat(),
            "s3_key": s3_key,
            # AI readiness 메타 (비이미지면 width/height/format=None)
            "width": _img_meta["width"],
            "height": _img_meta["height"],
            "sha256": _img_meta["sha256"],
            "format": _img_meta["format"],
        }
        attachments.append(new_att)
        state["attachments"] = attachments
        save_state(state)
        return new_att

    except HTTPException:
        # 의도된 4xx/5xx 는 그대로 통과
        raise
    except Exception as e:
        # 모든 예상치 못한 예외 → traceback 로깅 + 친절한 500 메시지 반환
        log.error(
            f"upload_task_file 실패 (task_id={task_id}, user_id={user_id}, "
            f"filename={getattr(file, 'filename', None)!r}): {type(e).__name__}: {e}\n"
            f"{_traceback.format_exc()}"
        )
        raise HTTPException(
            status_code=500,
            detail=f"파일 업로드 처리 실패: {type(e).__name__}: {e}",
        )


_IMAGE_EXT_MIME = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
    ".bmp": "image/bmp",
    ".svg": "image/svg+xml",
}


def _attachment_response_meta(filename: str, stored_name: str):
    """이미지 확장자면 inline + image/* MIME 으로 응답해 <img> 태그가 그대로 렌더되게 한다.
    그 외 파일은 기존대로 attachment 다운로드."""
    import urllib.parse
    ext = os.path.splitext(stored_name or filename or "")[1].lower()
    mime = _IMAGE_EXT_MIME.get(ext)
    if mime:
        encoded_filename = urllib.parse.quote(filename)
        return mime, f"inline; filename*=UTF-8''{encoded_filename}"
    encoded_filename = urllib.parse.quote(filename)
    return "application/octet-stream", f"attachment; filename*=UTF-8''{encoded_filename}"


@app.get("/api/tasks/{task_id}/files/{stored_name}/download")
def download_task_file(task_id: int, stored_name: str):
    """Download a file attachment from a task. S3 우선, 없으면 로컬 fallback.
    이미지 파일은 inline + image/* MIME 으로 응답하여 <img> 렌더링 가능."""
    from fastapi.responses import Response
    from app.utils.s3.s3_utils import download_from_s3, is_s3_configured

    state = load_state()
    att = next(
        (a for a in state.get("attachments", [])
         if int(a.get("task_id")) == task_id and a.get("stored_name") == stored_name),
        None
    )
    if not att:
        raise HTTPException(status_code=404, detail="File not found")

    filename = att.get("filename", stored_name)
    s3_key = att.get("s3_key", "")
    media_type, content_disposition = _attachment_response_meta(filename, stored_name)

    # 1) S3에서 다운로드 시도
    if s3_key and is_s3_configured():
        data = download_from_s3(s3_key)
        if data is not None:
            return Response(
                content=data,
                media_type=media_type,
                headers={
                    "Content-Disposition": content_disposition,
                    "Content-Length": str(len(data)),
                    "Cache-Control": "private, max-age=86400",
                },
            )

    # 2) s3_key가 없지만 S3 설정 + 메타데이터로 key 재구성 시도
    if not s3_key and is_s3_configured():
        from app.utils.s3.s3_utils import get_attachment_s3_key
        reconstructed_key = get_attachment_s3_key(filename, stored_name, "task", task_id)
        data = download_from_s3(reconstructed_key)
        if data is not None:
            return Response(
                content=data,
                media_type=media_type,
                headers={
                    "Content-Disposition": content_disposition,
                    "Content-Length": str(len(data)),
                    "Cache-Control": "private, max-age=86400",
                },
            )

    # 3) 로컬 fallback
    file_path = os.path.join(UPLOAD_DIR, "tasks", str(task_id), stored_name)
    if os.path.exists(file_path):
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type=media_type,
            content_disposition_type="inline" if media_type.startswith("image/") else "attachment",
        )

    raise HTTPException(status_code=404, detail="File not found")

# =========================
# Space Images (단발 일정 설명 등 task 없는 리치 에디터 이미지 업로드)
# =========================
@app.post("/api/spaces/{space_id}/images")
async def upload_space_image(
    space_id: int,
    file: UploadFile = FastAPIFile(...),
    user_id: int = Query(default=1),
    db: Session = Depends(get_db),
):
    """공간 단위 이미지 업로드. 캘린더 단발 일정 설명처럼 task 가 아직 없는 리치 에디터에서
    Ctrl+V 붙여넣기 이미지를 저장하는 용도. S3 우선, 실패 시 로컬 저장.
    권한은 공간 멤버십(_require_space_member)으로 통제한다."""
    from app.utils.s3.s3_utils import upload_attachment_bytes_to_s3, is_s3_configured
    import logging as _logging
    import traceback as _traceback

    log = _logging.getLogger("main")
    try:
        from app.utils.upload_policy import INLINE_IMAGE_POLICY, UploadValidationError

        space = db.query(Space).filter(Space.id == space_id, Space.is_active == True).first()
        if not space:
            raise HTTPException(status_code=404, detail="Space not found")
        _require_space_member(db, space_id, user_id)

        # 인라인 이미지 정책 1차 검증: 확장자 (파일 읽기 전)
        try:
            INLINE_IMAGE_POLICY.validate_type(file.filename, getattr(file, "content_type", "") or "")
        except UploadValidationError as ve:
            _log_upload_rejected("space-image", space_id, user_id, file.filename, str(ve))
            raise HTTPException(status_code=400, detail=str(ve))

        ext = os.path.splitext(file.filename or "")[1]
        stored_name = f"{uuid.uuid4().hex}{ext}"
        original_filename = file.filename or stored_name

        contents = await file.read()
        if not contents:
            raise HTTPException(status_code=400, detail="빈 파일입니다.")

        # 인라인 이미지 정책 2차 검증: 파일 크기 (S3 업로드 전)
        try:
            INLINE_IMAGE_POLICY.validate(
                filename=original_filename,
                size_bytes=len(contents),
                content_type=getattr(file, "content_type", "") or "",
            )
        except UploadValidationError as ve:
            _log_upload_rejected("space-image", space_id, user_id, original_filename, str(ve))
            raise HTTPException(status_code=400, detail=str(ve))

        s3_key = ""
        if is_s3_configured():
            try:
                s3_result = upload_attachment_bytes_to_s3(
                    data=contents,
                    original_filename=original_filename,
                    stored_name=stored_name,
                    context_type="space",
                    context_id=space_id,
                )
            except Exception as s3e:
                log.warning(f"S3 업로드 호출 자체 예외: {s3e}")
                s3_result = {"success": False, "s3_key": "", "error": str(s3e)}

            if s3_result.get("success"):
                s3_key = s3_result["s3_key"]
            else:
                log.warning(f"S3 업로드 실패, 로컬 저장으로 전환: {s3_result.get('error')}")
                space_dir = os.path.join(UPLOAD_DIR, "spaces", str(space_id))
                os.makedirs(space_dir, exist_ok=True)
                with open(os.path.join(space_dir, stored_name), "wb") as f:
                    f.write(contents)
        else:
            space_dir = os.path.join(UPLOAD_DIR, "spaces", str(space_id))
            os.makedirs(space_dir, exist_ok=True)
            with open(os.path.join(space_dir, stored_name), "wb") as f:
                f.write(contents)

        # width/height/sha256/format 캡처 (인라인 이미지 AI readiness/중복식별 기반).
        from app.utils.image_meta import compute_image_meta
        _img_meta = compute_image_meta(contents, getattr(file, "content_type", "") or "", original_filename)

        state = load_state()
        space_images = state.get("space_images", [])
        new_img = {
            "id": next_id(space_images),
            "space_id": space_id,
            "url": f"/api/spaces/{space_id}/images/{stored_name}/download",
            "filename": original_filename,
            "stored_name": stored_name,
            "type": "image",
            "size": len(contents),
            "uploader_id": user_id,
            "created_at": datetime.now().isoformat(),
            "s3_key": s3_key,
            # AI readiness 메타
            "content_type": getattr(file, "content_type", "") or "",
            "width": _img_meta["width"],
            "height": _img_meta["height"],
            "sha256": _img_meta["sha256"],
            "format": _img_meta["format"],
        }
        space_images.append(new_img)
        state["space_images"] = space_images
        save_state(state)
        return new_img

    except HTTPException:
        raise
    except Exception as e:
        log.error(
            f"upload_space_image 실패 (space_id={space_id}, user_id={user_id}, "
            f"filename={getattr(file, 'filename', None)!r}): {type(e).__name__}: {e}\n"
            f"{_traceback.format_exc()}"
        )
        raise HTTPException(
            status_code=500,
            detail=f"이미지 업로드 처리 실패: {type(e).__name__}: {e}",
        )


@app.get("/api/spaces/{space_id}/images/{stored_name}/download")
def download_space_image(space_id: int, stored_name: str):
    """공간 이미지 서빙. S3 우선, 없으면 로컬 fallback. 이미지면 inline 렌더."""
    from fastapi.responses import Response
    from app.utils.s3.s3_utils import download_from_s3, is_s3_configured, get_attachment_s3_key

    state = load_state()
    img = next(
        (a for a in state.get("space_images", [])
         if int(a.get("space_id")) == space_id and a.get("stored_name") == stored_name),
        None
    )
    if not img:
        raise HTTPException(status_code=404, detail="Image not found")

    filename = img.get("filename", stored_name)
    s3_key = img.get("s3_key", "")
    media_type, content_disposition = _attachment_response_meta(filename, stored_name)

    if s3_key and is_s3_configured():
        data = download_from_s3(s3_key)
        if data is not None:
            return Response(
                content=data,
                media_type=media_type,
                headers={
                    "Content-Disposition": content_disposition,
                    "Content-Length": str(len(data)),
                    "Cache-Control": "private, max-age=86400",
                },
            )

    if not s3_key and is_s3_configured():
        reconstructed_key = get_attachment_s3_key(filename, stored_name, "space", space_id)
        data = download_from_s3(reconstructed_key)
        if data is not None:
            return Response(
                content=data,
                media_type=media_type,
                headers={
                    "Content-Disposition": content_disposition,
                    "Content-Length": str(len(data)),
                    "Cache-Control": "private, max-age=86400",
                },
            )

    file_path = os.path.join(UPLOAD_DIR, "spaces", str(space_id), stored_name)
    if os.path.exists(file_path):
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type=media_type,
            content_disposition_type="inline" if media_type.startswith("image/") else "attachment",
        )

    raise HTTPException(status_code=404, detail="Image not found")

# =========================
# Project Files (실파일 업로드 / sidecar metadata)
# =========================
@app.get("/api/projects/{project_id}/files")
def get_project_files(project_id: int, user_id: Optional[int] = None, db: Session = Depends(get_db)):
    state = load_state()

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    if user_id:
        check_project_permission(db, state, project_id, user_id, "file_view")

    files = [f for f in state.get("project_files", []) if int(f.get("project_id")) == project_id]
    files.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return {"files": files}

@app.post("/api/projects/{project_id}/files")
async def upload_project_file(
    project_id: int,
    file: UploadFile = FastAPIFile(...),
    user_id: int = Query(default=1),
    db: Session = Depends(get_db),
):
    """프로젝트 파일 업로드. S3 우선, S3 실패 시 로컬 저장."""
    from app.utils.s3.s3_utils import upload_attachment_bytes_to_s3, is_s3_configured
    from app.utils.upload_policy import GENERAL_ATTACHMENT_POLICY, UploadValidationError
    import logging as _logging

    state = load_state()

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    # 권한 확인 (업로드는 file_upload 권한 정책을 따름; 기본값 all)
    if user_id and user_id > 0:
        check_project_permission(db, state, project_id, user_id, "file_upload")

    # 업로드 정책 1차 검증: 확장자/MIME (파일 읽기 전)
    try:
        GENERAL_ATTACHMENT_POLICY.validate_type(file.filename, getattr(file, "content_type", "") or "")
    except UploadValidationError as ve:
        _log_upload_rejected("project", project_id, user_id, file.filename, str(ve))
        raise HTTPException(status_code=400, detail=str(ve))

    ext = os.path.splitext(file.filename or "")[1]
    stored_name = f"{uuid.uuid4().hex}{ext}"
    original_filename = file.filename or stored_name

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    # 업로드 정책 2차 검증: 파일 크기 / project 첨부 개수 / project 총 용량 (S3 업로드 전)
    existing_pf = [
        f for f in state.get("project_files", [])
        if int(f.get("project_id") or 0) == project_id
    ]
    try:
        GENERAL_ATTACHMENT_POLICY.validate(
            filename=original_filename,
            size_bytes=len(contents),
            content_type=getattr(file, "content_type", "") or "",
            existing_count=len(existing_pf),
            existing_total_bytes=sum(int(f.get("size") or 0) for f in existing_pf),
        )
    except UploadValidationError as ve:
        _log_upload_rejected("project", project_id, user_id, original_filename, str(ve))
        raise HTTPException(status_code=400, detail=str(ve))

    s3_key = ""
    saved_to_local = False

    if is_s3_configured():
        # S3 업로드 시도
        s3_result = upload_attachment_bytes_to_s3(
            data=contents,
            original_filename=original_filename,
            stored_name=stored_name,
            context_type="project",
            context_id=project_id,
        )
        if s3_result["success"]:
            s3_key = s3_result["s3_key"]
        else:
            # S3 실패 → 로컬에 저장 (데이터 유실 방지)
            _logging.getLogger("main").warning(f"S3 업로드 실패, 로컬 저장으로 전환: {s3_result['error']}")
            project_dir = os.path.join(UPLOAD_DIR, str(project_id))
            os.makedirs(project_dir, exist_ok=True)
            with open(os.path.join(project_dir, stored_name), "wb") as f:
                f.write(contents)
            saved_to_local = True
    else:
        # S3 미설정 → 로컬 저장
        project_dir = os.path.join(UPLOAD_DIR, str(project_id))
        os.makedirs(project_dir, exist_ok=True)
        with open(os.path.join(project_dir, stored_name), "wb") as f:
            f.write(contents)
        saved_to_local = True

    # 이미지면 width/height/sha256/format 캡처 (AI readiness/중복식별 기반).
    from app.utils.image_meta import compute_image_meta
    _img_meta = compute_image_meta(contents, getattr(file, "content_type", "") or "", original_filename)

    project_files = state.get("project_files", [])
    new_file = {
        "id": next_id(project_files),
        "project_id": project_id,
        "filename": original_filename,
        "stored_name": stored_name,
        "size": len(contents),
        "size_bytes": len(contents),
        "content_type": getattr(file, "content_type", "") or "",
        "uploader_id": user_id,
        "created_at": datetime.now().isoformat(),
        "s3_key": s3_key,
        # AI readiness 메타 (비이미지면 width/height/format=None)
        "width": _img_meta["width"],
        "height": _img_meta["height"],
        "sha256": _img_meta["sha256"],
        "format": _img_meta["format"],
    }
    project_files.append(new_file)
    state["project_files"] = project_files
    save_state(state)

    return new_file

@app.get("/api/projects/{project_id}/files/{file_id}/download")
def download_project_file(project_id: int, file_id: int, user_id: Optional[int] = None, db: Session = Depends(get_db)):
    """프로젝트 파일 다운로드. S3 우선, 없으면 로컬 fallback."""
    from fastapi.responses import Response
    from app.utils.s3.s3_utils import download_from_s3, is_s3_configured, get_attachment_s3_key

    state = load_state()

    if user_id:
        check_project_permission(db, state, project_id, user_id, "file_download")

    pf = next(
        (f for f in state.get("project_files", []) if int(f.get("id")) == file_id and int(f.get("project_id")) == project_id),
        None
    )
    if not pf:
        raise HTTPException(status_code=404, detail="File not found")

    filename = pf["filename"]
    stored_name = pf["stored_name"]
    s3_key = pf.get("s3_key", "")

    # 1) S3에서 다운로드 시도
    if s3_key and is_s3_configured():
        data = download_from_s3(s3_key)
        if data is not None:
            import urllib.parse
            encoded_filename = urllib.parse.quote(filename)
            return Response(
                content=data,
                media_type="application/octet-stream",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                    "Content-Length": str(len(data)),
                },
            )

    # 2) s3_key 없지만 S3 설정 시 key 재구성
    if not s3_key and is_s3_configured():
        reconstructed_key = get_attachment_s3_key(filename, stored_name, "project", project_id)
        data = download_from_s3(reconstructed_key)
        if data is not None:
            import urllib.parse
            encoded_filename = urllib.parse.quote(filename)
            return Response(
                content=data,
                media_type="application/octet-stream",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                    "Content-Length": str(len(data)),
                },
            )

    # 3) 로컬 fallback
    file_path = os.path.join(UPLOAD_DIR, str(project_id), stored_name)
    if os.path.exists(file_path):
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/octet-stream",
        )

    raise HTTPException(status_code=404, detail="File not found")

@app.delete("/api/projects/{project_id}/files/{file_id}")
def delete_project_file(project_id: int, file_id: int):
    """프로젝트 파일 삭제. S3 + 로컬 모두에서 삭제."""
    from app.utils.s3.s3_utils import delete_from_s3, is_s3_configured, get_attachment_s3_key

    state = load_state()
    project_files = state.get("project_files", [])

    pf = next((f for f in project_files if int(f.get("id")) == file_id and int(f.get("project_id")) == project_id), None)
    if not pf:
        raise HTTPException(status_code=404, detail="File not found")

    # S3 삭제
    if is_s3_configured():
        s3_key = pf.get("s3_key", "")
        if not s3_key:
            s3_key = get_attachment_s3_key(pf["filename"], pf["stored_name"], "project", project_id)
        if s3_key:
            delete_from_s3(s3_key)

    # 로컬 파일도 있으면 삭제 (기존 파일 정리)
    file_path = os.path.join(UPLOAD_DIR, str(project_id), pf["stored_name"])
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception:
            pass

    state["project_files"] = [f for f in project_files if int(f.get("id")) != file_id]
    save_state(state)
    return {"message": "File deleted"}

# =========================
# List Order (B-1)
# =========================
@app.get("/api/projects/{project_id}/list/order")
def get_list_order(project_id: int):
    state = load_state()
    orders = state.get("list_orders", {})
    return {"order": orders.get(str(project_id), [])}

@app.get("/api/projects/{project_id}/list/all-orders")
def get_all_list_orders(project_id: int):
    """Return all order keys for a project (root tasks, subproject order, sp task orders)"""
    state = load_state()
    orders = state.get("list_orders", {})
    result = {
        "root": orders.get(str(project_id), []),
        "sp_order": orders.get(f"sp_{project_id}", []),
    }
    # Collect all sptask_* orders
    for key, val in orders.items():
        if key.startswith("sptask_"):
            result[key] = val
    return result

@app.put("/api/projects/{project_id}/list/order")
def save_list_order(project_id: int, body: ListOrderUpdate):
    state = load_state()
    if "list_orders" not in state:
        state["list_orders"] = {}
    state["list_orders"][str(project_id)] = body.order
    save_state(state)
    return {"message": "Order saved"}

@app.get("/api/projects/{project_id}/subprojects/order")
def get_subproject_order(project_id: int):
    state = load_state()
    orders = state.get("list_orders", {})
    return {"order": orders.get(f"sp_{project_id}", [])}

@app.put("/api/projects/{project_id}/subprojects/order")
def save_subproject_order(project_id: int, body: ListOrderUpdate):
    state = load_state()
    if "list_orders" not in state:
        state["list_orders"] = {}
    state["list_orders"][f"sp_{project_id}"] = body.order
    save_state(state)
    return {"message": "Subproject order saved"}

@app.get("/api/subprojects/{sub_id}/tasks/order")
def get_sp_task_order(sub_id: int):
    state = load_state()
    orders = state.get("list_orders", {})
    return {"order": orders.get(f"sptask_{sub_id}", [])}

@app.put("/api/subprojects/{sub_id}/tasks/order")
def save_sp_task_order(sub_id: int, body: ListOrderUpdate):
    state = load_state()
    if "list_orders" not in state:
        state["list_orders"] = {}
    state["list_orders"][f"sptask_{sub_id}"] = body.order
    save_state(state)
    return {"message": "SP task order saved"}

# =========================
# SubProjects (C-2: DB)
# =========================
def _subproject_dict(sp: SubProjectModel) -> dict:
    return {
        "id": sp.id,
        "project_id": sp.project_id,
        "name": sp.name,
        "description": sp.description,
        "parent_id": sp.parent_id,
        "created_at": iso(sp.created_at),
    }

@app.get("/api/projects/{project_id}/subprojects")
def get_subprojects(project_id: int, db: Session = Depends(get_db)):
    subs = db.query(SubProjectModel).filter(SubProjectModel.project_id == project_id).all()
    return {"sub_projects": [_subproject_dict(s) for s in subs]}

@app.post("/api/projects/{project_id}/subprojects")
def create_subproject(project_id: int, sub: SubProjectCreate, db: Session = Depends(get_db)):
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    new_sub = SubProjectModel(
        project_id=project_id,
        name=sub.name,
        description=sub.description,
        parent_id=sub.parent_id,
    )
    db.add(new_sub)
    db.commit()
    db.refresh(new_sub)
    return _subproject_dict(new_sub)

@app.patch("/api/subprojects/{sub_id}")
def update_subproject(
    sub_id: int, updates: SubProjectUpdate,
    request: Request = None, db: Session = Depends(get_db),
):
    sp = db.query(SubProjectModel).filter(SubProjectModel.id == sub_id).first()
    if not sp:
        raise HTTPException(status_code=404, detail="SubProject not found")

    # 권한: Task 수정 권한과 동일 (owner/담당자/admin 가능, viewer 불가) — 토큰이 있으면 검증
    state = load_state()
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if caller_id:
        check_task_edit_permission(db, state, sp.project_id, caller_id)

    data = updates.model_dump(exclude_unset=True)

    # 이름 변경: trim + 빈값 금지 + 같은 Project 내 중복(자기 제외, 대소문자 무시) 금지
    if "name" in data:
        new_name = (data["name"] or "").strip()
        if not new_name:
            raise HTTPException(status_code=400, detail="Sub Project 이름을 입력해주세요.")
        siblings = db.query(SubProjectModel).filter(
            SubProjectModel.project_id == sp.project_id,
            SubProjectModel.id != sp.id,
        ).all()
        if any((s.name or "").strip().lower() == new_name.lower() for s in siblings):
            raise HTTPException(status_code=409, detail="이미 같은 이름의 Sub Project가 있습니다.")
        data["name"] = new_name

    for k, v in data.items():
        setattr(sp, k, v)
    db.commit()
    db.refresh(sp)
    return _subproject_dict(sp)

@app.delete("/api/subprojects/{sub_id}")
def delete_subproject(sub_id: int, db: Session = Depends(get_db)):
    sp = db.query(SubProjectModel).filter(SubProjectModel.id == sub_id).first()
    if not sp:
        raise HTTPException(status_code=404, detail="SubProject not found")

    # Unassign tasks from this subproject
    tasks = db.query(Task).filter(Task.sub_project_id == sub_id).all()
    for t in tasks:
        t.sub_project_id = None

    # Also clear sidecar task_meta references
    state = load_state()
    for key, meta in state.get("task_meta", {}).items():
        if meta.get("sub_project_id") == sub_id:
            meta["sub_project_id"] = None
    save_state(state)

    db.delete(sp)
    db.commit()
    return {"message": "SubProject deleted"}

# =========================
# Notes (C-5: DB + NoteMention)
# =========================
def _note_dict(n: NoteModel, mentioned_user_ids: List[int] = None) -> dict:
    return {
        "id": n.id,
        "project_id": n.project_id,
        "author_id": n.author_id,
        "content": n.content,
        "created_at": iso(n.created_at),
        "updated_at": iso(n.updated_at),
        "mentioned_user_ids": mentioned_user_ids or [],
    }

@app.get("/api/projects/{project_id}/notes")
def get_notes(project_id: int, user_id: Optional[int] = None, db: Session = Depends(get_db)):
    state = load_state()

    if user_id:
        check_project_access(db, state, project_id, user_id)

    # Auto-cleanup: delete notes older than 7 days
    from datetime import timedelta
    cutoff = datetime.utcnow() - timedelta(days=7)
    old_notes = db.query(NoteModel).filter(
        NoteModel.project_id == project_id,
        NoteModel.created_at < cutoff,
    ).all()
    if old_notes:
        old_ids = [n.id for n in old_notes]
        db.query(NoteMention).filter(NoteMention.note_id.in_(old_ids)).delete(synchronize_session=False)
        db.query(NoteModel).filter(NoteModel.id.in_(old_ids)).delete(synchronize_session=False)
        db.commit()
        # Also clean sidecar
        state = load_state()
        state["notes"] = [n for n in state.get("notes", []) if int(n.get("id")) not in set(old_ids)]
        save_state(state)

    db_notes = db.query(NoteModel).filter(NoteModel.project_id == project_id).order_by(NoteModel.created_at.desc()).all()

    users_map = {u.id: u for u in db.query(User).all()}
    result = []
    for n in db_notes:
        mentions = db.query(NoteMention).filter(NoteMention.note_id == n.id).all()
        d = _note_dict(n, [m.user_id for m in mentions])
        author = users_map.get(n.author_id or 0)
        d["author_name"] = author.username if author else "Unknown"
        d["author_color"] = author.avatar_color if author else "#ccc"
        result.append(d)

    # Also include legacy sidecar notes
    sidecar_notes = [nc.copy() for nc in state.get("notes", []) if int(nc.get("project_id")) == project_id]
    sidecar_ids = {n["id"] for n in result}
    for sn in sidecar_notes:
        if sn.get("id") not in sidecar_ids:
            author = users_map.get(int(sn.get("author_id", 0)))
            sn["author_name"] = author.username if author else "Unknown"
            sn["author_color"] = author.avatar_color if author else "#ccc"
            result.append(sn)

    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return {"notes": result}

@app.post("/api/projects/{project_id}/notes")
def create_note(
    project_id: int,
    note: NoteCreate,
    background_tasks: BackgroundTasks,
    user_id: int = Query(default=1),
    db: Session = Depends(get_db),
):
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    # C-5: Parse @mentions from content
    mentioned_user_ids = []
    mentioned_usernames = []
    mention_matches = re.findall(r'@(\S+)', note.content)
    if mention_matches:
        all_db_users = db.query(User).filter(User.is_active == True).all()
        for mention_text in mention_matches:
            mention_lower = mention_text.lower()
            for u in all_db_users:
                if (u.loginid and u.loginid.lower() == mention_lower) or \
                   (u.username and u.username.lower() == mention_lower):
                    if u.id not in mentioned_user_ids:
                        mentioned_user_ids.append(u.id)
                        mentioned_usernames.append(u.username or u.loginid)
                    break

    # C-5: Save note to DB
    new_note = NoteModel(
        project_id=project_id,
        author_id=user_id,
        content=note.content,
    )
    db.add(new_note)
    db.commit()
    db.refresh(new_note)

    # C-5: Save mentions to DB
    for uid in mentioned_user_ids:
        mention = NoteMention(note_id=new_note.id, user_id=uid)
        db.add(mention)
    if mentioned_user_ids:
        db.commit()

    result = _note_dict(new_note, mentioned_user_ids)
    result["mentioned_usernames"] = mentioned_usernames

    # Also save to sidecar for backward compat
    state = load_state()
    sidecar_note = {
        "id": new_note.id,
        "project_id": project_id,
        "author_id": user_id,
        "content": note.content,
        "created_at": iso(new_note.created_at),
        "updated_at": iso(new_note.updated_at),
        "mentioned_user_ids": mentioned_user_ids,
        "mentioned_usernames": mentioned_usernames,
    }
    notes_list = state.get("notes", [])
    notes_list.append(sidecar_note)
    state["notes"] = notes_list
    save_state(state)

    # Knox Messenger notification (best-effort, fire-and-forget)
    if mentioned_user_ids:
        try:
            from app.services import knox_messenger_service
            recipients = db.query(User).filter(User.id.in_(mentioned_user_ids)).all()
            author = db.query(User).filter(User.id == user_id).first()
            author_name = (author.username if author else None) or "Someone"
            project_name = p.name if getattr(p, "name", None) else f"프로젝트 #{project_id}"
            preview = re.sub(r'\s+', ' ', (note.content or '')).strip()
            if len(preview) > 120:
                preview = preview[:120] + "…"
            message = f"[Plan] {author_name} 님이 '{project_name}' 노트에서 회원님을 멘션했습니다.\n{preview}"
            metadata = {
                "project_id": project_id,
                "note_id": new_note.id,
                "author_id": user_id,
            }
            background_tasks.add_task(
                knox_messenger_service.send_mention_messages_for_users,
                recipients,
                message=message,
                link=None,
                source="plan",
                msg_type="mention",
                metadata=metadata,
                exclude_user_id=user_id,
            )
        except Exception:  # noqa: BLE001 — never break note save
            pass

    return {**result, "message": "메모가 등록되었습니다"}

@app.delete("/api/notes/{note_id}")
def delete_note(note_id: int, user_id: Optional[int] = None, db: Session = Depends(get_db)):
    note = db.query(NoteModel).filter(NoteModel.id == note_id).first()
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")

    # Permission: only the author or project member can delete
    if user_id and note.author_id != user_id:
        state = load_state()
        try:
            check_project_access(db, state, note.project_id, user_id)
        except HTTPException:
            raise HTTPException(status_code=403, detail="삭제 권한이 없습니다")

    db.query(NoteMention).filter(NoteMention.note_id == note_id).delete()
    db.query(NoteModel).filter(NoteModel.id == note_id).delete()
    db.commit()
    # Also clean sidecar
    state = load_state()
    state["notes"] = [n for n in state.get("notes", []) if int(n.get("id")) != note_id]
    save_state(state)
    return {"message": "메시지가 삭제되었습니다"}

# =========================
# Mentions (C-5: DB + sidecar fallback)
# =========================
@app.get("/api/mentions")
def get_mentions(user_id: int = Query(...), db: Session = Depends(get_db)):
    """사용자가 멘션된 모든 notes를 반환"""
    state = load_state()

    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user:
        return {"mentions": []}

    current_username_lower = (current_user.username or "").lower()
    current_loginid_lower = (current_user.loginid or "").lower()

    users_map = {u.id: u for u in db.query(User).all()}
    projects_map = {}
    for p in db.query(Project).filter(Project.archived_at.is_(None)).all():
        projects_map[p.id] = p
    # project_id → space_slug (딥링크/목록 클릭이 현재 store space와 무관하게 이동하도록)
    spaces_slug_map = {s.id: s.slug for s in db.query(Space.id, Space.slug).all()}

    result = []
    seen_ids = set()

    # C-5: Check DB NoteMention first
    db_mentions = db.query(NoteMention).filter(NoteMention.user_id == user_id).all()
    for nm in db_mentions:
        note = db.query(NoteModel).filter(NoteModel.id == nm.note_id).first()
        if not note:
            continue
        seen_ids.add(note.id)
        all_mentions = db.query(NoteMention).filter(NoteMention.note_id == note.id).all()
        d = _note_dict(note, [m.user_id for m in all_mentions])
        author = users_map.get(note.author_id or 0)
        project = projects_map.get(note.project_id)
        d["author_name"] = author.username if author else "Unknown"
        d["author_color"] = author.avatar_color if author else "#ccc"
        d["project_name"] = project.name if project else "Unknown"
        result.append(d)

    # Fallback: sidecar notes
    all_notes = state.get("notes", [])
    for n in all_notes:
        if n.get("id") in seen_ids:
            continue

        mentioned_ids = n.get("mentioned_user_ids", [])
        mentioned_names = n.get("mentioned_usernames", [])

        is_mentioned = user_id in mentioned_ids

        if not is_mentioned and mentioned_names:
            for name in mentioned_names:
                if name.lower() == current_username_lower or name.lower() == current_loginid_lower:
                    is_mentioned = True
                    break

        if not is_mentioned and not mentioned_ids and not mentioned_names:
            content = n.get("content", "")
            mentions_in_content = re.findall(r'@(\S+)', content)
            for m in mentions_in_content:
                if m.lower() == current_username_lower or m.lower() == current_loginid_lower:
                    is_mentioned = True
                    break

        if is_mentioned:
            author = users_map.get(int(n.get("author_id", 0)))
            pid = int(n.get("project_id", 0))
            project = projects_map.get(pid)
            result.append({
                **n,
                "author_name": author.username if author else "Unknown",
                "author_color": author.avatar_color if author else "#ccc",
                "project_name": project.name if project else "Unknown",
            })

    # ── TaskActivity 멘션 추가 (체크 완료된 항목은 제외) ──
    activity_mentions = db.query(TaskActivityMention).filter(TaskActivityMention.user_id == user_id).all()
    for am in activity_mentions:
        activity = db.query(TaskActivityModel).filter(TaskActivityModel.id == am.activity_id).first()
        if not activity:
            continue
        # 체크박스 타입이고 완료된 항목은 멘션에서 제외
        if (activity.block_type or "checkbox") == "checkbox" and activity.checked:
            continue
        task = db.query(Task).filter(Task.id == activity.task_id, Task.archived_at.is_(None)).first()
        if not task:
            continue
        project = projects_map.get(task.project_id)
        if not project:
            continue
        # Strip HTML for display
        import re as _re_html
        plain_content = _re_html.sub(r'<[^>]+>', '', activity.content or '')
        author_id = task.project_id  # 작업노트는 author가 없으므로 task 정보로 대체
        result.append({
            "id": f"activity-{activity.id}",
            "project_id": task.project_id,
            "author_id": 0,
            "content": plain_content,
            "created_at": activity.created_at.isoformat() if activity.created_at else "",
            "author_name": f"작업노트 [{task.title}]",
            "author_color": "#7C3AED",
            "project_name": project.name if project else "Unknown",
            "mentioned_user_ids": [user_id],
            "source": "activity",
            "task_id": task.id,
            "task_title": task.title,
        })

    # ── Task Comment 멘션 추가 (삭제된 댓글 / 보관된 Task 제외) ──
    comment_mentions = db.query(TaskCommentMention).filter(
        TaskCommentMention.mentioned_user_id == user_id
    ).all()
    for cm in comment_mentions:
        comment = db.query(TaskComment).filter(
            TaskComment.id == cm.comment_id, TaskComment.deleted_at.is_(None)
        ).first()
        if not comment:
            continue
        task = db.query(Task).filter(Task.id == comment.task_id, Task.archived_at.is_(None)).first()
        if not task:
            continue
        project = projects_map.get(task.project_id)
        if not project:
            continue
        author = users_map.get(comment.author_user_id or 0)
        import re as _re_html2
        plain_content = _re_html2.sub(r'<[^>]+>', '', comment.content or '')
        result.append({
            "id": f"comment-{comment.id}",
            "project_id": task.project_id,
            "author_id": comment.author_user_id or 0,
            "content": plain_content,
            "created_at": comment.created_at.isoformat() if comment.created_at else "",
            "author_name": author.username if author else "알 수 없음",
            "author_color": author.avatar_color if author else "#7C3AED",
            "project_name": project.name if project else "Unknown",
            "mentioned_user_ids": [user_id],
            "source": "task_comment",
            "source_label": "작업댓글",
            "task_id": task.id,
            "task_title": task.title,
            "comment_id": comment.id,
        })

    # 각 항목에 space_slug 부여 → 프론트가 현재 선택 space가 아닌 멘션의 space로 이동
    for item in result:
        if item.get("space_slug"):
            continue
        proj = projects_map.get(item.get("project_id"))
        if proj is not None:
            item["space_slug"] = spaces_slug_map.get(proj.space_id)

    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return {"mentions": result}


# =========================
# Mention detail (딥링크: /mentions?mentionId={ref}) + 사용자 알림설정
# =========================
def _mention_preview(content: str, limit: int = 280) -> str:
    plain = re.sub(r"<[^>]+>", " ", content or "")
    plain = re.sub(r"\s+", " ", plain).strip()
    return plain[:limit].rstrip() + "…" if len(plain) > limit else plain


@app.get("/api/mentions/{mention_ref}")
def get_mention_detail(mention_ref: str, request: Request = None, db: Session = Depends(get_db)):
    """멘션 딥링크용 상세 조회. mention_ref = 'comment-{id}' | 'activity-{id}'.

    권한: 호출자가 해당 멘션의 대상(TaskCommentMention/TaskActivityMention row 보유)이거나 admin.
    - 삭제된 댓글 / 보관된 Task / 미존재 → 404
    - 대상 아님(권한 없음) → 403
    """
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if not caller_id:
        raise HTTPException(status_code=401, detail="인증이 필요합니다.")
    is_admin = is_admin_like_role(get_user_role(db, caller_id))

    try:
        source_prefix, raw_id = mention_ref.split("-", 1)
        source_id = int(raw_id)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=404, detail="잘못된 mention 형식입니다.")

    def _space_info(project_id: Optional[int]):
        """(space_id, space_slug). 딥링크는 현재 store space와 무관하게 이 slug로 이동해야 한다."""
        if not project_id:
            return (None, None)
        row = (
            db.query(Space.id, Space.slug)
            .join(Project, Project.space_id == Space.id)
            .filter(Project.id == project_id)
            .first()
        )
        return (row[0], row[1]) if row else (None, None)

    if source_prefix == "comment":
        c = db.query(TaskComment).filter(
            TaskComment.id == source_id, TaskComment.deleted_at.is_(None)
        ).first()
        if not c:
            raise HTTPException(status_code=404, detail="삭제되었거나 존재하지 않는 언급입니다.")
        task = db.query(Task).filter(Task.id == c.task_id, Task.archived_at.is_(None)).first()
        if not task:
            raise HTTPException(status_code=404, detail="연결된 Task를 찾을 수 없습니다.")
        if not is_admin:
            is_target = db.query(TaskCommentMention.id).filter(
                TaskCommentMention.comment_id == c.id,
                TaskCommentMention.mentioned_user_id == caller_id,
            ).first()
            if not is_target:
                raise HTTPException(status_code=403, detail="이 언급에 접근할 권한이 없습니다.")
        author = db.query(User).filter(User.id == c.author_user_id).first() if c.author_user_id else None
        c_space_id, c_space_slug = _space_info(task.project_id)
        return {
            "id": mention_ref,
            "source_type": "task_comment",
            "space_id": c_space_id,
            "space_slug": c_space_slug,
            "project_id": task.project_id,
            "task_id": task.id,
            "comment_id": c.id,
            "work_note_id": None,
            "content_preview": _mention_preview(c.content or ""),
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "author": (
                {"id": author.id, "username": author.username, "loginid": author.loginid}
                if author else None
            ),
        }

    if source_prefix == "activity":
        a = db.query(TaskActivityModel).filter(TaskActivityModel.id == source_id).first()
        if not a:
            raise HTTPException(status_code=404, detail="삭제되었거나 존재하지 않는 언급입니다.")
        task = db.query(Task).filter(Task.id == a.task_id, Task.archived_at.is_(None)).first()
        if not task:
            raise HTTPException(status_code=404, detail="연결된 Task를 찾을 수 없습니다.")
        if not is_admin:
            is_target = db.query(TaskActivityMention.id).filter(
                TaskActivityMention.activity_id == a.id,
                TaskActivityMention.user_id == caller_id,
            ).first()
            if not is_target:
                raise HTTPException(status_code=403, detail="이 언급에 접근할 권한이 없습니다.")
        a_space_id, a_space_slug = _space_info(task.project_id)
        return {
            "id": mention_ref,
            "source_type": "task_work_note",
            "space_id": a_space_id,
            "space_slug": a_space_slug,
            "project_id": task.project_id,
            "task_id": task.id,
            "comment_id": None,
            "work_note_id": a.id,
            "content_preview": _mention_preview(a.content or ""),
            "created_at": a.created_at.isoformat() if a.created_at else None,
            "author": None,  # 작업노트는 작성자 필드가 없음
        }

    raise HTTPException(status_code=404, detail="지원하지 않는 mention 형식입니다.")


class NotificationPreferenceUpdate(BaseModel):
    mention_email_enabled: Optional[bool] = None
    task_mention_email_enabled: Optional[bool] = None
    work_note_mention_email_enabled: Optional[bool] = None
    task_comment_email_enabled: Optional[bool] = None


def _notification_pref_dict(pref) -> dict:
    return {
        "user_id": pref.user_id,
        "mention_email_enabled": bool(pref.mention_email_enabled),
        "task_mention_email_enabled": bool(pref.task_mention_email_enabled),
        "work_note_mention_email_enabled": bool(pref.work_note_mention_email_enabled),
        "task_comment_email_enabled": bool(pref.task_comment_email_enabled),
    }


def _resolve_me_user_id(request, db, user_id_q: Optional[int]) -> int:
    caller_id = resolve_user_id_from_token(request, db) if request else None
    if not caller_id:
        caller_id = user_id_q
    if not caller_id:
        raise HTTPException(status_code=401, detail="인증이 필요합니다.")
    return int(caller_id)


@app.get("/api/users/me/notification-preferences")
def get_my_notification_preferences(
    request: Request = None, user_id: Optional[int] = Query(default=None), db: Session = Depends(get_db)
):
    from app.services.notification_processor import get_or_create_preference
    uid = _resolve_me_user_id(request, db, user_id)
    pref = get_or_create_preference(db, uid)
    return _notification_pref_dict(pref)


@app.patch("/api/users/me/notification-preferences")
def update_my_notification_preferences(
    payload: NotificationPreferenceUpdate,
    request: Request = None, user_id: Optional[int] = Query(default=None), db: Session = Depends(get_db),
):
    from app.services.notification_processor import get_or_create_preference
    uid = _resolve_me_user_id(request, db, user_id)
    pref = get_or_create_preference(db, uid)
    for field in (
        "mention_email_enabled", "task_mention_email_enabled",
        "work_note_mention_email_enabled", "task_comment_email_enabled",
    ):
        val = getattr(payload, field, None)
        if val is not None:
            setattr(pref, field, bool(val))
    db.commit()
    db.refresh(pref)
    return _notification_pref_dict(pref)


# =========================
# Roadmap APIs (DB tasks + sidecar sub_projects/task_meta)
# =========================
@app.get("/api/roadmap")
def get_roadmap(
    project_id: int = Query(...),
    view: str = Query(default="month"),
    from_date: Optional[str] = Query(default=None, alias="from"),
    to_date: Optional[str] = Query(default=None, alias="to"),
    assignee_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    state = load_state()
    today_str = date.today().isoformat()

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    project = project_dict(p, state)
    sub_projects = get_subprojects_from_db(db, project_id)

    all_task_rows = db.query(Task).filter(Task.project_id == project_id, Task.archived_at.is_(None)).all()
    all_tasks = [task_dict(t, state) for t in all_task_rows]
    tasks = list(all_tasks)
    wf_map = _workflow_label_map(db, project_id)

    if assignee_id:
        tasks = [t for t in tasks if assignee_id in (t.get("assignee_ids") or [])]
    if status:
        tasks = [t for t in tasks if t.get("status") == status]
    if from_date:
        tasks = [
            t for t in tasks
            if (t.get("due_date") or "9999") >= from_date or (t.get("start_date") or "9999") >= from_date
        ]
    if to_date:
        tasks = [t for t in tasks if (t.get("start_date") or "0000") <= to_date]

    # 프로젝트 진행률
    active_tasks = [t for t in all_tasks if t.get("status") != "hold"]
    total = len(active_tasks)
    done = len([t for t in active_tasks if t.get("status") == "done"])
    if total > 0:
        progress_sum = sum(100 if t.get("status") == "done" else (t.get("progress", 0) or 0) for t in active_tasks)
        project_progress = round(progress_sum / total)
    else:
        project_progress = 0

    start_dates = [t["start_date"] for t in all_tasks if t.get("start_date")]
    due_dates = [t["due_date"] for t in all_tasks if t.get("due_date")]

    project_item = {
        "id": f"project-{project_id}",
        "type": "project",
        "name": project.get("name", ""),
        "start_date": min(start_dates) if start_dates else None,
        "due_date": max(due_dates) if due_dates else None,
        "status": "done" if project_progress == 100 and total > 0 else ("in_progress" if project_progress > 0 else "todo"),
        "progress": project_progress,
        "overdue": bool(due_dates and max(due_dates) < today_str and project_progress < 100),
        "children": [],
    }

    for sp in sub_projects:
        sp_tasks_filtered = [t for t in tasks if t.get("sub_project_id") == sp["id"]]
        sp_all_tasks = [t for t in all_tasks if t.get("sub_project_id") == sp["id"]]

        sp_active = [t for t in sp_all_tasks if t.get("status") != "hold"]
        sp_total = len(sp_active)
        sp_done = len([t for t in sp_active if t.get("status") == "done"])
        if sp_total > 0:
            sp_progress_sum = sum(100 if t.get("status") == "done" else (t.get("progress", 0) or 0) for t in sp_active)
            sp_progress = round(sp_progress_sum / sp_total)
        else:
            sp_progress = 0

        sp_starts = [t["start_date"] for t in sp_all_tasks if t.get("start_date")]
        sp_dues = [t["due_date"] for t in sp_all_tasks if t.get("due_date")]

        sp_item = {
            "id": f"subproject-{sp['id']}",
            "type": "subproject",
            "name": sp.get("name", ""),
            "start_date": min(sp_starts) if sp_starts else None,
            "due_date": max(sp_dues) if sp_dues else None,
            "status": "done" if sp_progress == 100 and sp_total > 0 else ("in_progress" if sp_progress > 0 else "todo"),
            "progress": sp_progress,
            "overdue": bool(sp_dues and max(sp_dues) < today_str and sp_progress < 100),
            "children": [],
        }

        for t in sp_tasks_filtered:
            t_overdue = bool(t.get("due_date") and t["due_date"] < today_str and t.get("status") != "done")
            _wf = wf_map.get(t.get("workflow_column_id"))
            sp_item["children"].append({
                "id": f"task-{t['id']}",
                "type": "task",
                # 프론트 다단계 트리(parent_task_id 재귀) 구성을 위한 숫자 id/부모 id.
                "task_id": t["id"],
                "parent_task_id": t.get("parent_task_id"),
                "name": t.get("title", ""),
                "start_date": t.get("start_date"),
                "due_date": t.get("due_date"),
                "start_date_tbd": t.get("start_date_tbd", False),
                "due_date_tbd": t.get("due_date_tbd", False),
                "status": t.get("status", "todo"),
                "progress": t.get("progress", 0),
                "overdue": t_overdue,
                "assignee_ids": t.get("assignee_ids", []),
                "workflow_label": _wf.label if _wf else None,
                "workflow_color": _wf.color if _wf else None,
            })

        project_item["children"].append(sp_item)

    root_tasks = [t for t in tasks if not t.get("sub_project_id")]
    for t in root_tasks:
        t_overdue = bool(t.get("due_date") and t["due_date"] < today_str and t.get("status") != "done")
        _wf = wf_map.get(t.get("workflow_column_id"))
        project_item["children"].append({
            "id": f"task-{t['id']}",
            "type": "task",
            # 프론트 다단계 트리(parent_task_id 재귀) 구성을 위한 숫자 id/부모 id.
            "task_id": t["id"],
            "parent_task_id": t.get("parent_task_id"),
            "name": t.get("title", ""),
            "start_date": t.get("start_date"),
            "due_date": t.get("due_date"),
            "start_date_tbd": t.get("start_date_tbd", False),
            "due_date_tbd": t.get("due_date_tbd", False),
            "status": t.get("status", "todo"),
            "progress": t.get("progress", 0),
            "overdue": t_overdue,
            "assignee_ids": t.get("assignee_ids", []),
            "workflow_label": _wf.label if _wf else None,
            "workflow_color": _wf.color if _wf else None,
        })

    return {
        "view": view,
        "from": from_date,
        "to": to_date,
        "items": [project_item],
    }

@app.get("/api/roadmap/global")
def get_global_roadmap(
    user_id: int = Query(...),
    view: str = Query(default="month"),
    space_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
):
    state = load_state()
    today_str = date.today().isoformat()

    # 현재 공간 기준으로 프로젝트 필터링
    pq = db.query(Project).filter(Project.archived_at.is_(None))
    if space_id:
        pq = pq.filter(Project.space_id == space_id)
    rows = pq.all()
    all_projects = [project_dict(p, state) for p in rows]

    # 사용자가 멤버인 프로젝트만 표시 (admin이어도 멤버 기준)
    authorized_pids = get_user_project_ids(db, state, user_id)
    authorized_pids |= get_user_public_project_ids(db, state, user_id)
    accessible = authorized_pids
    projects = [p for p in all_projects if p["id"] in accessible]

    task_rows = db.query(Task).filter(Task.archived_at.is_(None)).all()
    all_tasks = [task_dict(t, state) for t in task_rows]
    all_sub_projects = get_subprojects_from_db(db)

    items = []

    for project in projects:
        pid = project["id"]
        p_tasks = [t for t in all_tasks if t.get("project_id") == pid]
        p_subs = [s for s in all_sub_projects if int(s.get("project_id")) == pid]

        # CUSTOM 프로젝트만 workflow 라벨 오버레이 (DEFAULT는 canonical 유지).
        # workflow_label이 채워진 task는 프론트가 그대로 사용 → 별도 mode 플래그 불필요.
        is_custom = (project.get("workflow_mode") or "DEFAULT") == "CUSTOM"
        wf_map = _workflow_label_map(db, pid) if is_custom else {}

        active_tasks = [t for t in p_tasks if t.get("status") != "hold"]
        total = len(active_tasks)
        done = len([t for t in active_tasks if t.get("status") == "done"])
        if total > 0:
            progress_sum = sum(
                100 if t.get("status") == "done" else (t.get("progress", 0) or 0)
                for t in active_tasks
            )
            project_progress = round(progress_sum / total)
        else:
            project_progress = 0

        start_dates = [t["start_date"] for t in p_tasks if t.get("start_date")]
        due_dates = [t["due_date"] for t in p_tasks if t.get("due_date")]

        project_item = {
            "id": f"project-{pid}",
            "type": "project",
            "name": project.get("name", ""),
            "start_date": min(start_dates) if start_dates else None,
            "due_date": max(due_dates) if due_dates else None,
            "status": "done" if project_progress == 100 and total > 0 else ("in_progress" if project_progress > 0 else "todo"),
            "progress": project_progress,
            "overdue": bool(due_dates and max(due_dates) < today_str and project_progress < 100),
            "children": [],
        }

        for sp in p_subs:
            sp_tasks = [t for t in p_tasks if t.get("sub_project_id") == sp["id"]]
            sp_active = [t for t in sp_tasks if t.get("status") != "hold"]
            sp_total = len(sp_active)
            sp_done = len([t for t in sp_active if t.get("status") == "done"])

            if sp_total > 0:
                sp_progress = round(
                    sum(100 if t.get("status") == "done" else (t.get("progress", 0) or 0) for t in sp_active) / sp_total
                )
            else:
                sp_progress = 0

            sp_starts = [t["start_date"] for t in sp_tasks if t.get("start_date")]
            sp_dues = [t["due_date"] for t in sp_tasks if t.get("due_date")]

            sp_item = {
                "id": f"subproject-{sp['id']}",
                "type": "subproject",
                "name": sp.get("name", ""),
                "start_date": min(sp_starts) if sp_starts else None,
                "due_date": max(sp_dues) if sp_dues else None,
                "status": "done" if sp_progress == 100 and sp_total > 0 else ("in_progress" if sp_progress > 0 else "todo"),
                "progress": sp_progress,
                "overdue": bool(sp_dues and max(sp_dues) < today_str and sp_progress < 100),
                "children": [],
            }

            for t in sp_tasks:
                t_overdue = bool(t.get("due_date") and t["due_date"] < today_str and t.get("status") != "done")
                _wf = wf_map.get(t.get("workflow_column_id"))
                sp_item["children"].append({
                    "id": f"task-{t['id']}",
                    "type": "task",
                    "name": t.get("title", ""),
                    "start_date": t.get("start_date"),
                    "due_date": t.get("due_date"),
                    "status": t.get("status", "todo"),
                    "progress": t.get("progress", 0),
                    "overdue": t_overdue,
                    "assignee_ids": t.get("assignee_ids", []),
                    "workflow_label": _wf.label if _wf else None,
                    "workflow_color": _wf.color if _wf else None,
                })

            project_item["children"].append(sp_item)

        root_tasks = [t for t in p_tasks if not t.get("sub_project_id")]
        for t in root_tasks:
            t_overdue = bool(t.get("due_date") and t["due_date"] < today_str and t.get("status") != "done")
            _wf = wf_map.get(t.get("workflow_column_id"))
            project_item["children"].append({
                "id": f"task-{t['id']}",
                "type": "task",
                "name": t.get("title", ""),
                "start_date": t.get("start_date"),
                "due_date": t.get("due_date"),
                "status": t.get("status", "todo"),
                "progress": t.get("progress", 0),
                "overdue": t_overdue,
                "assignee_ids": t.get("assignee_ids", []),
                "workflow_label": _wf.label if _wf else None,
                "workflow_color": _wf.color if _wf else None,
            })

        items.append(project_item)

    return {"view": view, "items": items}

# =========================
# Dashboard Stats (DB + task_meta)
# =========================
@app.get("/api/stats")
def get_stats(user_id: Optional[int] = None, space_id: Optional[int] = None, db: Session = Depends(get_db)):
    state = load_state()

    pq = db.query(Project).filter(Project.archived_at.is_(None))
    if space_id:
        pq = pq.filter(Project.space_id == space_id)
    project_rows = pq.all()
    project_ids_in_scope = {p.id for p in project_rows}
    projects = [project_dict(p, state) for p in project_rows]

    tq = db.query(Task).filter(Task.archived_at.is_(None))
    if space_id:
        tq = tq.filter(Task.project_id.in_(project_ids_in_scope))
    all_task_rows = tq.all()
    tasks = [task_dict(t, state) for t in all_task_rows]

    # 권한 필터: 프로젝트 멤버(viewer 제외)인 경우만 표시
    if user_id:
        if is_admin_like_role(get_user_role(db, user_id)):
            # admin은 전체 접근
            pass
        else:
            # viewer가 아닌 멤버 역할의 프로젝트만 필터
            member_pids = set()
            db_memberships = db.query(ProjectMemberModel).filter(
                ProjectMemberModel.user_id == int(user_id),
                ProjectMemberModel.role != 'viewer'
            ).all()
            for m in db_memberships:
                member_pids.add(m.project_id)
            # sidecar fallback
            for m in state.get("project_members", []):
                if int(m.get("user_id")) == int(user_id) and m.get("role") != "viewer":
                    member_pids.add(int(m.get("project_id")))
            # owner
            for p in projects:
                meta = get_project_meta(state, p["id"])
                if int(meta.get("owner_id") or 0) == int(user_id):
                    member_pids.add(p["id"])
            # 공개 프로젝트는 미참여 멤버에게도 노출 (캘린더 가시성).
            # private 프로젝트는 그대로 멤버 한정.
            public_pids = {p["id"] for p in projects if (p.get("visibility") == "public")}
            visible_pids = member_pids | public_pids
            tasks = [t for t in tasks if t.get("project_id") in visible_pids]
            projects = [p for p in projects if p["id"] in visible_pids]

    total = len(tasks)
    in_progress = len([t for t in tasks if t.get("status") == "in_progress"])
    done = len([t for t in tasks if t.get("status") == "done"])
    todo = len([t for t in tasks if t.get("status") == "todo"])
    hold = len([t for t in tasks if t.get("status") == "hold"])

    project_stats = []
    for p in projects:
        # 시스템 프로젝트는 통계 행에서 제외 (단발 일정 task 는 위 집계엔 그대로 포함)
        if p.get("is_system"):
            continue
        p_tasks = [t for t in tasks if t.get("project_id") == p["id"]]
        p_total = len(p_tasks)
        p_done = len([t for t in p_tasks if t.get("status") == "done"])
        progress = round((p_done / p_total * 100) if p_total > 0 else 0)
        project_stats.append({
            "id": p["id"],
            "name": p.get("name", ""),
            "total": p_total,
            "done": p_done,
            "in_progress": len([t for t in p_tasks if t.get("status") == "in_progress"]),
            "todo": len([t for t in p_tasks if t.get("status") == "todo"]),
            "progress": progress,
        })

    today_str = date.today().isoformat()

    overdue = [t for t in tasks if t.get("due_date") and t["due_date"] < today_str and t.get("status") != "done"]
    upcoming = [t for t in tasks if t.get("due_date") and t["due_date"] >= today_str and t.get("status") != "done"]
    # 캘린더 가시성을 위해 upcoming 은 assignee 한정 필터를 두지 않는다.
    # (개인용 본인 task 는 my_tasks 로 별도 제공.)
    upcoming.sort(key=lambda x: x.get("due_date", ""))

    my_tasks = []
    if user_id:
        my_tasks = [t for t in tasks if user_id in (t.get("assignee_ids") or [])]

    return {
        "total": total,
        "in_progress": in_progress,
        "done": done,
        "todo": todo,
        "hold": hold,
        "project_stats": project_stats,
        "all_tasks": tasks,
        "overdue": overdue[:10],
        "upcoming": upcoming[:10],
        "my_tasks": my_tasks,
    }

# =========================
# Task Activities (Checklist)
# =========================
def _task_activity_dict(a: TaskActivityModel) -> dict:
    return {
        "id": a.id,
        "task_id": a.task_id,
        "block_type": a.block_type or "checkbox",
        "order_index": a.order_index,
        "content": a.content,
        "checked": a.checked,
        "checked_at": a.checked_at.isoformat() if a.checked_at else None,
        "style": a.style,
        "created_at": a.created_at.isoformat() if getattr(a, "created_at", None) else None,
        # 단계 완료 체크포인트 메타
        "is_stage_checkpoint": bool(getattr(a, "is_stage_checkpoint", False)),
        "checkpoint_stage_id": getattr(a, "checkpoint_stage_id", None),
        "checkpoint_required": bool(getattr(a, "checkpoint_required", False)),
    }


@app.get("/api/tasks/{task_id}/activities")
def get_task_activities(task_id: int, db: Session = Depends(get_db)):
    activities = db.query(TaskActivityModel).filter(
        TaskActivityModel.task_id == task_id
    ).order_by(TaskActivityModel.order_index.asc(), TaskActivityModel.id.asc()).all()
    return {"activities": [_task_activity_dict(a) for a in activities]}

def _sync_activity_mentions(db: Session, activity_id: int, content: str,
                            explicit_user_ids: Optional[List[int]] = None) -> set[int]:
    """작업노트의 @멘션 대상을 TaskActivityMention 에 동기화한다.

    - explicit_user_ids 가 주어지면(프론트 피커/검증된 id) 그 user_id 집합을 저장한다.
    - 없으면 기존처럼 본문 @loginid / @username 을 파싱해 매칭한다(하위호환).
    Returns the set of user_ids that were newly added by this call.
    """
    import re as _re
    matched_user_ids: set[int] = set()
    if explicit_user_ids is not None:
        for raw in explicit_user_ids:
            try:
                matched_user_ids.add(int(raw))
            except (TypeError, ValueError):
                continue
        if not matched_user_ids:
            db.query(TaskActivityMention).filter(TaskActivityMention.activity_id == activity_id).delete()
            return set()
    else:
        # Strip HTML tags to get plain text
        plain = _re.sub(r'<[^>]+>', ' ', content or '')
        mention_texts = _re.findall(r'@(\S+)', plain)
        if not mention_texts:
            db.query(TaskActivityMention).filter(TaskActivityMention.activity_id == activity_id).delete()
            return set()
        # Find matching users
        for mt in mention_texts:
            mt_lower = mt.lower()
            user = db.query(User).filter(
                User.is_active == True,
                (func.lower(User.loginid) == mt_lower) | (func.lower(User.username) == mt_lower)
            ).first()
            if user:
                matched_user_ids.add(user.id)
    # Sync: remove old, add new
    existing = {m.user_id for m in db.query(TaskActivityMention).filter(TaskActivityMention.activity_id == activity_id).all()}
    newly_added = matched_user_ids - existing
    for uid in newly_added:
        db.add(TaskActivityMention(activity_id=activity_id, user_id=uid))
    for uid in existing - matched_user_ids:
        db.query(TaskActivityMention).filter(TaskActivityMention.activity_id == activity_id, TaskActivityMention.user_id == uid).delete()
    return newly_added

@app.post("/api/tasks/{task_id}/activities")
def create_task_activity(
    task_id: int,
    background_tasks: BackgroundTasks,
    body: dict = Body(...),
    request: Request = None,
    db: Session = Depends(get_db),
):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    max_order = db.query(func.max(TaskActivityModel.order_index)).filter(
        TaskActivityModel.task_id == task_id
    ).scalar() or 0
    insert_order = body.get("order_index")
    if insert_order is not None:
        # Shift existing items down to make room
        db.query(TaskActivityModel).filter(
            TaskActivityModel.task_id == task_id,
            TaskActivityModel.order_index >= insert_order,
        ).update({TaskActivityModel.order_index: TaskActivityModel.order_index + 1})
        db.flush()
    # 단계 완료 체크포인트 메타 — 켤 때 stage 미지정이면 Task 현재 단계로 기본 설정
    cp_on = bool(body.get("is_stage_checkpoint", False))
    cp_stage = body.get("checkpoint_stage_id")
    if cp_on and cp_stage is None:
        cp_stage = getattr(task, "workflow_column_id", None)
    activity = TaskActivityModel(
        task_id=task_id,
        block_type=body.get("block_type", "checkbox"),
        order_index=insert_order if insert_order is not None else max_order + 1,
        content=body.get("content", ""),
        checked=body.get("checked", False),
        style=body.get("style"),
        is_stage_checkpoint=cp_on,
        checkpoint_required=bool(body.get("checkpoint_required", cp_on)),
        checkpoint_stage_id=(cp_stage if cp_on else None),
    )
    db.add(activity)
    db.commit()
    db.refresh(activity)
    _sync_task_progress(db, task_id)
    # ── 멘션 대상 확정(user_id 기준) + 작업노트 멘션 알림 이벤트 생성 ──
    from app.services.mention_resolver import (
        resolve_mention_targets, create_mention_notifications, DEFAULT_SEND_MENTION_EMAIL,
    )
    # 작성자(actor)는 세션 토큰에서 우선 확정(댓글과 동일) — 프론트 body author_id 는 fallback.
    # 이렇게 해야 actor_user_id 가 안정적으로 채워져 멘션 메일에 실제 작성자명이 표시된다("누군가" 방지).
    _actor_id = resolve_user_id_from_token(request, db) if request else None
    if not _actor_id:
        _actor_id = body.get("author_id") or body.get("user_id")
    try:
        _actor_id = int(_actor_id) if _actor_id else None
    except (TypeError, ValueError):
        _actor_id = None
    _explicit = body.get("mentioned_user_ids")
    _space_id = db.query(Project.space_id).filter(Project.id == task.project_id).scalar()
    _res = resolve_mention_targets(
        db, text=activity.content, explicit_user_ids=_explicit,
        space_id=_space_id, project_id=task.project_id, task_id=task_id,
    )
    if _explicit is not None:
        new_mention_user_ids = _sync_activity_mentions(
            db, activity.id, activity.content, explicit_user_ids=_res["resolved_user_ids"])
    else:
        new_mention_user_ids = _sync_activity_mentions(db, activity.id, activity.content)
    # 작업노트 정책: @멘션(표시)과 발송을 분리. 기본 OFF — 프론트가 send_mention_email=true 를
    # 명시했을 때만(전용 "멘션 남기기" 입력줄) 알림 이벤트/메일을 만든다. 블록 편집기의 @text 는
    # 단순 표시로만 저장되고 어떤 채널로도 알림을 보내지 않는다.
    _send_email = bool(body.get("send_mention_email"))
    if _send_email:
        create_mention_notifications(
            db, source_type="task_work_note", source_id=activity.id, task=task, space_id=_space_id,
            actor_user_id=_actor_id, resolved_user_ids=_res["resolved_user_ids"], send_email=True,
        )
    db.commit()

    # Knox Messenger notification (best-effort, fire-and-forget) — only newly added mentions.
    # 메일과 동일하게 opt-in(send_mention_email) 일 때만 발송한다.
    if _send_email and new_mention_user_ids:
        try:
            from app.services import knox_messenger_service
            author_id = _actor_id or body.get("author_id") or body.get("user_id")
            recipients = db.query(User).filter(User.id.in_(new_mention_user_ids)).all()
            author = db.query(User).filter(User.id == author_id).first() if author_id else None
            author_name = (author.username if author else None) or "Someone"
            task_title = task.title if getattr(task, "title", None) else f"태스크 #{task_id}"
            plain = re.sub(r'<[^>]+>', ' ', activity.content or '')
            preview = re.sub(r'\s+', ' ', plain).strip()
            if len(preview) > 120:
                preview = preview[:120] + "…"
            message = f"[Plan] {author_name} 님이 '{task_title}'에서 회원님을 멘션했습니다.\n{preview}"
            metadata = {
                "task_id": task_id,
                "activity_id": activity.id,
                "author_id": author_id,
            }
            background_tasks.add_task(
                knox_messenger_service.send_mention_messages_for_users,
                recipients,
                message=message,
                link=None,
                source="plan",
                msg_type="mention",
                metadata=metadata,
                exclude_user_id=author_id,
            )
        except Exception:  # noqa: BLE001
            pass

    return {
        "id": activity.id,
        "task_id": activity.task_id,
        "block_type": activity.block_type or "checkbox",
        "order_index": activity.order_index,
        "content": activity.content,
        "checked": activity.checked,
        "checked_at": activity.checked_at.isoformat() if activity.checked_at else None,
        "style": activity.style,
        "created_at": activity.created_at.isoformat() if activity.created_at else None,
        "is_stage_checkpoint": bool(activity.is_stage_checkpoint),
        "checkpoint_stage_id": activity.checkpoint_stage_id,
        "checkpoint_required": bool(activity.checkpoint_required),
    }


def _apply_checkpoint_meta(db: Session, activity: TaskActivityModel, body: dict) -> None:
    """요청 body 의 단계 완료 체크포인트 메타를 activity 에 반영.

    - is_stage_checkpoint 를 켜면 stage 미지정 시 Task 현재 단계로 기본 설정, required 기본 True.
    - 끄면 stage/required 초기화.
    """
    if "is_stage_checkpoint" in body:
        on = bool(body["is_stage_checkpoint"])
        activity.is_stage_checkpoint = on
        if on:
            activity.checkpoint_required = bool(body.get("checkpoint_required", True))
            sid = body.get("checkpoint_stage_id")
            if sid is None:
                t = db.query(Task).filter(Task.id == activity.task_id).first()
                sid = getattr(t, "workflow_column_id", None) if t else None
            activity.checkpoint_stage_id = sid
        else:
            activity.checkpoint_required = False
            activity.checkpoint_stage_id = None
    else:
        if "checkpoint_required" in body:
            activity.checkpoint_required = bool(body["checkpoint_required"])
        if "checkpoint_stage_id" in body:
            activity.checkpoint_stage_id = body["checkpoint_stage_id"]


@app.patch("/api/activities/{activity_id}")
def update_task_activity(
    activity_id: int,
    background_tasks: BackgroundTasks,
    body: dict = Body(...),
    request: Request = None,
    db: Session = Depends(get_db),
):
    activity = db.query(TaskActivityModel).filter(TaskActivityModel.id == activity_id).first()
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")

    if "content" in body:
        content = body["content"]
        # base64 이미지 직접 저장 시도 차단 (근본 원인 해결 유도)
        if "data:image/" in content and ";base64," in content:
            if len(content) > 100000: # 대략 100KB 이상
                raise HTTPException(
                    status_code=413,
                    detail="이미지 용량이 너무 큽니다. 이미지는 본문에 직접 저장할 수 없으며, 업로드 후 URL 방식으로 저장해야 합니다."
                )
        activity.content = content

    if "block_type" in body:
        activity.block_type = body["block_type"]
    if "checked" in body:
        activity.checked = body["checked"]
        if body["checked"]:
            activity.checked_at = datetime.now(KST)
        else:
            activity.checked_at = None
    if "style" in body:
        activity.style = body["style"]
    if "order_index" in body:
        activity.order_index = body["order_index"]
    db.commit()
    db.refresh(activity)
    task_progress = _sync_task_progress(db, activity.task_id)
    task_after = db.query(Task).filter(Task.id == activity.task_id).first()
    task_status = task_after.status if task_after else None
    new_mention_user_ids: set[int] = set()
    if "content" in body:
        from app.services.mention_resolver import (
            resolve_mention_targets, create_mention_notifications, DEFAULT_SEND_MENTION_EMAIL,
        )
        _actor_id = resolve_user_id_from_token(request, db) if request else None
        if not _actor_id:
            _actor_id = body.get("author_id") or body.get("user_id")
        try:
            _actor_id = int(_actor_id) if _actor_id else None
        except (TypeError, ValueError):
            _actor_id = None
        _explicit = body.get("mentioned_user_ids")
        _space_id = (
            db.query(Project.space_id).filter(Project.id == task_after.project_id).scalar()
            if task_after else None
        )
        _res = resolve_mention_targets(
            db, text=activity.content, explicit_user_ids=_explicit,
            space_id=_space_id, project_id=(task_after.project_id if task_after else None),
            task_id=activity.task_id,
        )
        if _explicit is not None:
            new_mention_user_ids = _sync_activity_mentions(
                db, activity.id, activity.content, explicit_user_ids=_res["resolved_user_ids"])
        else:
            new_mention_user_ids = _sync_activity_mentions(db, activity.id, activity.content)
        # 작업노트 정책: 발송은 opt-in(send_mention_email) 일 때만. 기본 OFF.
        _send_email = bool(body.get("send_mention_email"))
        if task_after is not None and _send_email:
            create_mention_notifications(
                db, source_type="task_work_note", source_id=activity.id, task=task_after, space_id=_space_id,
                actor_user_id=_actor_id, resolved_user_ids=_res["resolved_user_ids"], send_email=True,
            )
        db.commit()

    # Knox Messenger notification — only newly added mentions on this update (opt-in 일 때만).
    # new_mention_user_ids 는 content 미포함 요청에서 빈 set 이므로 먼저 평가해 _send_email 미정의를 회피.
    if new_mention_user_ids and bool(body.get("send_mention_email")):
        try:
            from app.services import knox_messenger_service
            author_id = _actor_id or body.get("author_id") or body.get("user_id")
            recipients = db.query(User).filter(User.id.in_(new_mention_user_ids)).all()
            author = db.query(User).filter(User.id == author_id).first() if author_id else None
            author_name = (author.username if author else None) or "Someone"
            task_title = task_after.title if task_after and getattr(task_after, "title", None) else f"태스크 #{activity.task_id}"
            plain = re.sub(r'<[^>]+>', ' ', activity.content or '')
            preview = re.sub(r'\s+', ' ', plain).strip()
            if len(preview) > 120:
                preview = preview[:120] + "…"
            message = f"[Plan] {author_name} 님이 '{task_title}'에서 회원님을 멘션했습니다.\n{preview}"
            metadata = {
                "task_id": activity.task_id,
                "activity_id": activity.id,
                "author_id": author_id,
            }
            background_tasks.add_task(
                knox_messenger_service.send_mention_messages_for_users,
                recipients,
                message=message,
                link=None,
                source="plan",
                msg_type="mention",
                metadata=metadata,
                exclude_user_id=author_id,
            )
        except Exception:  # noqa: BLE001
            pass

    return {
        "id": activity.id,
        "task_id": activity.task_id,
        "block_type": activity.block_type or "checkbox",
        "order_index": activity.order_index,
        "content": activity.content,
        "checked": activity.checked,
        "checked_at": activity.checked_at.isoformat() if activity.checked_at else None,
        "style": activity.style,
        "is_stage_checkpoint": bool(activity.is_stage_checkpoint),
        "checkpoint_stage_id": activity.checkpoint_stage_id,
        "checkpoint_required": bool(activity.checkpoint_required),
        # v3.12: Task Status 즉시 반영용 — 체크 해제로 progress 가 100% 미만이 되면
        #        프론트에서 selectedTask/store 캐시를 즉시 패치한다.
        "task_progress": task_progress,
        "task_status": task_status,
    }

@app.put("/api/tasks/{task_id}/activities/reorder")
def reorder_task_activities(task_id: int, body: dict = Body(...), db: Session = Depends(get_db)):
    order = body.get("order", [])  # list of activity IDs in new order
    for idx, activity_id in enumerate(order):
        db.query(TaskActivityModel).filter(
            TaskActivityModel.id == activity_id,
            TaskActivityModel.task_id == task_id,
        ).update({"order_index": idx})
    db.commit()
    return {"ok": True}

@app.delete("/api/activities/{activity_id}")
def delete_task_activity(activity_id: int, db: Session = Depends(get_db)):
    activity = db.query(TaskActivityModel).filter(TaskActivityModel.id == activity_id).first()
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    task_id = activity.task_id
    db.delete(activity)
    db.commit()
    _sync_task_progress(db, task_id)
    return {"ok": True}


# =========================
# Task Comments (담당자 외 인원 의견 제시 / 논의 기록)
#   - CRUD + @멘션 저장 + notification event 생성(발송은 보류).
#   - 실제 Knox 메일/메신저 발송은 미구현. docs/NOTIFICATION_KNOX_MAIL_TODO.md 참고.
# =========================

class TaskCommentCreate(BaseModel):
    content: str = ""
    attachment_ids: List[int] = []  # 붙여넣기-먼저-업로드된 pending 이미지 링크 대상
    # 프론트 멘션 피커가 고른 대상 user_id (Phase B). None 이면 본문 @token 파싱으로 fallback.
    mentioned_user_ids: Optional[List[int]] = None
    # 작성자 측 "메일 알림 보내기" (None 이면 DEFAULT_SEND_MENTION_EMAIL).
    send_mention_email: Optional[bool] = None

class TaskCommentUpdate(BaseModel):
    content: str
    mentioned_user_ids: Optional[List[int]] = None
    send_mention_email: Optional[bool] = None


def _task_comment_attachment_dict(a: TaskCommentAttachment) -> dict:
    return {
        "id": a.id,
        "comment_id": a.comment_id,
        "task_id": a.task_id,
        "filename": a.filename,
        "stored_name": a.stored_name,
        "content_type": a.content_type,
        "file_size": a.file_size,
        "width": a.width,
        "height": a.height,
        # 서빙은 task 스코프 + 불투명 stored_name(uuid) 기준 (task 첨부/VOC 첨부와 동일 정책)
        "url": f"/api/tasks/{a.task_id}/comment-images/{a.stored_name}/download",
        "created_at": iso(a.created_at) if a.created_at else None,
    }


def _comment_attachments_map(db: Session, comment_ids: List[int]) -> Dict[int, List[TaskCommentAttachment]]:
    if not comment_ids:
        return {}
    rows = (
        db.query(TaskCommentAttachment)
        .filter(TaskCommentAttachment.comment_id.in_(comment_ids))
        .order_by(TaskCommentAttachment.id.asc())
        .all()
    )
    out: Dict[int, List[TaskCommentAttachment]] = {}
    for r in rows:
        out.setdefault(r.comment_id, []).append(r)
    return out


def _task_comment_dict(c: TaskComment, users_map: Dict[int, User],
                       mentioned_user_ids: List[int] = None,
                       caller_id: int = None, can_manage: bool = False,
                       is_admin: bool = False,
                       attachments: List[TaskCommentAttachment] = None) -> dict:
    author = users_map.get(c.author_user_id)
    is_deleted = c.deleted_at is not None
    is_author = caller_id is not None and c.author_user_id == caller_id
    return {
        "id": c.id,
        "task_id": c.task_id,
        "author_user_id": c.author_user_id,
        "author_name": (author.username if author else None) if author else None,
        "author_color": (author.avatar_color if author else None) if author else None,
        # 삭제된 댓글은 본문을 감춘다.
        "content": "" if is_deleted else c.content,
        "deleted": is_deleted,
        "mentioned_user_ids": mentioned_user_ids or [],
        "attachments": [] if is_deleted else [_task_comment_attachment_dict(a) for a in (attachments or [])],
        "created_at": iso(c.created_at) if c.created_at else None,
        "updated_at": iso(c.updated_at) if c.updated_at else None,
        "edited_at": iso(c.edited_at) if c.edited_at else None,
        # 프론트 버튼 노출용(백엔드에서도 다시 검증한다)
        "can_edit": (not is_deleted) and (is_author or is_admin),
        "can_delete": (not is_deleted) and (is_author or can_manage or is_admin),
    }


def _sync_comment_mentions(db: Session, comment_id: int, content: str,
                           explicit_user_ids: Optional[List[int]] = None) -> set:
    """댓글의 @멘션 대상을 TaskCommentMention 에 동기화한다.

    - explicit_user_ids 가 주어지면(프론트 피커/검증된 id) 그 user_id 집합을 그대로 저장한다.
    - 없으면 기존처럼 본문에서 @loginid / @username 을 파싱해 매칭한다(하위호환).
    반환: 이번 호출로 '새로 추가된' mentioned_user_id 집합 (알림 이벤트 대상 선별용).
    """
    matched_user_ids: set = set()
    if explicit_user_ids is not None:
        for raw in explicit_user_ids:
            try:
                matched_user_ids.add(int(raw))
            except (TypeError, ValueError):
                continue
    else:
        plain = re.sub(r"<[^>]+>", " ", content or "")
        mention_texts = re.findall(r"@(\S+)", plain)
        for mt in mention_texts:
            mt_lower = mt.strip().rstrip(".,;:!?").lower()
            if not mt_lower:
                continue
            user = db.query(User).filter(
                User.is_active == True,
                (func.lower(User.loginid) == mt_lower) | (func.lower(User.username) == mt_lower),
            ).first()
            if user:
                matched_user_ids.add(user.id)
    existing = {
        m.mentioned_user_id
        for m in db.query(TaskCommentMention).filter(TaskCommentMention.comment_id == comment_id).all()
    }
    newly_added = matched_user_ids - existing
    for uid in newly_added:
        db.add(TaskCommentMention(comment_id=comment_id, mentioned_user_id=uid))
    for uid in existing - matched_user_ids:
        db.query(TaskCommentMention).filter(
            TaskCommentMention.comment_id == comment_id,
            TaskCommentMention.mentioned_user_id == uid,
        ).delete()
    return newly_added


def _comment_mentions_map(db: Session, comment_ids: List[int]) -> Dict[int, List[int]]:
    if not comment_ids:
        return {}
    rows = db.query(TaskCommentMention).filter(TaskCommentMention.comment_id.in_(comment_ids)).all()
    out: Dict[int, List[int]] = {}
    for r in rows:
        out.setdefault(r.comment_id, []).append(r.mentioned_user_id)
    return out


@app.get("/api/tasks/{task_id}/comments")
def list_task_comments(task_id: int, request: Request = None, db: Session = Depends(get_db)):
    """Task 댓글 목록 (오래된순). Task 조회 권한 필요."""
    t = db.query(Task).filter(Task.id == task_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")

    caller_id = resolve_user_id_from_token(request, db) if request else None
    if not caller_id:
        raise HTTPException(status_code=401, detail="인증이 필요합니다.")
    state = load_state()
    check_project_access(db, state, t.project_id, caller_id)  # 조회 권한(멤버/공개 프로젝트 멤버)

    is_admin = is_admin_like_role(get_user_role(db, caller_id))
    can_manage = can_manage_project(db, state, t.project_id, caller_id)

    rows = (
        db.query(TaskComment)
        .filter(TaskComment.task_id == task_id, TaskComment.deleted_at.is_(None))  # 삭제 댓글 제외
        .order_by(TaskComment.created_at.asc(), TaskComment.id.asc())
        .all()
    )
    author_ids = {r.author_user_id for r in rows if r.author_user_id}
    users_map = (
        {u.id: u for u in db.query(User).filter(User.id.in_(author_ids)).all()}
        if author_ids else {}
    )
    comment_ids = [r.id for r in rows]
    mentions_map = _comment_mentions_map(db, comment_ids)
    attachments_map = _comment_attachments_map(db, comment_ids)
    return {
        "items": [
            _task_comment_dict(c, users_map, mentions_map.get(c.id, []), caller_id, can_manage, is_admin,
                               attachments_map.get(c.id, []))
            for c in rows
        ],
        "can_comment": True,  # Task 조회 권한이 있으면 댓글 작성 가능(VOC: 타 인원 의견 제시)
    }


@app.post("/api/tasks/{task_id}/comments")
def create_task_comment(task_id: int, payload: TaskCommentCreate,
                        request: Request = None, db: Session = Depends(get_db)):
    """Task 댓글 작성. Task 조회 권한이 있으면 누구나 작성 가능(viewer 포함).

    작성 시 담당자/멘션 대상에게 notification event 를 남긴다(실제 발송은 보류).
    """
    t = db.query(Task).filter(Task.id == task_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")

    caller_id = resolve_user_id_from_token(request, db) if request else None
    if not caller_id:
        raise HTTPException(status_code=401, detail="인증이 필요합니다.")
    state = load_state()
    check_project_access(db, state, t.project_id, caller_id)

    content = (payload.content or "").strip()
    # pending 이미지(붙여넣기-먼저-업로드) 링크 대상 검증 — 이 task 스코프의 아직 미링크 첨부만 허용
    attachment_ids = [int(a) for a in (payload.attachment_ids or []) if str(a).strip()]
    pending_atts: List[TaskCommentAttachment] = []
    if attachment_ids:
        pending_atts = (
            db.query(TaskCommentAttachment)
            .filter(
                TaskCommentAttachment.id.in_(attachment_ids),
                TaskCommentAttachment.task_id == task_id,
                TaskCommentAttachment.comment_id.is_(None),
            )
            .all()
        )
    # 텍스트도 이미지도 없으면 등록 불가
    if not content and not pending_atts:
        raise HTTPException(status_code=400, detail="내용 또는 이미지를 입력해 주세요.")
    content = _strip_dangerous_html(content)

    comment = TaskComment(task_id=task_id, author_user_id=caller_id, content=content)
    db.add(comment)
    db.commit()
    db.refresh(comment)

    # pending 이미지를 이 댓글에 링크
    for a in pending_atts:
        a.comment_id = comment.id
    if pending_atts:
        db.commit()

    # ── 멘션 대상 확정(user_id 기준) + 알림 이벤트 생성 ──
    from app.services.mention_resolver import (
        resolve_mention_targets, create_mention_notifications, DEFAULT_SEND_MENTION_EMAIL,
    )
    space_id = db.query(Project.space_id).filter(Project.id == t.project_id).scalar()
    res = resolve_mention_targets(
        db, text=content, explicit_user_ids=payload.mentioned_user_ids,
        space_id=space_id, project_id=t.project_id, task_id=task_id,
    )
    # 페이지 표시용 저장: explicit 이면 검증된 resolved 로, 아니면 기존 텍스트 파싱 유지(@나를 언급 무영향)
    if payload.mentioned_user_ids is not None:
        mentioned_ids = _sync_comment_mentions(db, comment.id, content, explicit_user_ids=res["resolved_user_ids"])
    else:
        mentioned_ids = _sync_comment_mentions(db, comment.id, content)

    send_email = DEFAULT_SEND_MENTION_EMAIL if payload.send_mention_email is None else bool(payload.send_mention_email)
    targeted: set = set()
    # 1) 멘션 대상 (task_comment_mentioned) — **DB에 실제 저장된 task_comment_mentions 기준**으로 이벤트 생성.
    #    (예전엔 res["resolved_user_ids"](project/space 멤버 한정)만 썼는데, _sync_comment_mentions 는
    #     전역 매칭이라 멘션은 저장됐지만 이벤트가 0건이 되는 케이스가 있었다 — 운영 미발송 버그.)
    #    newly_added(mentioned_ids) 가 아니라 저장된 전체 멘션을 다시 조회해 사용(재저장/타이밍에 견고).
    #    self-mention / 중복 / 접근권한 게이트는 create_mention_notifications·processor 가 담당(로그로 관측).
    #    ⚠️ SessionLocal 은 autoflush=False 이므로, 방금 add 한 TaskCommentMention 이 아래 query 에
    #       보이려면 반드시 flush() 해야 한다(안 하면 empty → 이벤트 0건 버그).
    db.flush()
    persisted_mention_ids = {
        m.mentioned_user_id
        for m in db.query(TaskCommentMention).filter(TaskCommentMention.comment_id == comment.id).all()
        if m.mentioned_user_id is not None
    }
    logging.getLogger("mention").info(
        "[mention] comment=%s author=%s persisted_mentions=%s newly_added=%s send_email=%s",
        comment.id, caller_id, sorted(persisted_mention_ids), sorted(mentioned_ids), send_email,
    )
    targeted.update(create_mention_notifications(
        db, source_type="task_comment", source_id=comment.id, task=t, space_id=space_id,
        actor_user_id=caller_id, resolved_user_ids=persisted_mention_ids, send_email=send_email,
    ))
    # 2) 담당자 대상 (task_comment_created) — 본인 작성이면 skip, 멘션과 중복 skip
    for uid in (t.assignee_ids or []):
        try:
            uid = int(uid)
        except (TypeError, ValueError):
            continue
        if uid == caller_id or uid in targeted:
            continue
        create_notification_event(
            db, event_type="task_comment_created",
            task_id=task_id, project_id=t.project_id, space_id=space_id,
            actor_user_id=caller_id, target_user_id=uid,
            payload={"comment_id": comment.id},
        )
        targeted.add(uid)

    if space_id:
        emit_realtime_event(
            db, space_id=space_id, event_type="task_comment_created",
            entity_type="task_comment", entity_id=comment.id,
            project_id=t.project_id, task_id=task_id, actor_user_id=caller_id,
            payload={"task_id": task_id},
        )
    db.commit()
    db.refresh(comment)

    is_admin = is_admin_like_role(get_user_role(db, caller_id))
    can_manage = can_manage_project(db, state, t.project_id, caller_id)
    caller = db.query(User).filter(User.id == caller_id).first()
    users_map = {caller.id: caller} if caller else {}
    linked_atts = _comment_attachments_map(db, [comment.id]).get(comment.id, [])
    return _task_comment_dict(comment, users_map, list(mentioned_ids), caller_id, can_manage, is_admin, linked_atts)


@app.patch("/api/tasks/{task_id}/comments/{comment_id}")
def update_task_comment(task_id: int, comment_id: int, payload: TaskCommentUpdate,
                        request: Request = None, db: Session = Depends(get_db)):
    """댓글 수정. 작성자 본인 또는 admin/super_admin 만 가능."""
    c = db.query(TaskComment).filter(
        TaskComment.id == comment_id, TaskComment.task_id == task_id,
    ).first()
    if not c:
        raise HTTPException(status_code=404, detail="댓글을 찾을 수 없습니다.")
    if c.deleted_at is not None:
        raise HTTPException(status_code=400, detail="삭제된 댓글은 수정할 수 없습니다.")

    caller_id = resolve_user_id_from_token(request, db) if request else None
    if not caller_id:
        raise HTTPException(status_code=401, detail="인증이 필요합니다.")
    is_admin = is_admin_like_role(get_user_role(db, caller_id))
    if c.author_user_id != caller_id and not is_admin:
        raise HTTPException(status_code=403, detail="이 댓글을 수정할 권한이 없습니다.")

    content = (payload.content or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="내용을 입력해 주세요.")
    content = _strip_dangerous_html(content)

    c.content = content
    c.edited_at = datetime.now()
    c.updated_at = datetime.now()

    t = db.query(Task).filter(Task.id == task_id).first()
    space_id = db.query(Project.space_id).filter(Project.id == t.project_id).scalar() if t else None

    # 멘션 재동기화 + 새로 추가된 멘션에 대해서만 알림 이벤트 생성(기존 이벤트는 dedup)
    from app.services.mention_resolver import (
        resolve_mention_targets, create_mention_notifications, DEFAULT_SEND_MENTION_EMAIL,
    )
    res = resolve_mention_targets(
        db, text=content, explicit_user_ids=payload.mentioned_user_ids,
        space_id=space_id, project_id=(t.project_id if t else None), task_id=task_id,
    )
    if payload.mentioned_user_ids is not None:
        mentioned_ids = _sync_comment_mentions(db, c.id, content, explicit_user_ids=res["resolved_user_ids"])
    else:
        mentioned_ids = _sync_comment_mentions(db, c.id, content)
    if t is not None:
        send_email = DEFAULT_SEND_MENTION_EMAIL if payload.send_mention_email is None else bool(payload.send_mention_email)
        # DB에 저장된 전체 멘션 기준으로 이벤트 생성(신규만 추가 — 기존은 create_mention_notifications 내부 dedup).
        # ⚠️ autoflush=False 라 방금 add/delete 한 멘션 반영을 위해 flush 필요(안 하면 empty → 이벤트 0건).
        db.flush()
        persisted_mention_ids = {
            m.mentioned_user_id
            for m in db.query(TaskCommentMention).filter(TaskCommentMention.comment_id == c.id).all()
            if m.mentioned_user_id is not None
        }
        logging.getLogger("mention").info(
            "[mention] (update) comment=%s author=%s persisted_mentions=%s newly_added=%s",
            c.id, caller_id, sorted(persisted_mention_ids), sorted(mentioned_ids),
        )
        create_mention_notifications(
            db, source_type="task_comment", source_id=c.id, task=t, space_id=space_id,
            actor_user_id=caller_id, resolved_user_ids=persisted_mention_ids, send_email=send_email,
        )

    if space_id:
        emit_realtime_event(
            db, space_id=space_id, event_type="task_comment_updated",
            entity_type="task_comment", entity_id=c.id,
            project_id=(t.project_id if t else None), task_id=task_id, actor_user_id=caller_id,
            payload={"task_id": task_id},
        )
    db.commit()
    db.refresh(c)

    state = load_state()
    can_manage = can_manage_project(db, state, t.project_id, caller_id) if t else False
    all_mentions = _comment_mentions_map(db, [c.id]).get(c.id, list(mentioned_ids))
    author = db.query(User).filter(User.id == c.author_user_id).first()
    users_map = {author.id: author} if author else {}
    linked_atts = _comment_attachments_map(db, [c.id]).get(c.id, [])
    return _task_comment_dict(c, users_map, all_mentions, caller_id, can_manage, is_admin, linked_atts)


@app.delete("/api/tasks/{task_id}/comments/{comment_id}")
def delete_task_comment(task_id: int, comment_id: int,
                        request: Request = None, db: Session = Depends(get_db)):
    """댓글 삭제(soft delete). 작성자 본인 / 프로젝트 소유자·담당자 / admin 가능."""
    c = db.query(TaskComment).filter(
        TaskComment.id == comment_id, TaskComment.task_id == task_id,
    ).first()
    if not c:
        raise HTTPException(status_code=404, detail="댓글을 찾을 수 없습니다.")
    if c.deleted_at is not None:
        return {"message": "이미 삭제된 댓글입니다."}

    caller_id = resolve_user_id_from_token(request, db) if request else None
    if not caller_id:
        raise HTTPException(status_code=401, detail="인증이 필요합니다.")
    t = db.query(Task).filter(Task.id == task_id).first()
    state = load_state()
    is_admin = is_admin_like_role(get_user_role(db, caller_id))
    can_manage = can_manage_project(db, state, t.project_id, caller_id) if t else False
    if c.author_user_id != caller_id and not can_manage and not is_admin:
        raise HTTPException(status_code=403, detail="이 댓글을 삭제할 권한이 없습니다.")

    c.deleted_at = datetime.now()
    space_id = db.query(Project.space_id).filter(Project.id == t.project_id).scalar() if t else None
    if space_id:
        emit_realtime_event(
            db, space_id=space_id, event_type="task_comment_deleted",
            entity_type="task_comment", entity_id=c.id,
            project_id=(t.project_id if t else None), task_id=task_id, actor_user_id=caller_id,
            payload={"task_id": task_id},
        )
    db.commit()
    return {"message": "삭제되었습니다."}


# ── Task 댓글 이미지 (Ctrl+V 붙여넣기) ───────────────────────────────
#   base64 DB 저장 X. 원본 보관 X — 리사이즈+webp 최적화본만 S3(우선)/로컬에 저장.
#   붙여넣기-먼저-업로드: comment_id 는 NULL(task 스코프 pending)로 올라간 뒤 댓글 등록 시 링크된다.
def _store_task_comment_image_bytes(task_id: int, raw: bytes, original_filename: str) -> dict:
    """댓글 이미지 원본을 최적화 후 S3(우선)/로컬에 저장하고 TaskCommentAttachment 메타를 돌려준다.
    이미지 최적화 로직은 VOC 첨부와 동일(_optimize_voc_image) 재사용."""
    from app.utils.s3.s3_utils import upload_attachment_bytes_to_s3, is_s3_configured
    import logging as _logging
    _log = _logging.getLogger("main")

    optimized, stored_ext, content_type, w, h = _optimize_voc_image(raw, original_filename)
    stored_name = f"{uuid.uuid4().hex}{stored_ext}"

    s3_key = ""
    saved = False
    if is_s3_configured():
        try:
            s3_result = upload_attachment_bytes_to_s3(
                data=optimized,
                original_filename=original_filename,
                stored_name=stored_name,
                context_type="task-comment",
                context_id=task_id,
            )
        except Exception as s3e:
            _log.warning(f"Task 댓글 이미지 S3 업로드 예외: {s3e}")
            s3_result = {"success": False, "s3_key": "", "error": str(s3e)}
        if s3_result.get("success"):
            s3_key = s3_result["s3_key"]
            saved = True
        else:
            _log.warning(f"Task 댓글 이미지 S3 실패, 로컬 저장 전환: {s3_result.get('error')}")

    if not saved:
        cdir = os.path.join(UPLOAD_DIR, "task-comments", str(task_id))
        os.makedirs(cdir, exist_ok=True)
        with open(os.path.join(cdir, stored_name), "wb") as f:
            f.write(optimized)

    return {
        "stored_name": stored_name,
        "content_type": content_type,
        "file_size": len(optimized),
        "width": w or None,
        "height": h or None,
        "s3_key": s3_key or None,
    }


@app.post("/api/tasks/{task_id}/comment-images")
async def upload_task_comment_image(
    task_id: int,
    file: UploadFile = FastAPIFile(...),
    request: Request = None,
    db: Session = Depends(get_db),
):
    """Task 댓글에 붙일 이미지 업로드(Ctrl+V). 댓글 등록 전 pending(comment_id=NULL)으로 저장.
    권한은 댓글 작성 권한과 동일(Task 조회 권한). 이미지당 10MB / png·jpeg·webp 만 허용."""
    t = db.query(Task).filter(Task.id == task_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")

    caller_id = resolve_user_id_from_token(request, db) if request else None
    if not caller_id:
        raise HTTPException(status_code=401, detail="인증이 필요합니다.")
    state = load_state()
    check_project_access(db, state, t.project_id, caller_id)

    from app.utils.upload_policy import VOC_IMAGE_POLICY, UploadValidationError
    original_filename = file.filename or "image"
    try:
        VOC_IMAGE_POLICY.validate_type(original_filename, getattr(file, "content_type", "") or "")
    except UploadValidationError as ve:
        _log_upload_rejected("task-comment", task_id, caller_id, original_filename, str(ve))
        raise HTTPException(status_code=400, detail=str(ve))

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")
    try:
        VOC_IMAGE_POLICY.validate(
            filename=original_filename,
            size_bytes=len(raw),
            content_type=getattr(file, "content_type", "") or "",
        )
    except UploadValidationError as ve:
        _log_upload_rejected("task-comment", task_id, caller_id, original_filename, str(ve))
        raise HTTPException(status_code=400, detail=str(ve))

    meta = _store_task_comment_image_bytes(task_id, raw, original_filename)
    att = TaskCommentAttachment(
        task_id=task_id,
        comment_id=None,  # 댓글 등록 시 링크
        uploader_id=caller_id,
        filename=original_filename[:255],
        **meta,
    )
    db.add(att)
    db.commit()
    db.refresh(att)
    return _task_comment_attachment_dict(att)


@app.get("/api/tasks/{task_id}/comment-images/{stored_name}/download")
def download_task_comment_image(task_id: int, stored_name: str, db: Session = Depends(get_db)):
    """댓글 이미지 서빙. S3 우선, 없으면 로컬 fallback. 이미지면 inline 렌더.
    (task 첨부/VOC 첨부와 동일하게 불투명 uuid stored_name 기준 서빙)"""
    from fastapi.responses import Response
    from app.utils.s3.s3_utils import download_from_s3, is_s3_configured

    att = (
        db.query(TaskCommentAttachment)
        .filter(TaskCommentAttachment.task_id == task_id,
                TaskCommentAttachment.stored_name == stored_name)
        .first()
    )
    if not att:
        raise HTTPException(status_code=404, detail="File not found")

    filename = att.filename or stored_name
    media_type = att.content_type or _IMAGE_EXT_MIME.get(
        os.path.splitext(stored_name)[1].lower(), "application/octet-stream"
    )
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename)
    content_disposition = (
        f"inline; filename*=UTF-8''{encoded_filename}"
        if media_type.startswith("image/")
        else f"attachment; filename*=UTF-8''{encoded_filename}"
    )

    if att.s3_key and is_s3_configured():
        data = download_from_s3(att.s3_key)
        if data is not None:
            return Response(
                content=data, media_type=media_type,
                headers={
                    "Content-Disposition": content_disposition,
                    "Content-Length": str(len(data)),
                    "Cache-Control": "private, max-age=86400",
                },
            )

    local_path = os.path.join(UPLOAD_DIR, "task-comments", str(task_id), stored_name)
    if os.path.exists(local_path):
        with open(local_path, "rb") as f:
            data = f.read()
        return Response(
            content=data, media_type=media_type,
            headers={
                "Content-Disposition": content_disposition,
                "Content-Length": str(len(data)),
                "Cache-Control": "private, max-age=86400",
            },
        )
    raise HTTPException(status_code=404, detail="File not found")


def _resolve_assignee_col(template) -> int | None:
    """SheetTemplate 의 column_role_mapping.assignee.col 위치를 반환. 없으면 None.

    structure.column_roles 에만 매핑이 있는 옛 양식도 폴백 확인.
    """
    if not template:
        return None
    roles = getattr(template, "column_role_mapping", None) or {}
    if not roles and isinstance(getattr(template, "structure", None), dict):
        roles = template.structure.get("column_roles") or {}
    if not isinstance(roles, dict):
        return None
    ar = roles.get("assignee")
    if isinstance(ar, dict) and isinstance(ar.get("col"), int):
        return ar["col"]
    return None


def _get_user_display_name(db: Session, user_id: int) -> str | None:
    """담당자 셀에 자동 기록할 사용자 표시명. username 우선, loginid 폴백."""
    if not user_id:
        return None
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        return None
    return (getattr(u, "username", None) or getattr(u, "loginid", None) or None)


def _fill_assignee_cell_if_empty(
    db: Session,
    execution_id: int,
    items_by_rc: dict[tuple[int, int], "SheetExecutionItem"],
    row_idx: int,
    assignee_col: int,
    display_name: str,
) -> None:
    """(row_idx, assignee_col) 셀에 사용자명을 기록한다. 셀에 이미 값이 있으면 유지.

    items_by_rc 는 호출자가 보유한 (row, col)→item 매핑 (없을 수도 있음).
    """
    def _col_letter(col_idx: int) -> str:
        s = ""
        n = col_idx + 1
        while n > 0:
            n -= 1
            s = chr(65 + (n % 26)) + s
            n //= 26
        return s

    existing = items_by_rc.get((row_idx, assignee_col)) if items_by_rc is not None else None
    if existing is None:
        existing = db.query(SheetExecutionItem).filter(
            SheetExecutionItem.execution_id == execution_id,
            SheetExecutionItem.row_idx == row_idx,
            SheetExecutionItem.col_idx == assignee_col,
        ).first()

    if existing is None:
        new_item = SheetExecutionItem(
            execution_id=execution_id,
            cell_ref=f"{_col_letter(assignee_col)}{row_idx + 1}",
            row_idx=row_idx,
            col_idx=assignee_col,
            checked=False,
            value=display_name,
        )
        db.add(new_item)
        if items_by_rc is not None:
            items_by_rc[(row_idx, assignee_col)] = new_item
        return

    # 이미 값이 있으면 유지 (사용자 수기 입력 우선).
    if (existing.value or "").strip():
        return
    existing.value = display_name


def _recompute_sheet_progress(db: Session, execution: "SheetExecution") -> tuple[int, int, int]:
    """Recompute (checked_count, total, progress) for a sheet execution, counting only
    items at template's checkable_cells positions. This guards against historical data
    corruption where mark-all overwrote progress_date/remark items with checked=True,
    which would otherwise cap progress at 100% even after a status flip back to X.

    Mutates execution.checked_items / execution.progress in place. Caller must commit.
    Returns (checked_count, effective_total, progress).
    """
    template = db.query(SheetTemplate).filter(SheetTemplate.id == execution.template_id).first()
    structure = (template.structure if template else {}) or {}
    checkable_cells = structure.get("checkable_cells") or []
    checkable_positions = {(int(c["row"]), int(c["col"])) for c in checkable_cells if "row" in c and "col" in c}

    items = db.query(SheetExecutionItem).filter(
        SheetExecutionItem.execution_id == execution.id
    ).all()

    # Fallback: 옛 데이터(checkable_cells 미저장 등) 는 기존 단순 카운트 유지.
    if not checkable_positions:
        total = execution.total_items or len(items) or 1
        na_count = sum(1 for it in items if it.value == "N/A")
        checked_count = sum(1 for it in items if it.checked)
    else:
        total = len(checkable_positions)
        na_count = 0
        checked_count = 0
        for it in items:
            if (it.row_idx, it.col_idx) not in checkable_positions:
                continue
            if it.value == "N/A":
                na_count += 1
            elif it.checked:
                checked_count += 1

    effective_total = max(1, total - na_count)
    progress = min(100, int(checked_count / effective_total * 100))
    execution.checked_items = checked_count
    execution.progress = progress
    return checked_count, total, progress


def _sync_task_progress(
    db: Session, task_id: int, force_zero_if_empty: bool = False,
    actor_user_id: Optional[int] = None, _advance_sink: Optional[list] = None,
):
    """Recalculate task progress from checkbox activities and auto-sync status.

    진행률/파생 status 정책은 _derive_task_status/_compute_task_progress 참고.
    체크박스 활동이 있으면 그 기준으로 동기화한다.
    v3.4: 체크박스가 없고 연결된 SheetExecution이 있으면 그 시트들의 진행률
          평균으로 task progress를 동기화한다 (점검표만 수행하는 task 케이스).
    v3.5: "내용이 비어있는 체크박스"는 작업노트로 간주하지 않고 시트 폴백을 허용.
          (block_type 기본값이 'checkbox'라 빈 placeholder가 잘못 카운트되는 케이스 대응)
    v3.6: force_zero_if_empty — Check Sheet 연결 해제 등에서 source가 모두 사라진 경우
          이전 progress(직전 시트 잔재)가 남는 문제를 방지하기 위해 명시적으로 0% 처리.

    Returns:
        int | None: 동기화 후 task.progress. 동기화 대상이 없으면 None.
    """
    if task_id is None:
        return None
    activities = db.query(TaskActivityModel).filter(TaskActivityModel.task_id == task_id).all()
    progress, cb_total = _compute_task_progress(activities)

    # v3.5: "실제로 의미있는" 체크박스만 집계 — content가 비어 있는 항목은 제외.
    #       이렇게 하면 빈 placeholder activities가 시트 폴백을 막는 일이 없다.
    meaningful_checkboxes = [
        a for a in activities
        if (a.block_type or "checkbox") == "checkbox" and (a.content or "").strip()
    ]
    has_meaningful_work_note = len(meaningful_checkboxes) > 0

    if not has_meaningful_work_note:
        # 체크박스가 없거나 모두 빈 placeholder일 때 시트 진행률로 폴백.
        # 관리형 시트(sheet_type != "inspection") 는 진행률 개념이 없으므로 task progress 폴백
        # 대상에서 제외 — 그렇지 않으면 progress=0 으로 영구 잠겨서 task 가 자동 done 처리되지 않는다.
        sheets = db.query(SheetExecution).filter(
            SheetExecution.task_id == task_id,
            SheetExecution.status.in_(("in_progress", "completed")),
            SheetExecution.sheet_type == "inspection",
        ).all()
        if not sheets:
            # v3.6: 호출자가 force_zero_if_empty=True를 주면 명시적으로 0% 동기화.
            #       (예: Check Sheet 연결 해제 직후 — 직전 시트 progress 잔재 제거)
            if force_zero_if_empty:
                progress = 0
            else:
                # 체크박스도 없고 시트도 없으면 동기화하지 않음 (수동 progress 유지)
                if cb_total == 0:
                    return None
                # cb_total > 0인데 모두 빈 placeholder인 경우엔 progress=0으로 안전 처리
                progress = 0
        else:
            total = 0
            for s in sheets:
                if s.status == "completed":
                    total += 100
                else:
                    total += int(s.progress or 0)
            progress = round(total / len(sheets))

    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        return None
    updates: dict = {"progress": progress}
    # CUSTOM 워크플로우(auto_progress_from_notes=False) 프로젝트는 작업노트 체크율로
    # status/컬럼을 자동 이동시키지 않는다 — progress(진행률 표시용)만 동기화한다.
    proj = db.query(Project).filter(Project.id == task.project_id).first()
    auto_progress = bool(getattr(proj, "auto_progress_from_notes", True)) if proj else True
    if auto_progress:
        expected_status = _derive_task_status(task.status, progress)
        if expected_status != task.status:
            updates["status"] = expected_status
    db.query(Task).filter(Task.id == task_id).update(updates)
    db.commit()
    # Also sync to sidecar task_meta
    state = load_state()
    set_task_meta(state, task_id, {"progress": progress})
    save_state(state)
    # 4차: progress 기준 자동 이동은 제거됨. progress 는 표시용 보조 지표로만 사용한다.
    # 단계 이동은 사용자 '단계 완료 선언'(POST /tasks/{id}/stage-completion)에서만 발생.
    return progress

# =========================
# Graph
# =========================
@app.get("/api/projects/{project_id}/graph")
def get_project_graph(project_id: int, db: Session = Depends(get_db)):
    state = load_state()

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    project = project_dict(p, state)
    nodes = []
    edges = []

    nodes.append({"id": f"project-{project_id}", "type": "project", "label": project["name"]})

    for sp in get_subprojects_from_db(db, project_id):
        sp_id = f"subproject-{sp['id']}"
        nodes.append({"id": sp_id, "type": "subproject", "label": sp.get("name", "")})
        edges.append({"source": f"project-{project_id}", "target": sp_id})

    task_rows = db.query(Task).filter(Task.project_id == project_id, Task.archived_at.is_(None)).all()
    tasks = [task_dict(t, state) for t in task_rows]
    wf_map = _workflow_label_map(db, project_id)

    for t in tasks:
        t_id = f"task-{t['id']}"
        node = {"id": t_id, "type": "task", "label": t["title"], "status": t.get("status")}
        wf_col = wf_map.get(t.get("workflow_column_id"))
        if wf_col is not None:
            node["workflow_label"] = wf_col.label
            node["workflow_color"] = wf_col.color
        nodes.append(node)

        # 상하 관계: 하위 Task 는 상위 Task 아래에 연결 (Project → Sub Project → Parent Task → Child Task)
        if t.get("parent_task_id"):
            edges.append({"source": f"task-{t['parent_task_id']}", "target": t_id})
        elif t.get("sub_project_id"):
            edges.append({"source": f"subproject-{t['sub_project_id']}", "target": t_id})
        else:
            edges.append({"source": f"project-{project_id}", "target": t_id})

        # Extract image URLs from task activities (Work Note)
        activities = db.query(TaskActivityModel).filter(TaskActivityModel.task_id == t["id"]).all()
        content_img_urls = set()
        img_regex = re.compile(r'<img[^>]+src=["\']([^">]+)["\']', re.IGNORECASE)
        for act in activities:
            if act.content:
                matches = img_regex.findall(act.content)
                for m in matches:
                    # Normalize URL (remove query params or absolute parts if needed for matching)
                    # But usually exact match or path match is enough
                    content_img_urls.add(m)

        def is_image_ext(filename):
            if not filename: return False
            ext = filename.split(".")[-1].lower()
            return ext in ["png", "jpg", "jpeg", "gif", "webp", "svg", "bmp"]

        # Attachments from DB
        db_attachments = db.query(AttachmentModel).filter(AttachmentModel.task_id == t["id"]).all()
        for a in db_attachments:
            is_img = is_image_ext(a.filename) or (a.type == "url" and is_image_ext(a.url))
            
            # If it's an image, check if it's in the content
            if is_img:
                # Try to find a match in content_img_urls
                found = False
                for c_url in content_img_urls:
                    if (a.url and a.url in c_url) or (a.filename and a.filename in c_url):
                        found = True
                        break
                if not found:
                    continue # Skip deleted/unreferenced images

            a_id = f"attachment-{a.id}"
            nodes.append({
                "id": a_id,
                "type": "attachment",
                "label": a.filename or a.url or "",
                "attachment_type": a.type or "url",
                "url": a.url or "",
            })
            edges.append({"source": t_id, "target": a_id})

        # Fallback: sidecar attachments (filter images here too if possible)
        seen_att_ids = {a.id for a in db_attachments}
        for a in state.get("attachments", []):
            if int(a.get("task_id")) == t["id"] and a.get("id") not in seen_att_ids:
                fname = a.get("filename")
                aurl = a.get("url", "")
                is_img = is_image_ext(fname) or (a.get("type") == "url" and is_image_ext(aurl))

                if is_img:
                    found = False
                    for c_url in content_img_urls:
                        if (aurl and aurl in c_url) or (fname and fname in c_url):
                            found = True
                            break
                    if not found:
                        continue

                a_id = f"attachment-{a['id']}"
                nodes.append({
                    "id": a_id,
                    "type": "attachment",
                    "label": fname or aurl,
                    "attachment_type": a.get("type", "url"),
                    "url": aurl,
                })
                edges.append({"source": t_id, "target": a_id})

    # C-5: Notes from DB
    db_notes = db.query(NoteModel).filter(NoteModel.project_id == project_id).all()
    seen_note_ids = set()
    for n in db_notes:
        content = n.content or ""
        label = content[:30] + ("..." if len(content) > 30 else "")
        n_id = f"note-{n.id}"
        nodes.append({"id": n_id, "type": "note", "label": label})
        edges.append({"source": f"project-{project_id}", "target": n_id})
        seen_note_ids.add(n.id)

    # Fallback: sidecar notes
    for n in state.get("notes", []):
        if int(n.get("project_id")) == project_id and n.get("id") not in seen_note_ids:
            content = n.get("content", "")
            label = content[:30] + ("..." if len(content) > 30 else "")
            n_id = f"note-{n['id']}"
            nodes.append({"id": n_id, "type": "note", "label": label})
            edges.append({"source": f"project-{project_id}", "target": n_id})

    return {"nodes": nodes, "edges": edges}

# =========================
# AI LLM Helper
# =========================
def _normalize_selected_model(value) -> str:
    """
    AI Settings/Report로 들어온 모델 식별값을 안전하게 문자열 key로 정규화.
    - str: 양끝 공백 제거 후 그대로 반환
    - dict: {"key"} → {"model"} → {"name"} 순서로 추출
    - 객체: .key → .model → .name 속성 순서로 추출
    - None / 추출 실패: "" 반환 → 호출부에서 dsllm_default_model_key()로 fallback
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict):
        for k in ("key", "model", "name"):
            v = value.get(k)
            if isinstance(v, str) and v.strip():
                return v.strip()
        return ""
    for attr in ("key", "model", "name"):
        v = getattr(value, attr, None)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def call_llm_api(api_url: str, model_name: str, system_prompt: str, user_prompt: str) -> str:
    # api_url 인자는 하위호환을 위해 남겨두지만 무시한다. BASE_URL은 env 전용.
    # db 컨텍스트가 없는 레거시 헬퍼라 env 기본 provider 를 사용한다.
    provider = llm_router.env_default_provider()
    return llm_router.chat(
        provider=provider,
        base_url=None,
        model_name=llm_router.resolve_model(provider, _normalize_selected_model(model_name)),
        system_prompt=system_prompt,
        user_prompt=user_prompt,
        temperature=0.3,
        max_tokens=4096,
    )

# =========================
# AI Settings & Report Generation (sidecar + DB data)
# =========================
def get_or_create_ai_setting(db: Session) -> AiSetting:
    row = db.query(AiSetting).order_by(AiSetting.id.asc()).first()
    if not row:
        row = AiSetting(api_url="", model_name="", api_key=None)  # api_key는 이제 사용 안 함
        db.add(row)
        db.commit()
        db.refresh(row)
    return row

def _active_provider(row: AiSetting) -> str:
    """저장된 provider(없으면 env 기본값)를 정규화해 반환."""
    return llm_router.resolve_active_provider(getattr(row, "provider", None))


@app.get("/api/settings/ai/models")
def get_ai_models(
    user_id: int = Query(...),
    provider: str | None = Query(None),
    db: Session = Depends(get_db),
):
    require_super_admin(db, user_id)
    # provider 미지정이면 저장된/기본 provider 기준 모델 목록을 반환.
    row = get_or_create_ai_setting(db)
    prov = llm_router.resolve_active_provider(provider or getattr(row, "provider", None))
    models = llm_router.available_models(prov)
    return {
        "provider": prov,
        "models": [m["key"] for m in models],
        "default_model": llm_router.default_model_key(prov),
        "model_details": models,
    }

@app.get("/api/settings/ai")
def get_ai_settings(
    user_id: int = Query(...),
    provider: str | None = Query(None),
    db: Session = Depends(get_db),
):
    require_super_admin(db, user_id)
    from app.config import settings as app_settings
    from app import config as app_config
    row = get_or_create_ai_setting(db)
    saved_prov = _active_provider(row)
    # 연결 상태(BASE URL / API KEY)는 "화면에서 보고 있는 provider" 기준으로 계산한다.
    #   드롭다운에서 OpenAI 를 골랐는데 상태는 저장된 DSLLM 기준이라 'API KEY 미설정' 으로
    #   뜨던 표시 불일치를 없앤다. 저장/라우팅 값(saved_prov)은 그대로 유지.
    _req = (provider or "").strip().lower()
    view_prov = _req if _req in ("dsllm", "openai") else saved_prov
    conn = llm_router.connection_state(view_prov)
    return {
        "provider": saved_prov,
        "view_provider": view_prov,
        "available_providers": llm_router.list_providers(),
        "openai_warning": llm_router.OPENAI_WARNING,
        "model_name": row.model_name or llm_router.default_model_key(saved_prov),
        "env_mode": app_settings.env_mode,
        "base_url_configured": conn["base_url_configured"],
        "base_url_display": conn["base_url_display"],
        "api_key_configured": conn["api_key_configured"],
        # ── 운영자 자가진단용 (민감정보/키 원문 없음) ──
        #   "UI Provider 는 OpenAI인데 report/generate 가 DSLLM 을 부른다" 같은 불일치를
        #   운영자가 바로 확인할 수 있도록, backend 가 실제 읽은 env 파일 경로와
        #   OPENAI_API_KEY 감지 여부(os.environ 기준)를 함께 반환한다.
        "debug": {
            "env_file": getattr(app_config, "ENV_FILE_DISPLAY", ""),
            "env_file_exists": bool(getattr(app_config, "ENV_FILE_EXISTS", False)),
            "openai_key_in_env": bool((os.environ.get("OPENAI_API_KEY") or "").strip()),
            "openai_provider_selectable": llm_router.openai_enabled(),
        },
    }

@app.get("/api/debug/knox/search")
async def debug_knox_search(
    query: str = Query(..., description="검색어 (loginId 또는 이름)"),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """super_admin 전용 KNOX 진단 엔드포인트. token 원문은 절대 반환하지 않는다."""
    require_super_admin(db, user_id)

    from app.services.knox_client import (
        knox_search_employees,
        knox_config_state,
        KnoxError,
    )

    q = (query or "").strip()
    cfg = knox_config_state()
    try:
        # ID 우선 검색, 없으면 이름 검색 (라우터 /employees와 동일한 직관)
        items = await knox_search_employees(userIds=q)
        if not items:
            items = await knox_search_employees(fullName=q)
        return {"ok": True, "query": q, "count": len(items), "items": items, **cfg}
    except KnoxError as e:
        return {
            "ok": False,
            "query": q,
            "error_type": e.error_type,
            "status_code": e.status_code,
            "message": str(e),
            **cfg,
        }
    except Exception as e:
        return {
            "ok": False,
            "query": q,
            "error_type": type(e).__name__,
            "message": str(e),
            **cfg,
        }


@app.put("/api/settings/ai")
def save_ai_settings(body: AiSettingsUpdate, user_id: int = Query(...), db: Session = Depends(get_db)):
    require_super_admin(db, user_id)
    row = get_or_create_ai_setting(db)

    # ── provider 결정/검증 (지정 안 하면 기존값 유지) ──
    if body.provider is not None:
        requested = (body.provider or "").strip().lower()
        if requested not in ("dsllm", "openai"):
            raise HTTPException(400, f"Unknown AI provider: {requested}")
        if requested == "openai" and not llm_router.openai_enabled():
            raise HTTPException(
                400,
                "OpenAI API Key가 설정되어 있지 않습니다. "
                "backend env 파일에 OPENAI_API_KEY를 설정한 뒤 다시 시도해주세요.",
            )
        row.provider = requested

    # 접속정보(api_url/api_key)는 env에서만 사용. model_name/provider만 저장.
    row.model_name = body.model_name.strip()
    db.add(row)
    db.commit()
    db.refresh(row)
    return {
        "message": "AI settings saved",
        "settings": {"model_name": row.model_name, "provider": _active_provider(row)},
    }


@app.post("/api/settings/ai/test")
def test_ai_connection(body: AiSettingsUpdate, user_id: int = Query(...), db: Session = Depends(get_db)):
    """현재(또는 요청에 지정된) provider/model 로 짧은 연결 테스트.

    ⚠️ 프로젝트/Task 데이터는 절대 넣지 않는다. 비민감 고정 문구만 사용한다.
    """
    require_super_admin(db, user_id)
    row = get_or_create_ai_setting(db)

    prov = llm_router.resolve_active_provider(body.provider or getattr(row, "provider", None))
    model_name = llm_router.resolve_model(prov, _normalize_selected_model(body.model_name or row.model_name))

    import logging as _logging
    _log = _logging.getLogger("main")

    # ── OpenAI: 단계별 진단(api_key → /v1/models → 모델 접근 → /v1/responses) ──
    if prov == "openai":
        from app.llm import openai_adapter
        result = openai_adapter.run_connection_test(model_name)
        _log.info(
            "AI connection test (openai): ok=%s stage=%s status=%s model=%s",
            result.get("ok"), result.get("stage"), result.get("status_code"), result.get("model"),
        )
        if result.get("ok"):
            return {
                "ok": True, "provider": prov, "model": result.get("model", model_name),
                "answer": result.get("answer", ""),
                "model_found": result.get("model_found"),
                "models_count": result.get("models_count"),
                "base_url": result.get("base_url"),
            }
        return {
            "ok": False, "provider": prov, "model": result.get("model", model_name),
            "stage": result.get("stage"), "status_code": result.get("status_code"),
            "model_found": result.get("model_found"), "base_url": result.get("base_url"),
            "error": result.get("message", "연결 실패: 호출에 실패했습니다."),
        }

    # ── DSLLM(사내): 기존 단순 테스트 ──
    try:
        content = llm_router.chat(
            provider=prov,
            base_url=None,
            model_name=model_name,
            system_prompt="You are a helpful assistant. 한 문장으로만 응답하세요.",
            user_prompt="안녕하세요. 한 문장으로 응답해주세요.",
            temperature=0.2,
            max_tokens=64,
        )
        answer = (content or "").strip()
        if len(answer) > 200:
            answer = answer[:200] + "…"
        return {"ok": True, "provider": prov, "model": model_name, "answer": answer}
    except Exception as e:
        # 사용자 친화 메시지 — API Key 등 민감정보는 노출하지 않는다.
        _log.warning(
            "AI connection test failed: provider=%s, model=%s, error=%s",
            prov, model_name, type(e).__name__,
        )
        return {
            "ok": False,
            "provider": prov,
            "model": model_name,
            "error": "연결 실패: 호출에 실패했거나 모델 설정을 확인해주세요.",
        }

# =========================================================
# Vision AI 설정 (super_admin 전용) — Text AI(/api/settings/ai)와 완전 분리
#   이미지 포함 AI Report(Vision Report) 전용. Text 경로(report/generate, ai-query)는
#   절대 건드리지 않는다. 저장은 vision_ai_settings 테이블, 해석은 vision_ai_settings 모듈.
# =========================================================
def _vision_ai_connection_payload(db: Session) -> dict:
    """Vision AI effective config + openai 연결상태 + 경고. (GET/저장 응답 공용)"""
    from app.llm import openai_adapter
    from app.services.vision_ai_settings import resolve_vision_ai_config

    cfg = resolve_vision_ai_config(db)
    conn = llm_router.connection_state("openai")
    return {
        **cfg,
        "openai_configured": openai_adapter.is_configured(),
        "openai_warning": llm_router.OPENAI_WARNING,
        "base_url_configured": conn["base_url_configured"],
        "base_url_display": conn["base_url_display"],
        "api_key_configured": conn["api_key_configured"],
    }


@app.get("/api/settings/vision-ai")
def get_vision_ai_settings(user_id: int = Query(...), db: Session = Depends(get_db)):
    require_super_admin(db, user_id)
    return _vision_ai_connection_payload(db)


@app.put("/api/settings/vision-ai")
def save_vision_ai_settings(
    body: VisionAiSettingsUpdate, user_id: int = Query(...), db: Session = Depends(get_db)
):
    require_super_admin(db, user_id)
    from app.services.vision_ai_settings import get_or_create_vision_ai_setting

    row = get_or_create_vision_ai_setting(db)

    if body.provider is not None:
        requested = (body.provider or "").strip().lower()
        # openai(외부) / dsllm(사내) 만 허용. dsllm 은 DS/Gemma4 vision 후보(현재 payload 미구현).
        if requested and requested not in ("openai", "dsllm"):
            raise HTTPException(400, "Vision AI provider는 openai 또는 dsllm만 지원합니다.")
        if requested == "openai" and not llm_router.openai_enabled():
            raise HTTPException(
                400,
                "OpenAI API Key가 설정되어 있지 않습니다. "
                "backend env 파일에 OPENAI_API_KEY를 설정한 뒤 다시 시도해주세요.",
            )
        row.provider = requested or "openai"

    if body.model_name is not None:
        row.model_name = (body.model_name or "").strip() or None
    if body.enabled is not None:
        row.enabled = bool(body.enabled)
    if body.max_output_tokens is not None:
        row.max_output_tokens = max(1, int(body.max_output_tokens))
    if body.batch_size is not None:
        row.batch_size = max(1, int(body.batch_size))
    if body.timeout_seconds is not None:
        row.timeout_seconds = max(1, int(body.timeout_seconds))

    # ── Fallback Text Model — Primary Vision Model 과 별개 저장(둘 다 저장 가능, §9) ──
    if body.fallback_text_provider is not None:
        fb_prov = (body.fallback_text_provider or "").strip().lower()
        if fb_prov and fb_prov not in ("openai", "dsllm"):
            raise HTTPException(400, "Fallback Text provider는 openai 또는 dsllm만 지원합니다.")
        if fb_prov == "openai" and not llm_router.openai_enabled():
            raise HTTPException(
                400,
                "OpenAI API Key가 설정되어 있지 않아 openai fallback을 저장할 수 없습니다.",
            )
        row.fallback_text_provider = fb_prov or None
    if body.fallback_text_model is not None:
        fb_model = (body.fallback_text_model or "").strip()
        # provider/model 조합 검증 — 모델 목록과 동일 registry(llm_router.available_models)로 확인.
        # 이번 요청에서 provider 를 바꿨으면 그 값을, 아니면 기존 저장값을, 그래도 없으면 dsllm 기준.
        if fb_model:
            eff_fb_prov = (
                (body.fallback_text_provider or "").strip().lower()
                or (getattr(row, "fallback_text_provider", "") or "").strip().lower()
                or "dsllm"
            )
            if eff_fb_prov not in ("openai", "dsllm"):
                eff_fb_prov = "dsllm"
            allowed = {m["key"] for m in llm_router.available_models(eff_fb_prov)}
            # 목록을 확인할 수 있을 때만 강제(빈 목록이면 검증 불가 → 통과).
            if allowed and fb_model not in allowed:
                raise HTTPException(
                    400,
                    "선택한 모델은 현재 Fallback Provider에서 사용할 수 없습니다.",
                )
        row.fallback_text_model = fb_model or None
    if body.fallback_text_enabled is not None:
        row.fallback_text_enabled = bool(body.fallback_text_enabled)
    if body.fallback_policy is not None:
        # 현재 정책은 ask_user_on_timeout 만 지원(자동 전환 금지).
        row.fallback_policy = (body.fallback_policy or "").strip() or None
    if body.soft_timeout_seconds is not None:
        row.soft_timeout_seconds = max(1, int(body.soft_timeout_seconds))

    db.add(row)
    db.commit()
    db.refresh(row)
    return {"message": "Vision AI settings saved", "settings": _vision_ai_connection_payload(db)}


@app.post("/api/settings/vision-ai/test")
def test_vision_ai_connection(
    body: VisionAiSettingsUpdate, user_id: int = Query(...), db: Session = Depends(get_db)
):
    """Vision AI 모델 텍스트 연결 테스트(이미지 없이). super_admin 전용.

    ⚠️ 프로젝트/회사 데이터는 절대 넣지 않는다(openai_adapter 의 고정 비민감 문구 사용).
    """
    require_super_admin(db, user_id)
    from app.llm import openai_adapter
    from app.services.vision_ai_settings import resolve_vision_ai_config

    cfg = resolve_vision_ai_config(db)
    # provider: 폼에서 고른 값(body) 우선, 없으면 저장된 effective provider.
    provider = (body.provider or "").strip().lower() or cfg["provider"]
    model_name = (body.model_name or "").strip() or cfg["model"]

    # ── DSLLM(사내): 텍스트 연결만 확인(이미지 미포함). vision payload 는 별도 테스트. ──
    if provider == "dsllm":
        try:
            content = llm_router.chat(
                provider="dsllm",
                base_url=None,
                model_name=model_name,
                system_prompt="You are a helpful assistant. 한 문장으로만 응답하세요.",
                user_prompt="안녕하세요. 한 문장으로 응답해주세요.",
                temperature=0.2,
                max_tokens=64,
            )
            answer = (content or "").strip()
            if len(answer) > 200:
                answer = answer[:200] + "…"
            return {"ok": True, "provider": "dsllm", "model": model_name, "answer": answer}
        except Exception:
            return {
                "ok": False, "provider": "dsllm", "model": model_name, "stage": "dsllm_call",
                "error": "연결 실패: 호출에 실패했거나 모델 설정을 확인해주세요.",
            }

    # ── OpenAI(외부): 기존 단계별 진단 ──
    if not openai_adapter.is_configured():
        return {
            "ok": False, "provider": "openai", "model": model_name, "stage": "api_key",
            "error": "OPENAI_API_KEY가 backend env에 설정되어 있지 않습니다.",
        }
    result = openai_adapter.run_connection_test(model_name)
    if result.get("ok"):
        return {
            "ok": True, "provider": "openai", "model": result.get("model", model_name),
            "answer": result.get("answer", ""), "models_count": result.get("models_count"),
            "base_url": result.get("base_url"),
        }
    return {
        "ok": False, "provider": "openai", "model": result.get("model", model_name),
        "stage": result.get("stage"), "status_code": result.get("status_code"),
        "base_url": result.get("base_url"),
        "error": result.get("message", "연결 실패: 호출에 실패했습니다."),
    }


def _solid_color_png(r: int, g: int, b: int, size: int = 16) -> bytes:
    """순색 RGB PNG 1장을 stdlib 만으로 생성(Pillow 불필요). vision 입력 테스트용."""
    import struct
    import zlib

    def _chunk(typ: bytes, data: bytes) -> bytes:
        return (
            struct.pack(">I", len(data))
            + typ
            + data
            + struct.pack(">I", zlib.crc32(typ + data) & 0xFFFFFFFF)
        )

    sig = b"\x89PNG\r\n\x1a\n"
    ihdr = struct.pack(">IIBBBBB", size, size, 8, 2, 0, 0, 0)  # 8-bit RGB
    row = b"\x00" + bytes([r, g, b]) * size
    idat = zlib.compress(row * size)
    return sig + _chunk(b"IHDR", ihdr) + _chunk(b"IDAT", idat) + _chunk(b"IEND", b"")


@app.post("/api/settings/vision-ai/vision-test")
def test_vision_ai_vision_input(
    body: VisionAiSettingsUpdate, user_id: int = Query(...), db: Session = Depends(get_db)
):
    """내장 테스트 이미지(순색) 1장을 모델에 넣어 vision 입력을 실제로 읽는지 검증.

    프로젝트/회사 데이터 미포함 — 서버에서 생성한 고정 합성 이미지(빨강)만 사용한다.
    호출이 성공하고 응답 텍스트가 있으면 vision 입력이 동작한다고 본다.
    """
    require_super_admin(db, user_id)
    import base64

    from app.llm import openai_adapter
    from app.services.vision_ai_settings import resolve_vision_ai_config, vision_supported

    cfg = resolve_vision_ai_config(db)
    provider = (body.provider or "").strip().lower() or cfg["provider"]
    model_name = (body.model_name or "").strip() or cfg["model"]

    # 내장 테스트 이미지(순색 빨강) — 프로젝트/회사 데이터 미포함, 서버 생성 고정 이미지.
    png = _solid_color_png(220, 38, 38)  # 빨강
    data_url = "data:image/png;base64," + base64.b64encode(png).decode("ascii")
    content = [
        {"type": "input_text", "text": "이 이미지에서 보이는 주요 색을 한국어 한 단어로만 답하세요."},
        {"type": "input_image", "image_url": data_url},
    ]
    instructions = "당신은 이미지를 보고 한 단어로만 답하는 도우미입니다."
    timeout = float(cfg.get("timeout_seconds") or 120)

    # ── 구조화 진단 로그(§1): Vision Report 호출 로그와 필드명을 맞춰 payload 를 직접 비교 가능하게 한다.
    #    base64 본문 / API key / 전체 response body 는 로그에 남기지 않고 "크기"만 기록한다. ──
    import time as _time
    import uuid as _uuid

    _vlog = logging.getLogger("main")
    request_id = _uuid.uuid4().hex[:8]
    _b64_chars = len(data_url) - len("data:image/png;base64,")
    _text_bytes = sum(
        len((c.get("text") or "").encode("utf-8", "ignore"))
        for c in content if c.get("type") == "input_text"
    )
    _payload_bytes = _text_bytes + len(data_url.encode("utf-8", "ignore"))
    try:
        _img_w = int.from_bytes(png[16:20], "big")
        _img_h = int.from_bytes(png[20:24], "big")
    except Exception:
        _img_w = _img_h = None
    _settings_max_tokens = 64
    _st0 = _time.perf_counter()

    _vlog.info(
        "[AiSettingsVisionTest] request_id=%s provider=%s model=%s stage=settings_test "
        "image_transport=data_url image_count=1 mime=image/png original_bytes=%d "
        "processed_bytes=%d base64_chars=%d payload_bytes=%d width=%s height=%s "
        "max_tokens=%d timeout_seconds=%.0f started",
        request_id, provider, model_name, len(png), len(png), _b64_chars, _payload_bytes,
        _img_w, _img_h, _settings_max_tokens, timeout,
    )

    try:
        _svu = db.query(User).filter(User.id == user_id).first()
        _settings_login_id = getattr(_svu, "login_id", None)
    except Exception:
        _settings_login_id = None
    _settings_endpoint = "/api/settings/vision-ai/vision-test"

    def _settings_log(http_status, error_type, success):
        elapsed_ms = int((_time.perf_counter() - _st0) * 1000)
        _vlog.info(
            "[AiSettingsVisionTest] request_id=%s provider=%s model=%s stage=settings_test "
            "payload_bytes=%d image_count=1 max_tokens=%d timeout_seconds=%.0f "
            "elapsed_ms=%d http_status=%s error_type=%s success=%s",
            request_id, provider, model_name, _payload_bytes, _settings_max_tokens, timeout,
            elapsed_ms, http_status, error_type, success,
        )
        # 기존 관리자 시스템 로그에도 기록: 실패=ERROR 진단, 성공=INFO(연결테스트 결과, §6/§13 Test A).
        _details = {
            "image_count": 1, "image_transport": "data_url", "mime": "image/png",
            "width": _img_w, "height": _img_h, "original_bytes": len(png),
            "processed_bytes": len(png), "base64_chars": _b64_chars,
            "payload_bytes": _payload_bytes, "max_tokens": _settings_max_tokens,
            "timeout_seconds": int(timeout), "elapsed_ms": elapsed_ms,
        }
        if success:
            try:
                import json as _json
                from app.services import vision_diagnostics as _vd
                from app.services.system_log import log_event as _le
                _diag = {
                    "failure_stage": "ai_settings_vision_test",
                    "failure_stage_label": _vd.failure_stage_label("ai_settings_vision_test"),
                    "error_type": None, "provider": provider, "model": model_name,
                    "http_status": http_status, "recovery_event": False,
                    "summary": "AI 설정 vision 입력 테스트 성공(내장 이미지 인식).",
                    "details": _vd.sanitize_details(_details),
                }
                _le(
                    "INFO", "VISION_AI", f"AI 설정 vision 입력 테스트 성공 — {model_name}",
                    detail="내장 테스트 이미지(빨강) 인식 성공.",
                    detail_json=_json.dumps(_diag, ensure_ascii=False),
                    endpoint=_settings_endpoint, status_code=http_status,
                    login_id=_settings_login_id, resolved=True,
                    dedupe_key="ai_settings_vision_test|success",
                )
            except Exception:
                pass
        else:
            _log_vision_diagnostic(
                "ERROR", "ai_settings_vision_test", error_type or "unknown_error",
                provider=provider, model=model_name, http_status=http_status,
                endpoint=_settings_endpoint, login_id=_settings_login_id, details=_details,
            )

    # ── DSLLM(사내): /v1/chat/completions 멀티모달(image_url) payload 로 실제 vision 입력 검증 ──
    if provider == "dsllm":
        from app.llm import dsllm_adapter

        try:
            res = dsllm_adapter.vision_chat(
                model_name=model_name,
                instructions=instructions,
                content=content,
                max_tokens=64,
                timeout=timeout,
            )
        except dsllm_adapter.DsllmVisionError as e:
            _settings_log(e.status_code, e.error_type, False)
            # 텍스트 연결은 되지만 image_url payload 처리 실패면 명확히 구분해 안내(§6).
            return {
                "ok": False, "model": model_name, "provider": "dsllm",
                "stage": "vision_payload_unsupported" if e.payload_unsupported else "vision_call",
                "status_code": e.status_code,
                "error_type": e.error_type,
                "vision_supported": "candidate",
                "error": e.user_message,
            }

        answer = (res.get("raw_text") or "").strip()
        _settings_log(200, None, bool(answer))
        return {
            "ok": bool(answer),
            "model": model_name,
            "provider": "dsllm",
            "answer": answer[:200],
            "vision_supported": "candidate",
            "expected": "빨강",
            "error": None if answer else "모델이 빈 응답을 반환했습니다(이미지 인식 실패 가능).",
        }

    # ── OpenAI(외부): Responses API vision 경로 ──
    if not openai_adapter.is_configured():
        _settings_log(None, "authentication_error", False)
        return {
            "ok": False, "model": model_name, "stage": "api_key",
            "error": "OPENAI_API_KEY가 backend env에 설정되어 있지 않습니다.",
        }

    try:
        res = openai_adapter.vision_chat(
            model_name=model_name,
            instructions=instructions,
            content=content,
            max_tokens=64,
            timeout=timeout,
        )
    except openai_adapter.OpenAIAdapterError as e:
        _settings_log(e.status_code, _std_openai_error_type(e), False)
        return {
            "ok": False, "model": model_name, "stage": "vision_call",
            "status_code": e.status_code, "error_type": e.error_type, "error": e.user_message,
            "vision_supported": vision_supported(model_name),
        }

    answer = (res.get("raw_text") or "").strip()
    _settings_log(200, None, bool(answer))
    return {
        "ok": bool(answer),
        "model": model_name,
        "provider": "openai",
        "answer": answer[:200],
        "vision_supported": vision_supported(model_name),
        "expected": "빨강",
        "error": None if answer else "모델이 빈 응답을 반환했습니다(이미지 인식 실패 가능).",
    }


class ProjectImageVisionTestRequest(BaseModel):
    user_id: int
    project_id: int
    image_id: str                      # canonical_image_id (예: task_activity:469:image:0)
    model_name: Optional[str] = None
    provider: Optional[str] = None
    input_mode: Optional[str] = None


@app.post("/api/settings/vision-ai/project-image-test")
def test_vision_ai_project_image(
    body: ProjectImageVisionTestRequest, db: Session = Depends(get_db)
):
    """super_admin 전용: 실제 프로젝트 이미지 1장을 **Vision Report Stage 1과 동일한 경로**로 진단.

    내장 빨강 PNG 테스트와 달리, 실제 attachment/작업노트 이미지를 골라 preprocess_image_for_ai →
    base64 Data URL → 동일 provider adapter → 동일 Stage 1 prompt 로 1회 전송한다(§3/§5).
    - 캐시 read/write 없음(§4): 항상 선택 이미지 bytes 를 모델에 실제 전송.
    - Stage 2 없음, DB(Observation/ContextLink/Report) 저장 없음.
    - 결과(성공 인식 or 실패 진단)는 기존 시스템 로그(VISION_AI)에도 기록한다(§8).
    """
    require_super_admin(db, body.user_id)
    state = load_state()
    check_project_access(db, state, body.project_id, body.user_id)

    import time as _time
    from app.services.system_log import current_request_id, log_event as _log_event
    from app.services.project_ai_context_builder import build_project_ai_context
    from app.services.project_ai_vision_report_test import (
        build_stage1_input, parse_stage1_evidence, normalize_evidence_type,
    )
    from app.services.project_ai_model_compatibility import INPUT_MODES, DEFAULT_INPUT_MODE
    from app.services.vision_ai_settings import resolve_vision_ai_config
    from app.services import vision_diagnostics as _vd

    p = db.query(Project).filter(
        Project.id == body.project_id, Project.archived_at.is_(None)
    ).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    cfg = resolve_vision_ai_config(db)
    provider = (body.provider or "").strip().lower() or cfg["provider"]
    model_name = (body.model_name or "").strip() or cfg["model"]
    input_mode = body.input_mode if body.input_mode in INPUT_MODES else DEFAULT_INPUT_MODE
    timeout = float(cfg.get("timeout_seconds") or 120)
    # Stage 1 단일 이미지 max_tokens = Report Stage 1 과 동일 산식(600 + 500*batch, batch=1).
    max_tokens = min(int(cfg.get("max_output_tokens") or 8000), 600 + 500 * 1)

    request_id = current_request_id()
    endpoint = "/api/settings/vision-ai/project-image-test"
    try:
        _vu = db.query(User).filter(User.id == body.user_id).first()
        login_id = getattr(_vu, "login_id", None)
    except Exception:
        login_id = None

    # 실제 Report 와 동일하게 프로젝트 context(image_manifest 포함) 구성 → 대상 이미지 찾기.
    ctx = build_project_ai_context(
        db, body.project_id, state=state, include_images=True, max_images=200
    )
    manifest = getattr(ctx, "image_manifest", None) or []
    item = next((it for it in manifest if str(it.get("image_id")) == str(body.image_id)), None)
    if item is None:
        raise HTTPException(status_code=404, detail="선택한 이미지를 프로젝트에서 찾지 못했습니다.")

    # 캐시 read/write 없음(§4): 바이트 직접 로딩 → 공통 helper 로 prepared 생성(Report Stage 1 동일).
    data, media_type = _load_manifest_image_bytes(state, item)
    prep = _prepare_vision_image_input(
        ctx.project_context, item, data, media_type, input_mode=input_mode
    )
    prep_meta = prep["prep_meta"]
    proc_media = prep["processed_media_type"]
    _b64_chars = len(prep["image_url"]) - len(f"data:{proc_media};base64,")

    # Stage 1 입력(동일 prompt/schema) — 단일 이미지 batch.
    batch = [{"image_id": item.get("image_id"), "image_url": prep["image_url"], "packet": prep["packet"]}]
    s1_instr, s1_content = build_stage1_input(batch, input_mode=input_mode)
    _payload_bytes = 0
    for c in s1_content:
        if not isinstance(c, dict):
            continue
        if c.get("type") == "input_text":
            _payload_bytes += len((c.get("text") or "").encode("utf-8", "ignore"))
        elif c.get("type") == "input_image":
            _payload_bytes += len((c.get("image_url") or "").encode("utf-8", "ignore"))

    def _metrics(elapsed_ms) -> dict:
        return {
            "project_id": body.project_id,
            "canonical_image_id": item.get("image_id"),
            "image_count": 1,
            "image_transport": "data_url",
            "original_mime": media_type,
            "mime": proc_media,
            "width": prep_meta.get("out_width") or prep_meta.get("width"),
            "height": prep_meta.get("out_height") or prep_meta.get("height"),
            "original_width": prep_meta.get("width"),
            "original_height": prep_meta.get("height"),
            "original_bytes": prep_meta.get("original_bytes"),
            "processed_bytes": prep_meta.get("processed_bytes"),
            "base64_chars": _b64_chars,
            "payload_bytes": _payload_bytes,
            "max_tokens": max_tokens,
            "timeout_seconds": int(timeout),
            "elapsed_ms": elapsed_ms,
        }

    def _fail(error_type: str, http_status, elapsed_ms: int) -> dict:
        """실패 진단 응답 + 시스템 로그(ERROR) 기록(§7/§8)."""
        metrics = _metrics(elapsed_ms)
        diag = _vd.build_diagnosis(
            "stage1_vision", error_type, provider=provider, model=model_name,
            http_status=http_status, details=metrics,
        )
        _log_vision_diagnostic(
            "ERROR", "stage1_vision", error_type, provider=provider, model=model_name,
            http_status=http_status, endpoint=endpoint, login_id=login_id, details=metrics,
        )
        return {
            "ok": False, "test_type": "project_image", "provider": provider, "model": model_name,
            "request_id": request_id, "diagnosis": diag, "metrics": diag["details"],
            "project_id": body.project_id, "task_id": item.get("task_id"),
            "task_title": item.get("task_title"), "canonical_image_id": item.get("image_id"),
        }

    # 실제 모델 1회 호출(provider 분기, Report 와 동일 adapter). Stage 2/캐시 없음.
    _t0 = _time.perf_counter()
    if provider == "dsllm":
        from app.llm import dsllm_adapter
        try:
            res = dsllm_adapter.vision_chat(
                model_name=model_name, instructions=s1_instr, content=s1_content,
                max_tokens=max_tokens, timeout=timeout,
            )
        except dsllm_adapter.DsllmVisionError as e:
            return _fail(e.error_type, e.status_code, int((_time.perf_counter() - _t0) * 1000))
        except Exception:
            logging.getLogger("main").exception("[ProjectImageVisionTest] dsllm 예외")
            return _fail("unknown_error", None, int((_time.perf_counter() - _t0) * 1000))
    else:
        from app.llm import openai_adapter
        if not openai_adapter.is_configured():
            return _fail("authentication_error", None, 0)
        try:
            res = openai_adapter.vision_chat(
                model_name=model_name, instructions=s1_instr, content=s1_content,
                max_tokens=max_tokens, timeout=timeout,
            )
        except openai_adapter.OpenAIAdapterError as e:
            return _fail(_std_openai_error_type(e), getattr(e, "status_code", None),
                         int((_time.perf_counter() - _t0) * 1000))
        except Exception:
            logging.getLogger("main").exception("[ProjectImageVisionTest] openai 예외")
            return _fail("unknown_error", None, int((_time.perf_counter() - _t0) * 1000))

    elapsed_ms = int((_time.perf_counter() - _t0) * 1000)
    raw_text = (res.get("raw_text") or "").strip()
    items = parse_stage1_evidence(raw_text)
    recognition = items[0] if items else None
    if recognition is not None:
        recognition["evidence_type"] = normalize_evidence_type(recognition.get("evidence_type"))

    metrics = _metrics(elapsed_ms)
    if not raw_text:
        # 응답은 왔지만 비어 있음 = 이미지 인식 실패로 간주(빈 응답).
        return _fail("malformed_response", 200, elapsed_ms)

    # 성공 시 INFO 시스템 로그(§8, category_detail=project_image, resolved=true).
    try:
        import json as _json
        _diag = {
            "failure_stage": "ai_settings_project_image_test",
            "failure_stage_label": "AI 설정 실제 프로젝트 이미지 테스트",
            "error_type": None, "provider": provider, "model": model_name,
            "http_status": 200, "recovery_event": False, "test_type": "project_image",
            "summary": "실제 프로젝트 이미지 Vision 입력 인식 성공.",
            "details": _vd.sanitize_details(metrics),
        }
        _log_event(
            "INFO", "VISION_AI",
            f"실제 프로젝트 이미지 Vision 입력 인식 성공 — {provider}/{model_name}",
            detail=f"project={body.project_id} image={item.get('image_id')} elapsed_ms={elapsed_ms}",
            detail_json=_json.dumps(_diag, ensure_ascii=False),
            endpoint=endpoint, status_code=200, login_id=login_id, resolved=True,
            dedupe_key=f"project_image_test|{item.get('image_id')}|success",
        )
    except Exception:
        pass

    return {
        "ok": True, "test_type": "project_image", "provider": provider, "model": model_name,
        "request_id": request_id, "recognition": recognition, "raw_present": bool(raw_text),
        "metrics": metrics, "project_id": body.project_id, "task_id": item.get("task_id"),
        "task_title": item.get("task_title"), "canonical_image_id": item.get("image_id"),
    }


# =========================================================
# AI Context Inspector (super_admin 검증 도구) — 부하/상한 정책
# =========================================================
AI_CONTEXT_IMAGES_DEFAULT_LIMIT = 50
AI_CONTEXT_IMAGES_MAX_LIMIT = 100


def _clamp_ai_context_max_images(max_images) -> int:
    """프론트 입력과 무관하게 서버에서 max_images 상한을 강제.

    - 값이 없거나 0 이하 → 기본 50
    - 100 초과 → 100 으로 clamp
    """
    try:
        n = int(max_images)
    except (TypeError, ValueError):
        return AI_CONTEXT_IMAGES_DEFAULT_LIMIT
    if n <= 0:
        return AI_CONTEXT_IMAGES_DEFAULT_LIMIT
    return min(n, AI_CONTEXT_IMAGES_MAX_LIMIT)


def _ai_context_warning_image_ids(manifest, warnings) -> set:
    """warnings 문자열에 image_id 가 언급된 이미지 집합."""
    ids = set()
    for w in warnings or []:
        for it in manifest:
            iid = it.get("image_id")
            if iid and iid in w:
                ids.add(iid)
    return ids


# role_confidence=low 단독으론 review_needed 가 아니지만, 아래 role 은 이미지 자체 판독
# 의존도가 높아 low confidence 와 결합되면 검토 대상으로 본다.
_REVIEW_LOW_CONF_ROLES = {"unknown", "text_crop", "partial_screen_capture"}


def _ai_context_image_is_normal(item: dict, warning_ids: set) -> bool:
    """'정상(검토 불필요)' 이미지 판정. 하나라도 어긋나면 review_needed 로 본다.

    정상 = readability ok + recommendation recommended + context_strength strong +
           nearby_text 존재(context_text_source=nearby_text, fallback 미사용) +
           original 사용 가능 + semantic_cache_key 존재 + 관련 warning 없음.

    role_confidence=low 는 **단독으로는** review_needed 사유가 아니다(정상 이미지 노이즈
    방지). 단 image_role 이 unknown/text_crop/partial_screen_capture 처럼 자체 판독
    의존도가 높은 경우에만 low confidence 와 결합해 검토 대상으로 본다.
    """
    q = item.get("image_quality") or {}
    card = item.get("evidence_card") or {}

    # 1) 가독성 / 추천도 / 맥락 신뢰도 (major signal)
    if q.get("readability_status") != "ok":
        return False
    if q.get("recommendation_level") != "recommended":
        return False
    if card.get("context_strength") != "strong":
        return False

    # 2) warning / nearby_text / fallback
    if item.get("image_id") in warning_ids:
        return False
    if not (item.get("nearby_text") or "").strip():
        return False
    if item.get("context_text_source") != "nearby_text":
        return False
    if (item.get("fallback_context_text") or "").strip():
        return False

    # 3) AI 입력 적합성 (original 사용 가능 + 의미 캐시 키 존재)
    original_available = ((item.get("ai_image_variants") or {}).get("original") or {}).get("available")
    if original_available is not True:
        return False
    if not ((item.get("analysis_identity") or {}).get("semantic_cache_key_preview") or ""):
        return False

    # 4) role_confidence=low 는 단독 사유 아님 — 의심 role 과 결합될 때만 검토 대상
    if (
        item.get("image_role_confidence") in ("low", "none")
        and item.get("image_role") in _REVIEW_LOW_CONF_ROLES
    ):
        return False

    return True


def _ai_context_apply_scope(scope: str, manifest: list, warning_ids: set):
    """image_scope 별 서버 필터. (normalized_scope, filtered_list) 반환.

    지원: all / review_needed / recommended / conditional / not_recommended /
          warnings / fallback_context / low_confidence. 미지원 → all.
    """
    def _level(it):
        return (it.get("image_quality") or {}).get("recommendation_level")

    predicates = {
        "review_needed": lambda it: not _ai_context_image_is_normal(it, warning_ids),
        "recommended": lambda it: _level(it) == "recommended",
        "conditional": lambda it: _level(it) == "conditional",
        "not_recommended": lambda it: _level(it) == "not_recommended",
        "warnings": lambda it: it.get("image_id") in warning_ids,
        "fallback_context": lambda it: bool((it.get("fallback_context_text") or "").strip())
        and it.get("context_text_source") != "nearby_text",
        "low_confidence": lambda it: it.get("image_role_confidence") in ("low", "none"),
    }
    pred = predicates.get(scope)
    if pred is None:
        return "all", list(manifest)
    return scope, [it for it in manifest if pred(it)]


def _redact_storage_ref(manifest) -> None:
    """프론트로 raw storage key / 경로 노출 방지. 이미지 preview 는 image_id 로만 접근한다.

    stored_name 은 builder 가 의도적으로 노출하는 불투명 식별자(uuid)라 유지하되,
    s3_key / local_path 는 응답에서 제거한다.
    """
    for it in manifest:
        sr = it.get("storage_ref")
        if isinstance(sr, dict):
            sr["s3_key"] = None
            sr["local_path"] = None


def _load_manifest_image_bytes(state: dict, item: dict) -> tuple[bytes, str]:
    """manifest item 1건의 실제 이미지 바이트를 서버 내부에서 로딩해 (bytes, media_type) 반환.

    S3 우선, 없으면 로컬 fallback. content-type 이 이미지가 아니면 거부.
    경로/스토리지 키는 호출자(응답)로 노출하지 않는다 — image_id 중심 접근 정책.
    (ai-context-preview/image 응답과 vision-test 모델 입력이 공유하는 로더.)
    """
    from app.utils.s3.s3_utils import download_from_s3, is_s3_configured, get_attachment_s3_key

    storage_ref = item.get("storage_ref") or {}
    stored_name = storage_ref.get("stored_name")
    s3_key = storage_ref.get("s3_key")
    filename = item.get("filename") or stored_name or "image"

    media_type, _disp = _attachment_response_meta(filename, stored_name or filename)
    if not media_type.startswith("image/"):
        raise HTTPException(status_code=415, detail="Not an image")

    if not stored_name and not s3_key:
        raise HTTPException(status_code=404, detail="Image source not available")

    # 원본 sidecar 레코드에서 소유 컨텍스트(local fallback 경로 + s3_key 보강) 재확인
    owner_type, owner_id = None, None
    if stored_name:
        for rec in state.get("space_images", []) or []:
            if rec.get("stored_name") == stored_name:
                owner_type, owner_id = "space", rec.get("space_id")
                s3_key = s3_key or rec.get("s3_key")
                break
        if owner_type is None:
            for rec in state.get("attachments", []) or []:
                if rec.get("stored_name") == stored_name:
                    owner_type, owner_id = "task", rec.get("task_id")
                    s3_key = s3_key or rec.get("s3_key")
                    break
        if owner_type is None:
            for rec in state.get("project_files", []) or []:
                if rec.get("stored_name") == stored_name:
                    owner_type, owner_id = "project", rec.get("project_id")
                    s3_key = s3_key or rec.get("s3_key")
                    break

    # 1) S3 우선
    if is_s3_configured():
        data = None
        if s3_key:
            data = download_from_s3(s3_key)
        if data is None and owner_type and owner_id is not None and stored_name:
            reconstructed = get_attachment_s3_key(filename, stored_name, owner_type, int(owner_id))
            data = download_from_s3(reconstructed)
        if data is not None:
            return data, media_type

    # 2) 로컬 fallback (owner_type 별 디렉터리 규칙)
    local_path = None
    if stored_name and owner_id is not None:
        if owner_type == "task":
            local_path = os.path.join(UPLOAD_DIR, "tasks", str(owner_id), stored_name)
        elif owner_type == "space":
            local_path = os.path.join(UPLOAD_DIR, "spaces", str(owner_id), stored_name)
        elif owner_type == "project":
            local_path = os.path.join(UPLOAD_DIR, str(owner_id), stored_name)
    if local_path and os.path.exists(local_path):
        with open(local_path, "rb") as f:
            return f.read(), media_type

    raise HTTPException(status_code=404, detail="Image file not found")


def _serve_manifest_image(state: dict, item: dict):
    """manifest item 1건의 실제 이미지 바이트를 응답. 경로/스토리지 키는 프론트에 노출하지 않는다."""
    from fastapi.responses import Response

    data, media_type = _load_manifest_image_bytes(state, item)
    return Response(
        content=data,
        media_type=media_type,
        headers={"Cache-Control": "private, max-age=300", "Content-Length": str(len(data))},
    )


def _prepare_vision_image_input(project_context: dict, item: dict, data: bytes, media_type: str,
                                *, input_mode: str) -> dict:
    """이미 로딩된 이미지 bytes → 모델 전송용 prepared input(전처리+base64 Data URL+packet).

    Vision Report Stage 1 과 AI 설정의 "실제 프로젝트 이미지 테스트"가 **동일한 전처리/스키마**
    를 쓰도록 이 결정론 변환을 공통화한다(§3). 바이트 로딩(`_load_manifest_image_bytes`)은 이미
    공통이며, 여기서는 role 기반 preprocess → base64 → model_input_packet 만 만든다(캐시/DB 무관).

    Returns:
        {"image_url": data URL, "packet": model_input_packet, "prep_meta": preprocess meta,
         "processed_media_type": str}
    """
    from app.services.vision_image_preprocess import preprocess_image_for_ai
    from app.services.project_ai_model_compatibility import build_model_input_packet
    import base64 as _b64

    proc_data, proc_media, prep_meta = preprocess_image_for_ai(
        data, media_type, role=item.get("image_role")
    )
    packet = build_model_input_packet(
        project_context, item, provider="generic", input_mode=input_mode
    )
    return {
        "image_url": f"data:{proc_media};base64,{_b64.b64encode(proc_data).decode('ascii')}",
        "packet": packet,
        "prep_meta": prep_meta,
        "processed_media_type": proc_media,
    }


@app.get("/api/projects/{project_id}/ai-context-preview")
def get_project_ai_context_preview(
    project_id: int,
    include_images: bool = Query(default=False),
    max_images: int = Query(default=AI_CONTEXT_IMAGES_DEFAULT_LIMIT),
    image_scope: Optional[str] = Query(default=None),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """프로젝트가 AI 에게 어떤 맥락으로 전달될 수 있는지 확인하는 Preview (디버깅/기반용).

    - **LLM 을 호출하지 않는다.** 기존 AI 요약 결과를 생성하지 않는다.
    - 프로젝트 조회 권한을 동일하게 적용한다(check_project_access → admin/super_admin 은
      모든 project, 일반 사용자는 본인 권한 project 만, 그 외 403).
    - text_context 일부 preview + image_manifest + counts + warnings 를 제공한다.
    - **이미지 바이트는 포함하지 않는다**(base64 금지). 실제 이미지는 image_id 단위
      ai-context-preview/image 로 1장씩 lazy 조회한다.
    - max_images 는 서버에서 상한(100) 강제. image_scope=review_needed 면 검토 필요
      이미지만 반환(summary/counts 는 전체 기준 유지).
    - VOC 데이터는 포함하지 않는다(AI 요약 대상은 프로젝트 단위로 한정).

    # Future multimodal path:
    # - 현재 모델은 text_context 만 사용. image_manifest 는 미래 vision 모델 연결을 위한
    #   구조화 데이터로만 생성하며, 여기서도 이미지 바이트/모델 호출은 하지 않는다.
    """
    state = load_state()
    # 권한: 기존 프로젝트 조회 권한과 동일(멤버/소유자/admin/공개 space 멤버)
    check_project_access(db, state, project_id, user_id)

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    from app.services.project_ai_context_builder import (
        build_project_ai_context,
        CONTEXT_SCHEMA_VERSION,
        IMAGE_MANIFEST_VERSION,
        READINESS_VERSION,
        FUTURE_IMAGE_ANALYSIS_PROMPT_VERSION,
        EVIDENCE_CARD_VERSION,
    )
    from app.services.ai_image_storage_policy import policy_descriptor

    safe_max_images = _clamp_ai_context_max_images(max_images)
    ctx = build_project_ai_context(
        db,
        project_id,
        state=state,
        include_images=bool(include_images),
        max_images=safe_max_images,
    )

    manifest = ctx.image_manifest
    warning_ids = _ai_context_warning_image_ids(manifest, ctx.warnings)

    # image_scope 서버 필터(summary/counts 는 전체 기준 유지). 미지정/미지원 → all.
    scope = ((image_scope or "").strip().lower()) or "all"
    total_images = len(manifest)
    scope, shown_manifest = _ai_context_apply_scope(scope, manifest, warning_ids)

    # 응답에서 raw storage key/경로 제거 (image preview 는 image_id 로만 접근)
    _redact_storage_ref(shown_manifest)

    return {
        "project_id": ctx.project_id,
        "project_name": ctx.project_name,
        # 스키마 버전 — 캐시/모델 교체/결과 비교 시 추적용
        "context_schema_version": CONTEXT_SCHEMA_VERSION,
        "image_manifest_version": IMAGE_MANIFEST_VERSION,
        "readiness_version": READINESS_VERSION,
        "future_image_analysis_prompt_version": FUTURE_IMAGE_ANALYSIS_PROMPT_VERSION,
        "evidence_card_version": EVIDENCE_CARD_VERSION,
        "counts": ctx.counts,
        # 전체 text_context 는 길 수 있으므로 preview 만 제공
        "text_context_preview": ctx.text_context[:2000],
        "image_manifest": shown_manifest,
        # Image Readability / AI-Ready Image Layer 요약 (AI 입력 준비도, 전체 기준)
        "image_readiness_summary": ctx.image_readiness_summary,
        # 이미지 분석 결과 재사용 정책 요약 (의미 분석은 versioned attempt 로 관리)
        "analysis_policy_summary": ctx.analysis_policy_summary,
        # variant 저장 정책(경로 규칙) — 현재 파일 미생성, 규칙만
        "variant_storage_policy": policy_descriptor(),
        "warnings": ctx.warnings,
        # 부하/스코프 메타
        "image_scope": scope,
        "image_scope_total": total_images,
        "image_scope_shown": len(shown_manifest),
        "max_images_applied": safe_max_images,
        "max_images_limit": AI_CONTEXT_IMAGES_MAX_LIMIT,
        "not_used_by_llm": True,
    }


@app.get("/api/projects/{project_id}/ai-context-preview/image")
def get_project_ai_context_preview_image(
    project_id: int,
    image_id: str = Query(...),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """AI Context Inspector 단일 이미지 preview (검증용, lazy 1장).

    - **vision/OCR/LLM 미호출.** 선택된 image_id 1장만 서버 내부에서 파일을 찾아 반환.
    - 권한: ai-context-preview 와 동일(check_project_access).
    - image_id 가 해당 project 의 image_manifest 안에 실제로 존재해야 한다(아니면 404).
    - 프론트엔 local_path / storage key 를 노출하지 않는다(image_id 만 입력).
    - 전체 이미지 일괄 조회 기능은 제공하지 않는다(항상 1장 단위).
    """
    state = load_state()
    check_project_access(db, state, project_id, user_id)

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    from app.services.project_ai_context_builder import build_project_ai_context

    # 상한(100)까지 빌드해 manifest 에서 image_id 검증 — 프론트가 보던 범위 내라면 항상 포함됨.
    ctx = build_project_ai_context(
        db,
        project_id,
        state=state,
        include_images=True,
        max_images=AI_CONTEXT_IMAGES_MAX_LIMIT,
    )
    item = next((it for it in ctx.image_manifest if it.get("image_id") == image_id), None)
    if not item:
        raise HTTPException(status_code=404, detail="image_id not found in project manifest")

    return _serve_manifest_image(state, item)


@app.get("/api/projects/{project_id}/ai-context-preview/model-packet")
def get_project_ai_context_model_packet(
    project_id: int,
    image_id: str = Query(...),
    provider: str = Query(default="generic"),
    input_mode: str = Query(default="image_with_evidence_card"),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """Model Compatibility Layer preview — image_manifest item 1건을 사내 모델 API 가
    이해하기 쉬운 표준 입력 패킷(model_input_packet) + provider payload 로 변환해 반환.

    - **vision/OCR/LLM 미호출.** payload 구조만 생성한다(실제 모델 호출은 별도 Phase).
    - Evidence Card 는 **정답이 아니라 힌트** 다 → image_role 은 image_role_hint 로 전달.
    - 권한: super_admin 전용(Inspector 와 동일 사용 맥락) + project 접근 권한 확인.
    - image_id 가 해당 project 의 image_manifest 안에 실제로 존재해야 한다(아니면 404).
    - **이미지 binary/base64/storage path 를 포함하지 않는다**(image_ref 는 image_id 중심).
    - input_mode: image_only / image_with_nearby_text / image_with_evidence_card (A/B/C 비교).
    """
    require_super_admin(db, user_id)
    state = load_state()
    check_project_access(db, state, project_id, user_id)

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    from app.services.project_ai_context_builder import build_project_ai_context
    from app.services.project_ai_model_compatibility import (
        build_model_input_packet,
        build_provider_payload,
        MODEL_INPUT_PACKET_VERSION,
        SUPPORTED_PROVIDERS,
        INPUT_MODES,
        DEFAULT_INPUT_MODE,
    )

    safe_provider = (provider or "generic").lower()
    if safe_provider not in SUPPORTED_PROVIDERS:
        safe_provider = "generic"
    safe_input_mode = input_mode if input_mode in INPUT_MODES else DEFAULT_INPUT_MODE

    # 상한(100)까지 빌드해 manifest 에서 image_id 검증 — 전체 이미지를 한 번에 처리하지 않고
    # 선택된 image_id 1개에 대해서만 packet 을 생성한다.
    ctx = build_project_ai_context(
        db,
        project_id,
        state=state,
        include_images=True,
        max_images=AI_CONTEXT_IMAGES_MAX_LIMIT,
    )
    item = next((it for it in ctx.image_manifest if it.get("image_id") == image_id), None)
    if not item:
        raise HTTPException(status_code=404, detail="image_id not found in project manifest")

    packet = build_model_input_packet(
        ctx.project_context,
        item,
        provider=safe_provider,
        input_mode=safe_input_mode,
    )
    provider_payload = build_provider_payload(packet, provider=safe_provider)

    return {
        "image_id": image_id,
        "provider": safe_provider,
        "input_mode": safe_input_mode,
        "packet_version": MODEL_INPUT_PACKET_VERSION,
        "packet": packet,
        "provider_payload": provider_payload,
        # 이미지 바이트/경로는 포함하지 않으며 실제 모델 호출도 하지 않음을 명시
        "not_used_by_llm": True,
    }


class VisionReportTestRequest(BaseModel):
    """super_admin 전용 이미지 포함 AI Report 테스트(PoC) 요청.

    기존 텍스트 종합보고서(/api/report/generate)와 완전히 분리. DB 저장 없음.
    """
    user_id: int
    model: Optional[str] = None
    provider: str = "openai"
    # 기본 UI 는 아래 image 파라미터를 노출하지 않는다(고급 설정 전용, optional).
    #   input_mode  : 기본 image_with_evidence_card (텍스트+이미지+Evidence Card 종합)
    #   image_scope : 기본 task_images_all (프로젝트 Task/작업노트 이미지 전체)
    #   max_images  : None → 서버 안전 상한까지 "프로젝트 전체 이미지" 포함
    input_mode: str = "image_with_evidence_card"
    image_scope: str = "task_images_all"
    max_images: Optional[int] = None
    image_ids: List[str] = Field(default_factory=list)
    include_text_context: bool = True
    include_image_evidence: bool = True
    save_result: bool = False  # PoC: 항상 무시(저장 안 함)


@app.get("/api/projects/{project_id}/ai-report/vision-test/meta")
def get_vision_report_test_meta(
    project_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """Vision Report Test 탭 UI용 메타(모델 목록/제한/모드). super_admin 전용.

    - 실제 모델 호출 없음. 드롭다운/제한값만 제공.
    """
    require_super_admin(db, user_id)
    state = load_state()
    check_project_access(db, state, project_id, user_id)

    from app.llm import openai_adapter
    from app.services.project_ai_vision_report_test import (
        VISION_TEST_DEFAULT_MAX_IMAGES,
        VISION_TEST_HARD_MAX_IMAGES,
        IMAGE_SCOPES,
        DEFAULT_IMAGE_SCOPE,
        DEFAULT_INPUT_MODE,
    )
    from app.services.project_ai_model_compatibility import INPUT_MODES
    from app.services.vision_ai_settings import resolve_vision_ai_config

    # 모델 목록/기본값은 Text AI 가 아니라 **Vision AI 설정**에서 가져온다(완전 분리).
    cfg = resolve_vision_ai_config(db)

    return {
        "models": cfg["models"],
        "default_model": cfg["model"],
        "model_capabilities": cfg["model_capabilities"],
        "vision_enabled": cfg["enabled"],
        "using_text_fallback": cfg["using_text_fallback"],
        "fallback_reason": cfg["fallback_reason"],
        "default_max_images": VISION_TEST_DEFAULT_MAX_IMAGES,
        "hard_max_images": VISION_TEST_HARD_MAX_IMAGES,
        "input_modes": list(INPUT_MODES),
        "default_input_mode": DEFAULT_INPUT_MODE,
        "image_scopes": list(IMAGE_SCOPES),
        "default_image_scope": DEFAULT_IMAGE_SCOPE,
        "openai_configured": openai_adapter.is_configured(),
        # ── 지연 시 텍스트 전용 fallback(사용자 선택 modal) 관련 메타 ──
        "provider": cfg["provider"],
        "soft_timeout_seconds": cfg["soft_timeout_seconds"],
        "timeout_seconds": cfg["timeout_seconds"],
        "fallback_text_enabled": cfg["fallback_text_enabled"],
        "fallback_text_provider": cfg["fallback_text_provider"],
        "fallback_text_model": cfg["fallback_text_model"],
        "fallback_policy": cfg["fallback_policy"],
    }


def _build_vision_report_base(db: Session, state: dict, project_id: int, ctx) -> dict:
    """Vision Report 의 **deterministic base**(서버/DB가 만드는 구조·사실 데이터).

    모델이 다시 생성/추론하면 안 되는 값(프로젝트명/설명/팀원/진행률/상태 count/
    Task 표: 상태·우선순위·진행률·마감일·담당자·단계)을 DB 에서 그대로 만든다.
    기존 `/api/report/generate` 의 structured 와 동일 산식(미수정, 별도 계산).

    Returns:
        {"project_overview": {...}, "tasks": [...], "attachment_summary": [...]}
    """
    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    project = project_dict(p, state) if p else {"name": "", "description": ""}

    task_rows = db.query(Task).filter(Task.project_id == project_id, Task.archived_at.is_(None)).all()
    tasks = [task_dict(t, state) for t in task_rows]
    sub_projects = get_subprojects_from_db(db, project_id)
    all_attachments = state.get("attachments", []) or []
    users_map = {u.id: user_dict(u, state) for u in db.query(User).all()}

    # 진행률(Hold 제외) — generate_report 와 동일 산식
    active = [t for t in tasks if t.get("status") != "hold"]
    done = [t for t in tasks if t.get("status") == "done"]
    in_progress = [t for t in tasks if t.get("status") == "in_progress"]
    todo = [t for t in tasks if t.get("status") == "todo"]
    hold = [t for t in tasks if t.get("status") == "hold"]
    if active:
        progress_sum = sum(100 if t.get("status") == "done" else (t.get("progress", 0) or 0) for t in active)
        overall_progress = round(progress_sum / len(active), 1)
    else:
        overall_progress = 0.0

    # 팀원 (viewer 제외)
    db_members = db.query(ProjectMemberModel).filter(
        ProjectMemberModel.project_id == project_id,
        ProjectMemberModel.role != "viewer",
    ).all()
    member_names: list = []
    for m in db_members:
        u = users_map.get(int(m.user_id))
        if u:
            member_names.append(f'{u["username"]} ({m.role})')

    # Task 이미지 개수(manifest 기준, Task/작업노트 이미지만)
    task_image_count: dict = {}
    for it in (getattr(ctx, "image_manifest", None) or []):
        if it.get("source_entity_type") in ("task_description", "task_activity", "task_attachment"):
            tid = it.get("task_id")
            if tid is not None:
                task_image_count[int(tid)] = task_image_count.get(int(tid), 0) + 1

    # 작업노트(TaskActivity) 개수 — Task별 grouped count 1회 쿼리(근거 컬럼 📝).
    note_count_map: dict = {}
    note_rows = (
        db.query(TaskActivityModel.task_id, func.count(TaskActivityModel.id))
        .join(Task, Task.id == TaskActivityModel.task_id)
        .filter(Task.project_id == project_id, Task.archived_at.is_(None))
        .group_by(TaskActivityModel.task_id)
        .all()
    )
    for tid_, cnt_ in note_rows:
        if tid_ is not None:
            note_count_map[int(tid_)] = int(cnt_ or 0)

    sp_by_id = {s["id"]: s for s in sub_projects}
    base_tasks: list = []
    attachment_summary: list = []
    for t in tasks:
        tid = t["id"]
        assignees = [users_map.get(a, {}).get("username", f"User {a}") for a in (t.get("assignee_ids") or [])]
        sp_name = (sp_by_id.get(t.get("sub_project_id")) or {}).get("name", "") if t.get("sub_project_id") else ""
        img_n = task_image_count.get(tid, 0)
        # 첨부 자료: Task 단위로 url/file 묶음(이미지 id 나열 대신 요약 중심)
        t_atts = [a for a in all_attachments if int(a.get("task_id") or 0) == tid]
        urls = [a.get("url", "") for a in t_atts if a.get("type") == "url" and a.get("url")]
        files = [a.get("filename", "") for a in t_atts if a.get("type") == "file" and a.get("filename")]
        note_n = note_count_map.get(tid, 0)
        base_tasks.append({
            "task_id": tid,
            "task_title": t.get("title", ""),
            "status": t.get("status", "todo"),
            "priority": t.get("priority", "medium"),
            "progress": t.get("progress", 0) or 0,
            "due_date": t.get("due_date"),
            "assignees": assignees,
            "sub_project": sp_name or None,
            "image_count": img_n,
            # 근거 컬럼(📷/📝/🔗) — 표에서 한눈에 보일 Task별 근거 자료 수.
            "note_count": note_n,
            "url_count": len(urls),
            "file_count": len(files),
        })
        if urls or files or img_n:
            attachment_summary.append({
                "task_id": tid,
                "task_title": t.get("title", ""),
                "urls": urls,
                "files": files,
                "image_count": img_n,
            })

    status_breakdown = {
        "total": len(tasks),
        "active": len(active),
        "done": len(done),
        "in_progress": len(in_progress),
        "todo": len(todo),
        "hold": len(hold),
        "overall_progress": overall_progress,
    }

    return {
        "project_overview": {
            "title": project.get("name", ""),
            "description": project.get("description", "") or "",
            "team": member_names,
            "overall_progress": overall_progress,
            "status_breakdown": status_breakdown,
        },
        "tasks": base_tasks,
        "attachment_summary": attachment_summary,
    }


@app.post("/api/projects/{project_id}/ai-report/vision-test")
def run_vision_report_test(
    project_id: int,
    body: VisionReportTestRequest,
    db: Session = Depends(get_db),
):
    """super_admin 전용 이미지 포함 AI Report 테스트(PoC).

    프로젝트 내 Task/작업노트 캡처 이미지들을 text_context + nearby_text + Evidence Card
    와 함께 vision 모델(OpenAI)에 넣어 이미지별 해석 + 종합 요약을 테스트한다.

    안전 정책:
    - super_admin 전용 + project 접근 권한 확인.
    - 이미지 파일은 서버 내부에서만 로딩(프론트는 image_id 만 전달). 경로 비노출.
    - max_images 기본 3 / 서버 hard 5 강제. 전체 무제한 분석 금지.
    - DB 저장 없음(save_result 무시). 기존 report/ai-query 결과 미반영.
    - 실패해도 기존 텍스트 종합보고서에 영향 없음.
    """
    import base64
    import time
    _req_t0 = time.perf_counter()

    require_super_admin(db, body.user_id)
    state = load_state()
    check_project_access(db, state, project_id, body.user_id)

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    from app.llm import openai_adapter
    from app.services.project_ai_context_builder import build_project_ai_context
    from app.services.project_ai_model_compatibility import build_model_input_packet
    from app.services.project_ai_vision_report_test import (
        select_images,
        build_stage1_input,
        parse_stage1_evidence,
        reconcile_evidence_with_images,
        normalize_evidence_type,
        build_enhancement_input,
        parse_enhancements,
        validate_enhancements,
        build_vision_enhanced_report,
        looks_truncated_json,
        chunk_list,
        stage1_batch_size,
        merge_usage,
        resolve_max_images,
        normalize_input_mode,
        default_vision_model,
        VISION_TEST_HARD_MAX_IMAGES,
        FINAL_REPORT_SCHEMA_VERSION,
    )

    # ── Vision AI 설정(Text AI 와 분리). 모델/제한/토큰/배치/타임아웃/provider 는 여기서만 가져온다 ──
    from app.services.vision_ai_settings import resolve_vision_ai_config
    from app.services.vision_image_preprocess import preprocess_image_for_ai
    from app.services.vision_image_observation import compute_image_sha256
    vision_cfg = resolve_vision_ai_config(db)
    if not vision_cfg["enabled"]:
        raise HTTPException(
            status_code=400,
            detail="Vision AI가 비활성화되어 있습니다. AI Settings의 Vision AI 설정에서 활성화해주세요.",
        )

    # provider 분기: openai(외부)=Responses vision_chat, dsllm(사내)=chat/completions 멀티모달.
    vision_provider = vision_cfg["provider"]
    if vision_provider == "dsllm":
        from app.llm import dsllm_adapter

        if not dsllm_adapter.supports_vision_payload():
            model_hint = (body.model or "").strip() or vision_cfg["model"] or "DS/Gemma4"
            raise HTTPException(
                status_code=400,
                detail=(
                    f"{model_hint}가 Vision AI 모델로 선택되었지만, 현재 DSLLM vision image payload "
                    "처리가 아직 구현되지 않았습니다. 사내 Gemma4 vision adapter 구현 후 다시 테스트해주세요."
                ),
            )
    else:
        # vision 호출 경로(openai)는 OpenAI adapter 만 지원. key 미설정이면 명확히 안내.
        if not openai_adapter.is_configured():
            raise HTTPException(
                status_code=400,
                detail=(
                    "OpenAI API Key가 설정되어 있지 않습니다. "
                    "backend env 파일에 OPENAI_API_KEY를 설정한 뒤 다시 시도해주세요."
                ),
            )

    safe_input_mode = normalize_input_mode(body.input_mode)
    # 기본 UI 는 max_images 를 보내지 않는다(None) → 안전 상한까지 프로젝트 전체 이미지 포함.
    safe_max_images = resolve_max_images(body.max_images)

    # 전체 manifest 빌드(상한까지) 후 선택. 한 번에 전체를 모델에 넣지 않는다.
    ctx = build_project_ai_context(
        db,
        project_id,
        state=state,
        include_images=True,
        max_images=AI_CONTEXT_IMAGES_MAX_LIMIT,
    )
    warning_ids = _ai_context_warning_image_ids(ctx.image_manifest, ctx.warnings)

    selected, scope, missing_ids, sel_info = select_images(
        ctx.image_manifest,
        image_ids=body.image_ids,
        image_scope=body.image_scope,
        max_images=safe_max_images,
        warning_ids=warning_ids,
        is_normal=_ai_context_image_is_normal,
    )
    if not selected:
        raise HTTPException(
            status_code=400,
            detail="분석할 Task/작업노트 이미지가 없습니다. image_scope 또는 선택한 이미지를 확인해주세요.",
        )

    # ── 모델/토큰/타임아웃 결정(캐시 조회에도 필요하므로 이미지 준비 전에 확정) ──
    resolved_model = (body.model or "").strip() or vision_cfg["model"]
    if not resolved_model:
        from app.llm import dsllm_adapter as _ds
        resolved_model = _ds.default_model_key() if vision_provider == "dsllm" else default_vision_model()
    vision_max_tokens = vision_cfg["max_output_tokens"]
    vision_timeout = float(vision_cfg["timeout_seconds"])

    # ── 2b-2a Exact Match Read-through: 캐시 활성 여부(플래그 + canary) ──
    from app.config import settings as _cache_settings
    from app.services import vision_image_observation as _obs
    _canary_raw = (getattr(_cache_settings, "vision_image_cache_canary_project_ids", "") or "").strip()
    _canary_ids = {int(x) for x in _canary_raw.replace(" ", "").split(",") if x.isdigit()}
    _rollout_mode = (getattr(_cache_settings, "vision_image_cache_rollout_mode", "off") or "off").strip().lower()
    cache_read_active = _obs.is_cache_read_active(
        enabled=_cache_settings.vision_image_observation_enabled,
        cache_read_enabled=_cache_settings.vision_image_observation_cache_read_enabled,
        rollout_mode=_rollout_mode, canary_ids=_canary_ids, project_id=project_id,
    )
    # 활성화 진단: (1)플래그 off (2)rollout/canary 불일치 (3)env 미반영(재시작 안 됨) 을 구분할 수 있게 기록.
    logging.getLogger("main").info(
        "[VisionReport] cache activation: project_id=%s observation_enabled=%s cache_read_enabled=%s "
        "rollout_mode=%s canary_ids=%s cache_read_active=%s",
        project_id, _cache_settings.vision_image_observation_enabled,
        _cache_settings.vision_image_observation_cache_read_enabled,
        _rollout_mode, sorted(_canary_ids) or "(none)", cache_read_active,
    )
    cached_evidence: list = []

    # ── 2b-2b: stale ContextLink Text 재생성 준비(원본 이미지 없이 Text 모델 batch) ──
    _regen_text_provider = (vision_cfg.get("fallback_text_provider") or "dsllm").strip().lower()
    if _regen_text_provider not in ("openai", "dsllm"):
        _regen_text_provider = "dsllm"
    _regen_text_model = (vision_cfg.get("fallback_text_model") or "").strip()
    _regen_text_ok = bool(
        vision_cfg.get("fallback_text_enabled", True) and _regen_text_model
        and (_regen_text_provider != "openai" or llm_router.openai_enabled())
    )
    regen_enabled = bool(
        cache_read_active
        and _cache_settings.vision_image_context_link_regen_enabled
        and _regen_text_ok
    )
    _regen_batch_size = int(getattr(_cache_settings, "vision_image_context_link_regen_batch_size", 12) or 12)
    stale_items: list = []
    regenerated_evidence: list = []
    regen_stats = {
        "enabled": regen_enabled, "requested": 0, "batch_calls": 0, "succeeded": 0, "failed": 0,
        "saved": 0, "save_failed": 0, "vision_fallback_images": 0,
        "input_tokens": 0, "output_tokens": 0, "elapsed_ms": 0,
        "first_error_stage": None, "first_error_type": None, "first_error_message": None,
    }

    def _img_source_text_hash(item):
        # write(shadow/write-back)와 read 가 반드시 동일 입력을 써야 매칭된다.
        # task 제목 + task_id + 이미지 인접 텍스트(작업노트/설명) → 텍스트 변경 시 hash 변화 → miss.
        return _obs.compute_source_text_hash(
            item.get("task_title"), str(item.get("task_id")), item.get("nearby_text")
        )

    _cache_expected = dict(
        expected_provider=vision_provider, expected_model=resolved_model,
        expected_pipeline_version=_obs.IMAGE_PIPELINE_VERSION,
        expected_prompt_version=_obs.IMAGE_OBSERVATION_PROMPT_VERSION,
        expected_preprocessing_version=_obs._preprocess_version(),
        expected_detail_policy=_obs.DEFAULT_DETAIL_POLICY,
    )
    cache_stats = {
        "lookup_attempted": 0, "observation_hits": 0, "observation_misses": 0,
        "observation_usable": 0, "observation_warning_hits": 0,
        "observation_raw_review_required": 0, "observation_invalid": 0,
        "context_link_hits": 0, "context_link_misses": 0,
        # ContextLink 상태 세분화(§2): stale(문맥 오래됨=재생성 대상) / missing(row 없음) / invalid(손상).
        "context_link_stale": 0, "context_link_missing": 0, "context_link_invalid": 0,
        "exact_match_hits": 0, "fallback_images": 0, "vision_fallback_images": 0,
        "lookup_elapsed_ms": 0,
        "context_link_miss_reasons": {
            "source_text_hash_mismatch": 0, "identity_mismatch": 0,
            "invalid_evidence_type": 0, "invalid_task_relevance": 0,
            "prompt_version_mismatch": 0,
        },
    }
    _cache_t0 = time.perf_counter()

    def _try_cache_hit(item, iid, sha):
        """캐시 조회 결과를 (status, payload) 로 반환.

        - ("hit", legacy_evidence): Observation usable + ContextLink exact 일치.
        - ("stale", {observation, source_text_hash, task_id}): Observation usable 인데
          ContextLink 가 stale(문맥 텍스트/프롬프트 변경). 2b-2b Text 재생성 후보.
        - ("miss", None): 그 외(=원본 Vision 재분석 대상).
        """
        cache_stats["lookup_attempted"] += 1
        try:
            obs = _obs.find_valid_observation(
                db, image_sha256=sha, provider=vision_provider, model_name=resolved_model,
                pipeline_version=_obs.IMAGE_PIPELINE_VERSION,
                prompt_version=_obs.IMAGE_OBSERVATION_PROMPT_VERSION,
                preprocessing_version=_obs._preprocess_version(),
                detail_policy=_obs.DEFAULT_DETAIL_POLICY, statuses=("completed",),
            )
            if obs is None:
                cache_stats["observation_misses"] += 1
                return ("miss", None)
            cache_stats["observation_hits"] += 1
            usability = _obs.evaluate_observation_usability(obs, **_cache_expected)
            if usability == _obs.ObservationUsability.USABLE:
                cache_stats["observation_usable"] += 1
            elif usability == _obs.ObservationUsability.USABLE_WITH_WARNING:
                cache_stats["observation_warning_hits"] += 1
            elif usability == _obs.ObservationUsability.RAW_REVIEW_REQUIRED:
                cache_stats["observation_raw_review_required"] += 1
                return ("miss", None)
            else:
                cache_stats["observation_invalid"] += 1
                return ("miss", None)
            tid = item.get("task_id")
            if tid is None:
                cache_stats["context_link_misses"] += 1
                return ("miss", None)
            sth = _img_source_text_hash(item)
            link = _obs.find_context_link(
                db, observation_id=obs.id, task_id=int(tid),
                source_text_hash=sth, prompt_version=_obs.IMAGE_OBSERVATION_PROMPT_VERSION,
            )
            _reasons = cache_stats["context_link_miss_reasons"]
            if link is None:
                _r = _obs.diagnose_context_link_miss(
                    db, observation_id=obs.id, task_id=int(tid),
                    source_text_hash=sth, prompt_version=_obs.IMAGE_OBSERVATION_PROMPT_VERSION,
                )
                _reasons[_r] = _reasons.get(_r, 0) + 1
                cache_stats["context_link_misses"] += 1
                # 문맥 변경(hash/prompt)=stale(재생성 대상), row 없음=missing(§2).
                if _r == "identity_mismatch":
                    cache_stats["context_link_missing"] += 1
                    return ("miss", None)
                cache_stats["context_link_stale"] += 1
                return ("stale", {"observation": obs, "source_text_hash": sth, "task_id": int(tid)})
            if not _obs.is_usable_context_link(
                link, observation_id=obs.id, project_id=project_id, task_id=int(tid),
                canonical_image_id=iid, source_text_hash=sth,
                prompt_version=_obs.IMAGE_OBSERVATION_PROMPT_VERSION,
            ):
                # exact 매치는 있었지만 검증 실패 → 손상(invalid)으로 분류(§2).
                _et = getattr(link, "evidence_type", None) or ""
                _tr = getattr(link, "task_relevance", None) or ""
                if _et not in _obs._ALLOWED_EVIDENCE_TYPES:
                    _reasons["invalid_evidence_type"] += 1
                elif _tr not in _obs.TASK_RELEVANCE_VALUES:
                    _reasons["invalid_task_relevance"] += 1
                else:
                    _reasons["identity_mismatch"] += 1
                cache_stats["context_link_invalid"] += 1
                cache_stats["context_link_misses"] += 1
                return ("miss", None)
            cache_stats["context_link_hits"] += 1
            cache_stats["exact_match_hits"] += 1
            return ("hit", _obs.build_legacy_evidence_from_cache(
                image_id=iid, task_id=tid, task_title=item.get("task_title"),
                image_role=item.get("image_role"), observation=obs, context_link=link,
            ))
        except Exception:
            logging.getLogger("main").exception(
                "[VisionReport] cache lookup failed image=%s (fallback to vision)", iid
            )
            cache_stats["fallback_images"] += 1
            return ("miss", None)

    # 선택 이미지별: 서버 내부 파일 로딩 → (캐시 hit면 재사용) 또는 data URL + model_input_packet
    prepared: list = []
    load_errors: list = []
    for item in selected:
        iid = item.get("image_id")
        try:
            data, media_type = _load_manifest_image_bytes(state, item)
        except HTTPException as e:
            load_errors.append({"image_id": iid, "error": str(e.detail)})
            continue
        _sha = compute_image_sha256(data)
        # ── 2b-2a/2b-2b: 캐시 조회(활성 시). hit=재사용, stale=Text 재생성 후보, miss=Vision ──
        if cache_read_active:
            _status, _payload = _try_cache_hit(item, iid, _sha)
            if _status == "hit":
                cached_evidence.append(_payload)
                continue
            if _status == "stale" and regen_enabled:
                # Observation 은 재사용하고 ContextLink 만 Text 로 재생성한다. 실패 시 vision fallback
                # 을 위해 원본 bytes 를 보관한다.
                stale_items.append({
                    "item": item, "iid": iid, "sha": _sha,
                    "data": data, "media_type": media_type,
                    "observation": _payload["observation"],
                    "source_text_hash": _payload["source_text_hash"],
                    "task_id": _payload["task_id"],
                })
                continue
            # ("miss") 또는 (stale + regen off) → 아래 전처리+prepare(=기존 Vision 경로, 2b-2a)
        # AI 전송용 전처리: 긴 변 축소 + JPEG 재인코딩(표/문서/화면은 더 큰 한도). 표시용 원본은
        # 그대로 유지되고, 여기서 만든 축소본은 오직 모델 payload 로만 쓴다(provider 무관).
        # ⚠️ AI 설정 "실제 프로젝트 이미지 테스트"와 동일한 공통 helper(§3).
        _prep = _prepare_vision_image_input(
            ctx.project_context, item, data, media_type, input_mode=safe_input_mode
        )
        prep_meta = _prep["prep_meta"]
        prepared.append({
            "image_id": iid,
            "image_url": _prep["image_url"],
            "packet": _prep["packet"],
            # authoritative Task 매핑(교정용). LLM echo 대신 이 값으로 evidence 를 Task 에 연결한다.
            "task_id": item.get("task_id"),
            "task_title": item.get("task_title"),
            "_prep_meta": prep_meta,
            # 이미지 관찰 캐시용: 원본 바이트 SHA-256(전처리 전) + role + 문맥 텍스트 hash(write-back용).
            "sha256": _sha,
            "image_role": item.get("image_role"),
            "source_text_hash": _img_source_text_hash(item),
        })

    cache_stats["lookup_elapsed_ms"] = int((time.perf_counter() - _cache_t0) * 1000)

    # ── 2b-2b: stale ContextLink 를 원본 이미지 없이 Text batch 로 재생성 ──
    def _append_stale_to_prepared(s):
        """재생성 실패 항목 → 원본 이미지로 Vision fallback(전처리 후 prepared 에 추가)."""
        _it = s["item"]
        _p = _prepare_vision_image_input(
            ctx.project_context, _it, s["data"], s["media_type"], input_mode=safe_input_mode
        )
        prepared.append({
            "image_id": s["iid"], "image_url": _p["image_url"], "packet": _p["packet"],
            "task_id": _it.get("task_id"), "task_title": _it.get("task_title"),
            "_prep_meta": _p["prep_meta"], "sha256": s["sha"], "image_role": _it.get("image_role"),
            "source_text_hash": s["source_text_hash"],
        })

    _regen_elapsed_ms = 0
    if regen_enabled and stale_items:
        from app.services import vision_context_link_regen as _regen
        from app.db_connections.sqlalchemy import SessionLocal as _RegenSession
        _regen_t0 = time.perf_counter()
        regen_inputs = [
            _regen.build_regen_input_item(
                canonical_image_id=s["iid"], project_id=project_id, task_id=s["task_id"],
                task_title=s["item"].get("task_title"),
                task_description=s["item"].get("description") or s["item"].get("task_description"),
                note_text=s["item"].get("nearby_text"),
                nearby_text=s["item"].get("nearby_text"),
                caption=s["item"].get("user_caption"),
                image_role=s["item"].get("image_role"),
                observation_json=getattr(s["observation"], "observation_json", None) or {},
            )
            for s in stale_items
        ]

        def _regen_text_chat(system, user):
            return llm_router.chat(
                provider=_regen_text_provider, base_url=None, model_name=_regen_text_model,
                system_prompt=system, user_prompt=user, temperature=0.2,
                max_tokens=min(vision_max_tokens, 2000),
            )

        try:
            _results_by_id, _batch_stats = _regen.regenerate_context_links_batch(
                regen_inputs, text_chat_fn=_regen_text_chat, batch_size=_regen_batch_size,
            )
        except Exception:
            logging.getLogger("main").exception("[ContextLinkRegen] batch orchestration 실패")
            _results_by_id, _batch_stats = {}, {"requested": len(regen_inputs), "batch_calls": 0,
                                               "succeeded": 0, "failed": len(regen_inputs)}
        for _k in ("requested", "batch_calls", "succeeded", "failed",
                   "first_error_stage", "first_error_type", "first_error_message"):
            if _k in _batch_stats:
                regen_stats[_k] = _batch_stats[_k]

        _regen_db = _RegenSession()
        try:
            for s in stale_items:
                res = _results_by_id.get(str(s["iid"]))
                if res is None:
                    _append_stale_to_prepared(s)   # 재생성 실패 → 해당 이미지만 Vision fallback
                    regen_stats["vision_fallback_images"] += 1
                    continue
                _o = s["observation"]
                try:
                    _obs.upsert_context_link(
                        _regen_db, observation_id=_o.id, project_id=project_id, task_id=s["task_id"],
                        canonical_image_id=s["iid"], source_text_hash=s["source_text_hash"],
                        evidence_type=res["evidence_type"],
                        prompt_version=_obs.IMAGE_OBSERVATION_PROMPT_VERSION,
                        task_relevance=res["task_relevance"],
                        supports_json=res.get("supports"), contradictions_json=res.get("contradictions"),
                        context_summary=res.get("context_summary"), confidence=res.get("confidence"),
                        model_name=_regen_text_model, commit=True,
                    )
                    regen_stats["saved"] += 1
                except Exception:
                    _regen_db.rollback()
                    regen_stats["save_failed"] += 1
                    logging.getLogger("main").exception(
                        "[ContextLinkRegen] ContextLink save 실패 image=%s", s["iid"]
                    )
                # 저장 성공 여부와 무관하게 재생성된 문맥은 이번 보고서에 사용한다.
                regenerated_evidence.append(_obs.build_legacy_evidence_from_regen(
                    image_id=s["iid"], task_id=s["task_id"], task_title=s["item"].get("task_title"),
                    image_role=s["item"].get("image_role"), observation=_o, regen_result=res,
                ))
        finally:
            _regen_db.close()
        _regen_elapsed_ms = int((time.perf_counter() - _regen_t0) * 1000)
        regen_stats["elapsed_ms"] = _regen_elapsed_ms

    # cache lookup 이후 원본이 Vision 으로 전달된 고유 이미지 수(=miss + 재생성실패 fallback)(§1).
    cache_stats["vision_fallback_images"] = len(prepared) if cache_read_active else 0
    # all-hit / 캐시+재생성만 있어도 정상 — Vision 호출 없이 진행한다.
    if not prepared and not cached_evidence and not regenerated_evidence:
        raise HTTPException(
            status_code=404,
            detail="선택된 이미지 파일을 불러오지 못했습니다.",
        )

    _vlog = logging.getLogger("main")
    _call_counter = {"n": 0}
    # 이 보고서 1건의 모든 Vision 호출을 묶는 request_id(§2). AI Settings 테스트 로그와 동일하게
    # request_id 를 남겨, 두 경로의 호출을 로그에서 서로 대조할 수 있게 한다.
    report_request_id = uuid.uuid4().hex[:8]

    # VISION_AI 시스템 로그(진단)용 컨텍스트: 운영자가 Request ID/Login 으로 검색·대조.
    _vision_endpoint = f"/api/projects/{project_id}/ai-report/vision-test"
    try:
        _vu = db.query(User).filter(User.id == body.user_id).first()
        _vision_login_id = getattr(_vu, "login_id", None)
    except Exception:
        _vision_login_id = None

    def _vision_error_details(*, batch_index, image_ids, image_metas, img_count,
                              payload_bytes, eff_max, elapsed_ms) -> dict:
        """진단 details(비민감 지표만) 구성. base64/프롬프트/키는 포함하지 않는다."""
        first = (image_metas or [{}])[0] if image_metas else {}
        return {
            "project_id": project_id,
            "canonical_image_id": (",".join(str(i) for i in (image_ids or [])) or None),
            "image_count": img_count,
            "image_transport": "data_url",
            "mime": first.get("mime"),
            "width": first.get("width"),
            "height": first.get("height"),
            "processed_bytes": first.get("processed_bytes"),
            "payload_bytes": payload_bytes,
            "max_tokens": eff_max,
            "timeout_seconds": int(vision_timeout),
            "elapsed_ms": elapsed_ms,
            "batch_index": batch_index,
        }

    def _vision_payload_stats(content) -> tuple:
        """(payload_bytes 근사, image_count, text_chars) 계산.

        원문/base64/키는 로그에 노출하지 않고 크기만 잰다. payload_bytes 는 텍스트+이미지
        URL(base64 포함) 바이트 합, text_chars 는 텍스트 파트 글자수(이미지 URL 제외).
        """
        if isinstance(content, str):
            b = content.encode("utf-8", "ignore")
            return len(b), 0, len(content)
        total = 0
        imgs = 0
        chars = 0
        for c in content or []:
            if not isinstance(c, dict):
                continue
            t = c.get("type")
            if t in ("input_text", "text"):
                txt = c.get("text") or ""
                total += len(txt.encode("utf-8", "ignore"))
                chars += len(txt)
            elif t in ("input_image", "image_url"):
                url = c.get("image_url")
                if isinstance(url, dict):
                    url = url.get("url")
                total += len((url or "").encode("utf-8", "ignore"))
                imgs += 1
        return total, imgs, chars

    def _format_image_metas(image_metas) -> str:
        """배치 이미지의 identity(전처리 후 mime/bytes/dims)를 안전한 요약 문자열로(§2).

        image_metas: [{"image_id","mime","processed_bytes","width","height"}, ...].
        base64/내용은 넣지 않고 크기/치수만 남긴다.
        """
        parts = []
        for m in image_metas or []:
            parts.append(
                "%s(mime=%s,bytes=%s,%sx%s)" % (
                    m.get("image_id"), m.get("mime"), m.get("processed_bytes"),
                    m.get("width"), m.get("height"),
                )
            )
        return ";".join(parts) or "-"

    def _vision_call(instructions, content, *, stage: str, max_tokens=None, image_ids=None,
                     batch_index=None, image_metas=None):
        # provider 분기: dsllm=chat/completions 멀티모달, openai=Responses vision.
        # 단계별 진단: 호출/입력/소요/결과를 구조화 로그로 남겨 병목을 "측정"한다.
        # (결론 확정 아님 — 실측 elapsed_ms 로 어느 단계가 느린지 데이터로 판단하기 위함)
        _call_counter["n"] += 1
        call_id = _call_counter["n"]
        eff_max = max_tokens if max_tokens is not None else vision_max_tokens
        payload_bytes, img_count, text_chars = _vision_payload_stats(content)
        instr_bytes = len((instructions or "").encode("utf-8", "ignore"))
        prompt_chars = text_chars + len(instructions or "")
        endpoint = "chat/completions" if vision_provider == "dsllm" else "responses"
        # vision 경로는 SDK 재시도를 끈 상태(dsllm). 실제 재시도가 없음을 로그로 명시한다.
        sdk_max_retries = "0" if vision_provider == "dsllm" else "sdk_default"
        ids_str = ",".join(str(i) for i in (image_ids or [])) or "-"
        _vlog.info(
            "[VisionReport] request_id=%s call=%d stage=%s batch_index=%s provider=%s model=%s "
            "endpoint=%s image_transport=data_url images=%d image_ids=%s image_identity=%s "
            "payload_bytes=%d prompt_chars=%d instr_bytes=%d max_tokens=%d "
            "timeout=%.0f sdk_max_retries=%s started",
            report_request_id, call_id, stage, batch_index, vision_provider, resolved_model,
            endpoint, img_count, ids_str, _format_image_metas(image_metas),
            payload_bytes, prompt_chars, instr_bytes, eff_max, vision_timeout, sdk_max_retries,
        )
        _t0 = time.perf_counter()
        try:
            if vision_provider == "dsllm":
                from app.llm import dsllm_adapter
                try:
                    res = dsllm_adapter.vision_chat(
                        model_name=resolved_model,
                        instructions=instructions,
                        content=content,
                        max_tokens=eff_max,
                        timeout=vision_timeout,
                    )
                except dsllm_adapter.DsllmVisionError as e:
                    elapsed_ms = int((time.perf_counter() - _t0) * 1000)
                    _vlog.warning(
                        "[VisionReport] request_id=%s call=%d stage=%s batch_index=%s result=error "
                        "kind=vision_error error_type=%s http_status=%s payload_unsupported=%s "
                        "elapsed_ms=%d",
                        report_request_id, call_id, stage, batch_index, e.error_type, e.status_code,
                        e.payload_unsupported, elapsed_ms,
                    )
                    _log_vision_diagnostic(
                        "ERROR", stage, e.error_type, provider="dsllm", model=resolved_model,
                        http_status=e.status_code, endpoint=_vision_endpoint,
                        login_id=_vision_login_id,
                        details=_vision_error_details(
                            batch_index=batch_index, image_ids=image_ids, image_metas=image_metas,
                            img_count=img_count, payload_bytes=payload_bytes, eff_max=eff_max,
                            elapsed_ms=elapsed_ms,
                        ),
                    )
                    raise HTTPException(status_code=e.status_code or 502, detail=e.user_message)
            else:
                try:
                    res = openai_adapter.vision_chat(
                        model_name=resolved_model,
                        instructions=instructions,
                        content=content,
                        max_tokens=eff_max,
                        timeout=vision_timeout,
                    )
                except openai_adapter.OpenAIAdapterError as e:
                    elapsed_ms = int((time.perf_counter() - _t0) * 1000)
                    _vlog.warning(
                        "[VisionReport] request_id=%s call=%d stage=%s batch_index=%s result=error "
                        "kind=openai_error error_type=%s http_status=%s elapsed_ms=%d",
                        report_request_id, call_id, stage, batch_index,
                        getattr(e, "error_type", None), getattr(e, "status_code", None), elapsed_ms,
                    )
                    _log_vision_diagnostic(
                        "ERROR", stage, _std_openai_error_type(e), provider="openai",
                        model=resolved_model, http_status=getattr(e, "status_code", None),
                        endpoint=_vision_endpoint, login_id=_vision_login_id,
                        details=_vision_error_details(
                            batch_index=batch_index, image_ids=image_ids, image_metas=image_metas,
                            img_count=img_count, payload_bytes=payload_bytes, eff_max=eff_max,
                            elapsed_ms=elapsed_ms,
                        ),
                    )
                    raise HTTPException(status_code=e.status_code or 502, detail=e.user_message)
        except HTTPException:
            raise
        except Exception:
            elapsed_ms = int((time.perf_counter() - _t0) * 1000)
            _vlog.exception(
                "[VisionReport] request_id=%s call=%d stage=%s batch_index=%s result=unexpected_error "
                "elapsed_ms=%d",
                report_request_id, call_id, stage, batch_index, elapsed_ms,
            )
            _log_vision_diagnostic(
                "ERROR", stage, "unknown_error", provider=vision_provider, model=resolved_model,
                endpoint=_vision_endpoint, login_id=_vision_login_id,
                details=_vision_error_details(
                    batch_index=batch_index, image_ids=image_ids, image_metas=image_metas,
                    img_count=img_count, payload_bytes=payload_bytes, eff_max=eff_max,
                    elapsed_ms=elapsed_ms,
                ),
            )
            raise
        elapsed_ms = int((time.perf_counter() - _t0) * 1000)
        usage = res.get("usage") or {}
        raw_text = res.get("raw_text") or ""
        _vlog.info(
            "[VisionReport] request_id=%s call=%d stage=%s batch_index=%s result=ok elapsed_ms=%d "
            "finish=%s truncated=%s output_tokens=%s response_chars=%d",
            report_request_id, call_id, stage, batch_index, elapsed_ms, res.get("status"),
            bool(res.get("truncated")), usage.get("output_tokens"), len(raw_text),
        )
        return res

    usage_total: dict = {}
    latency_total = 0.0
    stage1_vision_elapsed = 0.0   # Stage 1 이미지 Vision 호출 누적(캐시 hit 는 미포함)
    stage2_synthesis_elapsed = 0.0
    truncated = False
    max_output_tokens = None

    # ── Stage 1: 이미지 batch → Image Evidence 추출(내부 중간 결과) ──
    evidence_summaries: list = []
    stage1_raw_texts: list = []
    # batch 크기도 Vision AI 설정 우선(없으면 기존 env 기반 stage1_batch_size).
    # DSLLM(사내)은 1차 구현으로 이미지별 개별 호출(batch=1)을 강제한다(§5). 게이트웨이가
    # multi-image content 를 지원하지 않아도 안전하고, 실패 이미지 추적이 쉽다. openai 는 기존 배치.
    if vision_provider == "dsllm":
        stage1_batch = 1
    else:
        stage1_batch = vision_cfg.get("batch_size") or stage1_batch_size()
    batches = chunk_list(prepared, stage1_batch)
    # 설정 batch_size vs 실제 적용 batch 를 명시적으로 기록한다(§1·§2 확인용). dsllm 은
    # 현재 게이트웨이 multi-image 미검증으로 batch=1 을 강제하므로 configured 와 effective 가 다르다.
    _configured_batch = vision_cfg.get("batch_size") or stage1_batch_size()
    _all_image_ids = [str(pi.get("image_id")) for pi in prepared]
    _prep_orig = sum((pi.get("_prep_meta") or {}).get("original_bytes", 0) for pi in prepared)
    _prep_proc = sum((pi.get("_prep_meta") or {}).get("processed_bytes", 0) for pi in prepared)
    _prep_resized = sum(1 for pi in prepared if (pi.get("_prep_meta") or {}).get("resized"))
    _vlog.info(
        "[VisionReport] stage1 plan: provider=%s configured_batch_size=%s effective_batch=%d "
        "total_images=%d batches=%d preprocess_bytes=%d->%d resized=%d/%d image_ids=%s",
        vision_provider, _configured_batch, stage1_batch, len(prepared), len(batches),
        _prep_orig, _prep_proc, _prep_resized, len(prepared),
        ",".join(_all_image_ids) or "-",
    )
    for _batch_index, batch in enumerate(batches):
        s1_instr, s1_content = build_stage1_input(batch, input_mode=safe_input_mode)
        # Stage 1(이미지별 근거 추출)은 짧은 JSON 배열만 출력하므로 최종 합성(Stage 2)과 동일한
        # 8000 토큰 상한이 필요 없다. 상한만 낮추는 것이라 근거 JSON 은 이 한도로 잘리지 않으며,
        # 과도한 상한으로 인한 생성 지연/timeout 위험을 줄인다(이미지 수에 비례해 여유 배분).
        s1_max = min(vision_max_tokens, 600 + 500 * len(batch))
        _batch_ids = [str(b.get("image_id")) for b in batch]
        # 이미지 identity(전처리 후 mime/bytes/dims) — 어떤 이미지가 느리거나 큰지 로그로 식별(§2).
        _batch_metas = [
            {
                "image_id": b.get("image_id"),
                "mime": (b.get("_prep_meta") or {}).get("format"),
                "processed_bytes": (b.get("_prep_meta") or {}).get("processed_bytes"),
                "width": (b.get("_prep_meta") or {}).get("out_width")
                or (b.get("_prep_meta") or {}).get("width"),
                "height": (b.get("_prep_meta") or {}).get("out_height")
                or (b.get("_prep_meta") or {}).get("height"),
            }
            for b in batch
        ]
        started = time.perf_counter()
        s1 = _vision_call(
            s1_instr, s1_content, stage="stage1_vision", max_tokens=s1_max,
            image_ids=_batch_ids, batch_index=_batch_index, image_metas=_batch_metas,
        )
        _s1_dt = time.perf_counter() - started
        latency_total += _s1_dt
        stage1_vision_elapsed += _s1_dt
        usage_total = merge_usage(usage_total, s1.get("usage"))
        max_output_tokens = s1.get("max_output_tokens") or max_output_tokens
        s1_raw = s1.get("raw_text") or ""
        stage1_raw_texts.append(s1_raw)
        items = parse_stage1_evidence(s1_raw)
        # LLM 이 echo 한 task_id 는 신뢰하지 않고, backend 의 실제 image_id→task 매핑으로 교정한다.
        # (모델이 이미지는 잘 읽어도 Task 태깅이 틀려 visual coverage 0% 가 되던 근본 원인 수정)
        batch_meta = [
            {"image_id": b.get("image_id"), "task_id": b.get("task_id"), "task_title": b.get("task_title")}
            for b in batch
        ]
        items, recon_stats = reconcile_evidence_with_images(items, batch_meta)
        # evidence_type(직접/보조/참고/무관/판독불가) 정규화 — 과장 방지 위해 애매하면 reference.
        for it in items:
            it["evidence_type"] = normalize_evidence_type(it.get("evidence_type"))
        _vlog.info(
            "[VisionReport] stage1 evidence reconciled: batch_images=%d evidence=%d mapped=%d orphan=%d "
            "evidence_types=%s",
            len(batch), recon_stats["total"], recon_stats["mapped"], recon_stats["orphan"],
            ",".join(sorted({it.get("evidence_type") for it in items})) or "-",
        )
        evidence_summaries.extend(items)
        if bool(s1.get("truncated")) or (not items and looks_truncated_json(s1_raw)):
            truncated = True

    # ── 캐시 hit + 재생성(2b-2b) + fresh(vision miss) 를 manifest 순서로 병합(중복 없음) ──
    if cache_read_active and (cached_evidence or regenerated_evidence):
        _manifest_order = [it.get("image_id") for it in selected if it.get("image_id")]
        evidence_summaries = _obs.merge_evidence_by_manifest_order(
            _manifest_order, cached_evidence + regenerated_evidence, evidence_summaries
        )
        _vlog.info(
            "[VisionReport] cache read: cached=%d fresh=%d merged=%d exact_hits=%d "
            "obs_hits=%d obs_miss=%d ctx_hits=%d ctx_miss=%d vision_batches=%d lookup_ms=%d",
            len(cached_evidence), len(prepared), len(evidence_summaries),
            cache_stats["exact_match_hits"], cache_stats["observation_hits"],
            cache_stats["observation_misses"], cache_stats["context_link_hits"],
            cache_stats["context_link_misses"], len(batches), cache_stats["lookup_elapsed_ms"],
        )

    # ── deterministic base: 구조/사실 데이터는 DB로 서버가 확정(모델 미생성) ──
    report_base = _build_vision_report_base(db, state, project_id, ctx)

    # ── Stage 2(보강): base(사실) + Task별 이미지 evidence → 모델은 분석 포인트만 생성 ──
    s2_instr, s2_content = build_enhancement_input(
        report_base,
        ctx.text_context[:6000],  # 프롬프트 비대화 방지
        evidence_summaries,
    )
    started = time.perf_counter()
    s2 = _vision_call(s2_instr, s2_content, stage="stage2_synthesis", image_ids=[])
    _s2_dt = time.perf_counter() - started
    latency_total += _s2_dt
    stage2_synthesis_elapsed += _s2_dt
    usage_total = merge_usage(usage_total, s2.get("usage"))
    max_output_tokens = s2.get("max_output_tokens") or max_output_tokens

    stage2_raw_text = s2.get("raw_text") or ""
    enhancements = parse_enhancements(stage2_raw_text)
    # 보강 파싱은 실패해도 base(DB 구조/사실)로 보고서는 항상 렌더된다(분석 문장만 빔).
    enhancements_parse_ok = enhancements is not None
    if bool(s2.get("truncated")) or (not enhancements_parse_ok and looks_truncated_json(stage2_raw_text)):
        truncated = True

    # 이미지 사용 현황. 기본 UX 는 "프로젝트 전체 이미지"를 batch 로 모두 분석한다.
    # 캐시 hit + 재생성(2b-2b) + fresh(vision) 모두 이미지 근거로 쓰이므로 합산.
    images_from_cache = len(cached_evidence)
    images_context_regenerated = len(regenerated_evidence)
    images_vision_analyzed = len(prepared)
    images_analyzed = images_from_cache + images_context_regenerated + images_vision_analyzed
    images_total = sel_info.get("candidate_total", images_analyzed)
    images_excluded = sel_info.get("excluded_count", 0)
    images_failed = len(load_errors)
    batch_count = len(batches)
    _vlog.info(
        "[VisionReport] pipeline done project=%s provider=%s model=%s images_analyzed=%d "
        "stage1_batches=%d total_model_calls=%d latency_ms=%d truncated=%s",
        project_id, vision_provider, resolved_model, images_analyzed,
        batch_count, _call_counter["n"], int(latency_total * 1000), truncated,
    )
    excluded_reason = (
        f"후속 배치 분석 대상(이번 보고서는 안전 상한 {VISION_TEST_HARD_MAX_IMAGES}장을 우선 분석)"
        if images_excluded else None
    )

    # §6/Test E: ContextLink 재생성이 실패했지만 Vision fallback 으로 보고서가 정상 완료된 경우 =
    # 복구 이벤트(WARNING, resolved=true → 미처리 오류에 포함하지 않음). 보고서는 여기까지 왔으므로 성공.
    _regen_fallback = int(regen_stats.get("vision_fallback_images") or 0)
    if _regen_fallback > 0:
        _log_vision_diagnostic(
            "WARNING", "context_link_regeneration", "context_regeneration_failed",
            provider=vision_provider, model=resolved_model, endpoint=_vision_endpoint,
            login_id=_vision_login_id, recovery=True,
            message=(
                f"ContextLink 텍스트 재생성 {_regen_fallback}건 실패 후 Vision 분석으로 복구 "
                "(보고서 정상 완료)"
            ),
            details={"project_id": project_id, "image_count": _regen_fallback},
        )

    # ── 금지 문구 validator(§13): 이미지 분석했는데 "이미지 근거 없음" 류 문구 제거 ──
    enhancements, validation_notes = validate_enhancements(enhancements, images_analyzed)

    # ── 최종 합성: base(사실) + 모델 enhancement(분석) + 서버 featured → report ──
    manifest_by_id = {
        str(it.get("image_id")): it for it in selected if it.get("image_id")
    }
    # 인수인계 상세 보기는 image_count 만큼 이미지를 펼쳐야 하므로, 분석 대상(selected)뿐 아니라
    # 프로젝트 전체 이미지 manifest 도 함께 넘겨 Task별 연결 이미지 목록을 완성한다(스펙 §3·§4).
    full_manifest_by_id = {
        str(it.get("image_id")): it
        for it in (ctx.image_manifest or [])
        if it.get("image_id")
    }
    report = build_vision_enhanced_report(
        report_base, enhancements, evidence_summaries, manifest_by_id, full_manifest_by_id
    )
    # report 는 base 기반이라 항상 dict. report_parse_ok 는 "분석 보강 성공" 의미.
    report_parse_ok = enhancements_parse_ok

    # ── Shadow Write / Write-back: 기존 결과는 그대로 쓰고, 같은 정보를 신규 캐시 구조에도 저장한다.
    #    격리 세션 + best-effort — 저장 실패가 보고서 생성을 절대 깨지 않는다(추가 LLM 호출 없음).
    #    shadow_mode: 전량 write. cache_read: fresh(miss) 결과만 write-back(cached 는 reuse 로 무시).
    shadow_stats = None
    try:
        from app.config import settings as _obs_settings
        if (
            _obs_settings.vision_image_observation_enabled
            and (_obs_settings.vision_image_observation_shadow_mode or cache_read_active)
        ):
            from app.db_connections.sqlalchemy import SessionLocal
            from app.services.vision_image_observation import shadow_write_from_evidence
            image_meta_by_id = {
                str(pi.get("image_id")): {
                    "sha256": pi.get("sha256"),
                    "image_role": pi.get("image_role"),
                    "source_text_hash": pi.get("source_text_hash"),
                    "original_bytes": (pi.get("_prep_meta") or {}).get("original_bytes"),
                    "original_width": (pi.get("_prep_meta") or {}).get("width"),
                    "original_height": (pi.get("_prep_meta") or {}).get("height"),
                    "processed_bytes": (pi.get("_prep_meta") or {}).get("processed_bytes"),
                    "processed_width": (pi.get("_prep_meta") or {}).get("out_width"),
                    "processed_height": (pi.get("_prep_meta") or {}).get("out_height"),
                }
                for pi in prepared if pi.get("image_id")
            }
            _shadow_db = SessionLocal()
            try:
                shadow_stats = shadow_write_from_evidence(
                    _shadow_db,
                    evidence_items=evidence_summaries,
                    image_meta_by_id=image_meta_by_id,
                    provider=vision_provider,
                    model_name=resolved_model,
                    project_id=project_id,
                    context_link_enabled=_obs_settings.vision_image_context_link_enabled,
                )
            finally:
                _shadow_db.close()
            _vlog.info(
                "[VisionReport] shadow cache write: images_attempted=%d obs_inserted=%d obs_reused=%d "
                "ctx_inserted=%d ctx_reused=%d obs_failures=%d ctx_failures=%d shadow_mapped=%d "
                "first_error=%s/%s elapsed_ms=%d",
                shadow_stats["images_attempted"], shadow_stats["observations_inserted"],
                shadow_stats["observations_reused"], shadow_stats["context_links_inserted"],
                shadow_stats["context_links_reused"], shadow_stats["observation_failures"],
                shadow_stats["context_link_failures"], shadow_stats["shadow_mapped"],
                shadow_stats["first_error_stage"], shadow_stats["first_error_type"],
                shadow_stats["shadow_write_elapsed_ms"],
            )
    except Exception:
        _vlog.exception("[VisionReport] shadow cache write failed (무시하고 보고서는 정상 반환)")

    uncertainty_notes: list = []
    if images_excluded:
        uncertainty_notes.append(
            f"전체 이미지 {images_total}장 중 {images_analyzed}장을 우선 분석했습니다. "
            f"{images_excluded}장은 후속 배치 분석 대상입니다."
        )
    if images_failed:
        uncertainty_notes.append(f"{images_failed}개 이미지는 파일을 불러오지 못했습니다.")
    if truncated:
        uncertainty_notes.append("응답이 출력 길이 제한으로 일부 잘렸을 가능성이 있습니다.")
    if not enhancements_parse_ok:
        uncertainty_notes.append("AI 분석 보강에 실패해 구조/사실 데이터만 표시됩니다(분석 문장 비어 있음).")
    uncertainty_notes.extend(validation_notes)
    report["vision_poc_meta"] = {
        "model": resolved_model,
        "images_total": images_total,
        "images_analyzed": images_analyzed,
        "images_used": images_analyzed,  # 하위호환(기존 FE 필드)
        "images_failed": images_failed,
        "images_excluded": images_excluded,
        "batches": batch_count,
        "uncertainty_notes": uncertainty_notes,
        "validation_notes": validation_notes,
        "save_result": False,
        # 2b-1 shadow write 통계(개발자용). None 이면 shadow mode 비활성.
        "image_cache_shadow": shadow_stats,
        # ── 2b-2a 캐시 read-through 지표(새 source of truth) ──
        "images_from_cache": images_from_cache,
        "images_vision_analyzed": images_vision_analyzed,
        "raw_images_sent": images_vision_analyzed,
        "vision_model_calls": batch_count,   # Stage 1 이미지 Vision 호출 수(=batches). all-hit 이면 0.
        "vision_batches": batch_count,
        "cache_read_active": cache_read_active,
        # 2b-2b: Observation 은 캐시 재사용하고 ContextLink 만 Text 로 재생성한 이미지 수 + 통계.
        "images_context_regenerated": images_context_regenerated,
        "context_link_regeneration": regen_stats if regen_enabled else None,
        # 단계별 시간 분리(§7/§12) — all-hit인데 total이 크면 병목은 캐시가 아니라 Stage 2.
        "timings": {
            "cache_lookup_elapsed_ms": cache_stats["lookup_elapsed_ms"],
            "context_link_regeneration_elapsed_ms": _regen_elapsed_ms,
            "stage1_vision_elapsed_ms": int(stage1_vision_elapsed * 1000),
            "stage2_synthesis_elapsed_ms": int(stage2_synthesis_elapsed * 1000),
            "total_latency_ms": int((time.perf_counter() - _req_t0) * 1000),
        },
        "image_cache": cache_stats if cache_read_active else None,
        "image_cache_quality": ({
            "observations_checked": cache_stats["observation_hits"],
            "usable": cache_stats["observation_usable"],
            "usable_with_warning": cache_stats["observation_warning_hits"],
            "raw_review_required": cache_stats["observation_raw_review_required"],
            "invalid": cache_stats["observation_invalid"],
            "context_links_checked": cache_stats["context_link_hits"] + cache_stats["context_link_misses"],
            "context_links_usable": cache_stats["context_link_hits"],
            "context_links_stale": cache_stats["context_link_stale"],
            "context_links_missing": cache_stats["context_link_missing"],
            "context_links_invalid": cache_stats["context_link_invalid"],
        } if cache_read_active else None),
    }

    if report_parse_ok:
        parse_message = None
    elif truncated:
        parse_message = (
            "AI 분석 보강(JSON) 파싱 실패: 응답이 중간에 잘렸을 가능성이 높습니다. "
            "구조/사실 데이터는 표시되며, 분석 문장만 비어 있습니다. "
            "max_output_tokens를 늘리거나(VISION_REPORT_MAX_OUTPUT_TOKENS) 이미지 수를 줄여 재시도하세요."
        )
    else:
        parse_message = (
            "AI 분석 보강(JSON) 파싱 실패: 모델이 JSON 형식으로 응답하지 않았습니다. "
            "구조/사실 데이터는 표시되며, 개발자용 원본(stage2)에서 raw_text를 확인하세요."
        )

    return {
        "project_id": project_id,
        "model": resolved_model,
        "provider": vision_provider,
        "input_mode": safe_input_mode,
        "image_scope": scope,
        "schema_version": FINAL_REPORT_SCHEMA_VERSION,
        # ── 보고서 생성 방식 메타(§8) — vision 성공 경로 ──
        "report_mode": "vision",
        "primary_vision_model": resolved_model,
        "actual_model_used": resolved_model,
        "fallback_used": False,
        "image_analysis_included": True,
        # 최종 보고서(기존 AI Report 호환 schema) — 메인 화면 렌더 대상
        "report": report,
        "report_parse_ok": report_parse_ok,
        "parse_message": parse_message,
        # Stage 1 중간 결과(이미지 evidence) — 개발자용 원본 영역에만 노출
        "image_evidence_summaries": evidence_summaries,
        "evidence_count": len(evidence_summaries),
        # 이미지 사용 현황(meta)
        "used_image_ids": [pi["image_id"] for pi in prepared],
        "missing_image_ids": missing_ids,
        "load_errors": load_errors,
        "image_count": images_analyzed,
        "images_total": images_total,
        "images_used": images_analyzed,
        "images_analyzed": images_analyzed,
        "images_failed": images_failed,
        "images_excluded": images_excluded,
        "batches": batch_count,
        # ── 2b-2a 캐시 지표(top-level, FE 관리자 디버그 표시용) ──
        "images_from_cache": images_from_cache,
        "images_context_regenerated": images_context_regenerated,
        "images_vision_analyzed": images_vision_analyzed,
        "vision_fallback_images": images_vision_analyzed,  # §1: cache 이후 Vision 전달 고유 이미지 수
        "vision_model_calls": batch_count,
        "vision_batches": batch_count,
        "cache_read_active": cache_read_active,
        "candidate_total": images_total,
        "excluded_count": images_excluded,
        "excluded_reason": excluded_reason,
        # 서버 확정 대표 이미지(보고서 파싱 실패 시 FE fallback 렌더용 top-level 노출)
        "featured_visual_evidence": (report.get("featured_visual_evidence") if isinstance(report, dict) else []),
        "max_images_applied": safe_max_images,
        "max_images_limit": VISION_TEST_HARD_MAX_IMAGES,
        # 진단/원본(개발자용)
        "stage1_batches": batch_count,
        "stage1_raw_texts": stage1_raw_texts,
        "stage2_raw_text": stage2_raw_text,
        "latency_ms": round(latency_total * 1000),
        "usage": usage_total or None,
        "truncated": truncated,
        "max_output_tokens": max_output_tokens,
        # PoC: 저장/반영 안 함을 명시
        "save_result": False,
        "not_persisted": True,
    }


class VisionTextFallbackRequest(BaseModel):
    """Gemma4 Vision 지연/실패 시 사용자가 선택하는 **텍스트 전용 fallback** 보고서 요청.

    ⚠️ fallback 은 **일회성 실행 옵션**이다 — 이 요청은 Vision/Text AI 설정 저장값을 바꾸지 않는다.
    provider/model 을 비워두면 Vision AI 설정의 fallback_text_provider/model 을 사용한다.
    """
    user_id: int
    provider: Optional[str] = None   # 미지정 시 vision_cfg 의 fallback_text_provider
    model: Optional[str] = None      # 미지정 시 vision_cfg 의 fallback_text_model
    source: str = "vision_timeout_user_selected"
    user_choice: str = "text_only_fallback"


@app.post("/api/projects/{project_id}/ai-report/text-fallback")
def run_vision_text_fallback_report(
    project_id: int,
    body: VisionTextFallbackRequest,
    db: Session = Depends(get_db),
):
    """텍스트 전용 fallback 보고서(이미지 분석 제외) — Gemma4 Vision 지연 시 사용자 선택.

    Vision Report 와 **동일한 deterministic base(구조/사실)** 를 쓰되, Stage 1(이미지 evidence)
    를 건너뛰고 fallback Text 모델(DS/GPT-OSS 등)로 Stage 2(분석 포인트)만 생성한다.

    안전 정책:
    - super_admin + project 접근 권한.
    - **설정 변경 없음**: fallback 실행은 이번 요청에만 적용, Vision/Text AI 설정 저장값 불변.
    - 이미지 파일 로딩/모델 이미지 입력 없음(텍스트 전용). image_analysis_included=false.
    - 실패해도 기존 report/ai-query/Vision 경로에 영향 없음.
    """
    require_super_admin(db, body.user_id)
    state = load_state()
    check_project_access(db, state, project_id, body.user_id)

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    from app.services.project_ai_context_builder import build_project_ai_context
    from app.services.project_ai_vision_report_test import (
        build_enhancement_input,
        parse_enhancements,
        validate_enhancements,
        build_vision_enhanced_report,
        looks_truncated_json,
        FINAL_REPORT_SCHEMA_VERSION,
    )
    from app.services.vision_ai_settings import resolve_vision_ai_config

    vision_cfg = resolve_vision_ai_config(db)
    if not vision_cfg["enabled"]:
        raise HTTPException(
            status_code=400,
            detail="Vision AI가 비활성화되어 있습니다. AI Settings의 Vision AI 설정에서 활성화해주세요.",
        )
    if not vision_cfg.get("fallback_text_enabled", True):
        raise HTTPException(
            status_code=400,
            detail="텍스트 전용 fallback이 비활성화되어 있습니다. Vision AI 설정에서 Fallback Text Model을 활성화해주세요.",
        )

    # 일회성 실행 값: 요청값 우선, 없으면 Vision AI 설정의 fallback 값. (저장하지 않는다)
    primary_vision_model = vision_cfg["model"]
    text_provider = (body.provider or vision_cfg.get("fallback_text_provider") or "dsllm").strip().lower()
    if text_provider not in ("openai", "dsllm"):
        text_provider = "dsllm"
    text_model = (body.model or vision_cfg.get("fallback_text_model") or "").strip()
    if not text_model:
        raise HTTPException(
            status_code=400,
            detail=(
                "Fallback Text 모델이 설정되어 있지 않습니다. Vision AI 설정에서 Fallback Text Model"
                "(예: DS/GPT-OSS)을 지정하거나 VISION_AI_FALLBACK_TEXT_MODEL env를 설정해주세요."
            ),
        )
    if text_provider == "openai" and not llm_router.openai_enabled():
        raise HTTPException(
            status_code=400,
            detail="OpenAI API Key가 설정되어 있지 않아 openai fallback을 사용할 수 없습니다.",
        )

    import time
    started = time.perf_counter()

    # 텍스트 전용: 이미지 evidence(Stage 1) 없음. base 의 image_count 만 정확히 채우기 위해
    # manifest 는 만들되(모델 호출 아님), 이미지 파일 로딩/이미지 입력은 하지 않는다.
    ctx = build_project_ai_context(
        db, project_id, state=state, include_images=True, max_images=AI_CONTEXT_IMAGES_MAX_LIMIT,
    )
    report_base = _build_vision_report_base(db, state, project_id, ctx)

    # Stage 2(분석 포인트)만 — evidence 빈 리스트로 텍스트 전용 합성.
    s2_instr, s2_content = build_enhancement_input(report_base, ctx.text_context[:6000], [])
    user_text = "\n".join(
        c.get("text", "") for c in s2_content if isinstance(c, dict) and c.get("type") == "input_text"
    )

    try:
        raw_text = llm_router.chat(
            provider=text_provider,
            base_url=None,
            model_name=text_model,
            system_prompt=s2_instr,
            user_prompt=user_text,
            temperature=0.3,
            max_tokens=vision_cfg["max_output_tokens"],
        )
    except Exception as e:
        raise HTTPException(
            status_code=502,
            detail=f"텍스트 전용 fallback 보고서 생성 실패 (provider={text_provider}, model={text_model}): {e}",
        )
    latency_total = time.perf_counter() - started

    enhancements = parse_enhancements(raw_text)
    enhancements_parse_ok = enhancements is not None
    truncated = (not enhancements_parse_ok) and looks_truncated_json(raw_text)
    # 이미지 분석이 없으므로 images_analyzed=0 → 금지문구 validator 는 no-op(그대로 통과).
    enhancements, validation_notes = validate_enhancements(enhancements, 0)

    # 이미지 evidence/manifest 없이 합성 → featured/Task 이미지 없음(텍스트 전용).
    report = build_vision_enhanced_report(report_base, enhancements, [], {}, {})

    uncertainty_notes: list = [
        "텍스트 전용 보고서입니다. 캡처 이미지 내용은 분석에 포함되지 않았습니다.",
    ]
    if not enhancements_parse_ok:
        uncertainty_notes.append("AI 분석 보강에 실패해 구조/사실 데이터만 표시됩니다(분석 문장 비어 있음).")
    uncertainty_notes.extend(validation_notes)

    report["vision_poc_meta"] = {
        "model": text_model,
        "report_mode": "text_only_fallback",
        "images_total": 0,
        "images_analyzed": 0,
        "images_used": 0,
        "images_failed": 0,
        "images_excluded": 0,
        "batches": 0,
        "uncertainty_notes": uncertainty_notes,
        "validation_notes": validation_notes,
        "save_result": False,
    }

    parse_message = None
    if not enhancements_parse_ok:
        parse_message = (
            "AI 분석 보강(JSON) 파싱 실패: 구조/사실 데이터는 표시되며, 분석 문장만 비어 있습니다."
            + ("응답이 중간에 잘렸을 수 있습니다." if truncated else "")
        )

    return {
        "project_id": project_id,
        "model": text_model,
        "provider": text_provider,
        "input_mode": "text_only",
        "image_scope": "none",
        "schema_version": FINAL_REPORT_SCHEMA_VERSION,
        # ── 보고서 생성 방식 메타(§8) — 텍스트 전용 fallback 경로 ──
        "report_mode": "text_only_fallback",
        "primary_vision_model": primary_vision_model,
        "actual_model_used": text_model,
        "fallback_used": True,
        "fallback_reason": body.source or "vision_timeout_user_selected",
        "image_analysis_included": False,
        "user_choice": body.user_choice or "text_only_fallback",
        "report": report,
        "report_parse_ok": enhancements_parse_ok,
        "parse_message": parse_message,
        # 이미지 관련 필드는 모두 0(텍스트 전용) — FE 호환용.
        "image_evidence_summaries": [],
        "evidence_count": 0,
        "used_image_ids": [],
        "missing_image_ids": [],
        "load_errors": [],
        "image_count": 0,
        "images_total": 0,
        "images_used": 0,
        "images_analyzed": 0,
        "images_failed": 0,
        "images_excluded": 0,
        "batches": 0,
        "candidate_total": 0,
        "excluded_count": 0,
        "featured_visual_evidence": [],
        "max_images_applied": 0,
        "max_images_limit": 0,
        "stage2_raw_text": raw_text,
        "latency_ms": round(latency_total * 1000),
        "truncated": truncated,
        "max_output_tokens": vision_cfg["max_output_tokens"],
        "save_result": False,
        "not_persisted": True,
    }


@app.post("/api/report/generate")
def generate_report(body: ReportRequest, db: Session = Depends(get_db)):
    # ✅ dsllm_chat 내부에서 requests를 쓰므로, 예외 매핑용으로만 import
    import requests

    # ✅ sidecar 데이터(sub_projects, notes 등)는 기존처럼 유지
    state = load_state()

    # ✅ DB에서 provider/모델 선택값만 읽기 (BASE_URL/API_KEY는 env 전용)
    row = get_or_create_ai_setting(db)
    provider = _active_provider(row)
    selected_model = _normalize_selected_model(row.model_name)
    model_name = llm_router.resolve_model(provider, selected_model)

    import logging as _logging
    _logging.getLogger("main").info(
        "AI Report request: provider=%s, selected_model=%r, resolved_model=%r",
        provider, selected_model, model_name,
    )

    if not model_name:
        raise HTTPException(400, "AI settings not configured. Please choose a model in AI Settings.")

    # ✅ 프로젝트 로드
    p = db.query(Project).filter(Project.id == body.project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    project = project_dict(p, state)

    # ✅ 데이터 수집 (기존 그대로)
    task_rows = db.query(Task).filter(Task.project_id == body.project_id, Task.archived_at.is_(None)).all()
    tasks = [task_dict(t, state) for t in task_rows]
    wf_map = _workflow_label_map(db, body.project_id)
    workflow_mode = getattr(p, "workflow_mode", "DEFAULT") or "DEFAULT"

    sub_projects = get_subprojects_from_db(db, body.project_id)
    all_attachments = state.get("attachments", [])
    # DB 기준 멤버 (viewer 제외)
    db_members = db.query(ProjectMemberModel).filter(
        ProjectMemberModel.project_id == body.project_id,
        ProjectMemberModel.role != 'viewer'
    ).all()
    members = [{"user_id": m.user_id, "role": m.role, "project_id": m.project_id} for m in db_members]
    notes = [n for n in state.get("notes", []) if int(n.get("project_id")) == body.project_id]
    project_files = [f for f in state.get("project_files", []) if int(f.get("project_id")) == body.project_id]

    users_map = {u.id: user_dict(u, state) for u in db.query(User).all()}

    # ✅ 진행률 계산 (Hold 제외)
    active_tasks = [t for t in tasks if t.get("status") != "hold"]
    hold_tasks = [t for t in tasks if t.get("status") == "hold"]
    done_tasks = [t for t in tasks if t.get("status") == "done"]
    in_progress_tasks = [t for t in tasks if t.get("status") == "in_progress"]
    todo_tasks = [t for t in tasks if t.get("status") == "todo"]

    if len(active_tasks) > 0:
        progress_sum = sum(
            100 if t.get("status") == "done" else (t.get("progress", 0) or 0)
            for t in active_tasks
        )
        overall_progress = round(progress_sum / len(active_tasks), 1)
    else:
        overall_progress = 0.0

    # ✅ task details (작업노트 포함)
    task_details = []
    for t in tasks:
        assignees = [users_map.get(a, {}).get("username", f"User {a}") for a in (t.get("assignee_ids") or [])]
        sp_name = ""
        if t.get("sub_project_id"):
            sp = next((s for s in sub_projects if s["id"] == t["sub_project_id"]), None)
            sp_name = sp["name"] if sp else ""

        task_attachments = [a for a in all_attachments if int(a.get("task_id")) == t["id"]]

        # 작업노트(activity) 요약 정보
        activities = db.query(TaskActivityModel).filter(TaskActivityModel.task_id == t["id"]).order_by(TaskActivityModel.order_index).all()
        checkbox_activities = [a for a in activities if (a.block_type or "checkbox") == "checkbox"]
        checked_count = sum(1 for a in checkbox_activities if a.checked)
        total_checkboxes = len(checkbox_activities)
        activity_progress = round(checked_count / total_checkboxes * 100) if total_checkboxes > 0 else 0
        activity_summary = f"{total_checkboxes}개 항목 중 {checked_count}개 완료 ({activity_progress}%)" if total_checkboxes > 0 else ""

        _wf = wf_map.get(t.get("workflow_column_id"))
        task_details.append({
            "id": t["id"],
            "title": t.get("title", ""),
            "description": t.get("description", ""),
            "status": t.get("status", "todo"),
            "priority": t.get("priority", "medium"),
            "progress": t.get("progress", 0) or 0,
            "start_date": t.get("start_date"),
            "due_date": t.get("due_date"),
            "start_date_tbd": t.get("start_date_tbd", False),
            "due_date_tbd": t.get("due_date_tbd", False),
            "assignees": assignees,
            "sub_project": sp_name,
            "tags": t.get("tags", []),
            # CUSTOM 워크플로우 — AI가 사용자 단계(workflow_label) 기준으로 설명하도록 제공
            "workflow_column_id": t.get("workflow_column_id"),
            "workflow_label": _wf.label if _wf else None,
            "note_progress": f"{checked_count}/{total_checkboxes}" if total_checkboxes > 0 else None,
            "attachments": [
                {
                    "id": a["id"],
                    "filename": a.get("filename", ""),
                    "url": a.get("url", ""),
                    "type": a.get("type", "url"),
                }
                for a in task_attachments
            ],
            "activity_summary": activity_summary,
        })

    # ✅ 멤버명 (viewer 제외, 담당자만)
    member_names = []
    for m in members:
        role = m.get("role", "member")
        if role == "viewer":
            continue
        uid = int(m.get("user_id"))
        u = users_map.get(uid)
        if u:
            member_names.append(f'{u["username"]} ({role})')

    status_breakdown = {
        "total": len(tasks),
        "active": len(active_tasks),
        "done": len(done_tasks),
        "in_progress": len(in_progress_tasks),
        "todo": len(todo_tasks),
        "hold": len(hold_tasks),
        "overall_progress": overall_progress,
    }

    # 일정 미정(TBD) 집계 — Dashboard/AI 요약용
    _sched_start_tbd = sum(1 for t in tasks if t.get("start_date_tbd"))
    _sched_due_tbd = sum(1 for t in tasks if t.get("due_date_tbd"))
    _sched_unset = sum(1 for t in tasks if t.get("start_date_tbd") or t.get("due_date_tbd"))
    _sched_confirmed = sum(
        1 for t in tasks
        if not t.get("start_date_tbd") and not t.get("due_date_tbd")
        and t.get("start_date") and t.get("due_date")
    )
    schedule_breakdown = {
        "confirmed": _sched_confirmed,
        "tbd": _sched_unset,
        "start_tbd": _sched_start_tbd,
        "due_tbd": _sched_due_tbd,
    }

    structured = {
        "project": {
            "name": project.get("name", ""),
            "description": project.get("description", ""),
            "created_at": project.get("created_at", ""),
            "workflow_mode": workflow_mode,
        },
        "status_breakdown": status_breakdown,
        "schedule_breakdown": schedule_breakdown,
        "tasks": task_details,
        "sub_projects": [{"name": sp.get("name", ""), "description": sp.get("description", "")} for sp in sub_projects],
        "members": member_names,
        "project_files": [
            {"id": pf["id"], "filename": pf.get("filename", ""), "size": pf.get("size", 0), "created_at": pf.get("created_at", "")}
            for pf in project_files
        ],
    }

    # ✅ Prompt (기존 그대로)
    task_lines = []
    for t in task_details:
        att_info = ""
        if t["attachments"]:
            att_names = ", ".join([a["filename"] or a["url"] for a in t["attachments"]])
            att_info = f" | 첨부파일: {att_names}"
        note_info = f' | 작업노트: {t["activity_summary"]}' if t.get("activity_summary") else ""
        _start_label = "미정(TBD)" if t.get("start_date_tbd") else (t.get("start_date") or "없음")
        _due_label = "미정(TBD)" if t.get("due_date_tbd") else (t.get("due_date") or "없음")
        task_lines.append(
            f'- {t["title"]} | 상태: {t["status"]} | 우선순위: {t["priority"]} '
            f'| 진행률: {t["progress"]}% | 시작일: {_start_label} | 마감일: {_due_label} '
            f'| 담당자: {", ".join(t["assignees"]) if t["assignees"] else "미배정"}'
            f'| 설명: {t["description"] or "없음"}{att_info}{note_info}'
        )

    prompt = f"""당신은 전문 프로젝트 매니저 보조 AI입니다. 아래 프로젝트 데이터를 분석하여 종합 보고서를 작성해주세요.

## 프로젝트 정보
- 이름: {project["name"]}
- 설명: {project.get("description", "없음")}
- 생성일: {project.get("created_at", "N/A")}
- 팀원: {", ".join(member_names) if member_names else "미배정"}

## 진행 현황
- 전체 Task 수: {len(tasks)}개
- 활성 Task(Hold 제외): {len(active_tasks)}개
- 완료: {len(done_tasks)}개 | 진행 중: {len(in_progress_tasks)}개 | 대기: {len(todo_tasks)}개 | 보류: {len(hold_tasks)}개
- 전체 진행률(Hold 제외): {overall_progress}%

## 일정(스케줄) 현황
- 일정 확정(시작·마감 모두 확정): {schedule_breakdown["confirmed"]}개
- 일정 미정 Task(시작 또는 마감이 미정): {schedule_breakdown["tbd"]}개
- 이 중 시작일(Start) 미정: {schedule_breakdown["start_tbd"]}개 (마감 역산으로 착수일 설정 필요)
- 이 중 마감일(Due) 미정: {schedule_breakdown["due_tbd"]}개 (우선순위·마감 판단이 어려움)

## Task 상세 목록
{chr(10).join(task_lines) if task_lines else "Task 없음"}

## 서브프로젝트
{chr(10).join([f'- {sp["name"]}: {sp.get("description", "")}' for sp in sub_projects]) if sub_projects else "없음"}

## 프로젝트 첨부파일
{chr(10).join([f'- {pf.get("filename", "")} (크기: {round(pf.get("size", 0)/1024, 1)}KB, 업로드일: {pf.get("created_at", "N/A")})' for pf in project_files]) if project_files else "첨부파일 없음"}

---
아래 4개 섹션으로 나눠서 분석 보고서를 작성해주세요. 마크다운 문법(#, **, -, ```)을 사용하지 마세요. 일반 텍스트로만 작성하세요.

[작성 스타일 가이드라인 - 매우 중요]
단순히 "1. 제목, 1) 내용" 같은 나열식(List) 구조로 작성하지 마세요.
보고서를 '이야기(서술)' 형식으로 작성하세요.
각 Task의 체크박스 항목(완료/미완료)과 텍스트 박스 내용(작업노트, 메모)은 서로 밀접하게 연관된 정보입니다.
이들을 분리하지 말고, 인과관계와 흐름이 보이도록 문장형으로 연결하세요.
예시: "A 작업을 수행하면서 B라는 특이사항이 발견되어 C 방식으로 처리하였다."
정돈되지 않은 작업 내용이라도, 프로젝트의 '과제(Task)'와 '성과' 관점으로 재해석하여 전문 보고서 느낌이 나도록 다듬으세요.
소제목과 설명 문단을 적절히 섞어서 가독성을 높이세요.

[섹션1: 프로젝트 개요]
프로젝트의 전체 목적을 한 문장으로 요약하고, 현재 전체 진행률과 상태를 서술하세요.
단순 수치 나열이 아니라, 프로젝트가 어느 단계에 와 있고 어떤 맥락인지 한 문단으로 설명하세요.

[섹션2: Task별 분석]
각 Task에 대해 현재 상태, 진행률, 그리고 현재까지 어떤 단계까지 진행되었는지를 설명하세요.
작업노트의 체크박스 항목과 텍스트 메모를 맥락적으로 통합하여, 해당 Task에서 무엇이 수행되었고 어떤 결과가 나왔는지를 하나의 흐름으로 서술하세요.
첨부 자료가 있는 경우, 해당 자료가 Task 진행에서 어떤 역할을 하는지도 설명하세요.

반드시 아래 형식 규칙을 지켜서 작성하세요:
- 각 Task의 첫 줄은 반드시 [Task: Task제목] 만 단독 줄로 출력하세요. 같은 줄에 다른 텍스트를 이어서 쓰지 마세요.
- 각 Task 블록 사이에는 반드시 빈 줄 1줄을 넣으세요.
- 하나의 Task 분석이 끝나면 줄바꿈 2번 후 다음 [Task: ...]를 시작하세요.

올바른 예시:
[Task: Data 확보]
이 작업은 현재 진행 중이며...
작업노트에 따르면...

[Task: 회의 대응]
이 작업은 1차 협업 회의를 통해...

잘못된 예시 (절대 이렇게 쓰지 마세요):
[Task: Data 확보] 이 작업은 현재 진행 중이며...
[Task: 회의 대응] 다음 작업은...

순서는 반드시 진행 중(in_progress) -> 대기(todo) -> 보류(hold) -> 완료(done) 순으로 작성하세요.
완료된 Task는 맨 마지막에 간략하게만 작성하세요.

[섹션3: 종합 현황 분석]
현재 프로젝트가 어떤 단계에 있는지를 서술형으로 정리하세요.
핵심 진행 작업, 완료 작업, 지연/보류 작업을 맥락적으로 연결하여 프로젝트 전체 흐름이 보이도록 작성하세요.
프로젝트에 첨부파일이 있으면, 어떤 파일이 포함되어 있는지 간략히 안내하세요.

[섹션4: 다음 단계 추천]
다음으로 가장 중요한 작업이 무엇인지, 어떤 순서로 진행하면 좋을지 제안하세요.
현재 진행 상황과 연계하여 왜 그 순서가 적절한지 맥락을 설명하세요.
또한 위 '일정 현황'을 참고하여, 일정이 미정(TBD)인 Task가 있으면 다음 관점으로 짚어주세요:
- 일정 미정 Task가 몇 개인지, 그로 인해 프로젝트 관리에 어떤 리스크(우선순위 판단·마감 관리 어려움)가 있는지 서술하세요.
- 마감일(Due)은 정해졌으나 시작일(Start)이 미정인 Task는, 마감 역산 기준으로 착수일을 언제쯤 잡으면 좋을지 구체적으로 제안하세요.
- 시작일은 있으나 마감일(Due)이 미정인 Task는, 우선순위와 진행률을 고려해 목표 마감일을 제안하세요.
- 일정이 아예 확정되지 않은(둘 다 미정) Task는, 일정 확정 자체를 우선 조치로 권고하세요.
일정 미정 Task가 없으면 이 부분은 생략하세요.

각 섹션은 [섹션1], [섹션2] 등의 태그로 시작해주세요. 반드시 한국어로 작성하세요."""

    # ✅ LLM 호출 — provider(dsllm/openai) 분기. BASE_URL/API_KEY는 env 전용, 모델만 선택값.
    try:
        content = llm_router.chat(
            provider=provider,
            base_url=None,
            model_name=model_name,
            system_prompt=(
                "당신은 전문 프로젝트 매니저 보조 AI입니다. 반드시 한국어로 작성하세요. "
                "마크다운 문법(#, **, ```, 표)을 사용하지 마세요. "
                "단순 나열식 리스트가 아닌, 맥락이 연결된 서술형 보고서를 작성하세요. "
                "체크박스 항목과 작업노트 내용은 분리하지 말고, 인과관계와 흐름이 보이도록 문장형으로 연결하세요. "
                "정돈되지 않은 작업 내용이라도, 프로젝트의 과제와 성과 관점으로 재해석하여 전문 보고서 느낌이 나도록 다듬으세요."
            ),
            user_prompt=prompt,
            temperature=0.3,
            max_tokens=4096,
        )

        content = sanitize_llm_text(content)

        # ✅ 섹션 파싱 (기존 로직 유지)
        sections = {"overview": "", "task_analysis": "", "status_analysis": "", "next_steps": ""}
        current_section = ""
        for line in content.split("\n"):
            stripped = line.strip()
            if "[섹션1" in stripped:
                current_section = "overview"
                continue
            elif "[섹션2" in stripped:
                current_section = "task_analysis"
                continue
            elif "[섹션3" in stripped:
                current_section = "status_analysis"
                continue
            elif "[섹션4" in stripped:
                current_section = "next_steps"
                continue

            if current_section:
                sections[current_section] += line + "\n"

        if not any(sections.values()):
            sections["overview"] = content

        # ✅ Task 블록 정규화 (task_analysis만)
        if sections.get("task_analysis"):
            sections["task_analysis"] = normalize_task_blocks(sections["task_analysis"])

        # ✅ DB에 저장
        db_report = ProjectAiReport(
            project_id=body.project_id,
            overview=sections.get("overview", ""),
            task_analysis=sections.get("task_analysis", ""),
            status_analysis=sections.get("status_analysis", ""),
            next_steps=sections.get("next_steps", ""),
            raw_response=content,
            structured_snapshot=structured,
            model=model_name,
            created_at=datetime.utcnow(),
        )
        db.add(db_report)
        db.commit()
        db.refresh(db_report)

        return {
            "project_id": body.project_id,
            "report": content,
            "sections": sections,
            "structured": structured,
            "model": model_name,
            "updated_at": db_report.created_at.isoformat() if db_report.created_at else None,
        }

    except requests.exceptions.ConnectionError:
        # requests 기반 예외는 사실상 DSLLM(사내) 경로에서만 발생한다.
        _log_dsllm_failure("AI Report 생성 실패: DSLLM 연결 불가", status_code=502, endpoint="/api/report/generate")
        raise HTTPException(status_code=502, detail="Cannot connect to DSLLM. Please check BASE_URL in .env.local/.env.production.")
    except requests.exceptions.Timeout:
        _log_dsllm_failure("AI Report 생성 실패: DSLLM timeout", status_code=504, endpoint="/api/report/generate")
        raise HTTPException(status_code=504, detail="DSLLM request timed out.")
    except requests.exceptions.HTTPError as e:
        # DSLLM 응답이 4xx/5xx로 온 경우
        _log_dsllm_failure("AI Report 생성 실패: DSLLM HTTP error", detail=str(e)[:500], status_code=502, endpoint="/api/report/generate")
        raise HTTPException(status_code=502, detail=f"DSLLM returned HTTP error: {str(e)}")
    except Exception as e:
        # 예: ENV 누락(RuntimeError), OpenAIAdapterError, 파싱 실패 등 — active provider 기준으로 보고.
        #   (Text AI provider=openai 인데 "DSLLM" 이라고만 로깅하면 운영자가 라우팅을 오해한다.)
        _prov_label = "OpenAI" if provider == "openai" else "DSLLM"
        _log_dsllm_failure(
            f"AI Report 생성 실패({_prov_label}): {type(e).__name__}",
            detail=str(e)[:500], status_code=500, endpoint="/api/report/generate",
        )
        raise HTTPException(
            status_code=500,
            detail=f"Report generation failed (Text AI provider={provider}): {str(e)}",
        )

@app.get("/api/report/data/{project_id}")
def get_report_data(
    project_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    state = load_state()
    check_project_access(db, state, project_id, user_id)

    # DB에서 최신 보고서 조회
    row = (
        db.query(ProjectAiReport)
        .filter(ProjectAiReport.project_id == project_id)
        .order_by(ProjectAiReport.created_at.desc())
        .first()
    )

    if not row:
        raise HTTPException(status_code=404, detail="Report not found. Please generate report first.")

    return {
        "project_id": row.project_id,
        "sections": {
            "overview": row.overview or "",
            "task_analysis": row.task_analysis or "",
            "status_analysis": row.status_analysis or "",
            "next_steps": row.next_steps or "",
        },
        "structured": row.structured_snapshot,
        "model": row.model or "",
        "updated_at": row.created_at.isoformat() if row.created_at else None,
    }


@app.delete("/api/report/data/{project_id}")
def delete_report_data(
    project_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    state = load_state()
    check_project_access(db, state, project_id, user_id)

    db.query(ProjectAiReport).filter(ProjectAiReport.project_id == project_id).delete()
    db.commit()

    return {"message": "Report deleted"}

# =========================
# AI Project Q&A
# =========================
# ═══════════════════════════════════════════════════════════════
# AI 자유 질문 - Structured Data Query + LLM Formatting
# ═══════════════════════════════════════════════════════════════

def _resolve_question_scope(query: str, task_details: list, today) -> dict:
    """질문 대상(scope) 판별: project / task / assignee / schedule / notes"""
    query_lower = query.lower().strip()

    _stopwords = {'의', '에', '는', '은', '을', '를', '이', '가', '과', '와', '도', '로', '으로',
                  '에서', '까지', '부터', '대해', '관해', '관련', '어떻게', '얼마나', '언제',
                  '현황', '상태', '진행', '알려줘', '보여줘', '뭐야', '뭔가요', '있어', '없어',
                  '해줘', '할', '한', '하는', '된', '되는', '좀', '다', '그', '저', '이', '것'}

    # 1) Task 이름 매칭
    query_matched_tasks = []
    for t in task_details:
        title = (t.get("title") or "").strip()
        if not title:
            continue
        title_lower = title.lower()
        if title_lower in query_lower or title in query:
            query_matched_tasks.append(t)
            continue
        query_words = [w for w in query_lower.split() if len(w) >= 2 and w not in _stopwords]
        if query_words and any(w in title_lower for w in query_words):
            query_matched_tasks.append(t)

    # 2) 기간 매칭
    window = _parse_time_window_from_query(query, today)

    # 2-1) 기간 키워드가 있으면 window가 없어도 schedule로 처리 시도
    schedule_keywords = ["일정", "마감", "스케줄", "schedule", "deadline", "기한"]
    has_schedule_keyword = any(kw in query_lower for kw in schedule_keywords)

    # "마감 일정" 같은 경우 이번달로 기본 설정
    if has_schedule_keyword and not window and not query_matched_tasks:
        # "다음주 마감" 같은 경우는 이미 window가 잡히므로 여기는 fallback
        # 기간 키워드만 있고 시간 표현이 없으면 이번달로 기본 설정
        window = _month_window(today.year, today.month)

    # 3) 전체 현황 키워드
    is_full_overview = any(kw in query_lower for kw in [
        "전체 현황", "프로젝트 현황", "전체 요약", "전체 상태", "전체 진행",
        "프로젝트 상태", "프로젝트 진행", "프로젝트 요약", "전반적",
    ])

    # 4) 담당자 관련 키워드
    is_assignee_query = any(kw in query_lower for kw in [
        "담당자", "담당", "누가", "맡고", "배정", "할당",
    ])

    # scope 결정
    if query_matched_tasks:
        scope_type = "task"
    elif window:
        scope_type = "schedule"
    elif is_full_overview:
        scope_type = "project"
    elif is_assignee_query:
        scope_type = "assignee"
    else:
        scope_type = "general"

    return {
        "scope_type": scope_type,
        "matched_tasks": query_matched_tasks,
        "window": window,
        "is_full_overview": is_full_overview,
        "is_assignee_query": is_assignee_query,
        "query_lower": query_lower,
    }


def _fetch_project_context(db, project, sub_projects, members, users_map, task_details, project_notes):
    """프로젝트 수준 컨텍스트: 프로젝트 전체 정보 수집"""
    member_lines = []
    for m in members:
        role = m.get("role", "member")
        if role == "viewer":
            continue
        uid = int(m.get("user_id"))
        u = users_map.get(uid)
        if u:
            dept = u.get("deptname", "") or ""
            dept_str = f" ({dept})" if dept else ""
            member_lines.append(f"  - {u['username']}{dept_str} [{role}]")

    sp_lines = []
    for sp in sub_projects:
        sp_name = sp.get("name", "")
        sp_desc = sp.get("description", "")
        sp_tasks = [t for t in task_details if t.get("sub_project") == sp_name]
        done_count = sum(1 for t in sp_tasks if t.get("status") == "done")
        sp_lines.append(f"  - {sp_name}: {len(sp_tasks)}개 task ({done_count}개 완료)")
        if sp_desc:
            sp_lines.append(f"    설명: {sp_desc}")

    note_lines = []
    for n in project_notes[:10]:
        author = n.get("author_name", "")
        content = (n.get("content", "") or "")[:300]
        created = n.get("created_at", "")[:10] if n.get("created_at") else ""
        note_lines.append(f"  - [{created}] {author}: {content}")

    # Task 요약 (전체 현황용)
    total = len(task_details)
    done = sum(1 for t in task_details if t.get("status") == "done")
    in_prog = sum(1 for t in task_details if t.get("status") == "in_progress")
    todo = sum(1 for t in task_details if t.get("status") == "todo")
    hold = sum(1 for t in task_details if t.get("status") == "hold")

    ctx = f"""[프로젝트 정보]
프로젝트명: {project.name}
설명: {project.description or "없음"}
전체 Task: {total}개 (완료 {done}, 진행중 {in_prog}, 대기 {todo}, 보류 {hold})

[팀원 ({len(member_lines)}명)]
{chr(10).join(member_lines) if member_lines else "  없음"}
"""
    if sp_lines:
        ctx += f"\n[서브프로젝트]\n{chr(10).join(sp_lines)}\n"
    if note_lines:
        ctx += f"\n[프로젝트 노트 (최근 10개)]\n{chr(10).join(note_lines)}\n"
    return ctx


def _fetch_task_context_detail(t: dict, users_map: dict, db, include_full_text=True) -> str:
    """개별 Task 상세 컨텍스트 (자유 질문용 - 풍부한 텍스트 포함)"""
    lines = []
    lines.append(f"[Task: {t['title']}]")
    lines.append(f"  상태: {t['status']}")
    lines.append(f"  우선순위: {t['priority']}")
    lines.append(f"  진행률: {t['progress']}%")
    lines.append(f"  시작일: {t.get('start_date') or '미정'}")
    lines.append(f"  마감일: {t.get('due_date') or '미정'}")

    # 담당자 (assignee 기준)
    if t.get("assignees"):
        lines.append(f"  담당자: {', '.join(t['assignees'])}")
    else:
        lines.append(f"  담당자: 미배정")

    # 서브프로젝트
    if t.get("sub_project"):
        lines.append(f"  서브프로젝트: {t['sub_project']}")

    # 태그
    if t.get("tags"):
        lines.append(f"  태그: {', '.join(t['tags'])}")

    # 설명 (description) - 전문 포함
    desc = (t.get("description") or "").strip()
    if desc:
        if include_full_text:
            lines.append(f"  설명:\n    {desc}")
        else:
            lines.append(f"  설명: {desc[:200]}{'...' if len(desc) > 200 else ''}")

    # 작업노트 (activity) - 전문 포함
    activity_items = t.get("activity_items", [])
    if activity_items:
        total_cb = t.get("activity_total", 0)
        checked_cb = t.get("activity_checked", 0)
        lines.append(f"  작업노트: 체크리스트 {total_cb}개 중 {checked_cb}개 완료")
        cb_idx = 0
        for item in activity_items:
            if item["type"] == "checkbox":
                cb_idx += 1
                mark = "완료" if item["checked"] else "미완료"
                content = (item.get("content") or "").strip()
                # HTML 태그 제거 (간단)
                import re
                content = re.sub(r'<[^>]+>', '', content)
                lines.append(f"    {cb_idx}. {content} ({mark})")
            else:
                content = (item.get("content") or "").strip()
                if content:
                    import re
                    content = re.sub(r'<[^>]+>', '', content)
                    if include_full_text:
                        lines.append(f"    (메모) {content}")
                    else:
                        lines.append(f"    (메모) {content[:200]}{'...' if len(content) > 200 else ''}")

    # 첨부파일/URL
    attachments = t.get("attachments", [])
    if attachments:
        lines.append(f"  첨부/참조자료:")
        for a in attachments:
            fname = a.get("filename") or a.get("url", "")
            url = a.get("url", "")
            if url:
                lines.append(f"    - {fname}: {url}")
            else:
                lines.append(f"    - {fname}")

    return "\n".join(lines)


def _build_ai_free_question_context(
    scope: dict, db, project, sub_projects, members, users_map,
    task_details, project_notes, today
) -> tuple:
    """질문 scope에 따라 컨텍스트 문자열, scope_hint, prompt_tasks, context_members 생성"""
    import re
    scope_type = scope["scope_type"]
    matched_tasks = scope["matched_tasks"]
    window = scope["window"]
    is_full_overview = scope["is_full_overview"]
    ws = we = None

    # ── scope별 prompt_tasks / context 결정 ──
    if scope_type == "task" and matched_tasks:
        prompt_tasks = matched_tasks
        scope_hint = (
            f"\n[범위 제약]\n"
            f"사용자가 특정 Task({', '.join(t['title'] for t in matched_tasks)})에 대해 질문했습니다.\n"
            f"해당 Task 중심으로만 상세하게 답변하세요. 다른 Task는 언급하지 마세요.\n"
            f"제목만 말하지 말고, 아래 컨텍스트에 포함된 설명/작업노트/메모/일정/담당자/참조자료를 최대한 활용해서 풍부하게 설명하세요.\n"
        )
        # Task 질문: 해당 task assignees만
        task_assignee_ids = set()
        for t in matched_tasks:
            for aid in (t.get("assignee_ids") or []):
                task_assignee_ids.add(int(aid))
        if task_assignee_ids:
            context_members = []
            for uid in task_assignee_ids:
                u = users_map.get(uid)
                if u:
                    dept = u.get("deptname", "") or ""
                    dept_str = f" ({dept})" if dept else ""
                    context_members.append(f"{u['username']}{dept_str}")
        else:
            context_members = ["미배정"]

        # 상세 컨텍스트 빌드
        task_ctx_lines = []
        for t in matched_tasks:
            task_ctx_lines.append(_fetch_task_context_detail(t, users_map, db, include_full_text=True))

        context_str = f"""프로젝트명: {project.name}
프로젝트 설명: {project.description or "없음"}

[질문 대상 Task 상세 정보]
{chr(10).join(task_ctx_lines)}

[담당자]
{', '.join(context_members)}
"""

    elif scope_type == "schedule" and window:
        ws, we = window
        # 기간에 겹치는 task + 날짜 미설정이지만 진행 중인 task도 포함
        filtered_by_time = [t for t in task_details if _task_overlaps_window(t, ws, we)]
        no_date_active = [
            t for t in task_details
            if not t.get("start_date") and not t.get("due_date")
            and t.get("status") in ("in_progress", "todo")
        ]
        prompt_tasks = filtered_by_time
        scope_hint = (
            f"\n[기간 제약]\n"
            f"사용자 질문이 요청한 기간은 {ws.isoformat()} ~ {(we - timedelta(days=1)).isoformat()} 입니다.\n"
            f"이 기간에 해당하는 Task를 중심으로 답변하세요.\n"
            f"각 Task의 설명/작업노트/일정/담당자를 상세하게 답변하세요.\n"
            f"날짜가 설정되지 않았지만 진행 중/대기 상태인 Task가 있다면 별도로 안내하세요.\n"
            f"\n[중요] 아래 컨텍스트에 있는 모든 Task를 빠짐없이 답변에 포함하세요.\n"
            f"일정이 있으면 날짜순으로 정렬해서 보여주세요.\n"
        )
        # 기간 질문: project members
        context_members = _get_project_member_names(members, users_map)

        task_ctx_lines = []
        for t in filtered_by_time[:50]:
            task_ctx_lines.append(_fetch_task_context_detail(t, users_map, db, include_full_text=True))

        # 날짜 미설정 활성 task도 별도 섹션으로 추가
        no_date_lines = []
        for t in no_date_active[:20]:
            no_date_lines.append(_fetch_task_context_detail(t, users_map, db, include_full_text=False))

        proj_ctx = _fetch_project_context(db, project, sub_projects, members, users_map, task_details, project_notes)
        context_str = f"""{proj_ctx}

[기간 내 Task 상세 정보 ({ws.isoformat()} ~ {(we - timedelta(days=1)).isoformat()})]
총 {len(filtered_by_time)}개 Task가 이 기간에 해당합니다.
{chr(10).join(task_ctx_lines) if task_ctx_lines else "해당 기간에 Task 없음"}
"""
        if no_date_lines:
            context_str += f"""
[날짜 미설정 활성 Task ({len(no_date_active)}개)]
아래 Task들은 일정이 미설정이지만 진행 중/대기 상태입니다.
{chr(10).join(no_date_lines)}
"""

    elif scope_type == "project" or is_full_overview:
        prompt_tasks = task_details
        scope_hint = (
            "\n[범위]\n"
            "사용자가 프로젝트 전체 현황을 요청했습니다.\n"
            "핵심 항목(진행 중, 임박, 지연)을 중심으로 요약하되, 각 주요 Task에 대해서는 설명/작업내용도 간략히 포함하세요.\n"
        )
        context_members = _get_project_member_names(members, users_map)

        proj_ctx = _fetch_project_context(db, project, sub_projects, members, users_map, task_details, project_notes)

        # 전체 현황: 모든 task를 간략 상세로
        task_ctx_lines = []
        for t in task_details[:40]:
            task_ctx_lines.append(_fetch_task_context_detail(t, users_map, db, include_full_text=False))

        context_str = f"""{proj_ctx}

[전체 Task 상세]
{chr(10).join(task_ctx_lines) if task_ctx_lines else "Task 없음"}
"""

    elif scope_type == "assignee":
        prompt_tasks = task_details
        scope_hint = (
            "\n[범위]\n"
            "사용자가 담당자 관련 질문을 했습니다.\n"
            "질문에서 언급된 담당자와 관련된 Task만 선별해서 답변하세요.\n"
            "각 Task의 담당자는 task.assignee 기준으로 판단하세요.\n"
        )
        context_members = _get_project_member_names(members, users_map)

        task_ctx_lines = []
        for t in task_details[:40]:
            task_ctx_lines.append(_fetch_task_context_detail(t, users_map, db, include_full_text=False))

        proj_ctx = _fetch_project_context(db, project, sub_projects, members, users_map, task_details, project_notes)
        context_str = f"""{proj_ctx}

[전체 Task 상세 (담당자 포함)]
{chr(10).join(task_ctx_lines) if task_ctx_lines else "Task 없음"}
"""

    else:
        # general: 질문과 관련된 Task를 선별
        prompt_tasks = task_details
        scope_hint = (
            "\n[범위]\n"
            "질문과 관련된 Task만 선별해서 답변하세요. 전체 Task를 나열하지 마세요.\n"
            "답변할 때 관련 Task의 설명/작업노트/메모/일정/담당자를 최대한 활용해서 상세하게 설명하세요.\n"
        )
        context_members = _get_project_member_names(members, users_map)

        proj_ctx = _fetch_project_context(db, project, sub_projects, members, users_map, task_details, project_notes)

        task_ctx_lines = []
        for t in task_details[:50]:
            task_ctx_lines.append(_fetch_task_context_detail(t, users_map, db, include_full_text=True))

        context_str = f"""{proj_ctx}

[전체 Task 상세]
{chr(10).join(task_ctx_lines) if task_ctx_lines else "Task 없음"}
"""

    return context_str, scope_hint, prompt_tasks, context_members, ws, we


def _get_project_member_names(members: list, users_map: dict) -> list:
    """프로젝트 멤버 이름 목록 (viewer 제외)"""
    names = []
    for m in members:
        role = m.get("role", "member")
        if role == "viewer":
            continue
        uid = int(m.get("user_id"))
        u = users_map.get(uid)
        if u:
            names.append(f'{u["username"]} ({role})')
    return names


@app.post("/api/projects/{project_id}/ai-query")
def generate_project_ai_query(
    project_id: int,
    req: ProjectAiQueryRequest,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    import requests
    import re

    state = load_state()
    check_project_access(db, state, project_id, user_id)

    row = get_or_create_ai_setting(db)
    provider = _active_provider(row)
    selected_model = _normalize_selected_model(row.model_name)
    model_name = llm_router.resolve_model(provider, selected_model)

    import logging as _logging
    _logging.getLogger("main").info(
        "AI Query request: project_id=%s, provider=%s, selected_model=%r, resolved_model=%r",
        project_id, provider, selected_model, model_name,
    )

    if not model_name:
        raise HTTPException(400, "AI settings not configured. Please choose a model in AI Settings.")

    p = db.query(Project).filter(Project.id == project_id, Project.archived_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    # ── 1. 전체 데이터 수집 ──
    task_rows = db.query(Task).filter(Task.project_id == project_id, Task.archived_at.is_(None)).all()
    tasks = [task_dict(t, state) for t in task_rows]

    db_members = db.query(ProjectMemberModel).filter(
        ProjectMemberModel.project_id == project_id,
        ProjectMemberModel.role != 'viewer'
    ).all()
    members = [{"user_id": m.user_id, "role": m.role, "project_id": m.project_id} for m in db_members]
    sub_projects = get_subprojects_from_db(db, project_id)

    users_map = {u.id: user_dict(u, state) for u in db.query(User).all()}

    # Task 상세 수집 (assignee_ids 포함)
    task_details = []
    for t in tasks:
        assignee_ids_raw = t.get("assignee_ids") or []
        assignees = [users_map.get(a, {}).get("username", f"User {a}") for a in assignee_ids_raw]

        sp_name = ""
        if t.get("sub_project_id"):
            sp = next((s for s in sub_projects if int(s["id"]) == int(t["sub_project_id"])), None)
            sp_name = sp["name"] if sp else ""

        # DB 기반 첨부파일 조회
        db_attachments = db.query(AttachmentModel).filter(AttachmentModel.task_id == int(t["id"])).all()
        task_attachments = [
            {"id": a.id, "filename": a.filename or "", "url": a.url or "", "type": a.type or "url"}
            for a in db_attachments
        ]

        # 작업노트(activity) 상세 정보
        activities = db.query(TaskActivityModel).filter(
            TaskActivityModel.task_id == int(t["id"])
        ).order_by(TaskActivityModel.order_index).all()
        activity_items = []
        for act in activities:
            block_type = act.block_type or "checkbox"
            if block_type == "checkbox":
                activity_items.append({"type": "checkbox", "content": act.content or "", "checked": bool(act.checked)})
            else:
                activity_items.append({"type": "text", "content": act.content or ""})
        checkbox_acts = [a for a in activity_items if a["type"] == "checkbox"]
        checked_count = sum(1 for a in checkbox_acts if a["checked"])
        total_checkboxes = len(checkbox_acts)

        task_details.append({
            "id": int(t["id"]),
            "title": t.get("title", "") or "",
            "description": t.get("description", "") or "",
            "status": t.get("status", "todo") or "todo",
            "priority": t.get("priority", "medium") or "medium",
            "progress": t.get("progress", 0) or 0,
            "start_date": t.get("start_date"),
            "due_date": t.get("due_date"),
            "assignee_ids": assignee_ids_raw,
            "assignees": assignees or [],
            "sub_project": sp_name or "",
            "tags": t.get("tags", []) or [],
            "attachments": task_attachments,
            "activity_items": activity_items,
            "activity_checked": checked_count,
            "activity_total": total_checkboxes,
        })

    # 프로젝트 노트 수집
    db_notes = db.query(NoteModel).filter(NoteModel.project_id == project_id).order_by(NoteModel.created_at.desc()).all()
    project_notes = []
    for n in db_notes:
        author = users_map.get(n.author_id)
        project_notes.append({
            "id": n.id,
            "content": n.content or "",
            "author_name": author["username"] if author else "",
            "created_at": n.created_at.isoformat() if n.created_at else "",
        })

    # ── 2. 질문 scope 판별 ──
    today = _today_kst()
    scope = _resolve_question_scope(req.query, task_details, today)

    # ── 3. scope별 컨텍스트 빌드 ──
    context_str, scope_hint, prompt_tasks, context_members, ws, we = _build_ai_free_question_context(
        scope, db, p, sub_projects, members, users_map, task_details, project_notes, today
    )

    # ── 4. LLM 프롬프트 구성 ──
    is_task_scope = scope["scope_type"] == "task" and scope["matched_tasks"]

    if is_task_scope:
        detail_instruction = (
            "이 질문은 특정 Task에 대한 질문입니다.\n"
            "제목만 말하지 말고, 컨텍스트에 있는 설명/작업노트/메모/일정/담당자/참조자료를 모두 활용해서 최대한 상세하게 답변하세요.\n"
            "사용자가 작성해둔 텍스트(설명, 작업노트, 메모)가 있으면 그 내용을 그대로 활용해서 풍부하게 설명하세요.\n"
        )
    else:
        detail_instruction = (
            "컨텍스트에 있는 설명/작업노트/메모/일정/담당자 정보를 활용해서 답변하세요.\n"
            "제목만 나열하지 말고, 관련 상세 내용을 함께 설명하세요.\n"
        )

    # 기간 질문인 경우 조회 범위 명시 추가
    date_range_note = ""
    if scope["scope_type"] == "schedule" and ws and we:
        date_range_note = (
            f"\n[조회 기준]\n"
            f"오늘 날짜: {today.isoformat()}\n"
            f"조회 기간: {ws.isoformat()} ~ {(we - timedelta(days=1)).isoformat()}\n"
            f"컨텍스트에 포함된 Task 수: {len(prompt_tasks)}개\n"
        )

    prompt = f"""당신은 전문 프로젝트 매니저 보조 AI입니다.
아래 컨텍스트를 바탕으로 사용자의 질문에 답변해주세요.
{scope_hint}{date_range_note}
[핵심 원칙]
- 질문 범위를 먼저 파악하고, 그 범위 안에서만 답변하세요.
- 기간 관련 질문이면 컨텍스트에 있는 모든 Task를 빠짐없이 포함하세요. 절대 임의로 생략하지 마세요.
- 일정 정리 요청이면 날짜순(시작일 또는 마감일 기준)으로 정렬하세요.
- {detail_instruction}
- 특정 Task 질문이면 해당 Task의 상태/일정/담당자/작업노트/설명/첨부자료를 중심으로 답변하세요.
- Task 질문일 때 담당자는 해당 Task의 assignee만 언급하세요. 프로젝트 전체 팀원을 보여주지 마세요.
- 답변 시작 부분에 "조회 기간" 또는 "기준"을 간단히 명시하세요.

[컨텍스트]
{context_str}

[질문]
{req.query}

[출력 규칙]
- 반드시 4개 섹션을 순서대로 작성
- 각 섹션 태그는 반드시 '단독 줄'로 출력 (예: [섹션1: 한줄요약])
- 섹션 내용은 줄바꿈으로 구분 (가급적 '한 줄 = 한 문장')
- 문장은 중간에 끊지 말고 반드시 마침표(또는 '다.')로 끝내기
- 마크다운 금지: #, **, ```, 표(|---|) 금지
- 작업이나 항목을 나열할 때는 반드시 숫자 번호(1. 2. 3.)를 사용하세요.
- 절대 "첫 번째", "두 번째" 같은 서수 표현을 사용하지 마세요.

[섹션1: 한줄요약]
결론을 한 문장으로만 작성.

[섹션2: 상세설명]
질문 대상에 대한 상세 정보를 아래 항목별로 구분해서 작성하세요.
각 항목은 "항목명:" 형식으로 시작하고, 해당 내용이 없으면 그 항목을 생략하세요.
텍스트를 그냥 이어붙이지 말고, 항목별로 명확히 나눠서 작성하세요.

과제: 질문 대상이 되는 Task명만 간결하게 작성하세요. 불필요하게 긴 설명을 붙이지 마세요.
기간: 시작일 ~ 마감일 정보.
담당자: 해당 Task의 assignee 또는 프로젝트 팀원 (질문 범위에 맞게).
작업노트: 사용자가 작성한 작업 내용, 메모를 충분히 반영. 각 항목을 개별적으로 나열하되, 쉼표(,)로 이어붙이지 말고 각 항목을 별도 줄에 작성하세요.
완료 항목: 완료된 체크리스트 항목들. 각 항목을 별도 줄에 하나씩 작성하세요. 쉼표(,)로 이어붙이지 마세요.
미완료 항목: 아직 완료되지 않은 체크리스트 항목들. 각 항목을 별도 줄에 하나씩 작성하세요. 쉼표(,)로 이어붙이지 마세요.
참고자료: 연결된 URL, 첨부파일, 참조 정보. 각 항목을 별도 줄에 작성하세요.
주의사항: 리스크, 지연, 주의해야 할 사항. 각 항목을 별도 줄에 작성하세요.

컨텍스트에 없는 정보는 추측하지 말고 해당 항목을 생략하세요.

[섹션3: 핵심 일정]
질문과 직접 관련된 Task만 최대 8개.
각 Task를 아래 형식으로 작성 (슬래시(/) 구분자를 쓰지 마세요):
Task명: OOO
담당자: OOO (해당 Task의 assignee만 표시)
진행률: OO%
일정: YYYY-MM-DD ~ YYYY-MM-DD (시작일이나 마감일이 없으면 "미정"으로 표시)
상태: 진행 중/대기/보류/완료
세부 작업이 있으면 번호를 매겨서 표시:
1. 항목명 (완료)
2. 항목명 (미완료)
Task 간에는 빈 줄로 구분하세요.
없으면 "없음" 한 줄.

[섹션4: 다음 액션]
다음 액션 3~6개.
각 액션은 번호를 매겨서 한 줄씩 작성. 예) 1. 액션 내용
"""

    system_prompt = """
당신은 전문 프로젝트 매니저 보조 AI입니다.
반드시 한국어로 답변하세요.
마크다운(##, **, 표, 코드블록)을 절대 사용하지 마세요.
각 섹션 태그([섹션1], [섹션2], [섹션3], [섹션4])는 반드시 단독 줄로 출력하세요.
섹션 내용은 줄바꿈으로 구분하며, 가급적 한 줄 = 한 문장으로 작성하세요.
문장 중간에 끊지 말고 반드시 마침표(또는 '다.')로 끝내세요.

[가장 중요한 규칙]
- 사용자의 질문 범위를 먼저 파악하세요.
- 질문 범위 밖의 Task는 절대 언급하지 마세요.
- 전체 Task를 나열하지 마세요. 질문과 관련된 것만 선별하세요.
- 특정 Task 질문에는 해당 Task만 답하세요.
- Task를 언급할 때는 반드시 [Task: Task제목] 형식으로 제목을 단독 줄에 작성하고, 그 아래에 분석 내용을 작성하세요.
- Task 질문에서 담당자는 해당 Task의 assignee만 표시하세요. 프로젝트 전체 팀원을 담당자로 보여주면 안 됩니다.
- Viewer는 절대 담당자로 표시하지 마세요.
- 제목만 말하지 말고, 컨텍스트에 있는 상세 내용(설명, 작업노트, 메모, 첨부자료)을 적극 활용해서 풍부하게 답변하세요.
- 사용자가 작성해둔 텍스트가 있으면 반드시 그 내용을 답변에 반영하세요.
""".strip()

    try:
        content = llm_router.chat(
            provider=provider,
            base_url=None,
            model_name=model_name,
            system_prompt=system_prompt,
            user_prompt=prompt,
            temperature=0.2,
            max_tokens=4096,
        )

        content = sanitize_llm_text_ai(content)

        # ── 섹션 파싱 ──
        parsed = {"one_liner": "", "details": "", "key_schedule": "", "next_actions": ""}
        current = ""
        for line in content.split("\n"):
            sline = line.strip()
            if "[섹션1" in sline:
                current = "one_liner"; continue
            if "[섹션2" in sline:
                current = "details"; continue
            if "[섹션3" in sline:
                current = "key_schedule"; continue
            if "[섹션4" in sline:
                current = "next_actions"; continue
            if current:
                parsed[current] += line + "\n"

        parsed = {k: (v.strip() if v else "") for k, v in parsed.items()}
        if not parsed["one_liner"]:
            parsed["one_liner"] = (content[:200].strip() + ("…" if len(content) > 200 else "")) if content else ""
            parsed["details"] = content

        # ── context_tasks 선별 ──
        matched_tasks = scope["matched_tasks"]
        window = scope["window"]

        schedule_text = parsed.get("key_schedule", "") or ""
        response_matched = []
        if schedule_text:
            for t in task_details:
                if t["title"] and t["title"] in schedule_text:
                    response_matched.append(t)

        if matched_tasks:
            context_tasks = matched_tasks
        elif window:
            filtered_by_time = [t for t in task_details if _task_overlaps_window(t, ws, we)] if ws and we else []
            context_tasks = filtered_by_time
        elif response_matched:
            context_tasks = response_matched
        elif scope["is_full_overview"]:
            context_tasks = prompt_tasks[:15]
        else:
            context_tasks = response_matched if response_matched else prompt_tasks[:8]

        active_tasks = [t for t in context_tasks if t.get("status") != "hold"]
        hold_tasks = [t for t in context_tasks if t.get("status") == "hold"]
        done_tasks = [t for t in context_tasks if t.get("status") == "done"]
        in_progress_tasks = [t for t in context_tasks if t.get("status") == "in_progress"]
        todo_tasks = [t for t in context_tasks if t.get("status") == "todo"]

        if len(active_tasks) > 0:
            progress_sum = sum(
                100 if t.get("status") == "done" else (t.get("progress", 0) or 0)
                for t in active_tasks
            )
            overall_progress = round(progress_sum / len(active_tasks), 1)
        else:
            overall_progress = 0.0

        db_record = ProjectAiQuery(
            project_id=project_id,
            user_id=user_id,
            query=req.query,
            raw_response=content,
            model=model_name,
            created_at=datetime.utcnow(),
        )
        db.add(db_record)
        db.commit()
        db.refresh(db_record)

        context = {
            "status_breakdown": {
                "total": len(context_tasks),
                "active": len(active_tasks),
                "done": len(done_tasks),
                "in_progress": len(in_progress_tasks),
                "todo": len(todo_tasks),
                "hold": len(hold_tasks),
                "overall_progress": overall_progress,
            },
            "members": context_members,
            "tasks": context_tasks,
            "filter": {
                "mode": scope["scope_type"],
                "window_start": ws.isoformat() if ws else None,
                "window_end": we.isoformat() if we else None,
            },
        }

        return {
            "id": db_record.id,
            "project_id": db_record.project_id,
            "user_id": db_record.user_id,
            "query": db_record.query,
            "raw_response": db_record.raw_response,
            "model": db_record.model,
            "created_at": db_record.created_at.isoformat() if db_record.created_at else None,
            "parsed_response": parsed,
            "context": context,
        }

    except requests.exceptions.ConnectionError:
        _log_dsllm_failure("AI Query 실패: DSLLM 연결 불가", status_code=502, endpoint=f"/api/projects/{project_id}/ai-query")
        raise HTTPException(status_code=502, detail="Cannot connect to DSLLM. Please check BASE_URL in .env.local/.env.production.")
    except requests.exceptions.Timeout:
        _log_dsllm_failure("AI Query 실패: DSLLM timeout", status_code=504, endpoint=f"/api/projects/{project_id}/ai-query")
        raise HTTPException(status_code=504, detail="DSLLM request timed out.")
    except requests.exceptions.HTTPError as e:
        _log_dsllm_failure("AI Query 실패: DSLLM HTTP error", detail=str(e)[:500], status_code=502, endpoint=f"/api/projects/{project_id}/ai-query")
        raise HTTPException(status_code=502, detail=f"DSLLM returned HTTP error: {str(e)}")
    except Exception as e:
        _log_dsllm_failure(f"AI Query 실패: {type(e).__name__}", detail=str(e)[:500], status_code=500, endpoint=f"/api/projects/{project_id}/ai-query")
        raise HTTPException(status_code=500, detail=f"AI query failed: {str(e)}")

# AI Query History 조회
@app.get("/api/projects/{project_id}/ai-queries")
def get_project_ai_queries(
    project_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    state = load_state()
    check_project_access(db, state, project_id, user_id)

    rows = (
        db.query(ProjectAiQuery)
        .filter(ProjectAiQuery.project_id == project_id)
        .order_by(ProjectAiQuery.created_at.desc())
        .all()
    )

    queries = []
    for r in rows:
        raw = getattr(r, "raw_response", None) or ""
        one_liner = getattr(r, "one_liner", None)
        details = getattr(r, "details", None)
        key_schedule = getattr(r, "key_schedule", None)
        next_actions = getattr(r, "next_actions", None)

        # ✅ parsed_response를 프론트가 기대하는 형태로 생성 (없으면 raw_response로 fallback)
        parsed_response = {
            "one_liner": one_liner or (raw[:200].strip() + ("…" if len(raw) > 200 else "")) if raw else "",
            "details": details or raw or "",
            "key_schedule": key_schedule or "",
            "next_actions": next_actions or "",
        }

        queries.append(
            {
                "id": r.id,
                "project_id": r.project_id,
                "user_id": r.user_id,
                "query": r.query,

                # ✅ 기존 호환 필드들 유지
                "response": getattr(r, "response", None),
                "raw_response": raw,
                "model": getattr(r, "model", None),
                "created_at": r.created_at.isoformat() if r.created_at else None,

                # ✅ 확장 컬럼(있을 때만)
                "one_liner": one_liner,
                "details": details,
                "key_schedule": key_schedule,
                "next_actions": next_actions,
                "context_snapshot": getattr(r, "context_snapshot", None),

                # ✅ 프론트에서 바로 쓰게 추가
                "parsed_response": parsed_response,
            }
        )

    return {"queries": queries}

# =========================
# Admin APIs
# =========================
def _mask_ip(ip: str) -> str:
    """IPv4 는 앞 2옥텟만 남기고 마스킹(10.12.xxx.xxx). 그 외는 일부만 노출."""
    if not ip:
        return "-"
    parts = ip.split(".")
    if len(parts) == 4:
        return f"{parts[0]}.{parts[1]}.xxx.xxx"
    return (ip[:6] + "***") if len(ip) > 6 else "***"


@app.get("/api/admin/visit-stats")
def admin_get_visit_stats(
    user_id: int = Query(...),
    days: int = Query(default=30, ge=1, le=365),
    db: Session = Depends(get_db),
):
    """visit_log 기반 방문 통계. (require_admin)

    방문 수(raw count)와 순 방문자 수(distinct user_id/username/ip)를 구분해 반환한다.
    IP 는 super_admin 에게만 전체 노출, 그 외 admin 은 마스킹한다.
    """
    require_admin(db, None, user_id)

    from sqlalchemy import distinct, cast, String as SAString
    from datetime import date as _date, timedelta as _timedelta

    # 순 방문자 = COALESCE(user_id, username, ip_address) distinct
    uniq_expr = func.count(distinct(func.coalesce(
        cast(VisitLog.user_id, SAString),
        VisitLog.username,
        VisitLog.ip_address,
    )))

    today = datetime.now(KST).date()
    today_str = today.isoformat()
    d7_str = (today - _timedelta(days=6)).isoformat()        # 최근 7일 (오늘 포함)
    d30_str = (today - _timedelta(days=29)).isoformat()      # 최근 30일 (오늘 포함)
    range_start = (today - _timedelta(days=days - 1)).isoformat()
    month_start = today.replace(day=1).isoformat()

    def _visits(d1=None, d2=None):
        q = db.query(func.count(VisitLog.id))
        if d1:
            q = q.filter(VisitLog.visit_date >= d1)
        if d2:
            q = q.filter(VisitLog.visit_date <= d2)
        return int(q.scalar() or 0)

    def _unique(d1=None, d2=None):
        q = db.query(uniq_expr)
        if d1:
            q = q.filter(VisitLog.visit_date >= d1)
        if d2:
            q = q.filter(VisitLog.visit_date <= d2)
        return int(q.scalar() or 0)

    summary = {
        "today_visits": _visits(today_str, today_str),
        "today_unique_visitors": _unique(today_str, today_str),
        "last_7_days_visits": _visits(d7_str),
        "last_7_days_unique_visitors": _unique(d7_str),
        "last_30_days_visits": _visits(d30_str),
        "last_30_days_unique_visitors": _unique(d30_str),
        "month_unique_visitors": _unique(month_start),
        "total_visits": _visits(),
    }

    # ── 일별 (range_start ~ today), 누락일은 0으로 채움 ──
    daily_rows = (
        db.query(VisitLog.visit_date, func.count(VisitLog.id), uniq_expr)
        .filter(VisitLog.visit_date >= range_start)
        .group_by(VisitLog.visit_date)
        .all()
    )
    daily_map = {r[0]: (int(r[1] or 0), int(r[2] or 0)) for r in daily_rows if r[0]}
    daily = []
    for i in range(days):
        d = (today - _timedelta(days=days - 1 - i)).isoformat()
        v, u = daily_map.get(d, (0, 0))
        daily.append({"date": d, "visits": v, "unique_visitors": u})

    # ── 부서별 TOP 10 (range_start 기준) ──
    dept_label = func.coalesce(func.nullif(VisitLog.deptname, ""), "미확인")
    dept_rows = (
        db.query(
            dept_label.label("dept"),
            func.count(VisitLog.id),
            uniq_expr,
            func.max(VisitLog.visit_date),
        )
        .filter(VisitLog.visit_date >= range_start)
        .group_by(dept_label)
        .order_by(uniq_expr.desc(), func.count(VisitLog.id).desc())
        .limit(10)
        .all()
    )
    total_dept_visits = sum(int(r[1] or 0) for r in dept_rows) or 1
    departments = [
        {
            "deptname": r[0] or "미확인",
            "visits": int(r[1] or 0),
            "unique_visitors": int(r[2] or 0),
            "last_visit_date": r[3] or "-",
            "ratio": round(int(r[1] or 0) * 100.0 / total_dept_visits, 1),
        }
        for r in dept_rows
    ]

    # ── 최근 방문 기록 (IP 마스킹 정책 적용) ──
    show_full_ip = is_super_admin_user(db, user_id)
    recent_rows = (
        db.query(VisitLog)
        .order_by(VisitLog.timestamp.desc())
        .limit(30)
        .all()
    )
    recent = []
    for r in recent_rows:
        ip = r.ip_address or ""
        recent.append({
            "timestamp": r.timestamp.strftime("%Y-%m-%d %H:%M:%S") if r.timestamp else "-",
            "username": r.username or "-",
            "deptname": r.deptname or "미확인",
            "ip_address": ip if show_full_ip else _mask_ip(ip),
        })

    return {
        "summary": summary,
        "daily": daily,
        "departments": departments,
        "recent": recent,
        "ip_masked": not show_full_ip,
        "range_days": days,
    }


@app.get("/api/admin/usage-stats")
def admin_get_usage_stats(
    user_id: int = Query(...),
    days: int = Query(default=30, ge=1, le=365),
    db: Session = Depends(get_db),
):
    """사용 통계(visit-stats 확장): 방문 지표 + 프로젝트/Task 생성 활동 지표. (require_admin)

    - 방문 수/순 방문자: visit_log 기준 (기존 visit-stats 와 동일 로직)
    - 프로젝트 생성 수: projects.created_at 기준. 시스템 프로젝트([시스템] 단발 업무)와
      보관(archived_at) 프로젝트는 제외.
    - Task 생성 수: tasks.created_at 기준. 보관 task 제외. (1차: 전체 task 카운트)
    - 부서별 프로젝트 생성 수: projects.created_by(없으면 owner_id)의 users.deptname 기준.
      Task 는 created_by 컬럼이 없어 부서별 집계 미제공(전역 카운트만).
    - 날짜 기준은 KST. created_at 은 DB 서버 시각(운영 MySQL=KST) 기준 func.date 로 집계.
    """
    require_admin(db, None, user_id)

    from sqlalchemy import distinct, cast, String as SAString
    from datetime import timedelta as _timedelta

    # ── 시스템 프로젝트 ID 집합 (sidecar project_meta 의 is_system 마커 + 이름 안전망) ──
    state = load_state()
    proj_meta = state.get("project_meta", {}) or {}
    system_ids = {
        int(pid) for pid, m in proj_meta.items()
        if isinstance(m, dict) and m.get("is_system")
    }

    # 순 방문자 = COALESCE(user_id, username, ip_address) distinct
    uniq_expr = func.count(distinct(func.coalesce(
        cast(VisitLog.user_id, SAString),
        VisitLog.username,
        VisitLog.ip_address,
    )))

    today = datetime.now(KST).date()
    today_str = today.isoformat()
    d7_str = (today - _timedelta(days=6)).isoformat()        # 최근 7일 (오늘 포함)
    d30_str = (today - _timedelta(days=29)).isoformat()      # 최근 30일 (오늘 포함)
    range_start = (today - _timedelta(days=days - 1)).isoformat()
    month_start = today.replace(day=1).isoformat()

    # ── 방문 지표 ──
    def _visits(d1=None, d2=None):
        q = db.query(func.count(VisitLog.id))
        if d1:
            q = q.filter(VisitLog.visit_date >= d1)
        if d2:
            q = q.filter(VisitLog.visit_date <= d2)
        return int(q.scalar() or 0)

    def _unique(d1=None, d2=None):
        q = db.query(uniq_expr)
        if d1:
            q = q.filter(VisitLog.visit_date >= d1)
        if d2:
            q = q.filter(VisitLog.visit_date <= d2)
        return int(q.scalar() or 0)

    # ── 활동 지표 (프로젝트/Task 생성) ──
    proj_date = func.date(Project.created_at)
    task_date = func.date(Task.created_at)

    def _proj_base():
        # 보관/시스템 프로젝트 제외 (이름 prefix 안전망 포함)
        q = db.query(Project).filter(
            Project.archived_at.is_(None),
            Project.name != ONEOFF_PROJECT_NAME,
        )
        if system_ids:
            q = q.filter(~Project.id.in_(system_ids))
        return q

    def _projects(d1=None, d2=None):
        q = _proj_base().with_entities(func.count(Project.id))
        if d1:
            q = q.filter(proj_date >= d1)
        if d2:
            q = q.filter(proj_date <= d2)
        return int(q.scalar() or 0)

    def _tasks(d1=None, d2=None):
        q = db.query(func.count(Task.id)).filter(Task.archived_at.is_(None))
        if d1:
            q = q.filter(task_date >= d1)
        if d2:
            q = q.filter(task_date <= d2)
        return int(q.scalar() or 0)

    summary = {
        "today_visits": _visits(today_str, today_str),
        "today_unique_visitors": _unique(today_str, today_str),
        "today_projects_created": _projects(today_str, today_str),
        "today_tasks_created": _tasks(today_str, today_str),

        "last_7_days_visits": _visits(d7_str),
        "last_7_days_unique_visitors": _unique(d7_str),
        "last_7_days_projects_created": _projects(d7_str),
        "last_7_days_tasks_created": _tasks(d7_str),

        "last_30_days_visits": _visits(d30_str),
        "last_30_days_unique_visitors": _unique(d30_str),
        "last_30_days_projects_created": _projects(d30_str),
        "last_30_days_tasks_created": _tasks(d30_str),

        "month_unique_visitors": _unique(month_start),
        "month_projects_created": _projects(month_start),
        "month_tasks_created": _tasks(month_start),

        "total_visits": _visits(),
        "total_projects": _projects(),
        "total_tasks": _tasks(),
    }

    # ── 일별 추이 (range_start ~ today), 누락일은 0으로 채움 ──
    visit_rows = (
        db.query(VisitLog.visit_date, func.count(VisitLog.id), uniq_expr)
        .filter(VisitLog.visit_date >= range_start)
        .group_by(VisitLog.visit_date)
        .all()
    )
    visit_map = {r[0]: (int(r[1] or 0), int(r[2] or 0)) for r in visit_rows if r[0]}

    proj_daily_rows = (
        _proj_base().with_entities(proj_date, func.count(Project.id))
        .filter(proj_date >= range_start)
        .group_by(proj_date)
        .all()
    )
    proj_map = {str(r[0])[:10]: int(r[1] or 0) for r in proj_daily_rows if r[0]}

    task_daily_rows = (
        db.query(task_date, func.count(Task.id))
        .filter(Task.archived_at.is_(None), task_date >= range_start)
        .group_by(task_date)
        .all()
    )
    task_map = {str(r[0])[:10]: int(r[1] or 0) for r in task_daily_rows if r[0]}

    daily = []
    for i in range(days):
        d = (today - _timedelta(days=days - 1 - i)).isoformat()
        v, u = visit_map.get(d, (0, 0))
        daily.append({
            "date": d,
            "visits": v,
            "unique_visitors": u,
            "projects_created": proj_map.get(d, 0),
            "tasks_created": task_map.get(d, 0),
        })

    # ── 부서별 TOP 10 (range_start 기준, 방문 기준 정렬) ──
    dept_label = func.coalesce(func.nullif(VisitLog.deptname, ""), "미확인")
    dept_rows = (
        db.query(
            dept_label.label("dept"),
            func.count(VisitLog.id),
            uniq_expr,
            func.max(VisitLog.visit_date),
        )
        .filter(VisitLog.visit_date >= range_start)
        .group_by(dept_label)
        .order_by(uniq_expr.desc(), func.count(VisitLog.id).desc())
        .limit(10)
        .all()
    )

    # 부서별 프로젝트 생성 수 (created_by 없으면 owner_id 기준 users.deptname)
    creator_id = func.coalesce(Project.created_by, Project.owner_id)
    creator_dept = func.coalesce(func.nullif(User.deptname, ""), "미확인")
    proj_dept_rows = (
        _proj_base()
        .with_entities(creator_dept.label("dept"), func.count(Project.id))
        .join(User, User.id == creator_id)
        .filter(proj_date >= range_start)
        .group_by(creator_dept)
        .all()
    )
    proj_dept_map = {r[0]: int(r[1] or 0) for r in proj_dept_rows if r[0]}

    total_dept_visits = sum(int(r[1] or 0) for r in dept_rows) or 1
    departments = [
        {
            "deptname": r[0] or "미확인",
            "visits": int(r[1] or 0),
            "unique_visitors": int(r[2] or 0),
            "projects_created": proj_dept_map.get(r[0] or "미확인", 0),
            "tasks_created": None,  # tasks.created_by 미존재 → 부서별 집계 불가(전역 카운트만)
            "last_visit_date": r[3] or "-",
            "ratio": round(int(r[1] or 0) * 100.0 / total_dept_visits, 1),
        }
        for r in dept_rows
    ]

    # ── 최근 방문 기록 (IP 마스킹 정책 적용) ──
    show_full_ip = is_super_admin_user(db, user_id)
    recent_rows = (
        db.query(VisitLog)
        .order_by(VisitLog.timestamp.desc())
        .limit(30)
        .all()
    )
    recent = []
    for r in recent_rows:
        ip = r.ip_address or ""
        recent.append({
            "timestamp": r.timestamp.strftime("%Y-%m-%d %H:%M:%S") if r.timestamp else "-",
            "username": r.username or "-",
            "deptname": r.deptname or "미확인",
            "ip_address": ip if show_full_ip else _mask_ip(ip),
        })

    return {
        "summary": summary,
        "daily": daily,
        "departments": departments,
        "recent": recent,
        "ip_masked": not show_full_ip,
        "range_days": days,
        "task_dept_supported": False,  # 부서별 Task 생성 수 미지원(created_by 부재)
    }


# =========================================================
# 공통 페이지네이션 헬퍼 (server-side pagination)
# =========================================================
def _norm_page(page, page_size, *, default_size: int = 20, max_size: int = 100):
    """page/page_size 정규화. page>=1, 1<=page_size<=max_size 로 보정."""
    try:
        page = int(page) if page is not None else 1
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = int(page_size) if page_size is not None else default_size
    except (TypeError, ValueError):
        page_size = default_size
    page = max(1, page)
    page_size = max(1, min(page_size, max_size))
    return page, page_size


def _pagination_meta(page: int, page_size: int, total: int) -> dict:
    total_pages = (total + page_size - 1) // page_size if page_size > 0 else 0
    return {
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": total_pages,
        "has_next": page < total_pages,
        "has_prev": page > 1,
    }


# 정렬 컬럼 화이트리스트 (사용자 입력을 SQL 컬럼에 그대로 넣지 않기 위함)
_USER_SORT_COLUMNS = {
    "id": User.id,
    "username": User.username,
    "login_id": User.loginid,
    "deptname": User.deptname,
    "role": User.role,
    "active": User.is_active,
    "created_at": User.created_at,
    "last_login_at": User.last_login_at,
}


@app.get("/api/admin/users")
def admin_get_users(
    user_id: int = Query(...),
    page: int = Query(1),
    page_size: int = Query(20),
    search: Optional[str] = Query(None),
    role: Optional[str] = Query(None),
    active: Optional[bool] = Query(None),
    sort_by: str = Query("id"),
    sort_order: str = Query("desc"),
    db: Session = Depends(get_db),
):
    """관리자 전용 구성원 목록 (server-side pagination + 검색/역할/활성 필터)."""
    state = load_state()
    require_admin(db, state, user_id)

    page, page_size = _norm_page(page, page_size, default_size=20, max_size=100)

    q = db.query(User)
    if search and search.strip():
        kw = f"%{search.strip()}%"
        q = q.filter(
            (User.username.like(kw))
            | (User.loginid.like(kw))
            | (User.deptname.like(kw))
            | (User.mail.like(kw))
        )
    if role and role.strip():
        q = q.filter(User.role == role.strip())
    if active is not None:
        q = q.filter(User.is_active == bool(active))

    total = q.count()

    sort_col = _USER_SORT_COLUMNS.get((sort_by or "id"), User.id)
    sort_col = sort_col.asc() if (sort_order or "desc").lower() == "asc" else sort_col.desc()
    rows = (
        q.order_by(sort_col, User.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "items": [user_dict(u, state) for u in rows],
        # 기존 호출부 호환용 별칭
        "users": [user_dict(u, state) for u in rows],
        "pagination": _pagination_meta(page, page_size, total),
    }


# =========================================================
# Admin: 시스템 이벤트 로그 (운영 확인용 요약 로그)
# =========================================================
def _parse_detail_json(raw):
    """detail_json(문자열) → dict. 파싱 실패/빈 값이면 None (하위호환: 구 로그는 None)."""
    if not raw:
        return None
    try:
        obj = json.loads(raw)
        return obj if isinstance(obj, dict) else None
    except (json.JSONDecodeError, ValueError, TypeError):
        return None


def _system_event_dict(row: SystemEventLog) -> dict:
    return {
        "id": row.id,
        "level": row.level,
        "category": row.category,
        "message": row.message,
        "detail": row.detail,
        # 구조화 진단(비민감 지표). VISION_AI 로그의 상세 Modal 렌더링용. 구 로그는 None.
        "diagnostic": _parse_detail_json(getattr(row, "detail_json", None)),
        "endpoint": row.endpoint,
        "method": row.method,
        "status_code": row.status_code,
        "user_id": row.user_id,
        "login_id": row.login_id,
        "request_id": row.request_id,
        "resolved": bool(row.resolved),
        "resolved_at": row.resolved_at.isoformat() if getattr(row, "resolved_at", None) else None,
        "resolved_by": getattr(row, "resolved_by", None),
        "resolution_note": getattr(row, "resolution_note", None),
        "created_at": row.created_at.isoformat() if row.created_at else None,
    }


@app.get("/api/admin/system-logs")
def admin_get_system_logs(
    user_id: int = Query(...),
    level: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),  # YYYY-MM-DD
    date_to: Optional[str] = Query(None),    # YYYY-MM-DD
    resolved: Optional[bool] = Query(None),
    keyword: Optional[str] = Query(None),
    endpoint: Optional[str] = Query(None),
    status_code: Optional[int] = Query(None),
    page: int = Query(1),
    page_size: int = Query(50),
    db: Session = Depends(get_db),
):
    """super_admin/admin 전용 시스템 이벤트 로그 조회 (필터 + server-side pagination).

    로그 특성상 page_size 기본 50, 최대 100. 정렬은 항상 created_at desc(최신 우선).
    """
    require_admin(db, None, user_id)

    page, page_size = _norm_page(page, page_size, default_size=50, max_size=100)
    limit = page_size
    offset = (page - 1) * page_size

    q = db.query(SystemEventLog)
    if level:
        q = q.filter(SystemEventLog.level == level.upper())
    if category:
        q = q.filter(SystemEventLog.category == category.upper())
    if endpoint:
        q = q.filter(SystemEventLog.endpoint == endpoint.strip())
    if status_code is not None:
        q = q.filter(SystemEventLog.status_code == int(status_code))
    if resolved is not None:
        q = q.filter(SystemEventLog.resolved == bool(resolved))
    if date_from:
        d = _parse_iso_or_ymd(date_from)
        if d:
            q = q.filter(SystemEventLog.created_at >= datetime(d.year, d.month, d.day))
    if date_to:
        d = _parse_iso_or_ymd(date_to)
        if d:
            # date_to 당일 포함 (다음날 00:00 미만)
            end = datetime(d.year, d.month, d.day) + timedelta(days=1)
            q = q.filter(SystemEventLog.created_at < end)
    if keyword:
        kw = f"%{keyword.strip()}%"
        q = q.filter(
            (SystemEventLog.message.like(kw))
            | (SystemEventLog.detail.like(kw))
            | (SystemEventLog.endpoint.like(kw))
        )

    total = q.count()
    rows = (
        q.order_by(SystemEventLog.created_at.desc(), SystemEventLog.id.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return {
        "items": [_system_event_dict(r) for r in rows],
        "total": total,
        "limit": limit,
        "offset": offset,
        "pagination": _pagination_meta(page, page_size, total),
    }


@app.get("/api/admin/system-logs/recent")
def admin_get_recent_system_logs(
    user_id: int = Query(...),
    limit: int = Query(50),
    db: Session = Depends(get_db),
):
    """최근 이벤트 빠른 조회 (대시보드 카드용)."""
    require_admin(db, None, user_id)
    limit = max(1, min(int(limit or 50), 200))
    rows = (
        db.query(SystemEventLog)
        .order_by(SystemEventLog.created_at.desc(), SystemEventLog.id.desc())
        .limit(limit)
        .all()
    )
    return {"items": [_system_event_dict(r) for r in rows]}


@app.patch("/api/admin/system-logs/{log_id}/resolve")
def admin_resolve_system_log(
    log_id: int,
    user_id: int = Query(...),
    resolved: bool = Query(True),
    db: Session = Depends(get_db),
):
    """이벤트 처리 완료 여부 토글/설정."""
    require_admin(db, None, user_id)
    row = db.query(SystemEventLog).filter(SystemEventLog.id == log_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="로그를 찾을 수 없습니다.")
    _set_log_resolved(row, bool(resolved), user_id)
    db.commit()
    db.refresh(row)
    return _system_event_dict(row)


# ─── 로그 삭제/일괄 처리 (안전장치 포함) ───
class SystemLogFilterSpec(BaseModel):
    level: Optional[str] = None
    category: Optional[str] = None
    endpoint: Optional[str] = None
    status_code: Optional[int] = None
    search: Optional[str] = None          # message/detail/endpoint LIKE 검색
    date_from: Optional[str] = None       # YYYY-MM-DD
    date_to: Optional[str] = None         # YYYY-MM-DD
    resolved: Optional[bool] = None


class SystemLogBulkIds(BaseModel):
    ids: List[int] = Field(default_factory=list)
    reason: Optional[str] = None


class SystemLogBulkResolve(BaseModel):
    ids: List[int] = Field(default_factory=list)
    resolved: bool = True
    resolution_note: Optional[str] = None
    reason: Optional[str] = None


class SystemLogFilterAction(BaseModel):
    filters: SystemLogFilterSpec = Field(default_factory=SystemLogFilterSpec)
    confirm_text: Optional[str] = None
    resolution_note: Optional[str] = None
    reason: Optional[str] = None


def _set_log_resolved(row: SystemEventLog, resolved: bool, user_id: int, note: Optional[str] = None):
    """단일 로그 resolve 상태 + 메타 설정 (commit은 호출부 책임)."""
    row.resolved = bool(resolved)
    if resolved:
        row.resolved_at = datetime.now(KST).replace(tzinfo=None)
        row.resolved_by = int(user_id) if user_id else None
        if note is not None:
            row.resolution_note = (note or "")[:4000]
    else:
        row.resolved_at = None
        row.resolved_by = None
        row.resolution_note = None


def _apply_system_log_filters(q, f: "SystemLogFilterSpec"):
    """SystemLogFilterSpec 를 SystemEventLog 쿼리에 적용. 적용된 필터 개수도 함께 판단 가능."""
    applied = 0
    if f.level:
        q = q.filter(SystemEventLog.level == f.level.upper())
        applied += 1
    if f.category:
        q = q.filter(SystemEventLog.category == f.category.upper())
        applied += 1
    if f.endpoint:
        q = q.filter(SystemEventLog.endpoint == f.endpoint.strip())
        applied += 1
    if f.status_code is not None:
        q = q.filter(SystemEventLog.status_code == int(f.status_code))
        applied += 1
    if f.resolved is not None:
        q = q.filter(SystemEventLog.resolved == bool(f.resolved))
        applied += 1
    if f.date_from:
        d = _parse_iso_or_ymd(f.date_from)
        if d:
            q = q.filter(SystemEventLog.created_at >= datetime(d.year, d.month, d.day))
            applied += 1
    if f.date_to:
        d = _parse_iso_or_ymd(f.date_to)
        if d:
            end = datetime(d.year, d.month, d.day) + timedelta(days=1)
            q = q.filter(SystemEventLog.created_at < end)
            applied += 1
    if f.search and f.search.strip():
        kw = f"%{f.search.strip()}%"
        q = q.filter(
            (SystemEventLog.message.like(kw))
            | (SystemEventLog.detail.like(kw))
            | (SystemEventLog.endpoint.like(kw))
        )
        applied += 1
    return q, applied


def _filter_summary_text(f: "SystemLogFilterSpec") -> str:
    """audit 로그용 필터 요약 문자열 (로그 본문 전체는 저장하지 않음)."""
    parts = []
    if f.level: parts.append(f"level:{f.level}")
    if f.category: parts.append(f"category:{f.category}")
    if f.endpoint: parts.append(f"endpoint:{f.endpoint}")
    if f.status_code is not None: parts.append(f"status:{f.status_code}")
    if f.search: parts.append(f"search:{f.search}")
    if f.date_from: parts.append(f"from:{f.date_from}")
    if f.date_to: parts.append(f"to:{f.date_to}")
    if f.resolved is not None: parts.append(f"resolved:{f.resolved}")
    return ", ".join(parts) if parts else "(none)"


def _audit_admin_log(message: str, detail: str, user_id: int, login_id: Optional[str]):
    """관리자 행위 audit 로그 (category=ADMIN). 삭제된 로그 원문은 저장하지 않는다."""
    try:
        log_event(
            "INFO", "ADMIN", message,
            detail=detail, user_id=user_id, login_id=login_id,
        )
    except Exception:
        pass


@app.delete("/api/admin/system-logs/{log_id}")
def admin_delete_system_log(
    log_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """개별 로그 삭제 (admin 이상)."""
    user = require_admin_and_get_user(db, user_id)
    row = db.query(SystemEventLog).filter(SystemEventLog.id == log_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="로그를 찾을 수 없습니다.")
    db.delete(row)
    db.commit()
    _audit_admin_log(
        "System log deleted",
        f"deleted_count=1, ids=[{log_id}]",
        user_id, user.loginid,
    )
    return {"deleted": 1}


@app.post("/api/admin/system-logs/bulk-delete")
def admin_bulk_delete_system_logs(
    payload: SystemLogBulkIds,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """선택한 로그 일괄 삭제 (admin 이상)."""
    user = require_admin_and_get_user(db, user_id)
    ids = [int(i) for i in (payload.ids or []) if i]
    if not ids:
        raise HTTPException(status_code=400, detail="삭제할 로그를 선택하세요.")
    if len(ids) > 5000:
        raise HTTPException(status_code=400, detail="한 번에 5000건까지만 삭제할 수 있습니다.")
    deleted = (
        db.query(SystemEventLog)
        .filter(SystemEventLog.id.in_(ids))
        .delete(synchronize_session=False)
    )
    db.commit()
    _audit_admin_log(
        "System logs bulk deleted",
        f"deleted_count={int(deleted or 0)}, reason={(payload.reason or '')[:200]}",
        user_id, user.loginid,
    )
    return {"deleted": int(deleted or 0)}


@app.post("/api/admin/system-logs/bulk-resolve")
def admin_bulk_resolve_system_logs(
    payload: SystemLogBulkResolve,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """선택한 로그 일괄 해결/미해결 처리 (admin 이상)."""
    user = require_admin_and_get_user(db, user_id)
    ids = [int(i) for i in (payload.ids or []) if i]
    if not ids:
        raise HTTPException(status_code=400, detail="처리할 로그를 선택하세요.")
    rows = db.query(SystemEventLog).filter(SystemEventLog.id.in_(ids)).all()
    for row in rows:
        _set_log_resolved(row, bool(payload.resolved), user_id, payload.resolution_note)
    db.commit()
    _audit_admin_log(
        "System logs bulk resolved" if payload.resolved else "System logs bulk unresolved",
        f"count={len(rows)}, resolved={payload.resolved}, reason={(payload.reason or '')[:200]}",
        user_id, user.loginid,
    )
    return {"updated": len(rows)}


@app.post("/api/admin/system-logs/filter-resolve")
def admin_filter_resolve_system_logs(
    payload: SystemLogFilterAction,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """현재 필터 조건에 해당하는 로그 일괄 해결 처리 (admin 이상)."""
    user = require_admin_and_get_user(db, user_id)
    q, applied = _apply_system_log_filters(db.query(SystemEventLog), payload.filters)
    if applied == 0:
        raise HTTPException(status_code=400, detail="조건 없는 전체 처리는 허용되지 않습니다. 필터를 1개 이상 지정하세요.")
    rows = q.all()
    for row in rows:
        _set_log_resolved(row, True, user_id, payload.resolution_note)
    db.commit()
    _audit_admin_log(
        "System logs filter-resolved",
        f"count={len(rows)}, filter={_filter_summary_text(payload.filters)}, reason={(payload.reason or '')[:200]}",
        user_id, user.loginid,
    )
    return {"updated": len(rows)}


@app.post("/api/admin/system-logs/filter-delete-preview")
def admin_filter_delete_preview(
    payload: SystemLogFilterAction,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """필터 결과 삭제 전 미리보기: 대상 건수 + 최신/최오래 + 샘플 (admin 이상)."""
    require_admin(db, None, user_id)
    q, applied = _apply_system_log_filters(db.query(SystemEventLog), payload.filters)
    if applied == 0:
        raise HTTPException(status_code=400, detail="조건 없는 전체 미리보기는 허용되지 않습니다. 필터를 1개 이상 지정하세요.")
    matched = q.count()
    oldest = q.order_by(SystemEventLog.created_at.asc(), SystemEventLog.id.asc()).first()
    newest = q.order_by(SystemEventLog.created_at.desc(), SystemEventLog.id.desc()).first()
    sample_rows = (
        q.order_by(SystemEventLog.created_at.desc(), SystemEventLog.id.desc()).limit(5).all()
    )
    return {
        "matched_count": int(matched),
        "oldest": oldest.created_at.isoformat() if oldest and oldest.created_at else None,
        "newest": newest.created_at.isoformat() if newest and newest.created_at else None,
        "sample": [
            {
                "id": r.id,
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "level": r.level,
                "message": r.message,
                "endpoint": r.endpoint,
                "status_code": r.status_code,
                "request_id": r.request_id,
            }
            for r in sample_rows
        ],
    }


@app.post("/api/admin/system-logs/filter-delete")
def admin_filter_delete_system_logs(
    payload: SystemLogFilterAction,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """필터 결과 일괄 삭제 (super_admin 전용, confirm_text='DELETE' 필요)."""
    require_super_admin(db, user_id)
    user = db.query(User).filter(User.id == user_id).first()
    if (payload.confirm_text or "").strip().upper() != "DELETE":
        raise HTTPException(status_code=400, detail="확인을 위해 DELETE 를 입력해야 합니다.")
    q, applied = _apply_system_log_filters(db.query(SystemEventLog), payload.filters)
    if applied == 0:
        raise HTTPException(status_code=400, detail="조건 없는 전체 삭제는 허용되지 않습니다. 필터를 1개 이상 지정하세요.")
    deleted = q.delete(synchronize_session=False)
    db.commit()
    _audit_admin_log(
        "System logs filter-deleted",
        f"deleted_count={int(deleted or 0)}, filter={_filter_summary_text(payload.filters)}, reason={(payload.reason or '')[:200]}",
        user_id, user.loginid if user else None,
    )
    return {"deleted": int(deleted or 0)}


@app.get("/api/admin/system-health")
def admin_get_system_health(user_id: int = Query(...), db: Session = Depends(get_db)):
    """운영 요약: 최근 level별 카운트, 미처리 오류 수, 최근 백업 상태."""
    require_admin(db, None, user_id)

    now = datetime.now(KST).replace(tzinfo=None)
    since_24h = now - timedelta(hours=24)
    since_7d = now - timedelta(days=7)

    def _level_counts(since):
        rows = (
            db.query(SystemEventLog.level, func.count(SystemEventLog.id))
            .filter(SystemEventLog.created_at >= since)
            .group_by(SystemEventLog.level)
            .all()
        )
        out = {"INFO": 0, "WARNING": 0, "ERROR": 0, "CRITICAL": 0}
        for lvl, cnt in rows:
            out[lvl or "ERROR"] = int(cnt)
        return out

    unresolved = (
        db.query(func.count(SystemEventLog.id))
        .filter(
            SystemEventLog.resolved == False,  # noqa: E712
            SystemEventLog.level.in_(["ERROR", "CRITICAL"]),
        )
        .scalar()
    ) or 0

    # ── 백업 상태(BACKUP 카테고리 최신 이벤트에서 파생) ──
    def _latest_backup(message_prefix):
        return (
            db.query(SystemEventLog)
            .filter(
                SystemEventLog.category == "BACKUP",
                SystemEventLog.message.like(f"%{message_prefix}%"),
            )
            .order_by(SystemEventLog.created_at.desc(), SystemEventLog.id.desc())
            .first()
        )

    def _backup_event_view(row):
        if not row:
            return None
        s3_key = None
        if row.detail and row.detail.startswith("s3_key="):
            s3_key = row.detail[len("s3_key="):]
        return {
            "time": row.created_at.isoformat() if row.created_at else None,
            "success": (row.level == "INFO"),
            "message": row.message,
            "s3_key": s3_key,
        }

    last_db_backup = _backup_event_view(_latest_backup("[DB]"))
    last_cleanup = _backup_event_view(_latest_backup("[DB-CLEANUP]"))

    recent_backup_rows = (
        db.query(SystemEventLog)
        .filter(SystemEventLog.category == "BACKUP")
        .order_by(SystemEventLog.created_at.desc(), SystemEventLog.id.desc())
        .limit(15)
        .all()
    )

    return {
        "counts_24h": _level_counts(since_24h),
        "counts_7d": _level_counts(since_7d),
        "unresolved_errors": int(unresolved),
        "last_db_backup": last_db_backup,
        "last_backup_cleanup": last_cleanup,
        "recent_backup_events": [_system_event_dict(r) for r in recent_backup_rows],
        "server_time": now.isoformat(),
    }

@app.patch("/api/admin/users/{target_id}/toggle-active")
def admin_toggle_user_active(target_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    state = load_state()
    require_admin(db, state, user_id)

    target = db.query(User).filter(User.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    # 슈퍼운영자 비활성화 방지
    if is_super_owner_loginid(target.loginid):
        raise HTTPException(status_code=400, detail="Super owner 계정은 비활성화할 수 없습니다.")

    target.is_active = not bool(target.is_active)
    db.commit()
    db.refresh(target)
    return user_dict(target, state)

@app.delete("/api/admin/users/{target_id}")
def admin_delete_user(target_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    """관리자: 사용자 완전 삭제 (hard delete)"""
    state = load_state()
    require_admin(db, state, user_id)
    target = db.query(User).filter(User.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if is_super_owner_loginid(target.loginid):
        raise HTTPException(status_code=400, detail="Super owner 계정은 삭제할 수 없습니다.")

    # tasks.assignee_ids(JSON)에서 제거
    tasks = db.query(Task).filter(Task.archived_at.is_(None)).all()
    for t in tasks:
        ids = list(t.assignee_ids or [])
        if target_id in ids:
            ids.remove(target_id)
            t.assignee_ids = ids

    # preferences / user_shortcuts 삭제
    db.query(UserPreference).filter(UserPreference.user_id == target_id).delete()
    db.query(UserShortcut).filter(UserShortcut.user_id == target_id).delete()

    # group memberships 삭제
    db.query(GroupMembership).filter(GroupMembership.user_id == target_id).delete()

    # project memberships 삭제
    db.query(ProjectMemberModel).filter(ProjectMemberModel.user_id == target_id).delete()

    # sidecar 정리
    state["project_members"] = [m for m in state.get("project_members", []) if int(m.get("user_id", 0)) != target_id]
    state["join_requests"] = [jr for jr in state.get("join_requests", []) if int(jr.get("user_id", 0)) != target_id]
    state["notes"] = [n for n in state.get("notes", []) if int(n.get("author_id", -1)) != target_id]
    state.get("user_meta", {}).pop(str(target_id), None)

    db.delete(target)
    db.commit()
    save_state(state)
    return {"message": "사용자가 삭제되었습니다."}

@app.patch("/api/admin/users/{target_id}/role")
def admin_update_user_role(target_id: int, body: dict, user_id: int = Query(...), db: Session = Depends(get_db)):
    """v1.2: 프론트엔드에서 호출하던 누락 엔드포인트"""
    state = load_state()
    require_admin(db, state, user_id)

    target = db.query(User).filter(User.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    new_role = body.get("role")
    if not new_role:
        raise HTTPException(status_code=400, detail="role is required")

    valid_roles = {"super_admin", "admin", "manager", "member", "viewer"}
    if new_role not in valid_roles:
        raise HTTPException(status_code=400, detail=f"role must be one of {valid_roles}")

    target.role = new_role
    db.commit()
    db.refresh(target)
    return user_dict(target, state)

# =========================
# Org Admin APIs (v1.2)
# =========================
@app.get("/api/admin/org/tree")
def admin_get_org_tree(user_id: int = Query(...), db: Session = Depends(get_db)):
    """조직 트리 조회"""
    state = load_state()
    require_admin(db, state, user_id)

    from app.models import Group as GroupModel
    groups = db.query(GroupModel).filter(GroupModel.is_active == True).order_by(GroupModel.sort_order, GroupModel.id).all()
    result = []
    for g in groups:
        members = db.query(GroupMembership).filter(GroupMembership.group_id == g.id).all()
        member_list = []
        for m in members:
            u = db.query(User).filter(User.id == m.user_id).first()
            if u:
                member_list.append({
                    "user_id": u.id,
                    "username": u.username,
                    "loginid": u.loginid,
                    "org_role": m.org_role,
                    "detail_level": m.detail_level,
                    "is_primary": m.is_primary,
                })
        result.append({
            "id": g.id,
            "name": g.name,
            "description": g.description,
            "group_type": g.group_type,
            "parent_id": g.parent_id,
            "sort_order": g.sort_order,
            "is_active": g.is_active,
            "members": member_list,
        })
    return {"tree": result}

@app.post("/api/admin/org/groups")
def admin_create_org_group(body: dict, user_id: int = Query(...), db: Session = Depends(get_db)):
    """조직 그룹 노드 생성"""
    state = load_state()
    require_admin(db, state, user_id)

    from app.models import Group as GroupModel
    name = body.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    existing = db.query(GroupModel).filter(GroupModel.name == name).first()
    if existing:
        raise HTTPException(status_code=400, detail="이미 존재하는 그룹명입니다")

    new_group = GroupModel(
        name=name,
        description=body.get("description"),
        group_type=body.get("group_type", "PART"),
        parent_id=body.get("parent_id"),
        sort_order=body.get("sort_order", 0),
        is_active=True,
    )
    db.add(new_group)
    db.commit()
    db.refresh(new_group)
    return {"id": new_group.id, "name": new_group.name, "group_type": new_group.group_type}

@app.patch("/api/admin/org/groups/{group_id}")
def admin_update_org_group(group_id: int, body: dict, user_id: int = Query(...), db: Session = Depends(get_db)):
    """조직 그룹 수정"""
    state = load_state()
    require_admin(db, state, user_id)

    from app.models import Group as GroupModel
    g = db.query(GroupModel).filter(GroupModel.id == group_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Group not found")

    for k in ["name", "description", "group_type", "parent_id", "sort_order", "is_active"]:
        if k in body:
            setattr(g, k, body[k])
    db.commit()
    db.refresh(g)
    return {"id": g.id, "name": g.name, "group_type": g.group_type}

@app.delete("/api/admin/org/groups/{group_id}")
def admin_delete_org_group(group_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    """조직 그룹 비활성화"""
    state = load_state()
    require_admin(db, state, user_id)

    from app.models import Group as GroupModel
    g = db.query(GroupModel).filter(GroupModel.id == group_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Group not found")
    g.is_active = False
    db.commit()
    return {"message": "Group deactivated"}

@app.post("/api/admin/org/users/{target_user_id}/assign")
def admin_assign_user_part(target_user_id: int, body: dict, user_id: int = Query(...), db: Session = Depends(get_db)):
    """사용자 파트 배정 (group_id, org_role, detail_level)"""
    state = load_state()
    require_admin(db, state, user_id)

    target = db.query(User).filter(User.id == target_user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    group_id = body.get("group_id")
    org_role = body.get("org_role", "MEMBER")
    detail_level = body.get("detail_level", "FULL_DETAIL")

    if group_id:
        existing = db.query(GroupMembership).filter(
            GroupMembership.user_id == target_user_id,
            GroupMembership.group_id == group_id
        ).first()

        if existing:
            existing.org_role = org_role
            existing.detail_level = detail_level
        else:
            new_membership = GroupMembership(
                user_id=target_user_id,
                group_id=group_id,
                org_role=org_role,
                detail_level=detail_level,
                is_primary=True,
            )
            db.add(new_membership)

        # Update user's primary_part_id
        from app.models import Group as GroupModel
        grp = db.query(GroupModel).filter(GroupModel.id == group_id).first()
        if grp and grp.group_type == "PART":
            target.primary_part_id = group_id
        elif grp and grp.group_type == "TEAM":
            target.primary_team_id = group_id

    db.commit()
    return {"message": "User assigned", "user_id": target_user_id, "group_id": group_id}

@app.post("/api/admin/org/projects/{project_id}/assign-part")
def admin_assign_project_part(project_id: int, body: dict, user_id: int = Query(...), db: Session = Depends(get_db)):
    """프로젝트 파트 배정"""
    state = load_state()
    require_admin(db, state, user_id)

    p = db.query(Project).filter(Project.id == project_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    part_id = body.get("part_id")
    p.part_id = part_id
    db.commit()
    return {"message": "Project part assigned", "project_id": project_id, "part_id": part_id}

@app.get("/api/admin/org/unassigned-users")
def admin_get_unassigned_users(user_id: int = Query(...), db: Session = Depends(get_db)):
    """PART 배정 안 된 사용자 목록"""
    state = load_state()
    require_admin(db, state, user_id)

    all_users = db.query(User).filter(User.is_active == True).all()
    assigned_user_ids = set(
        m.user_id for m in db.query(GroupMembership).all()
    )
    unassigned = [user_dict(u, state) for u in all_users if u.id not in assigned_user_ids]
    return {"users": unassigned}


# =========================
# Groups (sidecar)
# - 레거시 /api/groups + 신규 /api/admin/groups 둘 다 지원
# =========================
def _group_to_out(g: dict, db: Session, state: dict) -> dict:
    all_users = db.query(User).all()
    group_name = g.get("name", "")
    # Match by group_name (user_meta) OR deptname
    matched_count = 0
    for u in all_users:
        umeta = get_user_meta(state, u.id)
        if umeta.get("group_name") == group_name or (u.deptname and u.deptname.strip() == group_name):
            matched_count += 1
    return {**g, "matched_count": matched_count}

@app.get("/api/groups")
def get_groups(db: Session = Depends(get_db)):
    state = load_state()
    groups = state.get("groups", [])
    return {"groups": [_group_to_out(g, db, state) for g in groups]}

@app.post("/api/groups")
def create_group_legacy(group: GroupCreate, db: Session = Depends(get_db)):
    state = load_state()
    groups = state.get("groups", [])
    if any(g.get("name") == group.name for g in groups):
        raise HTTPException(status_code=400, detail="Group name already exists")

    new_group = {
        "id": next_id(groups),
        "name": group.name,
        "description": group.description,
        "is_active": True,
        "created_at": datetime.now().isoformat(),
    }
    groups.append(new_group)
    state["groups"] = groups
    save_state(state)
    return _group_to_out(new_group, db, state)

@app.patch("/api/groups/{group_id}")
def update_group_legacy(group_id: int, body: GroupUpdate, db: Session = Depends(get_db)):
    state = load_state()
    groups = state.get("groups", [])
    for i, g in enumerate(groups):
        if int(g.get("id")) == group_id:
            groups[i].update(body.model_dump(exclude_unset=True))
            state["groups"] = groups
            save_state(state)
            return _group_to_out(groups[i], db, state)
    raise HTTPException(status_code=404, detail="Group not found")

@app.get("/api/admin/groups")
def admin_get_groups(user_id: int = Query(...), db: Session = Depends(get_db)):
    state = load_state()
    require_admin(db, state, user_id)
    return get_groups(db)

@app.post("/api/admin/groups")
def admin_create_group(group: GroupCreate, user_id: int = Query(...), db: Session = Depends(get_db)):
    state = load_state()
    require_admin(db, state, user_id)

    groups = state.get("groups", [])
    if any(g.get("name") == group.name for g in groups):
        raise HTTPException(status_code=400, detail=f"그룹명 '{group.name}'이(가) 이미 등록되어 있습니다.")

    new_group = {
        "id": next_id(groups),
        "name": group.name,
        "description": group.description,
        "is_active": True,
        "created_at": datetime.now().isoformat(),
    }
    groups.append(new_group)
    state["groups"] = groups
    save_state(state)
    return _group_to_out(new_group, db, state)

@app.delete("/api/admin/groups/{group_id}")
def admin_delete_group(group_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    state = load_state()
    require_admin(db, state, user_id)
    state["groups"] = [g for g in state.get("groups", []) if int(g.get("id")) != group_id]
    save_state(state)
    return {"message": "Group deleted"}

@app.post("/api/admin/groups/{group_id}/apply")
def admin_apply_group(group_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    state = load_state()
    require_admin(db, state, user_id)

    groups = state.get("groups", [])
    group = next((g for g in groups if int(g.get("id")) == group_id), None)
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    target_group_name = group.get("name")
    activated = 0

    users = db.query(User).all()
    for u in users:
        umeta = get_user_meta(state, u.id)
        # Match by group_name (user_meta) OR deptname
        if umeta.get("group_name") == target_group_name or (u.deptname and u.deptname.strip() == target_group_name):
            if not bool(u.is_active):
                u.is_active = True
                activated += 1

    db.commit()

    total_matched = 0
    for u in users:
        umeta = get_user_meta(state, u.id)
        if umeta.get("group_name") == target_group_name or (u.deptname and u.deptname.strip() == target_group_name):
            total_matched += 1

    return {
        "message": f"Activated {activated} users (total matched: {total_matched})",
        "activated": activated,
        "total_matched": total_matched,
    }


# =========================
# Member Groups (DB)
# =========================
def get_or_create_user_from_knox(db: Session, knox_user: dict) -> Optional[User]:
    """Knox에서 가져온 사용자 정보로 users 테이블을 upsert.
    - loginid 기준 중복 생성 방지.
    - 이미 있으면 기존 user 반환 (비활성이면 활성화하지 않음 — 기존 정책 유지).
    - 없으면 새 user 생성 (role='member', is_active=True).
    knox_user는 {loginid, name, email, deptname} 형태."""
    loginid = (knox_user.get("loginid") or knox_user.get("userId") or "").strip()
    if not loginid:
        return None
    existing = db.query(User).filter(User.loginid == loginid).first()
    if existing:
        # 이미 등록된 사용자: 빈 필드만 보강 (정책: 기존 데이터 우선)
        name = knox_user.get("name") or knox_user.get("fullName")
        dept = knox_user.get("deptname") or knox_user.get("departmentName")
        mail = knox_user.get("email") or knox_user.get("mail")
        changed = False
        if name and not existing.username:
            existing.username = name
            changed = True
        if dept and not existing.deptname:
            existing.deptname = dept
            changed = True
        if mail and not existing.mail:
            existing.mail = mail
            changed = True
        if changed:
            db.flush()
        return existing
    name = (knox_user.get("name") or knox_user.get("fullName") or loginid).strip()
    dept = knox_user.get("deptname") or knox_user.get("departmentName")
    mail = knox_user.get("email") or knox_user.get("mail") or f"{loginid}@samsung.com"
    new_user = User(
        loginid=loginid,
        username=name,
        deptname=dept,
        mail=mail,
        role="member",
        is_active=True,
    )
    db.add(new_user)
    db.flush()
    return new_user


def _member_group_dict(g: MemberGroup, db: Session) -> dict:
    members = db.query(MemberGroupUser).filter(MemberGroupUser.group_id == g.id).all()
    user_ids = [m.user_id for m in members]
    users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
    return {
        "id": g.id,
        "name": g.name,
        "description": g.description,
        "created_by": g.created_by,
        "created_at": g.created_at.isoformat() if g.created_at else None,
        "member_count": len(user_ids),
        "members": [
            {
                "user_id": u.id,
                "username": u.username,
                "loginid": u.loginid,
                "avatar_color": u.avatar_color,
                "deptname": getattr(u, "deptname", None) or getattr(u, "group_name", None),
            }
            for u in users
        ],
    }

@app.get("/api/member-groups")
def get_member_groups(user_id: int = Query(...), db: Session = Depends(get_db)):
    groups = db.query(MemberGroup).filter(MemberGroup.created_by == user_id).order_by(MemberGroup.created_at.desc()).all()
    return {"groups": [_member_group_dict(g, db) for g in groups]}

@app.post("/api/member-groups")
def create_member_group(body: MemberGroupCreate, user_id: int = Query(...), db: Session = Depends(get_db)):
    g = MemberGroup(name=body.name, description=body.description, created_by=user_id)
    db.add(g)
    db.flush()
    final_ids: set = set()
    for uid in (body.member_user_ids or []):
        final_ids.add(uid)
    # Knox 사용자 upsert
    for kx in (body.knox_members or []):
        u = get_or_create_user_from_knox(db, kx.dict() if hasattr(kx, "dict") else dict(kx))
        if u:
            final_ids.add(u.id)
    for uid in final_ids:
        # 중복 방지 (UniqueConstraint가 있지만 사전 차단)
        exists = db.query(MemberGroupUser).filter(
            MemberGroupUser.group_id == g.id, MemberGroupUser.user_id == uid
        ).first()
        if not exists:
            db.add(MemberGroupUser(group_id=g.id, user_id=uid))
    db.commit()
    db.refresh(g)
    return _member_group_dict(g, db)

@app.patch("/api/member-groups/{group_id}")
def update_member_group(group_id: int, body: MemberGroupUpdate, user_id: int = Query(...), db: Session = Depends(get_db)):
    g = db.query(MemberGroup).filter(MemberGroup.id == group_id).first()
    if not g:
        raise HTTPException(404, "Group not found")
    if body.name is not None:
        g.name = body.name
    if body.description is not None:
        g.description = body.description
    # member_user_ids 가 명시되어 있으면 전체 교체. knox_members만 들어오면 추가.
    if body.member_user_ids is not None:
        db.query(MemberGroupUser).filter(MemberGroupUser.group_id == group_id).delete()
        merged: set = set(body.member_user_ids)
        for kx in (body.knox_members or []):
            u = get_or_create_user_from_knox(db, kx.dict() if hasattr(kx, "dict") else dict(kx))
            if u:
                merged.add(u.id)
        for uid in merged:
            db.add(MemberGroupUser(group_id=group_id, user_id=uid))
    elif body.knox_members:
        # knox_members만 추가 (기존 목록 유지)
        for kx in body.knox_members:
            u = get_or_create_user_from_knox(db, kx.dict() if hasattr(kx, "dict") else dict(kx))
            if not u:
                continue
            exists = db.query(MemberGroupUser).filter(
                MemberGroupUser.group_id == group_id, MemberGroupUser.user_id == u.id
            ).first()
            if not exists:
                db.add(MemberGroupUser(group_id=group_id, user_id=u.id))
    db.commit()
    db.refresh(g)
    return _member_group_dict(g, db)

@app.delete("/api/member-groups/{group_id}")
def delete_member_group(group_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    g = db.query(MemberGroup).filter(MemberGroup.id == group_id).first()
    if not g:
        raise HTTPException(404, "Group not found")
    db.query(MemberGroupUser).filter(MemberGroupUser.group_id == group_id).delete()
    db.delete(g)
    db.commit()
    return {"message": "Group deleted"}


async def _knox_search_safe(query: str) -> list:
    """Knox 검색 호출 (실패해도 빈 배열 반환)."""
    q = (query or "").strip()
    if not q:
        return []
    try:
        from app.services.knox_client import knox_search_employees
        is_id_like = bool(re.match(r'^[a-zA-Z0-9._\-]+$', q))
        if is_id_like:
            employees = await knox_search_employees(userIds=q)
            if not employees:
                employees = await knox_search_employees(fullName=q)
        else:
            employees = await knox_search_employees(fullName=q)
            if not employees:
                employees = await knox_search_employees(userIds=q)
        seen = set()
        deduped = []
        for e in employees or []:
            uid = e.get("userId") or e.get("loginid")
            if uid and uid not in seen:
                seen.add(uid)
                deduped.append(e)
        return deduped
    except Exception:
        return []


@app.get("/api/member-groups/member-candidates")
async def get_group_member_candidates(
    q: str = Query(..., min_length=1),
    user_id: int = Query(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    """그룹 생성/수정 모달에서 사용할 통합 검색 (pagination 지원).

    - 사이트 users 검색: page/page_size 로 누적 탐색 가능 (같은 이름이 많아도 끝까지 볼 수 있음).
    - Knox 사내 검색: paging 미지원이라 page 1 에서만 보조적으로 덧붙인다(중복 loginid 제거).
    - pagination 메타는 사이트 users 기준으로 계산한다.
    - 결과는 {items: [{type, source, ...}], pagination: {...}} 형태."""
    # 권한 가드: 유효한(활성) 사용자만 디렉토리 검색 허용
    caller = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not caller:
        raise HTTPException(403, "사용자 검색 권한이 없습니다.")
    items: list = []
    search_q = f"%{q.strip()}%"
    base = db.query(User).filter(
        User.is_active == True,
        (User.username.ilike(search_q) | User.loginid.ilike(search_q)
         | User.deptname.ilike(search_q) | User.mail.ilike(search_q))
    )
    total = base.count()
    offset = (page - 1) * page_size
    site_users = (
        base.order_by(User.username.asc(), User.id.asc())
        .offset(offset).limit(page_size).all()
    )
    site_loginids = set()
    for u in site_users:
        items.append({
            "type": "user",
            "source": "site",
            "user_id": u.id,
            "login_id": u.loginid,
            "name": u.username,
            "email": u.mail,
            "department": u.deptname,
            "avatar_color": u.avatar_color,
        })
        site_loginids.add(u.loginid)
    # Knox 검색 (실패 안전) — paging 불가하므로 첫 페이지에서만 보조적으로 노출
    if page == 1:
        knox_employees = await _knox_search_safe(q)
        for e in knox_employees[:50]:
            loginid = e.get("userId") or e.get("loginid")
            if not loginid or loginid in site_loginids:
                continue
            site_loginids.add(loginid)
            items.append({
                "type": "user",
                "source": "knox",
                "login_id": loginid,
                "name": e.get("fullName") or loginid,
                "email": e.get("mail") or e.get("email"),
                "department": e.get("departmentName") or e.get("deptname"),
            })
    total_pages = max(1, (total + page_size - 1) // page_size)
    has_next = page < total_pages
    return {
        "items": items,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "has_next": has_next,
        },
    }


# =========================
# Spaces
# =========================

class SpaceCreate(BaseModel):
    name: str
    slug: Optional[str] = None
    description: Optional[str] = None
    member_user_ids: Optional[List[int]] = None
    # 아직 사이트 구성원이 아닌 Knox/사내 디렉토리 사용자를 초대할 때 사용.
    # create_space 에서 users 테이블로 upsert(get_or_create_user_from_knox) 후 멤버 추가한다.
    knox_members: Optional[List[KnoxUserPayload]] = None
    purpose: Optional[str] = "project_management"
    # project_management / equipment_ops / process_change / sw_dev / integrated_ops / custom

class SpaceUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    purpose: Optional[str] = None

def _space_dict(s, db: Session) -> dict:
    members = db.query(SpaceMember, User).join(User, SpaceMember.user_id == User.id).filter(SpaceMember.space_id == s.id).all()
    project_count = db.query(Project).filter(Project.space_id == s.id, Project.archived_at.is_(None)).count()
    return {
        "id": s.id,
        "name": s.name,
        "slug": s.slug,
        "description": s.description,
        "created_by": s.created_by,
        "is_active": s.is_active,
        "created_at": iso(s.created_at) if s.created_at else None,
        "warned_at": iso(s.warned_at) if s.warned_at else None,
        "archived_at": iso(s.archived_at) if getattr(s, "archived_at", None) else None,
        "delete_scheduled_at": iso(s.delete_scheduled_at) if getattr(s, "delete_scheduled_at", None) else None,
        "cleanup_exempt": bool(getattr(s, "cleanup_exempt", False)),
        "last_activity_at": iso(s.last_activity_at) if getattr(s, "last_activity_at", None) else None,
        "purpose": getattr(s, "purpose", None) or "project_management",
        "project_count": project_count,
        "member_count": len(members),
        "members": [
            {"user_id": m.user_id, "role": m.role, "username": u.username, "loginid": u.loginid, "avatar_color": u.avatar_color}
            for m, u in members
        ],
    }

import re as _re_slug

@app.get("/api/spaces")
def get_spaces(user_id: int = Query(...), db: Session = Depends(get_db)):
    """List spaces the user is a member of."""
    member_space_ids = [sm.space_id for sm in db.query(SpaceMember).filter(SpaceMember.user_id == user_id).all()]
    spaces = db.query(Space).filter(Space.id.in_(member_space_ids), Space.is_active == True).order_by(Space.created_at).all()
    return {"spaces": [_space_dict(s, db) for s in spaces]}

@app.get("/api/spaces/all")
def get_all_spaces(user_id: int = Query(...), db: Session = Depends(get_db)):
    """List ALL active spaces with is_member flag for the user."""
    all_spaces = db.query(Space).filter(Space.is_active == True).order_by(Space.created_at).all()
    member_space_ids = set(sm.space_id for sm in db.query(SpaceMember).filter(SpaceMember.user_id == user_id).all())
    # Get user's role in each space
    user_memberships = {sm.space_id: sm.role for sm in db.query(SpaceMember).filter(SpaceMember.user_id == user_id).all()}
    # 사용자가 참여 요청을 보낸(승인 대기) 공간 — 정렬 우선순위 계산용
    pending_space_ids = set(
        jr.space_id for jr in db.query(SpaceJoinRequest).filter(
            SpaceJoinRequest.user_id == user_id,
            SpaceJoinRequest.status == "pending",
        ).all()
    )
    result = []
    for s in all_spaces:
        d = _space_dict(s, db)
        d["is_member"] = s.id in member_space_ids
        d["my_role"] = user_memberships.get(s.id)
        d["pending_request"] = s.id in pending_space_ids
        result.append(d)
    return {"spaces": result}

@app.get("/api/users/search")
def search_users(q: str = Query(..., min_length=1), user_id: int = Query(...), db: Session = Depends(get_db)):
    """Search users by name or loginid (for space member addition)."""
    search_q = f"%{q}%"
    users_found = db.query(User).filter(
        User.is_active == True,
        (User.username.ilike(search_q) | User.loginid.ilike(search_q))
    ).limit(20).all()
    return {"users": [
        {"id": u.id, "username": u.username, "loginid": u.loginid, "avatar_color": u.avatar_color, "deptname": u.deptname, "role": u.role}
        for u in users_found
    ]}

def _space_limits(db: Session, user_id: int) -> dict:
    """사용자가 '직접 생성/소유'한 active 공간 기준의 생성 가능 한도 계산.

    - 초대받은 공간/보관(archived)/비활성 공간은 카운트에서 제외한다.
    - super_admin: 무제한, admin: MAX_ACTIVE_SPACES_PER_ADMIN, 그 외: MAX_ACTIVE_SPACES_PER_USER.
    - 일일 제한은 최근 24시간 동안 본인이 생성한 공간 수 기준(타임존 안전, 스팸 방지).
    """
    role = get_user_role(db, user_id)
    if role == "super_admin":
        limit = None  # 무제한
    elif role == "admin":
        limit = MAX_ACTIVE_SPACES_PER_ADMIN
    else:
        limit = MAX_ACTIVE_SPACES_PER_USER

    created = db.query(func.count(Space.id)).filter(
        Space.created_by == user_id,
        Space.is_active == True,
        Space.archived_at.is_(None),
    ).scalar() or 0

    since = datetime.utcnow() - timedelta(hours=24)
    per_day_used = db.query(func.count(Space.id)).filter(
        Space.created_by == user_id,
        Space.created_at >= since,
    ).scalar() or 0
    per_day_limit = None if role in ("admin", "super_admin") else MAX_SPACE_CREATE_PER_DAY

    remaining = None if limit is None else max(0, limit - int(created))
    can_create = (limit is None or created < limit) and (per_day_limit is None or per_day_used < per_day_limit)
    near_limit = bool(remaining is not None and remaining <= 1)
    return {
        "role": role,
        "created": int(created),
        "limit": limit,
        "remaining": remaining,
        "per_day_used": int(per_day_used),
        "per_day_limit": per_day_limit,
        "can_create": can_create,
        "near_limit": near_limit,
    }


@app.get("/api/spaces/limits")
def get_space_limits(user_id: int = Query(...), db: Session = Depends(get_db)):
    """공간 생성 화면용: 내가 생성한 공간 수 / 한도 / 생성 가능 여부."""
    return _space_limits(db, user_id)


def _normalize_space_name(s: str) -> str:
    """중복 공간명 비교용 정규화: 공백/하이픈/언더스코어 제거 + 소문자."""
    import re
    return re.sub(r"[\s\-_]+", "", (s or "")).lower()


@app.get("/api/spaces/name-check")
def check_space_name(name: str = Query(...), user_id: int = Query(...), db: Session = Depends(get_db)):
    """비슷한 이름의 활성 공간이 이미 있는지 확인(생성 차단 아님, 경고용).

    정규화(공백/하이픈/언더스코어 제거+소문자) 후 완전 일치 또는 부분 포함이면 유사로 본다.
    """
    target = _normalize_space_name(name)
    if len(target) < 2:
        return {"similar": []}

    member_ids = set(
        sm.space_id for sm in db.query(SpaceMember).filter(SpaceMember.user_id == user_id).all()
    )
    spaces = db.query(Space).filter(Space.is_active == True).order_by(Space.created_at).all()
    similar = []
    for s in spaces:
        n = _normalize_space_name(s.name)
        if not n:
            continue
        if n == target or target in n or n in target:
            similar.append({
                "id": s.id,
                "name": s.name,
                "slug": s.slug,
                "is_member": s.id in member_ids,
                "exact": n == target,
            })
        if len(similar) >= 5:
            break
    # 완전 일치를 먼저 보이도록 정렬
    similar.sort(key=lambda x: (not x["exact"], x["name"]))
    return {"similar": similar}


@app.post("/api/spaces")
def create_space(body: SpaceCreate, user_id: int = Query(...), db: Session = Depends(get_db)):
    # 생성 한도 검사 (생성/소유 기준, 초대 공간 제외)
    limits = _space_limits(db, user_id)
    if not limits["can_create"]:
        if limits["limit"] is not None and limits["created"] >= limits["limit"]:
            raise HTTPException(
                403,
                f"생성 가능한 공간 수를 초과했습니다. 현재 최대 {limits['limit']}개의 공간만 생성할 수 있습니다. "
                f"추가 공간이 필요하면 관리자에게 문의해주세요.",
            )
        if limits["per_day_limit"] is not None and limits["per_day_used"] >= limits["per_day_limit"]:
            raise HTTPException(
                429,
                f"하루 공간 생성 한도를 초과했습니다. 최근 24시간 동안 최대 {limits['per_day_limit']}개까지 생성할 수 있습니다. "
                f"잠시 후 다시 시도하거나 관리자에게 문의해주세요.",
            )
        raise HTTPException(403, "공간을 생성할 수 없습니다.")

    # Use name as slug directly (supports Korean), or custom slug if provided
    slug = body.slug or _re_slug.sub(r'\s+', '-', body.name.strip())[:100]
    # Ensure unique slug
    existing = db.query(Space).filter(Space.slug == slug).first()
    if existing:
        slug = f"{slug}-{int(datetime.utcnow().timestamp()) % 10000}"
    space = Space(name=body.name, slug=slug, description=body.description, created_by=user_id, purpose=body.purpose or "project_management")
    db.add(space)
    db.commit()
    db.refresh(space)
    # Add creator as owner
    db.add(SpaceMember(space_id=space.id, user_id=user_id, role="owner"))
    # 추가 멤버 user id 집합 (local + Knox upsert 결과 병합, login_id 기준 중복은 user.id로 자연 dedupe)
    member_ids: set = set()
    for uid in (body.member_user_ids or []):
        member_ids.add(uid)
    # 아직 사이트 구성원이 아닌 Knox 사용자는 users 테이블로 upsert 후 멤버로 추가
    for kx in (body.knox_members or []):
        u = get_or_create_user_from_knox(db, kx.dict() if hasattr(kx, "dict") else dict(kx))
        if u:
            member_ids.add(u.id)
    for uid in member_ids:
        if uid != user_id:
            exists = db.query(SpaceMember).filter(
                SpaceMember.space_id == space.id, SpaceMember.user_id == uid
            ).first()
            if not exists:
                db.add(SpaceMember(space_id=space.id, user_id=uid, role="member"))
    db.commit()
    return _space_dict(space, db)

@app.get("/api/spaces/{space_id}")
def get_space(space_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    space = db.query(Space).filter(Space.id == space_id, Space.is_active == True).first()
    if not space:
        raise HTTPException(404, "Space not found")
    member = db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.user_id == user_id).first()
    if not member:
        raise HTTPException(403, "이 공간에 접근 권한이 없습니다")
    return _space_dict(space, db)

def _require_space_member(db: Session, space_id: int, user_id: int):
    """공간 멤버(또는 admin/super_admin)인지 확인. 단발 일정/캘린더 이벤트 생성·수정 권한."""
    if is_admin_like_role(get_user_role(db, user_id)):
        return None
    m = db.query(SpaceMember).filter(
        SpaceMember.space_id == space_id,
        SpaceMember.user_id == user_id,
    ).first()
    if not m:
        raise HTTPException(403, "이 공간에 접근 권한이 없습니다.")
    return m

def _ensure_assignee_in_space(db: Session, space_id: int, assignee_id):
    """담당자(assignee_id)가 해당 공간의 멤버인지 검증. 멤버가 아니면 400.

    assignee_id 가 None/빈값이면 '미지정'으로 보고 통과시킨다.
    프론트 필터링과 무관하게 API 직접 호출도 막기 위한 서버측 가드.
    """
    if assignee_id in (None, "", 0):
        return
    exists = db.query(SpaceMember).filter(
        SpaceMember.space_id == space_id,
        SpaceMember.user_id == assignee_id,
    ).first()
    if not exists:
        raise HTTPException(400, "담당자는 해당 공간의 멤버여야 합니다.")

def _require_space_admin(db: Session, space_id: int, user_id: int):
    """Check that user is owner, admin, or operator of the space."""
    m = db.query(SpaceMember).filter(
        SpaceMember.space_id == space_id,
        SpaceMember.user_id == user_id,
        SpaceMember.role.in_(["owner", "admin", "operator"]),
    ).first()
    if not m:
        raise HTTPException(403, "공간 소유자 또는 관리자만 이 작업을 수행할 수 있습니다")
    return m

@app.patch("/api/spaces/{space_id}")
def update_space(space_id: int, body: SpaceUpdate, user_id: int = Query(...), db: Session = Depends(get_db)):
    space = db.query(Space).filter(Space.id == space_id).first()
    if not space:
        raise HTTPException(404, "Space not found")
    _require_space_admin(db, space_id, user_id)
    if body.name is not None:
        space.name = body.name
        # Sync slug with name change
        new_slug = _re_slug.sub(r'\s+', '-', body.name.strip())[:100]
        existing = db.query(Space).filter(Space.slug == new_slug, Space.id != space_id).first()
        if existing:
            new_slug = f"{new_slug}-{int(datetime.utcnow().timestamp()) % 10000}"
        space.slug = new_slug
    if body.description is not None:
        space.description = body.description
    if body.purpose is not None:
        space.purpose = body.purpose
    db.commit()
    return _space_dict(space, db)

@app.delete("/api/spaces/{space_id}")
def delete_space(space_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    """Delete (deactivate) a space. Only owner can delete."""
    space = db.query(Space).filter(Space.id == space_id).first()
    if not space:
        raise HTTPException(404, "Space not found")
    owner = db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.user_id == user_id, SpaceMember.role == "owner").first()
    if not owner:
        raise HTTPException(403, "공간 소유자만 삭제할 수 있습니다")
    space.is_active = False
    space.archived_at = datetime.utcnow()
    db.commit()
    return {"message": "공간이 삭제되었습니다"}


# =========================
# Admin: 공간 사용 현황 / 빈 공간 관리
# =========================
@app.get("/api/admin/spaces")
def admin_list_spaces(
    user_id: int = Query(...),
    filter: str = Query("all"),   # all / no_project / empty / inactive_7d / archived / by_creator / by_purpose
    q: Optional[str] = None,
    creator_id: Optional[int] = None,
    purpose: Optional[str] = None,
    page: int = Query(1),
    page_size: int = Query(20),
    db: Session = Depends(get_db),
):
    """관리자용 전 공간 사용 현황 (server-side pagination).

    필터/검색/정렬은 DB 단계에서 적용하고(빈 공간/프로젝트 없음 판정은 EXISTS 서브쿼리),
    무거운 집계(프로젝트/Task/시트/캘린더/멤버 카운트)는 현재 page 대상 공간에 대해서만 계산한다.
    VOC 는 공간에 묶이지 않는 전역 데이터라 공간별 카운트에서 제외(N/A).
    """
    require_admin(db, None, user_id)
    now = datetime.utcnow()
    page, page_size = _norm_page(page, page_size, default_size=20, max_size=100)

    # ── EXISTS 서브쿼리 (Space.id 에 상관 correlate) — 빈/무프로젝트 판정을 DB 단계에서 ──
    proj_exists = (
        db.query(Project.id)
        .filter(
            Project.space_id == Space.id,
            Project.archived_at.is_(None),
            Project.name != ONEOFF_PROJECT_NAME,
        )
        .exists()
    )
    task_exists = (
        db.query(Task.id)
        .join(Project, Task.project_id == Project.id)
        .filter(Project.space_id == Space.id, Task.archived_at.is_(None))
        .exists()
    )
    sheet_exists = (
        db.query(SheetExecution.id).filter(SheetExecution.space_id == Space.id).exists()
    )
    cal_exists = (
        db.query(CalendarEvent.id)
        .filter(CalendarEvent.space_id == Space.id, CalendarEvent.archived_at.is_(None))
        .exists()
    )
    empty_conds = (~proj_exists, ~task_exists, ~sheet_exists, ~cal_exists)
    inactive_cutoff = now - timedelta(days=7)

    def apply_filter(query):
        if filter == "no_project":
            return query.filter(Space.is_active.is_(True), ~proj_exists)
        if filter == "empty":
            return query.filter(Space.is_active.is_(True), *empty_conds)
        if filter == "inactive_7d":
            return query.filter(
                Space.is_active.is_(True), *empty_conds,
                Space.created_at.isnot(None), Space.created_at <= inactive_cutoff,
            )
        if filter == "archived":
            return query.filter(Space.is_active.is_(False))
        if filter == "by_creator":
            if creator_id is None:
                return query.filter(sql_false())
            return query.filter(Space.created_by == creator_id)
        if filter == "by_purpose":
            if purpose is None:
                return query.filter(sql_false())
            return query.filter(Space.purpose == purpose)
        # all: 활성 공간만 (보관/삭제 공간은 archived 필터로)
        return query.filter(Space.is_active.is_(True))

    base = apply_filter(db.query(Space))
    if q and q.strip():
        base = base.filter(Space.name.like(f"%{q.strip()}%"))

    total = base.count()
    spaces = (
        base.order_by(Space.created_at.desc(), Space.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    page_ids = [s.id for s in spaces]

    # ── 현재 page 공간에 대해서만 집계 (N+1 회피, 전수 계산 회피) ──
    def _grp(rows):
        return {r[0]: int(r[1]) for r in rows if r[0] is not None}

    def _grp_max(rows):
        return {r[0]: r[1] for r in rows if r[0] is not None}

    if page_ids:
        project_counts = _grp(
            db.query(Project.space_id, func.count(Project.id))
            .filter(Project.archived_at.is_(None), Project.name != ONEOFF_PROJECT_NAME, Project.space_id.in_(page_ids))
            .group_by(Project.space_id).all()
        )
        task_counts = _grp(
            db.query(Project.space_id, func.count(Task.id))
            .join(Task, Task.project_id == Project.id)
            .filter(Task.archived_at.is_(None), Project.space_id.in_(page_ids))
            .group_by(Project.space_id).all()
        )
        checksheet_counts = _grp(
            db.query(SheetExecution.space_id, func.count(SheetExecution.id))
            .filter(SheetExecution.space_id.in_(page_ids))
            .group_by(SheetExecution.space_id).all()
        )
        calendar_counts = _grp(
            db.query(CalendarEvent.space_id, func.count(CalendarEvent.id))
            .filter(CalendarEvent.archived_at.is_(None), CalendarEvent.space_id.in_(page_ids))
            .group_by(CalendarEvent.space_id).all()
        )
        member_counts = _grp(
            db.query(SpaceMember.space_id, func.count(SpaceMember.id))
            .filter(SpaceMember.space_id.in_(page_ids))
            .group_by(SpaceMember.space_id).all()
        )
        proj_max = _grp_max(
            db.query(Project.space_id, func.max(Project.created_at))
            .filter(Project.space_id.in_(page_ids)).group_by(Project.space_id).all()
        )
        task_max = _grp_max(
            db.query(Project.space_id, func.max(Task.updated_at))
            .join(Task, Task.project_id == Project.id)
            .filter(Project.space_id.in_(page_ids)).group_by(Project.space_id).all()
        )
        sheet_max = _grp_max(
            db.query(SheetExecution.space_id, func.max(SheetExecution.started_at))
            .filter(SheetExecution.space_id.in_(page_ids)).group_by(SheetExecution.space_id).all()
        )
        cal_max = _grp_max(
            db.query(CalendarEvent.space_id, func.max(CalendarEvent.updated_at))
            .filter(CalendarEvent.space_id.in_(page_ids)).group_by(CalendarEvent.space_id).all()
        )
        owner_uids = {s.created_by for s in spaces if s.created_by}
        owners = {}
        for sm in db.query(SpaceMember).filter(SpaceMember.space_id.in_(page_ids), SpaceMember.role == "owner").all():
            owners.setdefault(sm.space_id, sm.user_id)
        owner_uids.update(owners.values())
        users = {u.id: u for u in db.query(User).filter(User.id.in_(owner_uids)).all()} if owner_uids else {}
    else:
        project_counts = task_counts = checksheet_counts = calendar_counts = member_counts = {}
        proj_max = task_max = sheet_max = cal_max = {}
        owners = {}
        users = {}

    result = []
    for s in spaces:
        pc = project_counts.get(s.id, 0)
        tc = task_counts.get(s.id, 0)
        cc = checksheet_counts.get(s.id, 0)
        calc = calendar_counts.get(s.id, 0)
        mc = member_counts.get(s.id, 0)
        is_empty = (pc == 0 and tc == 0 and cc == 0 and calc == 0)
        cand = [v for v in (s.created_at, proj_max.get(s.id), task_max.get(s.id), sheet_max.get(s.id), cal_max.get(s.id)) if v is not None]
        last_act = max(cand) if cand else None
        days_old = (now - s.created_at).days if s.created_at else None

        if not s.is_active:
            status = "보관됨"
        elif s.warned_at:
            status = "경고 발송됨"
        elif is_empty:
            status = "빈 공간"
        else:
            status = "활성"

        creator = users.get(s.created_by)
        owner_uid = owners.get(s.id) or s.created_by
        owner_u = users.get(owner_uid)

        result.append({
            "id": s.id,
            "name": s.name,
            "slug": s.slug,
            "purpose": s.purpose or "project_management",
            "created_by": s.created_by,
            "creator_name": creator.username if creator else None,
            "owner_id": owner_uid,
            "owner_name": owner_u.username if owner_u else None,
            "created_at": iso(s.created_at) if s.created_at else None,
            "days_old": days_old,
            "last_activity_at": iso(last_act) if last_act else None,
            "member_count": mc,
            "project_count": pc,
            "task_count": tc,
            "checksheet_count": cc,
            "calendar_count": calc,
            "is_active": bool(s.is_active),
            "archived_at": iso(s.archived_at) if s.archived_at else None,
            "warned_at": iso(s.warned_at) if s.warned_at else None,
            "delete_scheduled_at": iso(s.delete_scheduled_at) if s.delete_scheduled_at else None,
            "cleanup_exempt": bool(s.cleanup_exempt),
            "is_empty": is_empty,
            "has_project": pc > 0,
            "status": status,
        })

    # 요약 칩 — 전체 공간 기준 count 쿼리 (전수 객체 로딩 없이 집계)
    def _count(query):
        return int(query.with_entities(func.count(Space.id)).scalar() or 0)

    summary = {
        "total": _count(db.query(Space)),
        "active": _count(db.query(Space).filter(Space.is_active.is_(True))),
        "archived": _count(db.query(Space).filter(Space.is_active.is_(False))),
        "no_project": _count(db.query(Space).filter(Space.is_active.is_(True), ~proj_exists)),
        "empty": _count(db.query(Space).filter(Space.is_active.is_(True), *empty_conds)),
        "warned": _count(db.query(Space).filter(Space.is_active.is_(True), Space.warned_at.isnot(None))),
    }
    return {
        "spaces": result,
        "summary": summary,
        "filter": filter,
        "pagination": _pagination_meta(page, page_size, total),
    }


def _admin_get_space_or_404(db: Session, space_id: int) -> Space:
    space = db.query(Space).filter(Space.id == space_id).first()
    if not space:
        raise HTTPException(404, "Space not found")
    return space


@app.post("/api/admin/spaces/{space_id}/archive")
def admin_archive_space(space_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    """공간 보관(soft). 멤버는 보존한다."""
    actor = require_admin_and_get_user(db, user_id)
    space = _admin_get_space_or_404(db, space_id)
    space.is_active = False
    space.archived_at = datetime.utcnow()
    db.commit()
    from app.services.system_log import log_event
    log_event("WARNING", "SPACE", "Space archived by admin",
              detail=f"space_id={space_id} by={user_id}", user_id=user_id,
              login_id=getattr(actor, "loginid", None))
    return _space_dict(space, db)


@app.post("/api/admin/spaces/{space_id}/restore")
def admin_restore_space(space_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    """보관된 공간 복구."""
    actor = require_admin_and_get_user(db, user_id)
    space = _admin_get_space_or_404(db, space_id)
    space.is_active = True
    space.archived_at = None
    space.warned_at = None
    space.delete_scheduled_at = None
    db.commit()
    from app.services.system_log import log_event
    log_event("INFO", "SPACE", "Space restored by admin",
              detail=f"space_id={space_id} by={user_id}", user_id=user_id,
              login_id=getattr(actor, "loginid", None))
    return _space_dict(space, db)


@app.post("/api/admin/spaces/{space_id}/notify-owner")
def admin_notify_owner(space_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    """빈 공간 경고를 소유자에게 발송(공간 내 배너로 노출)."""
    actor = require_admin_and_get_user(db, user_id)
    space = _admin_get_space_or_404(db, space_id)
    space.warned_at = datetime.utcnow()
    if space.delete_scheduled_at is None:
        space.delete_scheduled_at = datetime.utcnow() + timedelta(days=SPACE_EMPTY_DELETE_AFTER_DAYS)
    db.commit()
    owner = db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.role == "owner").first()
    from app.services.system_log import log_event
    log_event("INFO", "SPACE", "Empty space warning sent (manual)",
              detail=f"space_id={space_id} owner_id={owner.user_id if owner else None} by={user_id}",
              user_id=user_id, login_id=getattr(actor, "loginid", None))
    return _space_dict(space, db)


@app.post("/api/admin/spaces/{space_id}/toggle-exempt")
def admin_toggle_exempt(space_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    """자동 정리 제외 토글."""
    require_admin(db, None, user_id)
    space = _admin_get_space_or_404(db, space_id)
    space.cleanup_exempt = not bool(space.cleanup_exempt)
    db.commit()
    return _space_dict(space, db)


@app.post("/api/admin/spaces/{space_id}/extend-delete")
def admin_extend_delete(space_id: int, days: int = Query(30), user_id: int = Query(...), db: Session = Depends(get_db)):
    """삭제 예정일 연장."""
    require_admin(db, None, user_id)
    space = _admin_get_space_or_404(db, space_id)
    base = space.delete_scheduled_at or datetime.utcnow()
    space.delete_scheduled_at = base + timedelta(days=max(1, int(days)))
    db.commit()
    return _space_dict(space, db)


class SpaceHardDeleteBody(BaseModel):
    confirm: str


@app.delete("/api/admin/spaces/{space_id}/hard-delete")
def admin_hard_delete_space(space_id: int, body: SpaceHardDeleteBody, user_id: int = Query(...), db: Session = Depends(get_db)):
    """완전 삭제(hard delete). 매우 위험하므로 다중 안전장치를 둔다.

    조건: confirm == "DELETE" + 공간이 보관 상태(is_active False) + 완전 빈 공간.
    """
    actor = require_admin_and_get_user(db, user_id)
    space = _admin_get_space_or_404(db, space_id)

    if (body.confirm or "").strip() != "DELETE":
        raise HTTPException(400, "확인 문구가 일치하지 않습니다. 'DELETE' 를 입력하세요.")
    if space.is_active:
        raise HTTPException(400, "활성 공간은 삭제할 수 없습니다. 먼저 보관 처리하세요.")
    act = _space_activity(db, space_id)
    if not act["is_empty"]:
        raise HTTPException(
            400,
            f"비어 있지 않은 공간은 삭제할 수 없습니다. (프로젝트 {act['project_count']}, Task {act['task_count']}, "
            f"체크시트 {act['checksheet_count']}, Calendar {act['calendar_count']})",
        )

    # 완전 빈 공간 확인됨 → 잔여 연결 데이터 정리 후 공간 삭제
    space_name = space.name  # 삭제 후 접근 불가하므로 미리 보관
    db.query(SheetExecution).filter(SheetExecution.space_id == space_id).delete(synchronize_session=False)
    db.query(CalendarEvent).filter(CalendarEvent.space_id == space_id).delete(synchronize_session=False)
    db.query(Project).filter(Project.space_id == space_id).delete(synchronize_session=False)
    db.query(SpaceJoinRequest).filter(SpaceJoinRequest.space_id == space_id).delete(synchronize_session=False)
    db.query(SpaceMember).filter(SpaceMember.space_id == space_id).delete(synchronize_session=False)
    db.delete(space)
    db.commit()
    from app.services.system_log import log_event
    log_event("WARNING", "SPACE", "Space hard-deleted by admin",
              detail=f"space_id={space_id} name={space_name!r} by={user_id}",
              user_id=user_id, login_id=getattr(actor, "loginid", None))
    return {"message": "공간이 완전히 삭제되었습니다", "space_id": space_id}


@app.get("/api/projects/unassigned")
def get_unassigned_projects(user_id: int = Query(...), db: Session = Depends(get_db)):
    """Get projects not assigned to any active space."""
    state = load_state()
    rows = db.query(Project).filter(
        Project.archived_at.is_(None),
        Project.space_id.is_(None),
    ).all()
    projects = [project_dict(p, state) for p in rows]
    if user_id:
        accessible = get_user_project_ids(db, state, user_id)
        accessible |= get_user_public_project_ids(db, state, user_id)
        projects = [p for p in projects if p["id"] in accessible]
    return {"projects": projects}

@app.get("/api/projects/movable")
def get_movable_projects(user_id: int = Query(...), db: Session = Depends(get_db)):
    """현재 사용자가 다른 공간으로 이동시킬 수 있는 프로젝트 목록.

    이동 권한 기준은 단일 이동(move-space)과 동일(can_manage_project):
      1) 내가 프로젝트 소유자
      2) 내가 프로젝트 관리 역할 멤버(담당자)
      3) admin / super_admin
    각 프로젝트의 현재 공간 정보와 내 역할을 함께 반환한다.
    """
    state = load_state()
    rows = db.query(Project).filter(Project.archived_at.is_(None)).all()
    space_map = {s.id: s.name for s in db.query(Space).filter(Space.is_active == True).all()}
    admin_like = is_admin_like_role(get_user_role(db, user_id))
    pm_roles = {
        pm.project_id: pm.role
        for pm in db.query(ProjectMemberModel).filter(ProjectMemberModel.user_id == int(user_id)).all()
    }
    result = []
    for p in rows:
        meta = get_project_meta(state, p.id)
        # 시스템 프로젝트([시스템] 단발 업무)는 이동 대상에서 제외
        if meta.get("is_system"):
            continue
        is_owner = int(meta.get("owner_id") or 0) == int(user_id)
        pm_role = pm_roles.get(p.id)
        is_manager_member = pm_role in PROJECT_MANAGER_ROLES
        if not (is_owner or is_manager_member or admin_like):
            continue
        if is_owner:
            my_role = "owner"
        elif is_manager_member:
            my_role = pm_role
        else:
            my_role = "admin"
        result.append({
            "id": p.id,
            "name": p.name,
            "description": p.description,
            "current_space_id": p.space_id,
            "current_space_name": space_map.get(p.space_id) if p.space_id else None,
            "my_project_role": my_role,
            "can_move": True,
        })
    return {"projects": result}

@app.post("/api/spaces/{space_id}/members")
def add_space_member(space_id: int, user_id: int = Query(...), target_user_id: int = Query(...), role: str = Query(default="member"), db: Session = Depends(get_db)):
    _require_space_admin(db, space_id, user_id)
    existing = db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.user_id == target_user_id).first()
    if existing:
        return {"message": "이미 멤버입니다"}
    db.add(SpaceMember(space_id=space_id, user_id=target_user_id, role=role))
    db.commit()
    return {"message": "멤버가 추가되었습니다"}


class SpaceKnoxMemberAdd(BaseModel):
    """공간 멤버 추가 — Knox 사용자 (사이트 DB에 없을 수 있음)"""
    loginid: str
    name: Optional[str] = None
    email: Optional[str] = None
    deptname: Optional[str] = None
    role: Optional[str] = "member"


class SpaceGroupMemberAdd(BaseModel):
    """공간에 그룹 구성원 일괄 추가"""
    group_id: int
    role: Optional[str] = "member"


@app.get("/api/spaces/{space_id}/member-candidates")
async def get_space_member_candidates(
    space_id: int,
    q: str = Query(..., min_length=1),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """공간 멤버 추가 모달용 통합 검색.
    - 사이트 users + Knox 사내 검색 + 사용자가 만든 member_groups."""
    _require_space_admin(db, space_id, user_id)
    items: list = []
    search_q = f"%{q.strip()}%"
    site_users = db.query(User).filter(
        User.is_active == True,
        (User.username.ilike(search_q) | User.loginid.ilike(search_q) | User.deptname.ilike(search_q))
    ).limit(20).all()
    site_loginids = set()
    for u in site_users:
        items.append({
            "type": "user",
            "source": "site",
            "user_id": u.id,
            "login_id": u.loginid,
            "name": u.username,
            "email": u.mail,
            "department": u.deptname,
            "avatar_color": u.avatar_color,
        })
        site_loginids.add(u.loginid)
    # Knox 검색 (실패 안전)
    knox_employees = await _knox_search_safe(q)
    for e in knox_employees[:20]:
        loginid = e.get("userId") or e.get("loginid")
        if not loginid or loginid in site_loginids:
            continue
        items.append({
            "type": "user",
            "source": "knox",
            "login_id": loginid,
            "name": e.get("fullName") or loginid,
            "email": e.get("mail") or e.get("email"),
            "department": e.get("departmentName") or e.get("deptname"),
        })
    # MemberGroup 검색 (이름 기준)
    groups = db.query(MemberGroup).filter(MemberGroup.name.ilike(search_q)).limit(10).all()
    for g in groups:
        member_count = db.query(MemberGroupUser).filter(MemberGroupUser.group_id == g.id).count()
        items.append({
            "type": "group",
            "group_id": g.id,
            "name": g.name,
            "description": g.description,
            "member_count": member_count,
        })
    return {"items": items}


@app.post("/api/spaces/{space_id}/members/knox")
def add_space_member_from_knox(
    space_id: int,
    body: SpaceKnoxMemberAdd,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """Knox 사용자를 공간 멤버로 추가 (필요시 users 테이블에 upsert)."""
    _require_space_admin(db, space_id, user_id)
    target = get_or_create_user_from_knox(db, body.dict())
    if not target:
        raise HTTPException(400, "Knox 사용자 정보가 유효하지 않습니다")
    existing = db.query(SpaceMember).filter(
        SpaceMember.space_id == space_id, SpaceMember.user_id == target.id
    ).first()
    if existing:
        db.commit()
        return {"message": "이미 멤버입니다", "user_id": target.id, "skipped": True}
    role = body.role if body.role in ("owner", "admin", "operator", "member") else "member"
    if role in ("owner", "admin", "operator"):
        # 안전: Knox 추가는 기본 member 만 허용 (역할 변경은 별도 API)
        role = "member"
    db.add(SpaceMember(space_id=space_id, user_id=target.id, role=role))
    db.commit()
    return {"message": "멤버가 추가되었습니다", "user_id": target.id, "skipped": False}


@app.post("/api/spaces/{space_id}/members/groups")
def add_space_members_from_group(
    space_id: int,
    body: SpaceGroupMemberAdd,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """MemberGroup 구성원을 공간 멤버로 일괄 추가. 이미 멤버는 skip."""
    _require_space_admin(db, space_id, user_id)
    g = db.query(MemberGroup).filter(MemberGroup.id == body.group_id).first()
    if not g:
        raise HTTPException(404, "그룹을 찾을 수 없습니다")
    role = body.role if body.role in ("member", "operator") else "member"
    group_users = db.query(MemberGroupUser, User).join(
        User, MemberGroupUser.user_id == User.id
    ).filter(MemberGroupUser.group_id == body.group_id).all()
    existing_member_ids = set(
        sm.user_id for sm in db.query(SpaceMember).filter(SpaceMember.space_id == space_id).all()
    )
    added_users: list = []
    skipped_users: list = []
    for _, u in group_users:
        if u.id in existing_member_ids:
            skipped_users.append({"user_id": u.id, "login_id": u.loginid, "name": u.username})
            continue
        db.add(SpaceMember(space_id=space_id, user_id=u.id, role=role))
        existing_member_ids.add(u.id)
        added_users.append({"user_id": u.id, "login_id": u.loginid, "name": u.username})
    db.commit()
    return {
        "group_id": g.id,
        "group_name": g.name,
        "added_count": len(added_users),
        "skipped_count": len(skipped_users),
        "added_users": added_users,
        "skipped_users": skipped_users,
    }


@app.delete("/api/spaces/{space_id}/members/{target_user_id}")
def remove_space_member(space_id: int, target_user_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    _require_space_admin(db, space_id, user_id)
    db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.user_id == target_user_id).delete()
    db.commit()
    return {"message": "멤버가 제거되었습니다"}

@app.patch("/api/spaces/{space_id}/members/{target_user_id}/role")
def update_space_member_role(space_id: int, target_user_id: int, role: str = Query(...), user_id: int = Query(...), db: Session = Depends(get_db)):
    """Change a member's role. Only owner can promote to admin/operator; owner/admin/operator can set member."""
    caller = db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.user_id == user_id).first()
    if not caller or caller.role not in ("owner", "admin", "operator"):
        raise HTTPException(403, "권한이 없습니다")
    if role in ("admin", "operator") and caller.role != "owner":
        raise HTTPException(403, "관리자/공간운영 지정은 소유자만 가능합니다")
    if role not in ("owner", "admin", "operator", "member"):
        raise HTTPException(400, "유효하지 않은 역할입니다")
    target = db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.user_id == target_user_id).first()
    if not target:
        raise HTTPException(404, "해당 멤버를 찾을 수 없습니다")
    if target.role == "owner":
        raise HTTPException(403, "소유자의 역할은 변경할 수 없습니다")
    target.role = role
    db.commit()
    return {"message": f"역할이 {role}로 변경되었습니다"}

@app.patch("/api/projects/{project_id}/move-space")
def move_project_space(project_id: int, space_id: int = Query(...), user_id: int = Query(...), db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")
    # 1) 현재 프로젝트에 대한 관리 권한(소유자/담당자) 필요
    state = load_state()
    if not can_manage_project(db, state, project_id, user_id):
        raise HTTPException(403, "프로젝트 이동 권한이 없습니다. (소유자 또는 담당자만 가능)")
    # 2) 이동 대상 공간의 멤버여야 함 (기존 정책 유지)
    member = db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.user_id == user_id).first()
    if not member:
        raise HTTPException(403, "대상 공간의 멤버가 아닙니다")
    _old_space = project.space_id
    project.space_id = space_id
    actlog.log_activity(
        db, user_id=user_id, action=actlog.ACTION_MOVED,
        entity_type=actlog.ENTITY_PROJECT, entity_id=project_id,
        message="공간 이동",
        meta={"changes": [{"field": "space_id", "label": "공간", "before": _old_space, "after": space_id}]},
    )
    db.commit()
    return {"message": "프로젝트가 이동되었습니다", "project_id": project_id, "space_id": space_id}


class BulkMoveProjects(BaseModel):
    """여러 프로젝트를 한 공간으로 일괄 이동"""
    project_ids: List[int]
    target_space_id: int


@app.post("/api/projects/move-space")
def move_projects_space_bulk(body: BulkMoveProjects, user_id: int = Query(...), db: Session = Depends(get_db)):
    """여러 프로젝트를 대상 공간으로 일괄 이동.

    프론트 목록을 신뢰하지 않고 프로젝트마다 권한을 재검증한다.
      1) 해당 프로젝트 이동 권한(소유자/담당자/admin)
      2) 대상 공간 멤버 여부(admin/super_admin은 예외)
      3) 대상 공간이 활성 상태
      4) 이미 대상 공간에 있는 프로젝트가 아닌지
    일부만 실패할 수 있으므로 프로젝트별 결과를 반환한다.
    """
    state = load_state()
    target_space = db.query(Space).filter(
        Space.id == body.target_space_id, Space.is_active == True
    ).first()
    admin_like = is_admin_like_role(get_user_role(db, user_id))
    target_member = db.query(SpaceMember).filter(
        SpaceMember.space_id == body.target_space_id, SpaceMember.user_id == user_id
    ).first()

    success, failed = [], []
    for pid in body.project_ids:
        project = db.query(Project).filter(Project.id == pid).first()
        name = project.name if project else None
        if not project:
            failed.append({"project_id": pid, "name": name, "reason": "프로젝트를 찾을 수 없습니다."})
            continue
        if not can_manage_project(db, state, pid, user_id):
            failed.append({"project_id": pid, "name": name, "reason": "프로젝트 이동 권한이 없습니다. (소유자 또는 담당자만 가능)"})
            continue
        if not target_space:
            failed.append({"project_id": pid, "name": name, "reason": "대상 공간이 존재하지 않거나 비활성 상태입니다."})
            continue
        if not (target_member or admin_like):
            failed.append({"project_id": pid, "name": name, "reason": "대상 공간의 멤버가 아닙니다."})
            continue
        if project.space_id == body.target_space_id:
            failed.append({"project_id": pid, "name": name, "reason": "이미 대상 공간에 있는 프로젝트입니다."})
            continue
        project.space_id = body.target_space_id
        success.append({"project_id": pid, "name": name})

    if success:
        db.commit()
    return {"success": success, "failed": failed}

# ── Space Join Requests ──
@app.get("/api/spaces/by-slug/{slug}")
def get_space_by_slug(slug: str, user_id: Optional[int] = None, db: Session = Depends(get_db)):
    """Public lookup by slug. Returns basic info + whether user is a member."""
    space = db.query(Space).filter(Space.slug == slug, Space.is_active == True).first()
    if not space:
        raise HTTPException(404, "공간을 찾을 수 없습니다")
    is_member = False
    if user_id:
        m = db.query(SpaceMember).filter(SpaceMember.space_id == space.id, SpaceMember.user_id == user_id).first()
        is_member = m is not None
    pending = False
    if user_id and not is_member:
        pr = db.query(SpaceJoinRequest).filter(
            SpaceJoinRequest.space_id == space.id, SpaceJoinRequest.user_id == user_id, SpaceJoinRequest.status == "pending"
        ).first()
        pending = pr is not None
    return {"id": space.id, "name": space.name, "slug": space.slug, "description": space.description, "is_member": is_member, "pending_request": pending}

@app.post("/api/spaces/{space_id}/join-request")
def request_join_space(space_id: int, user_id: int = Query(...), message: Optional[str] = None, db: Session = Depends(get_db)):
    existing = db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.user_id == user_id).first()
    if existing:
        return {"message": "이미 멤버입니다"}
    pending = db.query(SpaceJoinRequest).filter(
        SpaceJoinRequest.space_id == space_id, SpaceJoinRequest.user_id == user_id, SpaceJoinRequest.status == "pending"
    ).first()
    if pending:
        return {"message": "이미 신청 대기 중입니다"}
    req = SpaceJoinRequest(space_id=space_id, user_id=user_id, message=message)
    db.add(req)
    db.commit()
    return {"message": "접속 권한이 신청되었습니다"}

@app.get("/api/spaces/{space_id}/join-requests")
def get_space_join_requests(space_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    _require_space_admin(db, space_id, user_id)
    reqs = db.query(SpaceJoinRequest, User).join(User, SpaceJoinRequest.user_id == User.id).filter(
        SpaceJoinRequest.space_id == space_id, SpaceJoinRequest.status == "pending"
    ).all()
    return {"requests": [
        {"id": r.id, "user_id": r.user_id, "username": u.username, "loginid": u.loginid, "message": r.message, "created_at": iso(r.created_at)}
        for r, u in reqs
    ]}

@app.get("/api/spaces/{space_id}/join-requests-safe")
def get_space_join_requests_safe(space_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    """join-requests 조회 — 권한 없으면 빈 배열 반환 (기능 차단 방지)"""
    try:
        _require_space_admin(db, space_id, user_id)
    except HTTPException:
        return {"requests": []}
    reqs = db.query(SpaceJoinRequest, User).join(User, SpaceJoinRequest.user_id == User.id).filter(
        SpaceJoinRequest.space_id == space_id, SpaceJoinRequest.status == "pending"
    ).all()
    return {"requests": [
        {"id": r.id, "user_id": r.user_id, "username": u.username, "loginid": u.loginid, "message": r.message, "created_at": iso(r.created_at)}
        for r, u in reqs
    ]}

@app.post("/api/spaces/{space_id}/join-requests/{request_id}/approve")
def approve_join_request(space_id: int, request_id: int, action: str = Query(...), user_id: int = Query(...), db: Session = Depends(get_db)):
    """action: 'approve' or 'reject'"""
    _require_space_admin(db, space_id, user_id)
    req = db.query(SpaceJoinRequest).filter(SpaceJoinRequest.id == request_id, SpaceJoinRequest.space_id == space_id).first()
    if not req:
        raise HTTPException(404, "요청을 찾을 수 없습니다")
    if action == "approve":
        req.status = "approved"
        req.resolved_by = user_id
        req.resolved_at = datetime.utcnow()
        existing = db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.user_id == req.user_id).first()
        if not existing:
            db.add(SpaceMember(space_id=space_id, user_id=req.user_id, role="member"))
        db.commit()
        return {"message": "승인되었습니다"}
    elif action == "reject":
        req.status = "rejected"
        req.resolved_by = user_id
        req.resolved_at = datetime.utcnow()
        db.commit()
        return {"message": "거절되었습니다"}
    raise HTTPException(400, "action은 approve 또는 reject여야 합니다")


# =========================
# Backfill: project_members loginid/deptname
# =========================
@app.post("/api/admin/backfill-project-members")
def backfill_project_members(user_id: int = Query(...), db: Session = Depends(get_db)):
    state = load_state()
    require_admin(db, state, user_id)
    rows = db.query(ProjectMemberModel).filter(
        (ProjectMemberModel.loginid.is_(None)) | (ProjectMemberModel.deptname.is_(None))
    ).all()
    updated = 0
    for pm in rows:
        u = db.query(User).filter(User.id == pm.user_id).first()
        if u:
            pm.loginid = u.loginid
            pm.deptname = getattr(u, "deptname", None)
            updated += 1
    db.commit()
    return {"message": f"Backfilled {updated} project_member rows"}


# =========================
# Shortcuts (sidecar)
# =========================
SHORTCUT_VISIBILITIES = {"PRIVATE", "PUBLIC", "SHARED_USERS", "SHARED_GROUPS"}


def _normalize_admin_shortcut(sc: dict) -> dict:
    """state.json에 저장된 기존 admin 바로가기에 누락된 공유 필드를 채워준다.
    기존 데이터(공유 필드 없음)는 이전 동작과 동일하게 전체 노출되도록 PUBLIC 으로 백필.
    """
    if "visibility" not in sc or not sc.get("visibility"):
        sc["visibility"] = "PUBLIC"
    if "shared_user_ids" not in sc or sc.get("shared_user_ids") is None:
        sc["shared_user_ids"] = []
    if "shared_group_ids" not in sc or sc.get("shared_group_ids") is None:
        sc["shared_group_ids"] = []
    if "owner_id" not in sc:
        sc["owner_id"] = None
    return sc


def _user_member_group_ids(db: Session, user_id: int) -> set:
    """현재 사용자가 속한 MemberGroup id 집합."""
    rows = db.query(MemberGroupUser).filter(MemberGroupUser.user_id == user_id).all()
    return {r.group_id for r in rows}


def _parse_id_list(raw) -> set:
    """user_shortcuts.shared_*_ids 는 JSON 직렬화된 정수 리스트 문자열로 보관."""
    if raw is None or raw == "":
        return set()
    if isinstance(raw, (list, tuple, set)):
        try:
            return {int(x) for x in raw}
        except Exception:
            return set()
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return {int(x) for x in parsed}
    except Exception:
        pass
    return set()


def _serialize_id_list(ids) -> Optional[str]:
    if not ids:
        return None
    try:
        return json.dumps([int(x) for x in ids])
    except Exception:
        return None


def _admin_shortcut_visible_to(sc: dict, user_id: int, user_group_ids: set, is_admin: bool) -> bool:
    vis = sc.get("visibility") or "PUBLIC"
    if is_admin:
        return True  # 관리자는 모든 바로가기를 관리/조회 가능
    if vis == "PUBLIC":
        return True
    if sc.get("owner_id") and int(sc.get("owner_id")) == int(user_id):
        return True
    if vis == "SHARED_USERS":
        return int(user_id) in {int(x) for x in (sc.get("shared_user_ids") or [])}
    if vis == "SHARED_GROUPS":
        return bool(user_group_ids.intersection({int(x) for x in (sc.get("shared_group_ids") or [])}))
    # PRIVATE — owner 만 (위에서 처리됨)
    return False


@app.get("/api/shortcuts")
def get_shortcuts(user_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    """관리자 페이지에서 사용 — 전체 admin 바로가기 목록 반환.
    user_id 가 주어지고 admin 권한이면 전체 노출, 아니면 본인이 만든 것만 노출.
    user_id 미지정 시(legacy 호출) 전체 노출 — 단 visibility 필드는 정규화하여 반환.
    """
    state = load_state()
    shortcuts = [_normalize_admin_shortcut(dict(s)) for s in state.get("shortcuts", [])]
    if user_id is None or user_id <= 0:
        return {"shortcuts": shortcuts}
    u = db.query(User).filter(User.id == user_id).first()
    is_admin = bool(u) and (u.role in ("admin", "super_admin"))
    if is_admin:
        return {"shortcuts": shortcuts}
    # admin 권한 없는 호출자는 본인 소유분만
    return {"shortcuts": [s for s in shortcuts if s.get("owner_id") == user_id]}


@app.get("/api/shortcuts/my")
def get_my_shortcuts(user_id: int = Query(...), db: Session = Depends(get_db)):
    """Dashboard 위젯용 — 현재 사용자가 볼 수 있는 바로가기 통합 조회.
    - admin 바로가기: visibility 정책에 따라 필터
    - user 바로가기(per-user): 본인 소유 + 공유받은 것
    """
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=403, detail="user not found")
    is_admin = u.role in ("admin", "super_admin")
    user_group_ids = _user_member_group_ids(db, user_id)

    # admin 바로가기
    state = load_state()
    admin_list = [_normalize_admin_shortcut(dict(s)) for s in state.get("shortcuts", [])]
    visible_admin = [
        {**s, "source": "admin"}
        for s in admin_list
        if s.get("active", True) and _admin_shortcut_visible_to(s, user_id, user_group_ids, is_admin)
    ]

    # user 바로가기 (DB)
    user_list = db.query(UserShortcut).filter(UserShortcut.active == True).all()
    visible_user = []
    for s in user_list:
        if s.user_id == user_id:
            visible_user.append({**user_shortcut_dict(s), "source": "user"})
            continue
        vis = s.visibility or "PRIVATE"
        if vis == "PUBLIC":
            visible_user.append({**user_shortcut_dict(s), "source": "user"})
            continue
        if vis == "SHARED_USERS":
            ids = _parse_id_list(s.shared_user_ids)
            if user_id in ids:
                visible_user.append({**user_shortcut_dict(s), "source": "user"})
                continue
        if vis == "SHARED_GROUPS":
            gids = _parse_id_list(s.shared_group_ids)
            if user_group_ids.intersection(gids):
                visible_user.append({**user_shortcut_dict(s), "source": "user"})
                continue

    return {"shortcuts": visible_admin + visible_user}


@app.post("/api/shortcuts")
def create_shortcut(shortcut: ShortcutCreate, user_id: int = Query(...), db: Session = Depends(get_db)):
    state = load_state()
    require_admin(db, state, user_id)

    vis = (shortcut.visibility or "PRIVATE").upper()
    if vis not in SHORTCUT_VISIBILITIES:
        raise HTTPException(status_code=400, detail="invalid visibility")

    shortcuts = state.get("shortcuts", [])
    new_sc = {
        "id": next_id(shortcuts),
        "name": shortcut.name,
        "url": shortcut.url,
        "icon_text": shortcut.icon_text or (shortcut.name[:1].upper() if shortcut.name else "S"),
        "icon_color": shortcut.icon_color or "#2955FF",
        "order": shortcut.order if shortcut.order is not None else len(shortcuts),
        "open_new_tab": bool(shortcut.open_new_tab),
        "active": True,
        "created_at": datetime.now().isoformat(),
        "owner_id": user_id,
        "visibility": vis,
        "shared_user_ids": list(shortcut.shared_user_ids or []) if vis == "SHARED_USERS" else [],
        "shared_group_ids": list(shortcut.shared_group_ids or []) if vis == "SHARED_GROUPS" else [],
    }
    shortcuts.append(new_sc)
    state["shortcuts"] = shortcuts
    save_state(state)
    return new_sc

@app.patch("/api/shortcuts/{shortcut_id}")
def update_shortcut(shortcut_id: int, updates: ShortcutUpdate, user_id: int = Query(...), db: Session = Depends(get_db)):
    state = load_state()
    require_admin(db, state, user_id)

    shortcuts = state.get("shortcuts", [])
    sc = next((s for s in shortcuts if int(s.get("id")) == shortcut_id), None)
    if not sc:
        raise HTTPException(status_code=404, detail="Shortcut not found")

    payload = updates.model_dump(exclude_unset=True)

    # visibility 검증 및 공유 리스트 정리
    if "visibility" in payload:
        vis = (payload["visibility"] or "PRIVATE").upper()
        if vis not in SHORTCUT_VISIBILITIES:
            raise HTTPException(status_code=400, detail="invalid visibility")
        payload["visibility"] = vis
        if vis != "SHARED_USERS":
            payload["shared_user_ids"] = []
        if vis != "SHARED_GROUPS":
            payload["shared_group_ids"] = []

    for key, val in payload.items():
        sc[key] = val

    _normalize_admin_shortcut(sc)
    save_state(state)
    return sc

@app.delete("/api/shortcuts/{shortcut_id}")
def delete_shortcut(shortcut_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    state = load_state()
    require_admin(db, state, user_id)

    state["shortcuts"] = [s for s in state.get("shortcuts", []) if int(s.get("id")) != shortcut_id]
    save_state(state)
    return {"message": "Shortcut deleted"}


# =========================
# User Shortcuts (per-user, DB)
# =========================
class UserShortcutCreate(BaseModel):
    name: str
    url: str
    icon_text: Optional[str] = None
    icon_color: str = "#2955FF"
    order: int = 0
    open_new_tab: bool = True
    visibility: Optional[str] = "PRIVATE"
    shared_user_ids: Optional[List[int]] = None
    shared_group_ids: Optional[List[int]] = None

class UserShortcutUpdate(BaseModel):
    name: Optional[str] = None
    url: Optional[str] = None
    icon_text: Optional[str] = None
    icon_color: Optional[str] = None
    order: Optional[int] = None
    open_new_tab: Optional[bool] = None
    active: Optional[bool] = None
    visibility: Optional[str] = None
    shared_user_ids: Optional[List[int]] = None
    shared_group_ids: Optional[List[int]] = None

def user_shortcut_dict(s: UserShortcut) -> dict:
    return {
        "id": s.id,
        "user_id": s.user_id,
        "name": s.name,
        "url": s.url,
        "icon_text": s.icon_text,
        "icon_color": s.icon_color,
        "order": s.order,
        "open_new_tab": s.open_new_tab,
        "active": s.active,
        "visibility": s.visibility or "PRIVATE",
        "shared_user_ids": sorted(_parse_id_list(s.shared_user_ids)),
        "shared_group_ids": sorted(_parse_id_list(s.shared_group_ids)),
        "created_at": iso(s.created_at),
    }


def _user_is_admin(db: Session, user_id: int) -> bool:
    u = db.query(User).filter(User.id == user_id).first()
    return bool(u) and (u.role in ("admin", "super_admin"))


def _validate_user_visibility(vis: str, db: Session, user_id: int) -> str:
    """일반 사용자 바로가기 공개범위 검증. PUBLIC 은 admin 만 허용."""
    vis = (vis or "PRIVATE").upper()
    if vis not in SHORTCUT_VISIBILITIES:
        raise HTTPException(status_code=400, detail="invalid visibility")
    if vis == "PUBLIC" and not _user_is_admin(db, user_id):
        raise HTTPException(status_code=403, detail="전체 공개는 관리자만 가능합니다.")
    return vis

@app.get("/api/user-shortcuts")
def get_user_shortcuts(user_id: int = Query(...), db: Session = Depends(get_db)):
    shortcuts = db.query(UserShortcut).filter(
        UserShortcut.user_id == user_id,
        UserShortcut.active == True,
    ).order_by(UserShortcut.order).all()
    return {"shortcuts": [user_shortcut_dict(s) for s in shortcuts]}

@app.post("/api/user-shortcuts")
def create_user_shortcut(body: UserShortcutCreate, user_id: int = Query(...), db: Session = Depends(get_db)):
    vis = _validate_user_visibility(body.visibility or "PRIVATE", db, user_id)
    s = UserShortcut(
        user_id=user_id,
        name=body.name,
        url=body.url,
        icon_text=body.icon_text,
        icon_color=body.icon_color,
        order=body.order,
        open_new_tab=body.open_new_tab,
        visibility=vis,
        shared_user_ids=_serialize_id_list(body.shared_user_ids) if vis == "SHARED_USERS" else None,
        shared_group_ids=_serialize_id_list(body.shared_group_ids) if vis == "SHARED_GROUPS" else None,
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    return user_shortcut_dict(s)

@app.patch("/api/user-shortcuts/{shortcut_id}")
def update_user_shortcut(shortcut_id: int, body: UserShortcutUpdate, user_id: int = Query(...), db: Session = Depends(get_db)):
    s = db.query(UserShortcut).filter(UserShortcut.id == shortcut_id, UserShortcut.user_id == user_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shortcut not found")
    if body.name is not None:
        s.name = body.name
    if body.url is not None:
        s.url = body.url
    if body.icon_text is not None:
        s.icon_text = body.icon_text
    if body.icon_color is not None:
        s.icon_color = body.icon_color
    if body.order is not None:
        s.order = body.order
    if body.open_new_tab is not None:
        s.open_new_tab = body.open_new_tab
    if body.active is not None:
        s.active = body.active
    if body.visibility is not None:
        vis = _validate_user_visibility(body.visibility, db, user_id)
        s.visibility = vis
        if vis != "SHARED_USERS":
            s.shared_user_ids = None
        if vis != "SHARED_GROUPS":
            s.shared_group_ids = None
    if body.shared_user_ids is not None and (s.visibility or "PRIVATE") == "SHARED_USERS":
        s.shared_user_ids = _serialize_id_list(body.shared_user_ids)
    if body.shared_group_ids is not None and (s.visibility or "PRIVATE") == "SHARED_GROUPS":
        s.shared_group_ids = _serialize_id_list(body.shared_group_ids)
    db.commit()
    db.refresh(s)
    return user_shortcut_dict(s)

@app.delete("/api/user-shortcuts/{shortcut_id}")
def delete_user_shortcut(shortcut_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    s = db.query(UserShortcut).filter(UserShortcut.id == shortcut_id, UserShortcut.user_id == user_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shortcut not found")
    db.delete(s)
    db.commit()
    return {"message": "Shortcut deleted"}


# =========================
# 사내 추천 도구 스트립 (Internal Service Recommendations)
# - 광고가 아닌 "업무와 연결된 추천 도구" 영역.
# - 기본값은 OFF. global flag 가 False 면 항상 빈 목록을 반환한다.
# - sidecar JSON state 에 저장 (DB 스키마 변경 없음 → 최소 침습).
# =========================
RECOMMENDATION_PLACEMENTS = {"PROJECT_DASHBOARD_TOP", "EQUIPMENT_DASHBOARD_AFTER_KPI"}
RECOMMENDATION_EVENT_TYPES = {"impression", "click", "dismiss"}
_RECOMMENDATION_PUBLIC_FIELDS = (
    "id", "target_space_type", "placement", "label", "badge", "title",
    "description", "reason_text", "cta_label", "cta_url", "secondary_label", "order",
)


def _recommendation_flag(state: dict) -> bool:
    return bool(state.get("internal_service_recommendations_enabled", False))


def _public_recommendation(item: dict) -> dict:
    return {k: item.get(k) for k in _RECOMMENDATION_PUBLIC_FIELDS}


def _dismissal_active(entry: dict) -> bool:
    """숨김 엔트리가 아직 유효한지(만료 전인지) 판단."""
    if not isinstance(entry, dict):
        return False
    until = entry.get("until")
    if not until:
        return False
    try:
        return datetime.fromisoformat(until) > datetime.now(KST)
    except Exception:
        return False


@app.get("/api/internal-recommendations")
def list_internal_recommendations(
    placement: str = Query(...),
    user_id: Optional[int] = Query(None),
):
    """사용자 화면용 추천 항목 조회.

    - global flag 가 False 면 항상 {"enabled": False, "recommendations": []}.
    - active=False 항목은 제외.
    - 해당 사용자가 '나중에 보기/닫기'로 숨긴(만료 전) 항목은 제외.
    - 실패해도 Dashboard 가 깨지지 않도록 프론트가 방어하지만, 여기서도 빈 목록으로 안전 반환.
    """
    try:
        state = load_state()
        if not _recommendation_flag(state):
            return {"enabled": False, "recommendations": []}

        if placement not in RECOMMENDATION_PLACEMENTS:
            return {"enabled": True, "recommendations": []}

        dismissals = state.get("internal_recommendation_dismissals", {}) or {}
        user_dismissals = dismissals.get(str(user_id), {}) if user_id else {}

        items = []
        for it in state.get("internal_recommendations", []) or []:
            if not it.get("active", True):
                continue
            if it.get("placement") != placement:
                continue
            entry = user_dismissals.get(str(it.get("id")))
            if _dismissal_active(entry):
                continue
            items.append(_public_recommendation(it))

        items.sort(key=lambda x: (x.get("order") or 0, x.get("id") or 0))
        return {"enabled": True, "recommendations": items}
    except Exception:
        # 절대 화면을 막지 않는다.
        return {"enabled": False, "recommendations": []}


@app.post("/api/internal-recommendations/{item_id}/dismiss")
def dismiss_internal_recommendation(
    item_id: int,
    user_id: int = Query(...),
    mode: str = Query("today"),  # today | later
):
    """추천 항목 숨김 처리.
    - today: 오늘 하루 숨김 (다음 자정까지)
    - later: 7일 동안 숨김
    """
    mode = (mode or "today").lower()
    if mode not in ("today", "later"):
        raise HTTPException(status_code=400, detail="invalid mode")

    now = datetime.now(KST)
    if mode == "today":
        # 다음 날 자정까지
        until = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        until = now + timedelta(days=7)

    state = load_state()
    dismissals = state.setdefault("internal_recommendation_dismissals", {})
    if not isinstance(dismissals, dict):
        dismissals = {}
        state["internal_recommendation_dismissals"] = dismissals
    user_key = str(user_id)
    user_map = dismissals.setdefault(user_key, {})
    if not isinstance(user_map, dict):
        user_map = {}
        dismissals[user_key] = user_map
    user_map[str(item_id)] = {"mode": mode, "until": until.isoformat()}
    save_state(state)
    return {"ok": True, "mode": mode, "until": until.isoformat()}


@app.post("/api/internal-recommendations/events")
def log_internal_recommendation_event(event: InternalRecommendationEvent):
    """impression/click/dismiss 이벤트 적재. 실패가 화면을 막지 않도록 항상 200."""
    try:
        if event.type not in RECOMMENDATION_EVENT_TYPES:
            return {"ok": False}
        state = load_state()
        events = state.setdefault("internal_recommendation_events", [])
        if not isinstance(events, list):
            events = []
            state["internal_recommendation_events"] = events
        events.append({
            "item_id": event.item_id,
            "user_id": event.user_id,
            "type": event.type,
            "at": datetime.now(KST).isoformat(),
        })
        # 최근 5000개만 보존
        if len(events) > 5000:
            del events[: len(events) - 5000]
        save_state(state)
        return {"ok": True}
    except Exception:
        return {"ok": False}


@app.get("/api/admin/internal-recommendations")
def admin_list_internal_recommendations(user_id: int = Query(...), db: Session = Depends(get_db)):
    """관리자용 — flag 상태 + 전체 추천 항목(active 포함) 조회."""
    require_admin(db, None, user_id)
    state = load_state()
    return {
        "enabled": _recommendation_flag(state),
        "recommendations": state.get("internal_recommendations", []) or [],
    }


@app.put("/api/admin/internal-recommendations/flag")
def admin_set_internal_recommendation_flag(
    body: InternalRecommendationFlagUpdate,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """관리자용 — global feature flag ON/OFF."""
    require_admin(db, None, user_id)
    state = load_state()
    state["internal_service_recommendations_enabled"] = bool(body.enabled)
    save_state(state)
    return {"enabled": bool(body.enabled)}


@app.patch("/api/admin/internal-recommendations/{item_id}")
def admin_update_internal_recommendation(
    item_id: int,
    body: InternalRecommendationUpdate,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """관리자용 — 개별 추천 항목 수정(active 토글 및 문구/링크 편집)."""
    require_admin(db, None, user_id)
    state = load_state()
    items = state.get("internal_recommendations", []) or []
    target = next((it for it in items if int(it.get("id")) == item_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="Recommendation not found")
    payload = body.model_dump(exclude_unset=True)
    for k, v in payload.items():
        target[k] = v
    state["internal_recommendations"] = items
    save_state(state)
    return target


# =========================
# S3 Backup APIs
# =========================
@app.post("/api/admin/backup/db")
def manual_backup_db(user_id: int = Query(...), db: Session = Depends(get_db)):
    """수동 DB 백업 실행 (super admin만)"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user or (user.loginid or "").lower() not in [lid.lower() for lid in SUPER_ADMIN_LOGINIDS]:
        raise HTTPException(status_code=403, detail="Super admin only")

    from app.services.backup_scheduler import backup_database
    result = backup_database()
    return result

@app.post("/api/admin/backup/files")
def manual_backup_files(user_id: int = Query(...), db: Session = Depends(get_db)):
    """수동 첨부파일 S3 동기화 실행 (super admin만)"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user or (user.loginid or "").lower() not in [lid.lower() for lid in SUPER_ADMIN_LOGINIDS]:
        raise HTTPException(status_code=403, detail="Super admin only")

    from app.services.backup_scheduler import sync_uploads_to_s3
    result = sync_uploads_to_s3()
    return result

@app.get("/api/admin/backup/status")
def get_backup_status(user_id: int = Query(...), db: Session = Depends(get_db)):
    """백업 상태/로그 조회 (super admin만)"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user or (user.loginid or "").lower() not in [lid.lower() for lid in SUPER_ADMIN_LOGINIDS]:
        raise HTTPException(status_code=403, detail="Super admin only")

    from app.utils.s3.s3_utils import is_s3_configured, list_s3_files
    from app.services.backup_scheduler import BACKUP_LOG_DIR, BACKUP_HOUR, BACKUP_MINUTE

    # 최근 로그 파일 읽기
    recent_logs = []
    log_dir = BACKUP_LOG_DIR
    if os.path.exists(log_dir):
        log_files = sorted(
            [f for f in os.listdir(log_dir) if f.startswith("backup_") and f.endswith(".log")],
            reverse=True
        )[:3]
        for lf in log_files:
            try:
                with open(os.path.join(log_dir, lf), "r", encoding="utf-8") as f:
                    lines = f.readlines()[-20:]  # 최근 20줄
                    recent_logs.extend([l.strip() for l in lines])
            except Exception:
                pass

    # S3 백업 파일 목록
    s3_db_files = []
    s3_attachment_count = 0
    if is_s3_configured():
        try:
            s3_db_files = list_s3_files("db-backups/")[:20]
            s3_attachments = list_s3_files("files/")
            s3_attachment_count = len(s3_attachments)
        except Exception:
            pass

    return {
        "s3_configured": is_s3_configured(),
        "schedule": f"{BACKUP_HOUR:02d}:{BACKUP_MINUTE:02d} KST daily",
        "recent_logs": recent_logs,
        "s3_db_backups": s3_db_files,
        "s3_attachment_count": s3_attachment_count,
    }

@app.get("/api/admin/notifications/diagnostics")
def notification_diagnostics(user_id: int = Query(...), db: Session = Depends(get_db)):
    """멘션 메일 알림 진단 (super admin만).

    - mail provider / Knox mail env 구성 여부(토큰 원문 제외, bool 로만 노출)
    - notification_events 상태별 카운트(pending/sent/failed/skipped/processing)
    - 최근 이벤트 20건(에러 메시지 포함) — 발송 API 문제 vs 멘션감지 문제 구분용
    """
    user = db.query(User).filter(User.id == user_id).first()
    if not user or (user.loginid or "").lower() not in [lid.lower() for lid in SUPER_ADMIN_LOGINIDS]:
        raise HTTPException(status_code=403, detail="Super admin only")

    from app.services import mail_sender
    from app.services.notification_processor import _processor_enabled, _interval_seconds
    from sqlalchemy import func as _func

    provider = mail_sender.get_mail_provider()
    knox_cfg = mail_sender._knox_config()

    status_rows = (
        db.query(NotificationEvent.status, _func.count(NotificationEvent.id))
        .group_by(NotificationEvent.status)
        .all()
    )
    status_counts = {str(s): int(c) for s, c in status_rows}

    recent = (
        db.query(NotificationEvent)
        .order_by(NotificationEvent.created_at.desc())
        .limit(20)
        .all()
    )
    recent_out = [
        {
            "id": e.id,
            "event_type": e.event_type,
            "status": e.status,
            "actor_user_id": e.actor_user_id,
            "target_user_id": e.target_user_id,
            "task_id": e.task_id,
            "error_message": e.error_message,
            "created_at": e.created_at.isoformat() if e.created_at else None,
            "processed_at": e.processed_at.isoformat() if e.processed_at else None,
        }
        for e in recent
    ]

    return {
        "mail_provider": provider,
        "mail_enabled": mail_sender.is_mail_enabled(),
        "processor_enabled": _processor_enabled(),
        "processor_interval_seconds": _interval_seconds(),
        "knox_mail_config": {
            # 토큰 원문은 절대 노출하지 않는다 — 구성 여부(bool)만.
            "base_url": knox_cfg["base_url"],
            "path": knox_cfg["path"],
            "user_id_query": knox_cfg["user_id"],
            "system_id_header_set": bool(knox_cfg["system_id"]),
            "sender_email_set": bool(knox_cfg["sender_email"]),
            "token_set": bool(knox_cfg["token"]),
            "content_type": knox_cfg["content_type"],
            "doc_secu_type": knox_cfg["doc_secu_type"],
        },
        "event_status_counts": status_counts,
        "recent_events": recent_out,
    }


@app.post("/api/admin/notifications/process-now")
def notification_process_now(user_id: int = Query(...), db: Session = Depends(get_db)):
    """pending 멘션 이벤트를 즉시 1회 처리 (super admin만). 처리 카운트 요약을 반환."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user or (user.loginid or "").lower() not in [lid.lower() for lid in SUPER_ADMIN_LOGINIDS]:
        raise HTTPException(status_code=403, detail="Super admin only")

    from app.services.notification_processor import process_pending_notification_events
    return process_pending_notification_events()


@app.get("/api/admin/s3/files")
def list_s3_storage(
    prefix: str = Query(default=""),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """S3 파일 목록 조회 (super admin만)"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user or (user.loginid or "").lower() not in [lid.lower() for lid in SUPER_ADMIN_LOGINIDS]:
        raise HTTPException(status_code=403, detail="Super admin only")

    from app.utils.s3.s3_utils import list_s3_files, is_s3_configured
    if not is_s3_configured():
        return {"files": [], "message": "S3 not configured"}

    files = list_s3_files(prefix)
    return {"files": files, "total": len(files)}


# =========================
# v3.0 공간 목적 기반 Overview API
# =========================

@app.get("/api/spaces/{space_id}/overview")
def get_space_overview(space_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    """공간 목적에 맞는 Overview 데이터 반환"""
    space = db.query(Space).filter(Space.id == space_id, Space.is_active == True).first()
    if not space:
        raise HTTPException(404, "Space not found")
    member = db.query(SpaceMember).filter(SpaceMember.space_id == space_id, SpaceMember.user_id == user_id).first()
    if not member:
        raise HTTPException(403, "접근 권한 없음")

    purpose = getattr(space, "purpose", None) or "project_management"
    state = load_state()
    today = date.today().isoformat()

    # 공통: 프로젝트 목록 (단발 일정의 task 가시성을 위해 시스템 프로젝트도 task 집계엔 포함)
    projects = db.query(Project).filter(Project.space_id == space_id, Project.archived_at.is_(None)).all()
    project_ids = [p.id for p in projects]
    # total_projects 카운트는 일반 프로젝트 목록 정책과 동일하게 시스템 프로젝트 제외
    visible_project_count = sum(
        1 for p in projects if not get_project_meta(state, p.id).get("is_system")
    )

    # 공통: Task 목록
    tasks = db.query(Task).filter(Task.project_id.in_(project_ids), Task.archived_at.is_(None)).all() if project_ids else []

    # 통계 계산
    my_tasks = [t for t in tasks if user_id in (t.assignee_ids or [])]
    overdue_tasks = [t for t in tasks if t.due_date and t.due_date < today and t.status not in ("done", "hold")]
    today_tasks = [t for t in tasks if t.due_date == today]
    in_progress = [t for t in tasks if t.status == "in_progress"]
    todo_tasks = [t for t in tasks if t.status == "todo"]
    done_tasks = [t for t in tasks if t.status == "done"]
    hold_tasks = [t for t in tasks if t.status == "hold"]

    # 이번 주 완료
    from datetime import timedelta
    week_start = (date.today() - timedelta(days=date.today().weekday())).isoformat()
    week_done = [t for t in done_tasks if t.updated_at and t.updated_at.strftime("%Y-%m-%d") >= week_start]

    # 다음 주 예정
    next_week_start = (date.today() + timedelta(days=(7 - date.today().weekday()))).isoformat()
    next_week_end = (date.today() + timedelta(days=(14 - date.today().weekday()))).isoformat()
    next_week_tasks = [t for t in tasks if t.due_date and next_week_start <= t.due_date < next_week_end and t.status not in ("done",)]

    # 우선순위 높은 Task: priority=high 이거나 마감 7일 이내 (overdue 포함, done/hold 제외).
    # 프론트에서 overdue→today→7d→high 정렬과 D-N 라벨링을 수행한다.
    seven_ahead = (date.today() + timedelta(days=7)).isoformat()
    high_priority = [
        t for t in tasks
        if t.status not in ("done", "hold") and (
            t.priority == "high"
            or (t.due_date and t.due_date <= seven_ahead)
        )
    ]

    # Sheet 실행 현황
    # 진행 중 sheet 는 status=="in_progress" + progress<100 + task_id 가 있는 것만 집계.
    # - 100% 도달했는데 status 가 자동으로 completed 로 플립되지 않은 경우에도 즉시 제외
    # - Task 에서 연결 해제된(task_id NULL) sheet 는 Dashboard 진행 중에 표시하지 않음
    # - 관리형 시트(sheet_type != "inspection") 는 진행률 개념이 없으므로 Dashboard 카운트/목록에서 제외.
    #   기존 데이터(NULL/legacy)는 default="inspection" 으로 잡혀있어 안전하다.
    sheet_executions_active = db.query(SheetExecution).filter(
        SheetExecution.space_id == space_id,
        SheetExecution.status == "in_progress",
        SheetExecution.progress < 100,
        SheetExecution.task_id.isnot(None),
        SheetExecution.sheet_type == "inspection",
    ).all() if project_ids or True else []
    # v3.11: progress=100 인데 아직 status="in_progress" 인 시트도 별도 노출.
    #   "진행 중 Sheet" 카드 카운트(active_sheets.length)는 그대로 유지(<100 만 카운트)하되,
    #   Check Sheet 현황 목록의 "완료된 항목 숨기기 OFF" 시에는 100% 시트도 표시되도록.
    sheet_executions_near_completed = db.query(SheetExecution).filter(
        SheetExecution.space_id == space_id,
        SheetExecution.status == "in_progress",
        SheetExecution.progress >= 100,
        SheetExecution.task_id.isnot(None),
        SheetExecution.sheet_type == "inspection",
    ).all()
    sheet_recent_completed = db.query(SheetExecution).filter(
        SheetExecution.space_id == space_id,
        SheetExecution.status == "completed",
        SheetExecution.sheet_type == "inspection",
    ).order_by(SheetExecution.completed_at.desc()).limit(20).all()

    def _task_brief(t):
        meta = get_task_meta(state, t.id)
        return {
            "id": t.id, "title": t.title, "status": t.status, "priority": t.priority,
            "start_date": t.start_date, "due_date": t.due_date,
            "progress": meta.get("progress", 0),
            "assignee_ids": t.assignee_ids or [], "project_id": t.project_id,
        }

    def _exec_brief(e):
        pn = next((p.name for p in projects if p.id == e.project_id), None)
        tn = next((t.title for t in tasks if t.id == e.task_id), None)
        return {
            "id": e.id, "title": e.title, "template_id": e.template_id,
            "status": e.status, "progress": e.progress,
            "equipment_name": e.equipment_name,
            "sheet_type": e.sheet_type,
            # v3.9: project_id / task_id 도 함께 노출 — 프론트에서 project 단위 grouping 시
            #       project_name 문자열이 아닌 id 로 안전하게 묶기 위함 (동명 프로젝트 충돌 방지).
            "project_id": e.project_id, "task_id": e.task_id,
            "project_name": pn, "task_name": tn,
            "started_at": iso(e.started_at), "completed_at": iso(e.completed_at),
        }

    base = {
        "purpose": purpose,
        "stats": {
            "total_projects": visible_project_count,
            "total_tasks": len(tasks),
            "my_tasks": len(my_tasks),
            "todo": len(todo_tasks),
            "in_progress": len(in_progress),
            "done": len(done_tasks),
            "hold": len(hold_tasks),
            "overdue": len(overdue_tasks),
        },
        "overdue_tasks": [_task_brief(t) for t in overdue_tasks[:10]],
        "today_tasks": [_task_brief(t) for t in today_tasks[:10]],
        "high_priority_tasks": [_task_brief(t) for t in high_priority[:10]],
        "in_progress_tasks": [_task_brief(t) for t in in_progress[:10]],
        "week_done_count": len(week_done),
        "next_week_tasks": [_task_brief(t) for t in next_week_tasks[:10]],
        "active_sheets": [_exec_brief(e) for e in sheet_executions_active],
        # v3.11: 100% 도달했지만 아직 status="in_progress" 인 시트 — Check Sheet 현황 목록에서
        #        "완료된 항목 숨기기 OFF" 시 표시 대상에 포함된다. 진행 중 카운트엔 미포함.
        "near_completed_sheets": [_exec_brief(e) for e in sheet_executions_near_completed],
        "recent_completed_sheets": [_exec_brief(e) for e in sheet_recent_completed],
    }

    # 목적별 추가 데이터
    if purpose == "equipment_ops":
        # 설비 운영: 미완료/이월 작업 강조
        incomplete = [t for t in tasks if t.status in ("todo", "in_progress") and t.due_date and t.due_date < today]
        base["incomplete_carried_over"] = [_task_brief(t) for t in incomplete[:10]]
    elif purpose == "sw_dev":
        # SW 개발: 개발 중 / 완료 / 대기 구분
        base["dev_in_progress"] = [_task_brief(t) for t in in_progress[:15]]
        base["dev_done_recent"] = [_task_brief(t) for t in week_done[:10]]

    return base


# =========================
# v3.0 Sheet Template API
# =========================

class SheetTemplateCreate(BaseModel):
    name: str
    description: Optional[str] = None
    category: Optional[str] = "general"

@app.post("/api/sheet-templates/inspect")
async def inspect_sheet_file(
    file: UploadFile = FastAPIFile(...),
):
    """업로드 전 파일의 Sheet 목록만 조회 (multi-sheet xlsx 선택 UX 용)."""
    contents = await file.read()
    filename = file.filename or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        return {"multi": False, "sheet_names": ["(CSV)"], "suggested": None}
    if ext not in ("xlsx", "xls"):
        raise HTTPException(400, f"지원하지 않는 파일 형식: .{ext} (xlsx, csv만 지원)")

    try:
        import io as _io
        import openpyxl
        wb = openpyxl.load_workbook(_io.BytesIO(contents), data_only=True, read_only=True)
        names = list(wb.sheetnames)
        active_title = wb.active.title if wb.active is not None else (names[0] if names else None)
        wb.close()
    except Exception as e:
        raise HTTPException(400, f"파일 열기 실패: {str(e)}")

    try:
        from app.services.sheet_parser import parse_excel_to_structure
        _, meta = parse_excel_to_structure(contents, filename, active_title)
        suggested_type = meta.get("suggested_type", "inspection")
    except Exception:
        suggested_type = "inspection"

    return {
        "multi": len(names) > 1,
        "sheet_names": names,
        "suggested": active_title,
        "suggested_type": suggested_type,
    }


@app.post("/api/sheet-templates/upload")
async def upload_sheet_template(
    file: UploadFile = FastAPIFile(...),
    space_id: int = Query(...),
    name: Optional[str] = Query(None),
    description: Optional[str] = Query(None),
    category: Optional[str] = Query("general"),
    sheet_name: Optional[str] = Query(None),
    sheet_type: Optional[str] = Query("inspection"),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """Excel/CSV 파일을 업로드하여 Sheet 템플릿으로 파싱 저장"""
    from app.services.sheet_parser import parse_excel_to_structure

    contents = await file.read()
    original_filename = file.filename or "unknown.xlsx"
    _parser_ext = original_filename.rsplit(".", 1)[-1].lower() if "." in original_filename else ""
    _parser_type = "csv" if _parser_ext == "csv" else ("xlsx" if _parser_ext in ("xlsx", "xls") else _parser_ext or "unknown")

    # 파싱
    #   사용자에게는 간단한 메시지, 서버 로그에는 traceback 전체 + 파일 메타를 남긴다.
    #   (회사 서버에서만 나는 오류의 근본 원인 추적을 위해 filename/content_type/size/parser_type 기록)
    try:
        _slog = logging.getLogger("main")
        _slog.info(
            "[SHEET_UPLOAD] parser=%s filename=%s content_type=%s size=%s",
            _parser_type, original_filename, getattr(file, "content_type", None), len(contents),
        )
        structure, meta = parse_excel_to_structure(contents, original_filename, sheet_name)
    except Exception as e:
        logging.getLogger("main").exception(
            "[SHEET_UPLOAD] file parse failed: parser=%s filename=%s content_type=%s size=%s",
            _parser_type, original_filename, getattr(file, "content_type", None), len(contents),
        )
        raise HTTPException(400, f"파일 파싱 실패: {str(e)}")

    # v3.1: 같은 구조 해시의 이전 매핑이 있으면 자동 적용
    s_hash = meta.get("structure_hash", "")
    prev_mapping = None
    if s_hash:
        prev_tpl = db.query(SheetTemplate).filter(
            SheetTemplate.structure_hash == s_hash,
            SheetTemplate.column_role_mapping.isnot(None),
        ).order_by(SheetTemplate.created_at.desc()).first()
        if prev_tpl:
            prev_mapping = prev_tpl.column_role_mapping

    template = SheetTemplate(
        space_id=space_id,
        name=name or os.path.splitext(original_filename)[0],
        description=description,
        category=category,
        original_filename=original_filename,
        sheet_name=meta.get("sheet_name"),
        sheet_type=sheet_type or meta.get("suggested_type", "inspection"),
        structure=structure,
        row_count=meta.get("row_count", 0),
        col_count=meta.get("col_count", 0),
        checkable_count=meta.get("checkable_count", 0),
        structure_hash=s_hash,
        column_role_mapping=prev_mapping,  # 이전 매핑 자동 적용
        created_by=user_id,
    )
    db.add(template)
    db.commit()
    db.refresh(template)

    return {
        "id": template.id,
        "name": template.name,
        "category": template.category,
        "original_filename": template.original_filename,
        "sheet_name": template.sheet_name,
        "row_count": template.row_count,
        "col_count": template.col_count,
        "checkable_count": template.checkable_count,
        "structure_hash": template.structure_hash,
        "column_roles": structure.get("column_roles", {}),
        "column_role_mapping": template.column_role_mapping,
        "created_at": iso(template.created_at),
    }


@app.get("/api/sheet-templates")
def list_sheet_templates(
    space_id: int = Query(...),
    category: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """공간 내 Sheet 템플릿 목록"""
    q = db.query(SheetTemplate).filter(SheetTemplate.space_id == space_id)
    if category:
        q = q.filter(SheetTemplate.category == category)
    templates = q.order_by(SheetTemplate.created_at.desc()).all()
    return {"templates": [
        {
            "id": t.id, "name": t.name, "description": t.description,
            "category": t.category, "original_filename": t.original_filename,
            "sheet_name": t.sheet_name,
            "row_count": t.row_count, "col_count": t.col_count,
            "checkable_count": t.checkable_count,
            "created_by": t.created_by,
            "created_at": iso(t.created_at),
        } for t in templates
    ]}


@app.get("/api/sheet-templates/{template_id}")
def get_sheet_template(template_id: int, db: Session = Depends(get_db)):
    """Sheet 템플릿 상세 (구조 포함)"""
    t = db.query(SheetTemplate).filter(SheetTemplate.id == template_id).first()
    if not t:
        raise HTTPException(404, "Template not found")
    return {
        "id": t.id, "name": t.name, "description": t.description,
        "category": t.category, "original_filename": t.original_filename,
        "sheet_name": t.sheet_name,
        "structure": t.structure,
        "row_count": t.row_count, "col_count": t.col_count,
        "checkable_count": t.checkable_count,
        "column_role_mapping": t.column_role_mapping,
        "structure_hash": t.structure_hash,
        "created_by": t.created_by,
        "created_at": iso(t.created_at),
    }


@app.post("/api/sheet-templates/{template_id}/confirm-roles")
def confirm_sheet_roles(
    template_id: int,
    roles: Dict[str, Any] = Body(...),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """사용자가 확인/수정한 컬럼 역할 매핑을 저장"""
    t = db.query(SheetTemplate).filter(SheetTemplate.id == template_id).first()
    if not t:
        raise HTTPException(404, "Template not found")
    t.column_role_mapping = roles
    db.commit()
    return {"message": "Column role mapping saved", "column_role_mapping": roles}


@app.delete("/api/sheet-templates/{template_id}")
def delete_sheet_template(template_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    """Sheet 템플릿 삭제 (연결된 실행본과 항목/로그도 함께 정리)"""
    t = db.query(SheetTemplate).filter(SheetTemplate.id == template_id).first()
    if not t:
        raise HTTPException(404, "Template not found")

    # SheetExecution.template_id FK 에는 ON DELETE CASCADE 가 없으므로 수동으로 정리.
    # SheetExecutionItem / SheetExecutionLog 는 execution_id 에 ondelete=CASCADE 가 있지만
    # SQLite 등에서 cascade 가 꺼져 있을 수 있어 명시적으로 먼저 지운다.
    exec_rows = db.query(SheetExecution.id).filter(SheetExecution.template_id == template_id).all()
    exec_ids = [r[0] for r in exec_rows]
    if exec_ids:
        db.query(SheetExecutionLog).filter(SheetExecutionLog.execution_id.in_(exec_ids)).delete(synchronize_session=False)
        db.query(SheetExecutionItem).filter(SheetExecutionItem.execution_id.in_(exec_ids)).delete(synchronize_session=False)
        db.query(SheetExecution).filter(SheetExecution.template_id == template_id).delete(synchronize_session=False)

    db.delete(t)
    db.commit()
    return {"message": "Template deleted", "deleted_executions": len(exec_ids)}


# =========================
# v3.0 Sheet Execution API
# =========================

class SheetExecutionCreate(BaseModel):
    template_id: int
    project_id: Optional[int] = None
    task_id: Optional[int] = None
    title: Optional[str] = None
    equipment_name: Optional[str] = None

@app.post("/api/sheet-executions")
def create_sheet_execution(
    body: SheetExecutionCreate,
    space_id: int = Query(...),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """Sheet 실행 시작 — 템플릿 기반으로 실행본 + 체크 항목 생성"""
    template = db.query(SheetTemplate).filter(SheetTemplate.id == body.template_id).first()
    if not template:
        raise HTTPException(404, "Template not found")

    execution = SheetExecution(
        template_id=template.id,
        project_id=body.project_id,
        task_id=body.task_id,
        space_id=space_id,
        title=body.title or f"{template.name} - {date.today().isoformat()}",
        equipment_name=body.equipment_name,
        sheet_type=template.sheet_type,
        status="in_progress",
        total_items=template.checkable_count,
        checked_items=0,
        progress=0,
        started_by=user_id,
    )
    db.add(execution)
    db.flush()

    # 템플릿의 checkable_cells로부터 실행 항목 생성 또는 매핑 생성
    structure = template.structure or {}
    initial_checked_count = 0

    if template.sheet_type == "assignment_mapping":
        from app.models import SheetExecutionMapping
        cells = structure.get("cells", [])
        data_start_row = structure.get("data_start_row", 1)
        headers = structure.get("headers", [])
        column_roles = structure.get("column_roles", {})
        
        assignment_cols = []
        for h in headers:
            text = h["value"]
            if "설비" in text or "장비" in text or "부품" in text or "적용" in text:
                assignment_cols.append(h["col"])
        
        manager_col = column_roles.get("assignee", {}).get("col", -1)
        remark_col = column_roles.get("remark", {}).get("col", -1)

        cell_map = {(c["row"], c["col"]): c["value"] for c in cells}
        max_row = max((c["row"] for c in cells), default=0)
        
        for r in range(data_start_row, max_row + 1):
            master_name = cell_map.get((r, 0), "")
            if not str(master_name).strip():
                continue
            master_code = cell_map.get((r, 1), "")
            
            manager = str(cell_map.get((r, manager_col), "")) if manager_col >= 0 else ""
            note = str(cell_map.get((r, remark_col), "")) if remark_col >= 0 else ""
            
            for c_idx in assignment_cols:
                assigned_entity = str(cell_map.get((r, c_idx), "")).strip()
                if assigned_entity:
                    mapping = SheetExecutionMapping(
                        execution_id=execution.id,
                        master_name=str(master_name).strip(),
                        master_code=str(master_code).strip() if master_code else None,
                        assigned_entity=assigned_entity,
                        manager=manager.strip() if manager else None,
                        note=note.strip() if note else None,
                    )
                    db.add(mapping)

        # 매핑 유형은 total_items를 실제 매핑 수로 설정
        db.flush()
        mapping_count = db.query(SheetExecutionMapping).filter(
            SheetExecutionMapping.execution_id == execution.id
        ).count()
        execution.total_items = mapping_count
    else:
        # 기존 점검형: 템플릿의 checkable_cells로부터 실행 항목 생성
        checkable_cells = structure.get("checkable_cells", [])
        COMPLETED_INITIAL = {"완료", "OK", "ok", "PASS", "pass", "양호", "O", "o", "○", "●", "✓", "✔", "☑"}
        checked_at_now = datetime.utcnow()
        for cell in checkable_cells:
            init_val = cell.get("initial_value")
            init_val_str = (str(init_val).strip() if init_val is not None else "")
            parsed_status = (cell.get("parsed_status") or "").strip()
            parsed_note = (cell.get("parsed_note") or "").strip()
            status_for_check = parsed_status or init_val_str
            pre_checked = status_for_check in COMPLETED_INITIAL
            if pre_checked:
                initial_checked_count += 1
            item = SheetExecutionItem(
                execution_id=execution.id,
                cell_ref=cell.get("ref", ""),
                row_idx=cell.get("row", 0),
                col_idx=cell.get("col", 0),
                label=cell.get("label", ""),
                checked=pre_checked,
                value=(parsed_status or init_val_str) or None,
                memo=parsed_note or None,
                checked_by=user_id if pre_checked else None,
                checked_at=checked_at_now if pre_checked else None,
            )
            db.add(item)

    if initial_checked_count and execution.total_items:
        execution.checked_items = initial_checked_count
        execution.progress = round(initial_checked_count / execution.total_items * 100)

    # 시작 로그
    db.add(SheetExecutionLog(
        execution_id=execution.id, action="start", user_id=user_id,
        new_value=f"실행 시작: {execution.title}",
    ))

    db.commit()
    db.refresh(execution)

    return {
        "id": execution.id,
        "template_id": execution.template_id,
        "project_id": execution.project_id,
        "task_id": execution.task_id,
        "title": execution.title,
        "status": execution.status,
        "total_items": execution.total_items,
        "progress": execution.progress,
        "started_at": iso(execution.started_at),
    }


@app.get("/api/sheet-executions")
def list_sheet_executions(
    space_id: int = Query(...),
    project_id: Optional[int] = Query(None),
    task_id: Optional[int] = Query(None),
    template_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    equipment_name: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),  # YYYY-MM-DD
    date_to: Optional[str] = Query(None),    # YYYY-MM-DD
    user_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """Sheet 실행 목록 조회 (필터 지원)"""
    q = db.query(SheetExecution).filter(SheetExecution.space_id == space_id)
    if project_id is not None:
        q = q.filter(SheetExecution.project_id == project_id)
    if task_id is not None:
        q = q.filter(SheetExecution.task_id == task_id)
    if template_id is not None:
        q = q.filter(SheetExecution.template_id == template_id)
    if status:
        q = q.filter(SheetExecution.status == status)
    if equipment_name:
        q = q.filter(SheetExecution.equipment_name.like(f"%{equipment_name}%"))
    if date_from:
        from sqlalchemy import text as _text
        q = q.filter(SheetExecution.started_at >= date_from)
    if date_to:
        q = q.filter(SheetExecution.started_at <= f"{date_to} 23:59:59")
    execs = q.order_by(SheetExecution.started_at.desc()).limit(200).all()

    # v3.12: 프론트 Dashboard 위젯(예: 약품관리)이 sheet_type 과 template.category 로
    # 카테고리별 분류를 하므로 응답에 함께 노출한다. N+1 방지용 단발성 조회.
    template_ids = {e.template_id for e in execs if e.template_id is not None}
    template_map: dict[int, SheetTemplate] = {}
    if template_ids:
        for t in db.query(SheetTemplate).filter(SheetTemplate.id.in_(template_ids)).all():
            template_map[t.id] = t

    return {"executions": [
        {
            "id": e.id, "template_id": e.template_id, "project_id": e.project_id,
            "task_id": e.task_id,
            "title": e.title, "equipment_name": e.equipment_name,
            "sheet_type": e.sheet_type,
            "template_name": (template_map.get(e.template_id).name if template_map.get(e.template_id) else None),
            "template_category": (template_map.get(e.template_id).category if template_map.get(e.template_id) else None),
            "status": e.status, "total_items": e.total_items,
            "checked_items": e.checked_items, "progress": e.progress,
            "started_by": e.started_by, "started_at": iso(e.started_at),
            "completed_at": iso(e.completed_at), "completed_by": e.completed_by,
        } for e in execs
    ]}


@app.get("/api/sheet-executions/{execution_id}")
def get_sheet_execution(execution_id: int, db: Session = Depends(get_db)):
    """Sheet 실행 상세 + 항목 + 템플릿 구조 + (assignment_mapping 유형의 경우) 매핑 목록"""
    from app.models import SheetExecutionMapping

    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")

    items = db.query(SheetExecutionItem).filter(
        SheetExecutionItem.execution_id == execution_id
    ).order_by(SheetExecutionItem.row_idx, SheetExecutionItem.col_idx).all()

    template = db.query(SheetTemplate).filter(SheetTemplate.id == execution.template_id).first()

    mappings_data = []
    if execution.sheet_type == "assignment_mapping":
        mappings = db.query(SheetExecutionMapping).filter(
            SheetExecutionMapping.execution_id == execution_id
        ).order_by(SheetExecutionMapping.master_name, SheetExecutionMapping.id).all()
        mappings_data = [
            {
                "id": m.id,
                "master_name": m.master_name,
                "master_code": m.master_code,
                "assigned_entity": m.assigned_entity,
                "manager": m.manager,
                "last_checked_at": iso(m.last_checked_at) if m.last_checked_at else None,
                "note": m.note,
            } for m in mappings
        ]

    return {
        "id": execution.id,
        "template_id": execution.template_id,
        "project_id": execution.project_id,
        "task_id": execution.task_id,
        "title": execution.title,
        "equipment_name": execution.equipment_name,
        "sheet_type": execution.sheet_type,
        "status": execution.status,
        "total_items": execution.total_items,
        "checked_items": execution.checked_items,
        "progress": execution.progress,
        "started_by": execution.started_by,
        "started_at": iso(execution.started_at),
        "completed_at": iso(execution.completed_at),
        "completed_by": execution.completed_by,
        "template_structure": template.structure if template else {},
        "template_name": template.name if template else "",
        "hidden_cols": execution.hidden_cols or [],
        "hidden_rows": execution.hidden_rows or [],
        "structure_overlay": execution.structure_overlay or {
            "added_columns": [],
            "added_rows": [],
            "renamed_headers": {},
        },
        "items": [
            {
                "id": item.id,
                "cell_ref": item.cell_ref,
                "row_idx": item.row_idx,
                "col_idx": item.col_idx,
                "label": item.label,
                "checked": item.checked,
                "value": item.value,
                "memo": item.memo,
                "checked_by": item.checked_by,
                "checked_at": iso(item.checked_at) if item.checked_at else None,
            } for item in items
        ],
        "mappings": mappings_data,
    }


@app.get("/api/sheet-executions/{execution_id}/mappings")
def list_sheet_mappings(execution_id: int, db: Session = Depends(get_db)):
    """assignment_mapping 유형 실행의 매핑 목록"""
    from app.models import SheetExecutionMapping

    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")

    mappings = db.query(SheetExecutionMapping).filter(
        SheetExecutionMapping.execution_id == execution_id
    ).order_by(SheetExecutionMapping.master_name, SheetExecutionMapping.id).all()
    return {
        "mappings": [
            {
                "id": m.id,
                "master_name": m.master_name,
                "master_code": m.master_code,
                "assigned_entity": m.assigned_entity,
                "manager": m.manager,
                "last_checked_at": iso(m.last_checked_at) if m.last_checked_at else None,
                "note": m.note,
            } for m in mappings
        ]
    }


@app.patch("/api/sheet-executions/{execution_id}/items/{item_id}")
def update_sheet_execution_item(
    execution_id: int,
    item_id: int,
    checked: Optional[bool] = Body(None),
    value: Optional[str] = Body(None),
    memo: Optional[str] = Body(None),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """실행 항목 체크/메모 업데이트"""
    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")
    if execution.status == "completed":
        raise HTTPException(400, "이미 완료된 실행입니다")

    item = db.query(SheetExecutionItem).filter(
        SheetExecutionItem.id == item_id, SheetExecutionItem.execution_id == execution_id
    ).first()
    if not item:
        raise HTTPException(404, "Item not found")

    old_checked = item.checked
    old_value = item.value

    if checked is not None:
        item.checked = checked
        item.checked_by = user_id
        item.checked_at = datetime.now(KST) if checked else None
    if value is not None:
        item.value = value
    if memo is not None:
        item.memo = memo

    # v3.12: 단일 항목이 "진행" 으로 전환되면 같은 행의 담당자 셀에 현재 사용자명을 기록.
    # 전환 판정:
    #   - checked False→True 또는
    #   - value 가 'O' 로 바뀐 경우 (기존 값이 'O' 가 아니었어야 함)
    became_progressed = (
        (checked is True and not old_checked)
        or (value is not None and (value or "").strip().upper() == "O"
            and (old_value or "").strip().upper() != "O")
    )
    if became_progressed:
        template = db.query(SheetTemplate).filter(SheetTemplate.id == execution.template_id).first()
        assignee_col = _resolve_assignee_col(template)
        # 담당자 컬럼 자체를 수정하는 경우는 자동 입력하지 않음 (사용자 의도 보호)
        if assignee_col is not None and assignee_col != item.col_idx:
            display_name = _get_user_display_name(db, user_id)
            if display_name:
                _fill_assignee_cell_if_empty(
                    db, execution_id, None, item.row_idx, assignee_col, display_name,
                )

    # 이력 로그
    if checked is not None and checked != old_checked:
        db.add(SheetExecutionLog(
            execution_id=execution_id, item_id=item_id,
            action="check" if checked else "uncheck",
            old_value=str(old_checked), new_value=str(checked),
            user_id=user_id,
        ))
    if memo is not None:
        db.add(SheetExecutionLog(
            execution_id=execution_id, item_id=item_id,
            action="memo", new_value=memo[:200] if memo else "",
            user_id=user_id,
        ))

    # 진행률 재계산 — N/A는 모수에서 제외 (해당 없음 의미)
    total = execution.total_items or 1
    na_count = db.query(SheetExecutionItem).filter(
        SheetExecutionItem.execution_id == execution_id,
        SheetExecutionItem.value == "N/A",
    ).count()
    effective_total = max(1, total - na_count)
    checked_count = db.query(SheetExecutionItem).filter(
        SheetExecutionItem.execution_id == execution_id, SheetExecutionItem.checked == True
    ).count()
    execution.checked_items = checked_count
    execution.progress = min(100, int(checked_count / effective_total * 100))

    db.commit()

    # v3.4: 시트가 task에 연결되어 있고 task에 체크박스 활동이 없으면
    #       시트 진행률로 task.progress 자동 동기화
    if execution.task_id:
        try:
            _sync_task_progress(db, execution.task_id)
        except Exception:
            pass

    return {
        "id": item.id,
        "checked": item.checked,
        "value": item.value,
        "memo": item.memo,
        "checked_by": item.checked_by,
        "checked_at": iso(item.checked_at) if item.checked_at else None,
        "execution_progress": execution.progress,
        "execution_checked_items": execution.checked_items,
    }

class CellUpsertRequest(BaseModel):
    value: Optional[str] = None
    memo: Optional[str] = None
    checked: Optional[bool] = None

@app.patch("/api/sheet-executions/{execution_id}/cells/{cell_ref}")
def upsert_sheet_execution_cell(
    execution_id: int,
    cell_ref: str,
    body: CellUpsertRequest,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """특정 셀(cell_ref)의 값을 동적으로 Upsert (담당자/비고 등 임의의 컬럼 수정용)"""
    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")
    if execution.status == "completed":
        raise HTTPException(400, "이미 완료된 실행입니다")

    item = db.query(SheetExecutionItem).filter(
        SheetExecutionItem.execution_id == execution_id,
        SheetExecutionItem.cell_ref == cell_ref
    ).first()

    # 셀의 row, col 인덱스를 파싱 (임시 방편: A1 등 정규식으로 파싱 가능하지만, 없으면 0으로 처리)
    import re
    row_idx, col_idx = 0, 0
    match = re.match(r"^([A-Z]+)(\d+)$", cell_ref)
    if match:
        col_str = match.group(1)
        col_idx = 0
        for char in col_str:
            col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
        col_idx -= 1
        row_idx = int(match.group(2)) - 1

    old_value = None
    old_checked = False
    if not item:
        item = SheetExecutionItem(
            execution_id=execution_id,
            cell_ref=cell_ref,
            row_idx=row_idx,
            col_idx=col_idx,
            checked=False,
            value=body.value,
            memo=body.memo
        )
        if body.checked:
            item.checked = True
            item.checked_by = user_id
            item.checked_at = datetime.now(KST)
        db.add(item)
    else:
        old_value = item.value
        old_checked = item.checked
        if body.checked is not None:
            item.checked = body.checked
            if body.checked and not old_checked:
                item.checked_by = user_id
                item.checked_at = datetime.now(KST)
            elif not body.checked and old_checked:
                item.checked_at = None
        if body.value is not None:
            item.value = body.value
        if body.memo is not None:
            item.memo = body.memo

    # v3.12: 상태 dropdown 으로 'O' 진행 처리할 때(개별 진행)도 같은 행의 담당자 셀에
    # 현재 사용자명을 자동 기록한다. update_sheet_execution_item 과 동일 정책.
    became_progressed = (
        (body.checked is True and not old_checked)
        or (body.value is not None and (body.value or "").strip().upper() == "O"
            and (old_value or "").strip().upper() != "O")
    )
    if became_progressed:
        template = db.query(SheetTemplate).filter(SheetTemplate.id == execution.template_id).first()
        assignee_col = _resolve_assignee_col(template)
        # 담당자 컬럼 자체를 수정하는 경우는 자동 입력하지 않음 (사용자 의도 보호)
        if assignee_col is not None and assignee_col != item.col_idx:
            display_name = _get_user_display_name(db, user_id)
            if display_name:
                _fill_assignee_cell_if_empty(
                    db, execution_id, None, item.row_idx, assignee_col, display_name,
                )

    db.commit()

    # 진행률 재계산 — checkable_cells 위치의 항목만 모수에 포함 (N/A 제외)
    checked_count, total, progress = _recompute_sheet_progress(db, execution)
    db.commit()

    # v3.4: task progress 동기화 (시트만 있는 task 대응)
    # v3.6: 동기화된 task progress를 응답에 포함해서 프론트가 즉시 반영하게 함
    # v3.11: task_status 도 함께 응답 — Check Sheet 100% 완료 시 Task Details Status
    #        (Done) 가 새로고침 없이 즉시 반영되도록.
    task_progress: int | None = None
    task_status: str | None = None
    if execution.task_id:
        try:
            task_progress = _sync_task_progress(db, execution.task_id)
            task_after = db.query(Task).filter(Task.id == execution.task_id).first()
            if task_after:
                task_status = task_after.status
        except Exception:
            pass

    return {
        "status": "ok",
        "progress": progress,
        "checked_items": checked_count,
        "total_items": total,
        "task_id": execution.task_id,
        "task_progress": task_progress,
        "task_status": task_status,
    }


@app.patch("/api/sheet-executions/{execution_id}/hidden-cols")
def update_sheet_hidden_cols(
    execution_id: int,
    body: dict = Body(...),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """실행본에서 숨길 컬럼 인덱스 배열 저장 (template은 변경하지 않음).
    body: { "hidden_cols": [int, ...] }
    """
    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")
    if execution.status == "completed":
        raise HTTPException(400, "이미 완료된 실행입니다")

    raw = body.get("hidden_cols") or []
    if not isinstance(raw, list):
        raise HTTPException(400, "hidden_cols must be an array")
    cleaned = sorted({int(v) for v in raw if isinstance(v, (int, float)) and int(v) >= 0})

    # role 컬럼은 삭제 불가 (진행률 계산에 영향)
    template = db.query(SheetTemplate).filter(SheetTemplate.id == execution.template_id).first()
    structure = (template.structure if template else {}) or {}
    roles = structure.get("column_roles") or {}
    protected: set[int] = set()
    for key in ("check_status", "checked_at", "progress_date"):
        col = (roles.get(key) or {}).get("col")
        if isinstance(col, int) and col >= 0:
            protected.add(col)
    blocked = [c for c in cleaned if c in protected]
    if blocked:
        raise HTTPException(400, f"진행 상태/점검일자 컬럼은 삭제할 수 없습니다 (col={blocked})")

    execution.hidden_cols = cleaned
    db.commit()
    return {"hidden_cols": cleaned}


@app.patch("/api/sheet-executions/{execution_id}/hidden-rows")
def update_sheet_hidden_rows(
    execution_id: int,
    body: dict = Body(...),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """실행본에서 숨길 행 인덱스 배열 저장 (template은 변경하지 않음).
    body: { "hidden_rows": [int, ...] }
    """
    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")
    if execution.status == "completed":
        raise HTTPException(400, "이미 완료된 실행입니다")

    raw = body.get("hidden_rows") or []
    if not isinstance(raw, list):
        raise HTTPException(400, "hidden_rows must be an array")
    cleaned = sorted({int(v) for v in raw if isinstance(v, (int, float)) and int(v) >= 0})

    # 헤더 행은 삭제 불가
    template = db.query(SheetTemplate).filter(SheetTemplate.id == execution.template_id).first()
    structure = (template.structure if template else {}) or {}
    header_row_idx = structure.get("header_row_idx")
    if isinstance(header_row_idx, int) and header_row_idx in cleaned:
        raise HTTPException(400, f"헤더 행은 삭제할 수 없습니다 (row={header_row_idx})")

    execution.hidden_rows = cleaned
    db.commit()
    return {"hidden_rows": cleaned}


@app.patch("/api/sheet-executions/{execution_id}/structure-overlay")
def update_sheet_structure_overlay(
    execution_id: int,
    body: dict = Body(...),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """실행본 단위 구조 오버레이 저장 (전체 덮어쓰기).

    body: {
      "added_columns": [{"col_idx": int, "header": str}, ...],
      "added_rows":    [{"row_idx": int}, ...],
      "renamed_headers": {"<col_idx>": "<header>"},
    }

    이번 라운드에서는 added_columns만 실제로 사용된다. added_rows /
    renamed_headers 는 슬롯만 유지하므로 유효성 검증은 형 체크 수준에서만
    수행한다.
    """
    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")
    if execution.status == "completed":
        raise HTTPException(400, "이미 완료된 실행입니다")

    raw_added_cols = body.get("added_columns") or []
    raw_added_rows = body.get("added_rows") or []
    raw_renamed = body.get("renamed_headers") or {}
    raw_cols_order = body.get("added_columns_order") or []
    raw_rows_order = body.get("added_rows_order") or []
    # column_order / row_order: full render order INCLUDING base col_idx / row_idx so that
    # added cols/rows can be interleaved between base ones. added_columns_order /
    # added_rows_order remain for back-compat; older clients keep working.
    raw_column_order = body.get("column_order")
    raw_row_order = body.get("row_order")

    if not isinstance(raw_added_cols, list):
        raise HTTPException(400, "added_columns must be an array")
    if not isinstance(raw_added_rows, list):
        raise HTTPException(400, "added_rows must be an array")
    if not isinstance(raw_renamed, dict):
        raise HTTPException(400, "renamed_headers must be an object")
    if not isinstance(raw_cols_order, list):
        raise HTTPException(400, "added_columns_order must be an array")
    if not isinstance(raw_rows_order, list):
        raise HTTPException(400, "added_rows_order must be an array")
    if raw_column_order is not None and not isinstance(raw_column_order, list):
        raise HTTPException(400, "column_order must be an array")
    if raw_row_order is not None and not isinstance(raw_row_order, list):
        raise HTTPException(400, "row_order must be an array")

    # added_columns 정리: col_idx(int) + header(str) 만 유지, col_idx 중복 제거(나중에 들어온 게 우선).
    # kind: 'equipment' | 'plain' — 표시 방식을 구분 (legacy/누락 시 'equipment' 로 간주, 기존 호환).
    def _normalize_added_col(col_idx: int, header_raw, kind_raw) -> dict:
        h = header_raw if isinstance(header_raw, str) else (str(header_raw) if header_raw is not None else "")
        k = kind_raw if isinstance(kind_raw, str) and kind_raw in ("equipment", "plain") else None
        out = {"col_idx": col_idx, "header": h}
        if k:
            out["kind"] = k
        return out

    added_columns: list[dict] = []
    seen_col_idx: set[int] = set()
    for entry in raw_added_cols:
        if not isinstance(entry, dict):
            continue
        col_raw = entry.get("col_idx")
        if not isinstance(col_raw, (int, float)):
            continue
        col_idx = int(col_raw)
        if col_idx < 0:
            continue
        normalized = _normalize_added_col(col_idx, entry.get("header"), entry.get("kind"))
        if col_idx in seen_col_idx:
            # 동일 col_idx 가 두 번 들어오면 마지막 항목으로 덮어쓴다
            for i, existing in enumerate(added_columns):
                if existing["col_idx"] == col_idx:
                    added_columns[i] = normalized
                    break
            continue
        seen_col_idx.add(col_idx)
        added_columns.append(normalized)
    added_columns.sort(key=lambda x: x["col_idx"])

    # added_rows: row_idx(int) 만 유지
    added_rows: list[dict] = []
    seen_row_idx: set[int] = set()
    for entry in raw_added_rows:
        if not isinstance(entry, dict):
            continue
        row_raw = entry.get("row_idx")
        if not isinstance(row_raw, (int, float)):
            continue
        row_idx = int(row_raw)
        if row_idx < 0 or row_idx in seen_row_idx:
            continue
        seen_row_idx.add(row_idx)
        added_rows.append({"row_idx": row_idx})
    added_rows.sort(key=lambda x: x["row_idx"])

    # renamed_headers: key=col_idx(str), value=header(str)
    renamed_headers: dict[str, str] = {}
    for k, v in raw_renamed.items():
        try:
            col_idx = int(k)
        except (TypeError, ValueError):
            continue
        if col_idx < 0:
            continue
        if not isinstance(v, str):
            v = str(v) if v is not None else ""
        renamed_headers[str(col_idx)] = v

    # added_columns_order / added_rows_order: 정수 배열, 중복 제거.
    # added_columns / added_rows 에 실제로 존재하는 idx 만 살리고, 누락된 idx 는 끝에 자동 보충.
    valid_col_idxs = {c["col_idx"] for c in added_columns}
    valid_row_idxs = {r["row_idx"] for r in added_rows}

    def _normalize_order(raw: list, valid: set[int]) -> list[int]:
        seen: set[int] = set()
        out: list[int] = []
        for v in raw:
            if not isinstance(v, (int, float)):
                continue
            iv = int(v)
            if iv in seen or iv not in valid:
                continue
            seen.add(iv)
            out.append(iv)
        # 누락된 idx 는 오름차순으로 뒤에 보충 (신규 추가/삭제 시 일관성 유지)
        for iv in sorted(valid - seen):
            out.append(iv)
        return out

    added_columns_order = _normalize_order(raw_cols_order, valid_col_idxs)
    added_rows_order = _normalize_order(raw_rows_order, valid_row_idxs)

    # column_order / row_order: full render order. We don't know base col_idx range here
    # so we accept any non-negative int, dedupe, and let the frontend filter against the
    # actual sheet structure at render time. Added idxs not present in this payload are
    # appended at the end (asc) so a client that forgets to include them still gets a
    # stable result.
    def _normalize_full_order(raw: list | None, must_include: set[int]) -> list[int] | None:
        if raw is None:
            return None
        seen: set[int] = set()
        out: list[int] = []
        for v in raw:
            if not isinstance(v, (int, float)):
                continue
            iv = int(v)
            if iv < 0 or iv in seen:
                continue
            seen.add(iv)
            out.append(iv)
        for iv in sorted(must_include - seen):
            out.append(iv)
        return out

    column_order = _normalize_full_order(raw_column_order, valid_col_idxs)
    row_order = _normalize_full_order(raw_row_order, valid_row_idxs)

    overlay = {
        "added_columns": added_columns,
        "added_rows": added_rows,
        "renamed_headers": renamed_headers,
        "added_columns_order": added_columns_order,
        "added_rows_order": added_rows_order,
    }
    if column_order is not None:
        overlay["column_order"] = column_order
    if row_order is not None:
        overlay["row_order"] = row_order
    execution.structure_overlay = overlay
    db.commit()
    return {"structure_overlay": overlay}


@app.post("/api/sheet-executions/{execution_id}/copy")
def copy_sheet_execution(
    execution_id: int,
    title: str = Query(...),
    include_data: bool = Query(False),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """기존 실행본을 복제하여 새로운 실행본 생성"""
    original = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not original:
        raise HTTPException(404, "Execution not found")
        
    execution = SheetExecution(
        template_id=original.template_id,
        project_id=original.project_id,
        task_id=original.task_id,
        space_id=original.space_id,
        title=title,
        equipment_name=original.equipment_name,
        status="in_progress",
        total_items=original.total_items,
        checked_items=0,
        progress=0,
        started_by=user_id,
    )
    db.add(execution)
    db.flush()

    # 데이터 복제
    if include_data:
        original_items = db.query(SheetExecutionItem).filter(SheetExecutionItem.execution_id == execution_id).all()
        checked_count = 0
        for item in original_items:
            new_item = SheetExecutionItem(
                execution_id=execution.id,
                cell_ref=item.cell_ref,
                row_idx=item.row_idx,
                col_idx=item.col_idx,
                label=item.label,
                checked=item.checked,
                value=item.value,
                memo=item.memo,
                checked_by=user_id if item.checked else None,
                checked_at=datetime.now(KST) if item.checked else None,
            )
            if new_item.checked:
                checked_count += 1
            db.add(new_item)
            
        execution.checked_items = checked_count
        execution.progress = min(100, int((checked_count / (execution.total_items or 1)) * 100))
    
    db.commit()
    
    return {"id": execution.id}

@app.patch("/api/sheet-executions/{execution_id}/complete")
def complete_sheet_execution(
    execution_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """Sheet 실행 완료 처리"""
    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")

    execution.status = "completed"
    execution.completed_at = datetime.now(KST)
    execution.completed_by = user_id

    db.add(SheetExecutionLog(
        execution_id=execution_id, action="complete",
        new_value=f"완료 (진행률: {execution.progress}%)", user_id=user_id,
    ))

    db.commit()

    # v3.4: task progress 동기화 (시트만 있는 task의 경우 100%로 반영)
    if execution.task_id:
        try:
            _sync_task_progress(db, execution.task_id)
        except Exception:
            pass

    return {"message": "실행이 완료되었습니다", "progress": execution.progress}


@app.patch("/api/sheet-executions/{execution_id}/unlink-task")
def unlink_sheet_execution_task(
    execution_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """Sheet 실행과 Task의 연결만 해제 (sheet 자체는 보존).

    v3.9: 연결 해제 시 status="unlinked"로 명시 표시.
          Dashboard 진행 중 sheet count / Sheets 관리 화면에서 unlinked 를 별도로
          식별할 수 있게 한다 (이전엔 task_id=NULL만으로는 "원래 task 없이 시작한 시트"와
          "task에서 해제된 시트"를 구분할 수 없었음).
          단, completed 상태인 시트는 이력 보존을 위해 status를 바꾸지 않는다.
    """
    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")
    old_task_id = execution.task_id
    if old_task_id is None:
        return {"message": "이미 Task와 연결되어 있지 않습니다", "task_id": None}
    old_status = execution.status
    execution.task_id = None
    if execution.status not in ("completed", "cancelled"):
        execution.status = "unlinked"
    db.add(SheetExecutionLog(
        execution_id=execution_id, action="unlink_task",
        old_value=f"task_id={old_task_id},status={old_status}",
        new_value=f"task_id=None,status={execution.status}",
        user_id=user_id,
    ))
    db.commit()

    # v3.4: 연결 해제된 task의 progress도 재계산 (다른 시트가 남았는지 등)
    # v3.6: 다른 시트/체크박스 소스가 남아있지 않으면 명시적으로 0%로 동기화한다
    #       (이전 시트 progress가 잔재로 남는 문제 방지)
    task_progress: int | None = None
    try:
        task_progress = _sync_task_progress(db, old_task_id, force_zero_if_empty=True)
    except Exception:
        pass

    return {
        "message": "Task 연결이 해제되었습니다",
        "task_id": None,
        "old_task_id": old_task_id,
        "task_progress": task_progress,
    }


class MarkAllProgressRequest(BaseModel):
    keep_na: bool = True


@app.post("/api/sheet-executions/{execution_id}/mark-all-progress")
def mark_all_sheet_progress(
    execution_id: int,
    body: MarkAllProgressRequest = Body(default_factory=MarkAllProgressRequest),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """v3.9: 시트의 모든 항목을 한 번에 "진행" 처리한다.

    - keep_na=True (기본): value=="N/A" 인 항목은 건드리지 않는다 (사용자가 의도적으로
      N/A로 둔 항목을 강제로 진행 처리하지 않기 위해).
    - keep_na=False: N/A 도 진행 처리.
    - 이미 checked=True 인 항목은 skip (변경 카운트에서도 제외).
    - 진행률 재계산 후 _sync_task_progress 로 task progress/status 도 즉시 반영.
    - 응답에 task_progress/task_status 까지 포함시켜 프론트가 추가 fetch 없이
      Board / Dashboard / TaskDrawer 를 즉시 업데이트할 수 있게 한다.
    """
    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")
    if execution.status == "completed":
        raise HTTPException(400, "이미 완료된 실행입니다")

    items = db.query(SheetExecutionItem).filter(
        SheetExecutionItem.execution_id == execution_id
    ).all()

    # v3.11: 진행일자 컬럼 위치 (template.column_role_mapping.progress_date.col)
    #        ALL 진행으로 'O' 가 되는 행은 같은 row의 progress_date 컬럼 셀에 오늘 날짜를
    #        함께 기록한다. 개별 dropdown 변경 시 프론트가 같은 행위를 하므로 일관 유지.
    template = db.query(SheetTemplate).filter(SheetTemplate.id == execution.template_id).first()
    progress_date_col: int | None = None
    try:
        roles = (template.column_role_mapping if template else None) or {}
        # 일부 양식은 structure.column_roles 에만 매핑이 있을 수 있어 폴백 확인
        if not roles and template and isinstance(template.structure, dict):
            roles = template.structure.get("column_roles") or {}
        pd = roles.get("progress_date") if isinstance(roles, dict) else None
        if isinstance(pd, dict) and isinstance(pd.get("col"), int):
            progress_date_col = pd["col"]
    except Exception:
        progress_date_col = None

    # 담당자 컬럼 위치 + 현재 사용자 표시명 (진행으로 바뀐 행의 담당자 셀이 비어있을 때만 자동 기록)
    assignee_col = _resolve_assignee_col(template)
    assignee_name = _get_user_display_name(db, user_id) if assignee_col is not None else None

    # 체크 가능 셀 위치만 ALL 진행 대상이다. 진행일자/비고/담당자 등 비체크 셀은 건드리지 않는다.
    structure = (template.structure if template else {}) or {}
    checkable_cells = structure.get("checkable_cells") or []
    checkable_positions = {(int(c["row"]), int(c["col"])) for c in checkable_cells if "row" in c and "col" in c}

    def _col_letter(col_idx: int) -> str:
        s = ""
        n = col_idx + 1
        while n > 0:
            n -= 1
            s = chr(65 + (n % 26)) + s
            n //= 26
        return s

    today_ymd = datetime.now(KST).strftime("%Y-%m-%d")

    # 같은 execution 내 (row_idx, col_idx) → SheetExecutionItem 빠른 lookup
    items_by_rc: dict[tuple[int, int], SheetExecutionItem] = {
        (it.row_idx, it.col_idx): it for it in items
    }

    now = datetime.now(KST)
    updated_count = 0
    skipped_na_count = 0
    progressed_rows: set[int] = set()
    for it in items:
        # 체크 가능 셀이 아닌 항목(진행일자/비고 등)은 건너뛴다.
        if checkable_positions and (it.row_idx, it.col_idx) not in checkable_positions:
            continue
        if it.value == "N/A":
            if body.keep_na:
                skipped_na_count += 1
                continue
        if it.checked and (it.value or "").strip().upper() == "O":
            continue
        it.checked = True
        it.checked_by = user_id
        it.checked_at = now
        # 미진행/빈 값/null/X → 모두 'O'(진행) 로 통일.
        # 기존엔 value="X"(미진행) 인 행이 checked=True 인데 value="X" 로 남는 모순 발생.
        it.value = "O"
        updated_count += 1
        progressed_rows.add(it.row_idx)

    # 진행일자 컬럼이 있고 진행으로 바뀐 행이 있으면 해당 행의 진행일자 셀에 오늘 날짜 기록
    if progress_date_col is not None and progressed_rows:
        col_letter = _col_letter(progress_date_col)
        for row_idx in progressed_rows:
            existing = items_by_rc.get((row_idx, progress_date_col))
            if existing is not None:
                existing.value = today_ymd
            else:
                new_item = SheetExecutionItem(
                    execution_id=execution_id,
                    cell_ref=f"{col_letter}{row_idx + 1}",
                    row_idx=row_idx,
                    col_idx=progress_date_col,
                    checked=False,
                    value=today_ymd,
                )
                db.add(new_item)
                items_by_rc[(row_idx, progress_date_col)] = new_item

    # 담당자 컬럼이 있고 진행으로 바뀐 행이 있으면 같은 행의 담당자 셀에 현재 사용자명 기록
    # (이미 값이 있으면 유지 — 사용자 수기 입력 우선)
    if assignee_col is not None and assignee_name and progressed_rows:
        for row_idx in progressed_rows:
            _fill_assignee_cell_if_empty(
                db, execution_id, items_by_rc, row_idx, assignee_col, assignee_name,
            )

    if updated_count > 0:
        db.add(SheetExecutionLog(
            execution_id=execution_id,
            action="mark_all_progress",
            old_value=f"unchecked->checked count={updated_count}",
            new_value=f"keep_na={body.keep_na}",
            user_id=user_id,
        ))

    # 진행률 재계산 — checkable_cells 위치 항목만 모수 (N/A 제외)
    checked_count, total, _progress = _recompute_sheet_progress(db, execution)
    db.commit()

    # Task 동기화
    task_progress: int | None = None
    task_status: str | None = None
    if execution.task_id:
        try:
            task_progress = _sync_task_progress(db, execution.task_id)
            task = db.query(Task).filter(Task.id == execution.task_id).first()
            if task:
                task_status = task.status
        except Exception:
            pass

    return {
        "success": True,
        "sheet_execution_id": execution.id,
        "updated_count": updated_count,
        "skipped_na_count": skipped_na_count,
        "total_items": total,
        "checked_items": execution.checked_items,
        "progress": execution.progress,
        "task_id": execution.task_id,
        "task_progress": task_progress,
        "task_status": task_status,
    }


@app.delete("/api/sheet-executions/{execution_id}")
def delete_sheet_execution(
    execution_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """Sheet 실행본 영구 삭제 (운영자 직접 DB 삭제를 없애기 위한 안전 엔드포인트).

    v3.9: 다음 상태에서만 삭제 허용:
        - unlinked  : Task 에서 연결 해제됨 (Dashboard 잔존 정리용)
        - cancelled : 취소된 실행
        - completed : 완료된 실행 (이력 정리)
        - in_progress 이지만 task_id IS NULL (어떤 task 에도 연결 안된 진행중 시트)
    Task 에 연결된 in_progress 실행본은 삭제 불가 — 진행 중 작업의 진행률 소스를
    실수로 날리는 것을 방지한다. 필요 시 먼저 연결 해제(unlink)를 거친 뒤 삭제.

    SheetExecutionItem / SheetExecutionLog 는 ondelete=CASCADE 로 자동 정리되지만
    SheetExecutionMapping 은 ondelete=CASCADE 가 없을 수 있으므로 명시적으로 정리.
    """
    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")

    # in_progress + task_id 가 살아있는 task 를 가리킬 때만 차단.
    # task 가 이미 archive(soft-delete) 되었거나 사라졌다면 사실상 미연결이므로 삭제 허용.
    if execution.status == "in_progress" and execution.task_id is not None:
        linked_task = db.query(Task).filter(Task.id == execution.task_id).first()
        if linked_task is not None and linked_task.archived_at is None:
            raise HTTPException(
                400,
                "Task 에 연결된 진행 중 시트는 삭제할 수 없습니다. 먼저 연결 해제 후 삭제하세요.",
            )

    # SheetExecutionMapping 수동 정리 (cascade 없을 수 있음)
    try:
        from app.models import SheetExecutionMapping
        db.query(SheetExecutionMapping).filter(
            SheetExecutionMapping.execution_id == execution_id
        ).delete(synchronize_session=False)
    except Exception:
        pass

    # 감사 로그를 한 줄 남기고 삭제 (CASCADE 로 곧 사라지지만 다른 시트와의 분리 추적은 어려움)
    db.add(SheetExecutionLog(
        execution_id=execution_id, action="delete",
        old_value=f"status={execution.status},task_id={execution.task_id}",
        new_value="deleted",
        user_id=user_id,
    ))
    db.commit()

    db.delete(execution)
    db.commit()

    return {"message": "Sheet 실행본이 삭제되었습니다", "execution_id": execution_id}


@app.get("/api/sheet-executions/{execution_id}/export")
def export_sheet_execution_xlsx(
    execution_id: int,
    db: Session = Depends(get_db),
):
    """현재 웹에서 수정된 최신 상태를 xlsx 파일로 다운로드."""
    import io as _io
    from urllib.parse import quote as _quote
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(404, "Execution not found")
    template = db.query(SheetTemplate).filter(SheetTemplate.id == execution.template_id).first()
    if not template:
        raise HTTPException(404, "Template not found")

    structure = template.structure or {}
    cells = structure.get("cells", [])
    merges = structure.get("merges", [])
    col_widths = structure.get("col_widths", [])
    row_heights = structure.get("row_heights", [])
    total_rows = structure.get("total_rows", 0) or 0
    total_cols = structure.get("total_cols", 0) or 0
    column_roles = structure.get("column_roles") or {}

    items = db.query(SheetExecutionItem).filter(
        SheetExecutionItem.execution_id == execution_id
    ).all()
    # cell_ref → (value, memo, checked) 맵
    item_map = {}
    for it in items:
        item_map[(it.row_idx, it.col_idx)] = it

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (template.sheet_name or template.name or "Sheet")[:31]

    # 1) 원본 셀 값 + style 복원
    for cell_data in cells:
        r = cell_data.get("row", 0) + 1
        c = cell_data.get("col", 0) + 1
        if r < 1 or c < 1:
            continue
        # 사용자가 web에서 수정한 값이 있으면 우선 적용
        item = item_map.get((cell_data.get("row", 0), cell_data.get("col", 0)))
        value = cell_data.get("value", "")
        if item is not None:
            if item.value is not None and item.value != "":
                value = item.value
            elif item.checked:
                value = "O"
        try:
            target = ws.cell(row=r, column=c, value=value if value != "" else None)
        except Exception:
            continue
        # 폰트
        font_info = cell_data.get("font") or {}
        if font_info:
            target.font = Font(
                bold=bool(font_info.get("bold")),
                italic=bool(font_info.get("italic")),
                size=font_info.get("fontSize") or 11,
                color=font_info.get("fontColor", "FF000000").lstrip("#") if font_info.get("fontColor") else None,
            )
        # 정렬
        align = cell_data.get("align")
        if align or cell_data.get("wrapText"):
            target.alignment = Alignment(
                horizontal=align if align in ("left", "center", "right") else None,
                vertical="center",
                wrap_text=bool(cell_data.get("wrapText")),
            )
        # 배경
        bg = cell_data.get("bg")
        if bg:
            try:
                target.fill = PatternFill(start_color=bg.lstrip("#"), end_color=bg.lstrip("#"), fill_type="solid")
            except Exception:
                pass

    # 2) 사용자 수정값 중 원본 cell이 없던 위치(가상 컬럼 등) 추가
    seen_positions = {(c.get("row", 0), c.get("col", 0)) for c in cells}
    for it in items:
        pos = (it.row_idx, it.col_idx)
        if pos in seen_positions:
            continue
        try:
            v = it.value if it.value not in (None, "") else ("O" if it.checked else None)
            ws.cell(row=it.row_idx + 1, column=it.col_idx + 1, value=v)
        except Exception:
            continue

    # 2-1) 가상 컬럼 헤더 보강 (원본에 없는 가상 컬럼은 헤더 cell이 있어도 다시 적용)
    header_row_1based = (structure.get("header_row_idx") or 0) + 1
    for role_key, role_info in column_roles.items():
        if not role_info or not role_info.get("virtual"):
            continue
        col_1based = (role_info.get("col") or 0) + 1
        header_text = role_info.get("header") or role_key
        try:
            ws.cell(row=header_row_1based, column=col_1based, value=header_text).font = Font(bold=True, color="FF6B7280")
        except Exception:
            pass

    # 3) 병합셀
    for mg in merges:
        try:
            ws.merge_cells(
                start_row=mg["startRow"] + 1,
                start_column=mg["startCol"] + 1,
                end_row=mg["endRow"] + 1,
                end_column=mg["endCol"] + 1,
            )
        except Exception:
            continue

    # 4) 컬럼 너비
    from openpyxl.utils import get_column_letter as _col_letter
    for idx, w in enumerate(col_widths or []):
        if idx >= total_cols:
            break
        try:
            ws.column_dimensions[_col_letter(idx + 1)].width = float(w)
        except Exception:
            continue

    # 5) 행 높이
    for idx, h in enumerate(row_heights or []):
        if idx >= total_rows:
            break
        try:
            ws.row_dimensions[idx + 1].height = float(h)
        except Exception:
            continue

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    fname_base = (execution.title or template.name or "sheet").replace("/", "_").replace("\\", "_")
    encoded = _quote(f"{fname_base}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@app.get("/api/sheet-executions/{execution_id}/logs")
def get_sheet_execution_logs(execution_id: int, db: Session = Depends(get_db)):
    """실행 이력 로그 조회"""
    logs = db.query(SheetExecutionLog).filter(
        SheetExecutionLog.execution_id == execution_id
    ).order_by(SheetExecutionLog.created_at.desc()).limit(200).all()
    return {"logs": [
        {
            "id": l.id, "action": l.action, "item_id": l.item_id,
            "old_value": l.old_value, "new_value": l.new_value,
            "memo": l.memo, "user_id": l.user_id,
            "created_at": iso(l.created_at),
        } for l in logs
    ]}


@app.get("/api/tasks/{task_id}/sheet-summary")
def get_task_sheet_summary(task_id: int, db: Session = Depends(get_db)):
    """Task 에 직접 연결된 Sheet 실행 요약 (Task Details 패널용)"""
    execs = db.query(SheetExecution).filter(SheetExecution.task_id == task_id).order_by(SheetExecution.started_at.desc()).all()
    active = [e for e in execs if e.status == "in_progress"]
    completed = [e for e in execs if e.status == "completed"]
    return {
        "task_id": task_id,
        "total_executions": len(execs),
        "active_count": len(active),
        "completed_count": len(completed),
        "active_executions": [
            {"id": e.id, "title": e.title, "progress": e.progress, "template_id": e.template_id,
             "sheet_type": e.sheet_type, "started_at": iso(e.started_at)}
            for e in active
        ],
        "recent_completed": [
            {"id": e.id, "title": e.title, "progress": e.progress, "template_id": e.template_id,
             "sheet_type": e.sheet_type, "completed_at": iso(e.completed_at)}
            for e in completed[:10]
        ],
    }


@app.get("/api/projects/{project_id}/sheet-summary")
def get_project_sheet_summary(project_id: int, db: Session = Depends(get_db)):
    """프로젝트에 연결된 Sheet 실행 요약"""
    execs = db.query(SheetExecution).filter(SheetExecution.project_id == project_id).order_by(SheetExecution.started_at.desc()).all()
    active = [e for e in execs if e.status == "in_progress"]
    completed = [e for e in execs if e.status == "completed"]

    # 연결된 task 의 archive 상태를 한 번에 조회 — 프론트가 삭제 가능 여부를 백엔드와 동일 기준으로 판단할 수 있도록.
    linked_task_ids = {e.task_id for e in execs if e.task_id is not None}
    task_archived_map: dict[int, bool] = {}
    if linked_task_ids:
        for t in db.query(Task).filter(Task.id.in_(linked_task_ids)).all():
            task_archived_map[t.id] = t.archived_at is not None

    def _serialize(e: SheetExecution, *, include_completed: bool = False) -> dict:
        # task_id 가 가리키는 task 가 사라졌으면 archived 로 간주(미연결과 동일하게 취급).
        task_archived = True if e.task_id is not None and task_archived_map.get(e.task_id, True) else False
        d = {
            "id": e.id,
            "title": e.title,
            "progress": e.progress,
            "template_id": e.template_id,
            "task_id": e.task_id,
            "task_archived": task_archived,
            "started_at": iso(e.started_at),
        }
        if include_completed:
            d["completed_at"] = iso(e.completed_at)
        return d

    return {
        "project_id": project_id,
        "total_executions": len(execs),
        "active_count": len(active),
        "completed_count": len(completed),
        "active_executions": [_serialize(e) for e in active],
        "recent_completed": [_serialize(e, include_completed=True) for e in completed[:10]],
    }


class SheetMappingCreate(BaseModel):
    master_name: str
    assigned_entity: str

class SheetMappingUpdate(BaseModel):
    master_name: Optional[str] = None
    assigned_entity: Optional[str] = None
    manager: Optional[str] = None
    note: Optional[str] = None

@app.post("/api/sheet-executions/{execution_id}/mappings")
def add_sheet_mapping(execution_id: int, body: SheetMappingCreate, user_id: int = Query(...), db: Session = Depends(get_db)):
    from app.models import SheetExecutionMapping
    execution = db.query(SheetExecution).filter(SheetExecution.id == execution_id).first()
    if not execution: raise HTTPException(404, "Execution not found")

    mapping = SheetExecutionMapping(
        execution_id=execution_id,
        master_name=body.master_name,
        assigned_entity=body.assigned_entity,
        manager=None, note=None,
    )
    db.add(mapping)
    
    db.add(SheetExecutionLog(
        execution_id=execution_id, action="add_mapping", user_id=user_id,
        new_value=f"[{body.master_name}]에 [{body.assigned_entity}] 추가"
    ))
    db.commit()
    db.refresh(mapping)
    return mapping

@app.patch("/api/sheet-executions/{execution_id}/mappings/{mapping_id}")
def update_sheet_mapping(execution_id: int, mapping_id: int, body: SheetMappingUpdate, user_id: int = Query(...), db: Session = Depends(get_db)):
    from app.models import SheetExecutionMapping
    mapping = db.query(SheetExecutionMapping).filter(SheetExecutionMapping.id == mapping_id, SheetExecutionMapping.execution_id == execution_id).first()
    if not mapping: raise HTTPException(404, "Mapping not found")

    old_master = mapping.master_name
    old_assigned = mapping.assigned_entity

    log_msg = []
    if body.master_name is not None and body.master_name != mapping.master_name:
        mapping.master_name = body.master_name
        log_msg.append(f"마스터 변경: {old_master} -> {body.master_name}")
    if body.assigned_entity is not None and body.assigned_entity != mapping.assigned_entity:
        mapping.assigned_entity = body.assigned_entity
        log_msg.append(f"설비 변경: {old_assigned} -> {body.assigned_entity}")
    if body.manager is not None: mapping.manager = body.manager
    if body.note is not None: mapping.note = body.note

    mapping.last_checked_at = datetime.utcnow()

    if log_msg:
        db.add(SheetExecutionLog(
            execution_id=execution_id, action="update_mapping", user_id=user_id,
            old_value=f"[{old_master}] {old_assigned}",
            new_value=", ".join(log_msg)
        ))

    db.commit()
    db.refresh(mapping)
    return mapping

@app.delete("/api/sheet-executions/{execution_id}/mappings/{mapping_id}")
def delete_sheet_mapping(execution_id: int, mapping_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    from app.models import SheetExecutionMapping
    mapping = db.query(SheetExecutionMapping).filter(SheetExecutionMapping.id == mapping_id, SheetExecutionMapping.execution_id == execution_id).first()
    if not mapping: raise HTTPException(404, "Mapping not found")

    db.add(SheetExecutionLog(
        execution_id=execution_id, action="delete_mapping", user_id=user_id,
        old_value=f"[{mapping.master_name}]에 배정된 [{mapping.assigned_entity}] 삭제"
    ))
    db.delete(mapping)
    db.commit()
    return {"message": "Mapping deleted"}

# =========================
# VOC / 개선 요청 (전역 피드백)
# =========================
# 모든 로그인 사용자가 작성 가능. 본인은 본인 글 조회, 관리자/super_admin 은 전체 조회 + 상태 변경.
# 분류/관련 화면/우선순위/상태 값은 프론트와 합의된 키만 허용.
VOC_CATEGORIES = ("bug", "improvement", "feature", "ux", "design", "etc")
VOC_RELATED_SCREENS = ("dashboard", "space", "project", "settings", "sheet", "messenger", "ai_report", "etc")
VOC_PRIORITIES = ("low", "normal", "high", "urgent")
VOC_STATUSES = ("received", "reviewing", "in_progress", "completed", "on_hold")
# 공개 범위. public = 모든 로그인 사용자 조회 가능, private = 작성자/관리자만.
VOC_VISIBILITIES = ("public", "private")

class VocCreate(BaseModel):
    title: str
    category: str
    content: str
    related_screen: Optional[str] = None
    priority: Optional[str] = "normal"
    # 신규 작성분은 기본 공개. 민감한 내용이면 작성자가 비공개로 등록할 수 있다.
    visibility: Optional[str] = "public"

class VocUpdate(BaseModel):
    # 본문 수정(작성자 또는 관리자) — 어느 필드든 None 이면 변경 없음
    title: Optional[str] = None
    category: Optional[str] = None
    content: Optional[str] = None
    related_screen: Optional[str] = None
    # 관리자 전용 — 일반 사용자가 보내도 권한 체크에서 거절
    status: Optional[str] = None
    priority: Optional[str] = None
    admin_note: Optional[str] = None
    # 관리자 메모 저장 시 "작성자에게 메일로 관리자 메모 알림 보내기" 체크 여부(기본 OFF).
    # admin_note 가 실제로 변경되고 send_mail=true 일 때만 voc_admin_memo_mail_requested 이벤트 생성.
    send_mail: bool = False
    # 공개/비공개 전환은 관리자 전용 (개인정보/보안정보 보호)
    visibility: Optional[str] = None
    # 삭제할 첨부 id 목록 (선택)
    remove_attachment_ids: Optional[List[int]] = None

class VocCommentCreate(BaseModel):
    content: str
    # 추가 문의/댓글 입력은 메일 알림을 발생시키지 않는다.
    # (하위호환: 구버전 프론트가 send_mail 을 보내더라도 pydantic 이 무시하도록 필드를 두지 않는다.)

# 개선요청 첨부 이미지 최적화 파라미터 (장수/용량/확장자 제한은 공통 VOC_IMAGE_POLICY 사용)
VOC_ATTACHMENT_MAX_EDGE = 1920                 # 긴 변 기준 리사이즈
VOC_ATTACHMENT_WEBP_QUALITY = 80

# 본문은 contentEditable 에디터에서 만든 제한된 HTML 만 들어온다.
# 그래도 변조/직접 호출 대비로 위험한 요소만 강하게 제거 (allowlist 기반은 프론트가 담당).
_VOC_DANGEROUS_TAG_RE = re.compile(
    r"<\s*(script|style|iframe|object|embed|link|meta|form|input|button|svg)\b[^>]*>"
    r"(.*?<\s*/\s*\1\s*>)?",
    re.IGNORECASE | re.DOTALL,
)
_VOC_ON_ATTR_RE = re.compile(r"\son[a-z]+\s*=\s*(?:\"[^\"]*\"|'[^']*'|[^\s>]+)", re.IGNORECASE)
_VOC_JS_URL_RE = re.compile(r"(href|src)\s*=\s*([\"']?)\s*javascript:[^\"'>\s]*\2", re.IGNORECASE)

def _strip_dangerous_html(s: str) -> str:
    if not s:
        return s
    s = _VOC_DANGEROUS_TAG_RE.sub("", s)
    s = _VOC_ON_ATTR_RE.sub("", s)
    s = _VOC_JS_URL_RE.sub(r"\1=\2\2", s)
    return s

def _voc_attachment_dict(a: VocAttachment) -> dict:
    return {
        "id": a.id,
        "voc_id": a.voc_id,
        "filename": a.filename,
        "stored_name": a.stored_name,
        "content_type": a.content_type,
        "file_size": a.file_size,
        "width": a.width,
        "height": a.height,
        "url": f"/api/voc/{a.voc_id}/attachments/{a.stored_name}/download",
        "created_at": iso(a.created_at) if a.created_at else None,
    }

def _voc_dict(
    v: VocItem,
    users_map: Dict[int, User],
    attachments: Optional[List[VocAttachment]] = None,
    vote_count: int = 0,
    voted_by_me: bool = False,
) -> dict:
    author = users_map.get(v.author_id)
    resolver = users_map.get(v.resolved_by) if v.resolved_by else None
    editor = users_map.get(v.edited_by) if v.edited_by else None
    return {
        "id": v.id,
        "author_id": v.author_id,
        "author_name": author.username if author else None,
        "author_loginid": author.loginid if author else None,
        "author_color": author.avatar_color if author else None,
        "title": v.title,
        "category": v.category,
        "content": v.content,
        "related_screen": v.related_screen,
        "priority": v.priority,
        "status": v.status,
        "visibility": v.visibility or "public",
        "admin_note": v.admin_note,
        "resolved_by": v.resolved_by,
        "resolver_name": resolver.username if resolver else None,
        "edited_by": v.edited_by,
        "editor_name": editor.username if editor else None,
        "edited_at": iso(v.edited_at) if v.edited_at else None,
        "change_summary": v.change_summary,
        "created_at": iso(v.created_at) if v.created_at else None,
        "updated_at": iso(v.updated_at) if v.updated_at else None,
        "attachments": [_voc_attachment_dict(a) for a in (attachments or [])],
        "attachment_count": len(attachments) if attachments is not None else 0,
        "vote_count": vote_count,
        "voted_by_me": voted_by_me,
    }

def _is_voc_admin(db: Session, user_id: int) -> bool:
    role = get_user_role(db, user_id)
    return role in ("admin", "super_admin")

@app.post("/api/voc")
def create_voc(payload: VocCreate, user_id: int = Query(...), db: Session = Depends(get_db)):
    author = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not author:
        raise HTTPException(status_code=403, detail="로그인이 필요합니다.")

    title = (payload.title or "").strip()
    content = (payload.content or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="제목을 입력해 주세요.")
    if not content:
        raise HTTPException(status_code=400, detail="내용을 입력해 주세요.")
    if payload.category not in VOC_CATEGORIES:
        raise HTTPException(status_code=400, detail="유효하지 않은 분류입니다.")
    if payload.related_screen and payload.related_screen not in VOC_RELATED_SCREENS:
        raise HTTPException(status_code=400, detail="유효하지 않은 관련 화면입니다.")
    priority = payload.priority or "normal"
    if priority not in VOC_PRIORITIES:
        raise HTTPException(status_code=400, detail="유효하지 않은 우선순위입니다.")
    visibility = (payload.visibility or "public").lower()
    if visibility not in VOC_VISIBILITIES:
        raise HTTPException(status_code=400, detail="유효하지 않은 공개 설정입니다.")

    voc = VocItem(
        author_id=user_id,
        title=title[:200],
        category=payload.category,
        content=_strip_dangerous_html(content),
        related_screen=payload.related_screen,
        priority=priority,
        status="received",
        visibility=visibility,
    )
    db.add(voc)
    db.commit()
    db.refresh(voc)
    # VOC 는 전역 데이터(space 없음) → 전역 채널(0)로 전파. voc.id 확정 후 emit.
    emit_realtime_event(
        db, space_id=REALTIME_GLOBAL_CHANNEL, event_type="voc_created",
        entity_type="voc", entity_id=voc.id, actor_user_id=user_id,
    )
    db.commit()
    users_map = {author.id: author}
    return _voc_dict(voc, users_map, attachments=[])

@app.get("/api/voc")
def list_voc(
    user_id: int = Query(...),
    scope: Optional[str] = Query(None),   # "mine" | "all" — admin 만 all 가능
    status: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1),
    page_size: int = Query(20),
    sort_order: str = Query("desc"),
    db: Session = Depends(get_db),
):
    caller = db.query(User).filter(User.id == user_id).first()
    if not caller:
        raise HTTPException(status_code=403, detail="로그인이 필요합니다.")
    is_admin = _is_voc_admin(db, user_id)
    scope = (scope or "all").lower()
    page, page_size = _norm_page(page, page_size, default_size=20, max_size=100)

    q = db.query(VocItem)
    if scope == "mine":
        # 본인 글만 (일반 사용자/관리자 공통)
        q = q.filter(VocItem.author_id == user_id)
    elif not is_admin:
        # 일반 사용자 전체 보기: 공개 VOC + 본인 비공개 글
        q = q.filter((VocItem.visibility == "public") | (VocItem.author_id == user_id))
    # 관리자 + scope=all → 비공개 포함 전체 조회 (필터 없음)
    if status:
        if status not in VOC_STATUSES:
            raise HTTPException(status_code=400, detail="유효하지 않은 상태 필터입니다.")
        q = q.filter(VocItem.status == status)
    if category:
        if category not in VOC_CATEGORIES:
            raise HTTPException(status_code=400, detail="유효하지 않은 분류 필터입니다.")
        q = q.filter(VocItem.category == category)
    if search and search.strip():
        kw = f"%{search.strip()}%"
        q = q.filter((VocItem.title.like(kw)) | (VocItem.content.like(kw)))

    total = q.count()
    order_col = (
        VocItem.created_at.asc() if (sort_order or "desc").lower() == "asc"
        else VocItem.created_at.desc()
    )
    rows = (
        q.order_by(order_col, VocItem.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    user_ids = {r.author_id for r in rows}
    user_ids.update({r.resolved_by for r in rows if r.resolved_by})
    user_ids.update({r.edited_by for r in rows if r.edited_by})
    users_map = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()} if user_ids else {}

    # 첨부 일괄 조회 후 voc_id 별로 그룹핑 (N+1 회피)
    voc_ids = [r.id for r in rows]
    attachments_by_voc: Dict[int, List[VocAttachment]] = {}
    if voc_ids:
        att_rows = (
            db.query(VocAttachment)
            .filter(VocAttachment.voc_id.in_(voc_ids))
            .filter(VocAttachment.comment_scoped.isnot(True))  # 댓글 이미지는 본문 첨부에서 제외
            .order_by(VocAttachment.created_at.asc())
            .all()
        )
        for a in att_rows:
            attachments_by_voc.setdefault(a.voc_id, []).append(a)

    # 공감 수 + 내가 공감한 voc 일괄 집계 (N+1 회피)
    vote_count_by_voc: Dict[int, int] = {}
    my_voted_ids: set = set()
    if voc_ids:
        for vid, cnt in (
            db.query(VocVote.voc_id, func.count(VocVote.id))
            .filter(VocVote.voc_id.in_(voc_ids))
            .group_by(VocVote.voc_id)
            .all()
        ):
            vote_count_by_voc[vid] = cnt
        my_voted_ids = {
            r[0]
            for r in db.query(VocVote.voc_id)
            .filter(VocVote.voc_id.in_(voc_ids), VocVote.user_id == user_id)
            .all()
        }

    return {
        "items": [
            _voc_dict(
                r,
                users_map,
                attachments=attachments_by_voc.get(r.id, []),
                vote_count=vote_count_by_voc.get(r.id, 0),
                voted_by_me=r.id in my_voted_ids,
            )
            for r in rows
        ],
        "is_admin": is_admin,
        "pagination": _pagination_meta(page, page_size, total),
    }

@app.get("/api/voc/stats")
def voc_stats(user_id: int = Query(...), db: Session = Depends(get_db)):
    """VOC 처리율 통계. 일반 사용자는 공개 VOC + 본인 글 기준, 관리자는 전체 기준."""
    caller = db.query(User).filter(User.id == user_id).first()
    if not caller:
        raise HTTPException(status_code=403, detail="로그인이 필요합니다.")
    is_admin = _is_voc_admin(db, user_id)
    q = db.query(VocItem.status, func.count(VocItem.id))
    if not is_admin:
        q = q.filter((VocItem.visibility == "public") | (VocItem.author_id == user_id))
    by_status = {s: 0 for s in VOC_STATUSES}
    for s, c in q.group_by(VocItem.status).all():
        if s in by_status:
            by_status[s] = c
    total = sum(by_status.values())
    completed = by_status.get("completed", 0)
    completion_rate = round(completed / total * 100) if total else 0
    return {
        "total": total,
        "completed": completed,
        "completion_rate": completion_rate,
        "by_status": by_status,
    }


@app.get("/api/voc/{voc_id}")
def get_voc(voc_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    v = db.query(VocItem).filter(VocItem.id == voc_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="VOC를 찾을 수 없습니다.")
    if v.visibility != "public" and v.author_id != user_id and not _is_voc_admin(db, user_id):
        raise HTTPException(status_code=403, detail="이 VOC를 조회할 권한이 없습니다.")
    user_ids = {v.author_id}
    if v.resolved_by:
        user_ids.add(v.resolved_by)
    if v.edited_by:
        user_ids.add(v.edited_by)
    users_map = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()}
    attachments = (
        db.query(VocAttachment)
        .filter(VocAttachment.voc_id == voc_id)
        .filter(VocAttachment.comment_scoped.isnot(True))  # 댓글 이미지는 본문 첨부에서 제외
        .order_by(VocAttachment.created_at.asc())
        .all()
    )
    vote_count = db.query(func.count(VocVote.id)).filter(VocVote.voc_id == voc_id).scalar() or 0
    voted_by_me = (
        db.query(VocVote.id).filter(VocVote.voc_id == voc_id, VocVote.user_id == user_id).first() is not None
    )
    return _voc_dict(
        v, users_map, attachments=attachments,
        vote_count=vote_count, voted_by_me=voted_by_me,
    )

@app.patch("/api/voc/{voc_id}")
def update_voc(voc_id: int, payload: VocUpdate, user_id: int = Query(...), db: Session = Depends(get_db)):
    """본인은 본인 글 본문(제목/분류/내용/관련 화면) 및 첨부 삭제 가능.
    상태/우선순위/관리자 메모 변경은 관리자(admin/super_admin) 전용."""
    caller = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not caller:
        raise HTTPException(status_code=403, detail="로그인이 필요합니다.")

    v = db.query(VocItem).filter(VocItem.id == voc_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="VOC를 찾을 수 없습니다.")

    is_admin = _is_voc_admin(db, user_id)
    is_author = (v.author_id == user_id)
    # 관리자도 아니고 작성자도 아니면 어떤 수정도 불가
    if not is_admin and not is_author:
        raise HTTPException(status_code=403, detail="이 VOC를 수정할 권한이 없습니다.")

    # 관리자 전용 필드는 본인이라도 관리자 권한이 없으면 거절
    admin_only_attempted = (
        payload.status is not None
        or payload.priority is not None
        or payload.admin_note is not None
        or payload.visibility is not None
    )
    if admin_only_attempted and not is_admin:
        raise HTTPException(status_code=403, detail="상태/우선순위/공개 설정/관리자 메모 변경은 관리자만 가능합니다.")

    changed_fields: List[str] = []

    # ── 본문 수정 (작성자 또는 관리자) ──
    if payload.title is not None:
        title = (payload.title or "").strip()
        if not title:
            raise HTTPException(status_code=400, detail="제목을 입력해 주세요.")
        if v.title != title[:200]:
            v.title = title[:200]
            changed_fields.append("title")
    if payload.category is not None:
        if payload.category not in VOC_CATEGORIES:
            raise HTTPException(status_code=400, detail="유효하지 않은 분류입니다.")
        if v.category != payload.category:
            v.category = payload.category
            changed_fields.append("category")
    if payload.content is not None:
        content = (payload.content or "").strip()
        if not content:
            raise HTTPException(status_code=400, detail="내용을 입력해 주세요.")
        content = _strip_dangerous_html(content)
        if v.content != content:
            v.content = content
            changed_fields.append("content")
    if payload.related_screen is not None:
        rs = payload.related_screen or None
        if rs and rs not in VOC_RELATED_SCREENS:
            raise HTTPException(status_code=400, detail="유효하지 않은 관련 화면입니다.")
        if v.related_screen != rs:
            v.related_screen = rs
            changed_fields.append("related_screen")

    # ── 관리자 전용 필드 ──
    if payload.status is not None:
        if payload.status not in VOC_STATUSES:
            raise HTTPException(status_code=400, detail="유효하지 않은 상태입니다.")
        if v.status != payload.status:
            v.status = payload.status
            if payload.status == "completed":
                v.resolved_by = user_id
            changed_fields.append("status")
    if payload.priority is not None:
        if payload.priority not in VOC_PRIORITIES:
            raise HTTPException(status_code=400, detail="유효하지 않은 우선순위입니다.")
        if v.priority != payload.priority:
            v.priority = payload.priority
            changed_fields.append("priority")
    admin_note_changed = False
    if payload.admin_note is not None:
        if (v.admin_note or "") != payload.admin_note:
            v.admin_note = payload.admin_note
            changed_fields.append("admin_note")
            admin_note_changed = True
    if payload.visibility is not None:
        if payload.visibility not in VOC_VISIBILITIES:
            raise HTTPException(status_code=400, detail="유효하지 않은 공개 설정입니다.")
        if v.visibility != payload.visibility:
            v.visibility = payload.visibility
            changed_fields.append("visibility")

    # ── 첨부 삭제 (작성자 본인 글 또는 관리자) ──
    if payload.remove_attachment_ids:
        to_remove = (
            db.query(VocAttachment)
            .filter(
                VocAttachment.id.in_(payload.remove_attachment_ids),
                VocAttachment.voc_id == voc_id,
            )
            .all()
        )
        for att in to_remove:
            _delete_voc_attachment_storage(att)
            db.delete(att)
        if to_remove:
            changed_fields.append("attachments")

    if not changed_fields:
        return {"message": "변경 없음"}

    # 본문/첨부 등 사용자 인지 가능한 변경이 있으면 edit 메타 갱신
    user_visible = {"title", "category", "content", "related_screen", "attachments"}
    if any(f in user_visible for f in changed_fields):
        v.edited_by = user_id
        v.edited_at = datetime.now()
        v.change_summary = ",".join(sorted(set(changed_fields)))[:255]

    # ── 관리자 메모 메일 알림 (관리자 opt-in) ──
    # 관리자 메모가 실제로 변경되고 "메일 보내기"를 체크한 경우에만 outbox 이벤트 생성 → processor 가 발송.
    # self 여부와 무관하게 발송 허용(관리자가 명시적으로 선택). 수신자 이메일 없으면 processor 가 skip.
    if admin_note_changed and is_admin and bool(getattr(payload, "send_mail", False)):
        db.add(NotificationEvent(
            event_type="voc_admin_memo_mail_requested",
            task_id=None, project_id=None, space_id=None,
            actor_user_id=user_id, target_user_id=v.author_id,
            payload_json={
                "voc_id": v.id,
                "source": "voc_admin_memo",
                "send_mail": True,
            },
            status="pending",
        ))

    emit_realtime_event(
        db, space_id=REALTIME_GLOBAL_CHANNEL, event_type="voc_updated",
        entity_type="voc", entity_id=v.id, actor_user_id=user_id,
    )
    db.commit()
    db.refresh(v)

    user_ids = {v.author_id}
    if v.resolved_by:
        user_ids.add(v.resolved_by)
    if v.edited_by:
        user_ids.add(v.edited_by)
    users_map = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()}
    attachments = (
        db.query(VocAttachment)
        .filter(VocAttachment.voc_id == voc_id)
        .order_by(VocAttachment.created_at.asc())
        .all()
    )
    return _voc_dict(v, users_map, attachments=attachments)

@app.delete("/api/voc/{voc_id}")
def delete_voc(voc_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    v = db.query(VocItem).filter(VocItem.id == voc_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="VOC를 찾을 수 없습니다.")
    # 본인 글은 본인이, 그 외에는 관리자만 삭제 가능
    if v.author_id != user_id and not _is_voc_admin(db, user_id):
        raise HTTPException(status_code=403, detail="이 VOC를 삭제할 권한이 없습니다.")
    # 첨부 파일 스토리지 정리 (DB 레코드는 FK CASCADE 또는 명시 삭제)
    attachments = db.query(VocAttachment).filter(VocAttachment.voc_id == voc_id).all()
    for att in attachments:
        _delete_voc_attachment_storage(att)
        db.delete(att)
    db.delete(v)
    db.commit()
    return {"message": "삭제되었습니다."}


# =========================
# VOC 공감 ('나도 필요해요')
# =========================
# 조회 가능한 VOC(공개글 또는 본인/관리자)에만 공감 가능. 사용자당 1회.
def _voc_visible_to(v: VocItem, db: Session, user_id: int) -> bool:
    return v.visibility == "public" or v.author_id == user_id or _is_voc_admin(db, user_id)


@app.post("/api/voc/{voc_id}/vote")
def add_voc_vote(voc_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    from sqlalchemy.exc import IntegrityError
    caller = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not caller:
        raise HTTPException(status_code=403, detail="로그인이 필요합니다.")
    v = db.query(VocItem).filter(VocItem.id == voc_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="VOC를 찾을 수 없습니다.")
    if not _voc_visible_to(v, db, user_id):
        raise HTTPException(status_code=403, detail="이 VOC에 공감할 권한이 없습니다.")

    existing = db.query(VocVote).filter(VocVote.voc_id == voc_id, VocVote.user_id == user_id).first()
    if not existing:
        db.add(VocVote(voc_id=voc_id, user_id=user_id))
        try:
            db.commit()
        except IntegrityError:
            # 동시 중복 요청 — 이미 공감 처리됨
            db.rollback()
    vote_count = db.query(func.count(VocVote.id)).filter(VocVote.voc_id == voc_id).scalar() or 0
    return {"voc_id": voc_id, "vote_count": vote_count, "voted_by_me": True}


@app.delete("/api/voc/{voc_id}/vote")
def remove_voc_vote(voc_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    existing = db.query(VocVote).filter(VocVote.voc_id == voc_id, VocVote.user_id == user_id).first()
    if existing:
        db.delete(existing)
        db.commit()
    vote_count = db.query(func.count(VocVote.id)).filter(VocVote.voc_id == voc_id).scalar() or 0
    return {"voc_id": voc_id, "vote_count": vote_count, "voted_by_me": False}


# =========================
# VOC 댓글 / 추가 문의 (관리자 답변 이후 대화 이어가기)
# =========================
# 정책(1차):
#  - 조회: 해당 VOC 를 볼 수 있는 사용자(공개글 / 본인 / 관리자)면 댓글 조회 가능
#  - 작성: VOC 작성자 본인 또는 관리자(admin/super_admin) 만 가능
#           (공개 VOC 라도 제3의 일반 사용자는 1차에서 댓글 작성 불가)
#  - 삭제: 댓글 작성자 본인 또는 관리자
#  - 상태 변경: 1차에서는 댓글 작성이 VOC 상태를 자동으로 바꾸지 않음(관리자 수동 처리)

def _voc_comment_dict(c: VocComment, users_map: Dict[int, User]) -> dict:
    author = users_map.get(c.author_id)
    return {
        "id": c.id,
        "voc_id": c.voc_id,
        "author_id": c.author_id,
        "author_name": author.username if author else None,
        "author_color": author.avatar_color if author else None,
        "is_admin_reply": bool(c.is_admin_reply),
        "content": c.content,
        "created_at": iso(c.created_at) if c.created_at else None,
        "updated_at": iso(c.updated_at) if c.updated_at else None,
    }


def _can_comment_on_voc(v: VocItem, db: Session, user_id: int) -> bool:
    """1차 정책: VOC 작성자 본인 또는 관리자만 댓글 작성 가능."""
    return v.author_id == user_id or _is_voc_admin(db, user_id)


@app.get("/api/voc/{voc_id}/comments")
def list_voc_comments(voc_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    v = db.query(VocItem).filter(VocItem.id == voc_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="VOC를 찾을 수 없습니다.")
    if not _voc_visible_to(v, db, user_id):
        raise HTTPException(status_code=403, detail="이 VOC를 조회할 권한이 없습니다.")

    rows = (
        db.query(VocComment)
        .filter(VocComment.voc_id == voc_id)
        .order_by(VocComment.created_at.asc(), VocComment.id.asc())
        .all()
    )
    author_ids = {r.author_id for r in rows}
    users_map = (
        {u.id: u for u in db.query(User).filter(User.id.in_(author_ids)).all()}
        if author_ids else {}
    )
    return {
        "items": [_voc_comment_dict(c, users_map) for c in rows],
        "can_comment": _can_comment_on_voc(v, db, user_id),
        "is_admin": _is_voc_admin(db, user_id),
    }


@app.post("/api/voc/{voc_id}/comments")
def create_voc_comment(
    voc_id: int,
    payload: VocCommentCreate,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    caller = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not caller:
        raise HTTPException(status_code=403, detail="로그인이 필요합니다.")
    v = db.query(VocItem).filter(VocItem.id == voc_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="VOC를 찾을 수 없습니다.")
    if not _can_comment_on_voc(v, db, user_id):
        raise HTTPException(status_code=403, detail="이 VOC에 댓글을 작성할 권한이 없습니다.")

    content = (payload.content or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="내용을 입력해 주세요.")
    content = _strip_dangerous_html(content)

    is_admin = _is_voc_admin(db, user_id)
    comment = VocComment(
        voc_id=voc_id,
        author_id=user_id,
        content=content,
        is_admin_reply=is_admin,
    )
    db.add(comment)
    db.commit()
    db.refresh(comment)
    emit_realtime_event(
        db, space_id=REALTIME_GLOBAL_CHANNEL, event_type="voc_comment_created",
        entity_type="voc_comment", entity_id=comment.id, actor_user_id=user_id,
        payload={"voc_id": voc_id},
    )

    # 추가 문의/댓글 저장은 메일 알림을 발생시키지 않는다.
    # (작성자에게 메일로 알리는 경로는 "관리자 메모 저장" 시 opt-in 으로만 존재 → update_voc 참고)
    db.commit()
    users_map = {caller.id: caller}
    return _voc_comment_dict(comment, users_map)


@app.delete("/api/voc/{voc_id}/comments/{comment_id}")
def delete_voc_comment(voc_id: int, comment_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    c = (
        db.query(VocComment)
        .filter(VocComment.id == comment_id, VocComment.voc_id == voc_id)
        .first()
    )
    if not c:
        raise HTTPException(status_code=404, detail="댓글을 찾을 수 없습니다.")
    if c.author_id != user_id and not _is_voc_admin(db, user_id):
        raise HTTPException(status_code=403, detail="이 댓글을 삭제할 권한이 없습니다.")
    db.delete(c)
    db.commit()
    return {"message": "삭제되었습니다."}


# =========================
# VOC Attachments (이미지)
# =========================
def _voc_storage_dir(voc_id: int) -> str:
    return os.path.join(UPLOAD_DIR, "voc", str(voc_id))


def _delete_voc_attachment_storage(att: VocAttachment) -> None:
    """S3 + 로컬 파일 정리 (best-effort). 어떤 경우에도 호출자에게 예외가 전파되지 않게 한다.
    파일 삭제 실패가 VOC 삭제 자체를 막으면 안 되므로, 모든 단계를 개별 try 로 감싼다.
    """
    if not att.stored_name:
        return
    import logging as _logging
    _log = _logging.getLogger("main")

    # 1) S3 정리 — 어떤 예외도 밖으로 새지 않게
    try:
        from app.utils.s3.s3_utils import delete_from_s3, is_s3_configured, get_attachment_s3_key
        if is_s3_configured():
            s3_key = att.s3_key or get_attachment_s3_key(
                att.filename or att.stored_name, att.stored_name, "voc", int(att.voc_id)
            )
            if s3_key:
                try:
                    result = delete_from_s3(s3_key)
                    if isinstance(result, dict) and not result.get("success", True):
                        _log.warning(
                            f"VOC 첨부 S3 삭제 실패 (무시): {s3_key} - {result.get('error')}"
                        )
                except Exception as e:
                    _log.warning(f"VOC 첨부 S3 삭제 예외 (무시): {s3_key} - {e}")
    except Exception as e:
        _log.warning(f"VOC 첨부 S3 처리 예외 (무시): {e}")

    # 2) 로컬 정리 — 항상 시도, 실패해도 무시
    try:
        local_path = os.path.join(_voc_storage_dir(att.voc_id), att.stored_name)
        if os.path.exists(local_path):
            os.remove(local_path)
    except Exception as e:
        _log.warning(f"VOC 첨부 로컬 삭제 실패 (무시): {e}")


def _optimize_voc_image(raw: bytes, original_filename: str) -> tuple[bytes, str, str, int, int]:
    """이미지를 긴 변 기준 리사이즈 후 webp 로 변환. 실패 시 원본 그대로 반환.
    반환: (data, stored_ext, content_type, width, height)
    """
    try:
        from PIL import Image
        import io as _io
        img = Image.open(_io.BytesIO(raw))
        img.load()
        # 알파 보존: RGBA 는 유지, P/LA 도 RGBA 로, 그 외 대형 모드는 RGB 로
        if img.mode in ("P", "LA"):
            img = img.convert("RGBA")
        elif img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGB")

        w, h = img.size
        long_edge = max(w, h)
        if long_edge > VOC_ATTACHMENT_MAX_EDGE:
            ratio = VOC_ATTACHMENT_MAX_EDGE / float(long_edge)
            new_size = (max(1, int(w * ratio)), max(1, int(h * ratio)))
            try:
                resample = Image.Resampling.LANCZOS
            except AttributeError:
                resample = Image.LANCZOS  # type: ignore[attr-defined]
            img = img.resize(new_size, resample)
            w, h = img.size

        buf = _io.BytesIO()
        img.save(buf, format="WEBP", quality=VOC_ATTACHMENT_WEBP_QUALITY, method=6)
        return buf.getvalue(), ".webp", "image/webp", w, h
    except Exception as e:
        import logging as _logging
        _logging.getLogger("main").warning(
            f"VOC 이미지 최적화 실패 → 원본 저장: {original_filename}: {e}"
        )
        ext = (os.path.splitext(original_filename)[1] or ".bin").lower()
        mime = _IMAGE_EXT_MIME.get(ext, "application/octet-stream")
        return raw, ext, mime, 0, 0


def _store_voc_image_bytes(voc_id: int, raw: bytes, original_filename: str) -> dict:
    """원본 바이트를 최적화(리사이즈+webp)한 뒤 S3(우선)/로컬에 저장하고
    VocAttachment 생성에 필요한 메타를 돌려준다. (본문/댓글 이미지 공통 저장 경로)"""
    from app.utils.s3.s3_utils import upload_attachment_bytes_to_s3, is_s3_configured
    import logging as _logging
    _log = _logging.getLogger("main")

    optimized, stored_ext, content_type, w, h = _optimize_voc_image(raw, original_filename)
    stored_name = f"{uuid.uuid4().hex}{stored_ext}"

    s3_key = ""
    saved = False
    if is_s3_configured():
        try:
            s3_result = upload_attachment_bytes_to_s3(
                data=optimized,
                original_filename=original_filename,
                stored_name=stored_name,
                context_type="voc",
                context_id=voc_id,
            )
        except Exception as s3e:
            _log.warning(f"VOC S3 업로드 호출 예외: {s3e}")
            s3_result = {"success": False, "s3_key": "", "error": str(s3e)}
        if s3_result.get("success"):
            s3_key = s3_result["s3_key"]
            saved = True
        else:
            _log.warning(f"VOC S3 업로드 실패, 로컬 저장으로 전환: {s3_result.get('error')}")

    if not saved:
        voc_dir = _voc_storage_dir(voc_id)
        os.makedirs(voc_dir, exist_ok=True)
        with open(os.path.join(voc_dir, stored_name), "wb") as f:
            f.write(optimized)

    return {
        "stored_name": stored_name,
        "content_type": content_type,
        "file_size": len(optimized),
        "width": w or None,
        "height": h or None,
        "s3_key": s3_key or None,
    }


@app.get("/api/voc/{voc_id}/attachments")
def list_voc_attachments(voc_id: int, user_id: int = Query(...), db: Session = Depends(get_db)):
    v = db.query(VocItem).filter(VocItem.id == voc_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="VOC를 찾을 수 없습니다.")
    if v.author_id != user_id and not _is_voc_admin(db, user_id):
        raise HTTPException(status_code=403, detail="조회 권한이 없습니다.")
    rows = (
        db.query(VocAttachment)
        .filter(VocAttachment.voc_id == voc_id)
        .filter(VocAttachment.comment_scoped.isnot(True))  # 댓글 이미지는 본문 첨부 목록에서 제외
        .order_by(VocAttachment.created_at.asc())
        .all()
    )
    return {"attachments": [_voc_attachment_dict(a) for a in rows]}


@app.post("/api/voc/{voc_id}/attachments")
async def upload_voc_attachment(
    voc_id: int,
    file: UploadFile = FastAPIFile(...),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """개선요청 첨부 이미지 업로드. 작성자 또는 관리자만 가능.
    - 최대 5장 / 1장당 10MB / 이미지 확장자만 허용
    - 긴 변 1920px 로 리사이즈 후 webp 로 변환 저장 (원본 보관 X)
    - S3 우선, 실패 시 로컬 fallback
    """
    v = db.query(VocItem).filter(VocItem.id == voc_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="VOC를 찾을 수 없습니다.")
    if v.author_id != user_id and not _is_voc_admin(db, user_id):
        raise HTTPException(status_code=403, detail="이 VOC에 첨부할 권한이 없습니다.")

    from app.utils.upload_policy import VOC_IMAGE_POLICY, UploadValidationError

    original_filename = file.filename or "image"

    # 공통 이미지 정책 1차 검증: 확장자 (읽기 전)
    try:
        VOC_IMAGE_POLICY.validate_type(original_filename, getattr(file, "content_type", "") or "")
    except UploadValidationError as ve:
        _log_upload_rejected("voc", voc_id, user_id, original_filename, str(ve))
        raise HTTPException(status_code=400, detail=str(ve))

    # 본문 첨부 장수 제한은 본문(comment_scoped=False) 이미지만 카운트
    existing_count = (
        db.query(VocAttachment)
        .filter(VocAttachment.voc_id == voc_id, VocAttachment.comment_scoped.isnot(True))
        .count()
    )

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    # 공통 이미지 정책 2차 검증: 파일 크기 + 장수 (최적화/S3 전)
    try:
        VOC_IMAGE_POLICY.validate(
            filename=original_filename,
            size_bytes=len(raw),
            content_type=getattr(file, "content_type", "") or "",
            existing_count=existing_count,
        )
    except UploadValidationError as ve:
        _log_upload_rejected("voc", voc_id, user_id, original_filename, str(ve))
        raise HTTPException(status_code=400, detail=str(ve))

    meta = _store_voc_image_bytes(voc_id, raw, original_filename)

    att = VocAttachment(
        voc_id=voc_id,
        filename=original_filename[:255],
        comment_scoped=False,
        **meta,
    )
    db.add(att)

    # 첨부 추가도 사용자 인지 가능한 변경 → edit 메타 갱신
    v.edited_by = user_id
    v.edited_at = datetime.now()
    v.change_summary = "attachments"

    db.commit()
    db.refresh(att)
    return _voc_attachment_dict(att)


@app.post("/api/voc/{voc_id}/comment-images")
async def upload_voc_comment_image(
    voc_id: int,
    file: UploadFile = FastAPIFile(...),
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """VOC 댓글/추가 문의에 붙이는 이미지 업로드.
    - 권한은 댓글 작성 권한과 동일(작성자 또는 관리자). 본문 첨부와 동일하게 webp 최적화.
    - comment_scoped=True 로 저장해 본문 첨부 목록/수정 diff/edit 메타와 분리한다.
      (본문 편집 시 댓글 이미지가 삭제되거나 VOC 가 '수정됨'으로 표시되지 않도록)
    - 본문 5장 제한과 별개. 장수 제한은 프론트 에디터(댓글당 최대 5장)에서 통제한다.
    """
    v = db.query(VocItem).filter(VocItem.id == voc_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="VOC를 찾을 수 없습니다.")
    if not _can_comment_on_voc(v, db, user_id):
        raise HTTPException(status_code=403, detail="이 VOC에 댓글 이미지를 올릴 권한이 없습니다.")

    from app.utils.upload_policy import VOC_IMAGE_POLICY, UploadValidationError

    original_filename = file.filename or "image"

    try:
        VOC_IMAGE_POLICY.validate_type(original_filename, getattr(file, "content_type", "") or "")
    except UploadValidationError as ve:
        _log_upload_rejected("voc-comment", voc_id, user_id, original_filename, str(ve))
        raise HTTPException(status_code=400, detail=str(ve))

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    # 크기/확장자 검증 (장수 제한은 본문과 분리 — 댓글당 제한은 프론트에서 통제)
    try:
        VOC_IMAGE_POLICY.validate(
            filename=original_filename,
            size_bytes=len(raw),
            content_type=getattr(file, "content_type", "") or "",
        )
    except UploadValidationError as ve:
        _log_upload_rejected("voc-comment", voc_id, user_id, original_filename, str(ve))
        raise HTTPException(status_code=400, detail=str(ve))

    meta = _store_voc_image_bytes(voc_id, raw, original_filename)

    att = VocAttachment(
        voc_id=voc_id,
        filename=original_filename[:255],
        comment_scoped=True,
        **meta,
    )
    db.add(att)
    db.commit()
    db.refresh(att)
    return _voc_attachment_dict(att)


@app.get("/api/voc/{voc_id}/attachments/{stored_name}/download")
def download_voc_attachment(voc_id: int, stored_name: str, db: Session = Depends(get_db)):
    """이미지는 inline + image/* 로 응답. 인증은 별도 적용하지 않음(현재 task 첨부와 동일 정책)."""
    from fastapi.responses import Response
    from app.utils.s3.s3_utils import download_from_s3, is_s3_configured

    att = (
        db.query(VocAttachment)
        .filter(VocAttachment.voc_id == voc_id, VocAttachment.stored_name == stored_name)
        .first()
    )
    if not att:
        raise HTTPException(status_code=404, detail="File not found")

    filename = att.filename or stored_name
    media_type = att.content_type or _IMAGE_EXT_MIME.get(
        os.path.splitext(stored_name)[1].lower(), "application/octet-stream"
    )
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename)
    content_disposition = (
        f"inline; filename*=UTF-8''{encoded_filename}"
        if media_type.startswith("image/")
        else f"attachment; filename*=UTF-8''{encoded_filename}"
    )

    if att.s3_key and is_s3_configured():
        data = download_from_s3(att.s3_key)
        if data is not None:
            return Response(
                content=data,
                media_type=media_type,
                headers={
                    "Content-Disposition": content_disposition,
                    "Content-Length": str(len(data)),
                    "Cache-Control": "private, max-age=86400",
                },
            )

    file_path = os.path.join(_voc_storage_dir(voc_id), stored_name)
    if os.path.exists(file_path):
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type=media_type,
            content_disposition_type="inline" if media_type.startswith("image/") else "attachment",
        )

    raise HTTPException(status_code=404, detail="File not found")


@app.delete("/api/voc/{voc_id}/attachments/{attachment_id}")
def delete_voc_attachment(
    voc_id: int,
    attachment_id: int,
    user_id: int = Query(...),
    db: Session = Depends(get_db),
):
    att = (
        db.query(VocAttachment)
        .filter(VocAttachment.id == attachment_id, VocAttachment.voc_id == voc_id)
        .first()
    )
    if not att:
        raise HTTPException(status_code=404, detail="첨부를 찾을 수 없습니다.")
    v = db.query(VocItem).filter(VocItem.id == voc_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="VOC를 찾을 수 없습니다.")
    if v.author_id != user_id and not _is_voc_admin(db, user_id):
        raise HTTPException(status_code=403, detail="이 첨부를 삭제할 권한이 없습니다.")

    _delete_voc_attachment_storage(att)
    db.delete(att)
    v.edited_by = user_id
    v.edited_at = datetime.now()
    v.change_summary = "attachments"
    db.commit()
    return {"message": "삭제되었습니다."}

# =========================
# Run
# =========================
if __name__ == "__main__":
    import uvicorn

    # reload는 개발(development)에서만 켠다. 운영에서는 코드 변경 감시 프로세스가
    # 불필요한 부하/재시작을 유발하므로 ENV_MODE 기준으로 자동 비활성화한다.
    # (UVICORN_RELOAD 로 강제 override 가능)
    _env_mode = os.getenv("ENV_MODE", "development").lower()
    _reload_override = os.getenv("UVICORN_RELOAD")
    if _reload_override is not None:
        _reload = _reload_override.strip().lower() in ("1", "true", "yes", "on")
    else:
        _reload = _env_mode == "development"

    # ✅ 네 기존 사내 포트 유지(8085)
    uvicorn.run("main:app", host="0.0.0.0", port=8085, reload=_reload)